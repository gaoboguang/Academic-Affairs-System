from __future__ import annotations

from openpyxl import Workbook

from app.core.config import Settings
from app.utils.parsers import make_timestamped_filename, relative_to_project


def _append_analysis_insight_sheet(
    workbook: Workbook,
    rows: list[dict[str, object]],
) -> None:
    insight_sheet = workbook.create_sheet("摘要概览")
    insight_sheet.append(["标题", "摘要", "说明", "提示级别"])
    for row in rows:
        insight_sheet.append([
            row.get("title"),
            row.get("summary"),
            row.get("detail"),
            row.get("tone"),
        ])


def _growth_type_label(value: object) -> str:
    mapping = {
        "reward": "奖励记录",
        "discipline": "处分记录",
        "activity": "活动记录",
        "cadre": "干部任职",
        "interview": "谈话记录",
        "home_school": "家校沟通",
        "mental_health": "心理关注",
        "quality_eval": "综合素质评价",
        "other": "其他",
    }
    return mapping.get(str(value or ""), str(value or "-"))


def _build_growth_insight_rows(
    student_meta: dict[str, object],
    records: list[dict[str, object]],
) -> list[dict[str, object]]:
    attachment_count = sum(len(item.get("attachments") or []) for item in records)
    record_type_counts: dict[str, int] = {}
    for item in records:
        record_type = str(item.get("record_type") or "")
        record_type_counts.setdefault(record_type, 0)
        record_type_counts[record_type] += 1
    dominant_type = sorted(record_type_counts.items(), key=lambda item: item[1], reverse=True)[0] if record_type_counts else None
    latest_record = sorted(records, key=lambda item: str(item.get("occurred_on") or ""), reverse=True)[0] if records else None

    rows: list[dict[str, object]] = [
        {
            "title": "成长档案整体状态",
            "summary": f"{student_meta.get('student_name') or '该学生'} 当前共有 {len(records)} 条成长记录",
            "detail": f"成长类型 {len(record_type_counts)} 类，附件 {attachment_count} 个。导出前建议先确认是否需要补齐最近阶段的记录。",
            "tone": "info",
        }
    ]

    if dominant_type:
        rows.append(
            {
                "title": "高频成长类型",
                "summary": f"{_growth_type_label(dominant_type[0])} 当前记录最多",
                "detail": f"共 {dominant_type[1]} 条，适合在导出汇报时作为该学生当前成长档案的主要线索。",
                "tone": "info",
            }
        )

    if latest_record:
        latest_detail = _growth_type_label(latest_record.get("record_type"))
        if latest_record.get("owner_name"):
            latest_detail += f" / 责任人 {latest_record.get('owner_name')}"
        latest_detail += "。"
        rows.append(
            {
                "title": "最近一条记录",
                "summary": f"{latest_record.get('occurred_on')} · {latest_record.get('title') or '-'}",
                "detail": latest_detail,
                "tone": "success",
            }
        )

    if not records:
        rows.append(
            {
                "title": "当前暂无成长记录",
                "summary": "导出结果会以空档案形式呈现",
                "detail": "若准备正式汇报，建议先补录关键成长事件、谈话记录或奖励处分信息。",
                "tone": "warning",
            }
        )

    return rows


def export_growth_summary(
    settings: Settings,
    student_meta: dict[str, object],
    records: list[dict[str, object]],
) -> str:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "学生概况"
    summary_sheet.append(["学号", student_meta.get("student_no")])
    summary_sheet.append(["姓名", student_meta.get("student_name")])
    summary_sheet.append(["当前年级", student_meta.get("grade_name")])
    summary_sheet.append(["当前班级", student_meta.get("class_name")])
    summary_sheet.append(["记录数量", len(records)])

    detail_sheet = workbook.create_sheet("成长记录")
    detail_sheet.append(["日期", "类型", "标题", "内容", "责任人", "备注", "附件数"])
    for record in records:
        detail_sheet.append(
            [
                record.get("occurred_on"),
                _growth_type_label(record.get("record_type")),
                record.get("title"),
                record.get("content"),
                record.get("owner_name"),
                record.get("note"),
                len(record.get("attachments", [])),
            ]
        )

    _append_analysis_insight_sheet(workbook, _build_growth_insight_rows(student_meta, records))

    filename = make_timestamped_filename("growth_summary_report", ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)
