from __future__ import annotations

from openpyxl import Workbook

from app.core.config import Settings
from app.importers.pathway_profiles import (
    PATHWAY_PROFILE_TEMPLATE_HEADERS,
    PATHWAY_PROFILE_SAMPLE_ROW,
    format_candidate_type,
    format_exam_type,
    format_pathway_profile_bool,
)
from app.utils.parsers import make_timestamped_filename, relative_to_project


def build_pathway_profile_template() -> Workbook:
    workbook = Workbook()
    intro_sheet = workbook.active
    intro_sheet.title = "说明"
    intro_sheet.append(["升学画像导入模板"])
    intro_sheet.append(["学号为必填列，姓名用于辅助核对。"])
    intro_sheet.append(["空白单元格导入时会保留系统已有值，不会清空原画像。"])
    intro_sheet.append(["布尔字段支持 是/否、true/false、1/0。"])

    data_sheet = workbook.create_sheet("数据")
    data_sheet.append(PATHWAY_PROFILE_TEMPLATE_HEADERS)
    data_sheet.append(PATHWAY_PROFILE_SAMPLE_ROW)
    return workbook


def export_pathway_profile_template(settings: Settings) -> str:
    workbook = build_pathway_profile_template()
    path = settings.templates_dir / "student_pathway_profiles_import_template.xlsx"
    workbook.save(path)
    return relative_to_project(path, settings.project_root)


def export_pathway_profiles(settings: Settings, rows: list[dict[str, object]]) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(PATHWAY_PROFILE_TEMPLATE_HEADERS)
    for row in rows:
        body_limitations = row.get("known_body_limitations_json")
        body_note = body_limitations.get("note") if isinstance(body_limitations, dict) else None
        region_pref = row.get("region_preferences_json") if isinstance(row.get("region_preferences_json"), dict) else {}
        career_pref = row.get("career_preferences_json") if isinstance(row.get("career_preferences_json"), dict) else {}

        def _list_cell(value: object) -> str:
            return ",".join(value) if isinstance(value, list) else ""

        def _bool_cell(value: object) -> str | None:
            return format_pathway_profile_bool(value) if isinstance(value, bool) else None

        sheet.append(
            [
                row.get("student_no"),
                row.get("name"),
                row.get("class_name"),
                row.get("province"),
                format_candidate_type(row.get("candidate_type") if isinstance(row.get("candidate_type"), str) else None),
                format_exam_type(row.get("exam_type") if isinstance(row.get("exam_type"), str) else None),
                row.get("subject_combination"),
                row.get("spring_exam_category"),
                row.get("art_track"),
                row.get("art_professional_score"),
                row.get("art_professional_full_score"),
                row.get("art_score_source"),
                row.get("art_score_note"),
                row.get("sports_track"),
                _bool_cell(row.get("has_gaokao_registration")),
                _bool_cell(row.get("is_fresh_graduate")),
                _bool_cell(row.get("is_vocational_student")),
                _bool_cell(row.get("is_social_candidate")),
                _bool_cell(row.get("has_high_school_equivalent")),
                _bool_cell(row.get("accept_junior_college")),
                _bool_cell(row.get("accept_private_college")),
                _bool_cell(row.get("accept_sino_foreign")),
                _bool_cell(row.get("accept_outside_province")),
                _bool_cell(row.get("accept_early_batch")),
                _bool_cell(row.get("accept_service_commitment")),
                _bool_cell(row.get("accept_interview_or_physical_test")),
                _list_cell(region_pref.get("target_regions")),
                _list_cell(region_pref.get("school_level_tags")),
                region_pref.get("major_keyword") if isinstance(region_pref.get("major_keyword"), str) else "",
                row.get("primary_direction_name") or "",
                row.get("secondary_direction_name") or "",
                row.get("alternative_direction_name") or "",
                _list_cell(career_pref.get("priority_focuses")),
                _list_cell(career_pref.get("preferred_industries")),
                _list_cell(career_pref.get("preferred_job_types")),
                _list_cell(career_pref.get("target_employment_cities")),
                _bool_cell(career_pref.get("accepts_postgraduate")),
                _bool_cell(career_pref.get("accepts_public_service")),
                _bool_cell(career_pref.get("accepts_certificate")),
                _bool_cell(career_pref.get("accepts_long_training")),
                body_note,
                row.get("note"),
            ]
        )

    filename = make_timestamped_filename("student_pathway_profiles_export", ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)
