from __future__ import annotations

from openpyxl import Workbook

from app.core.config import Settings
from app.exporters._recommendation_export_formatters import (
    RECOMMENDATION_RISK_FLAG_LABELS,
    SHANDONG_2026_DATA_NOTICE,
    format_export_value as _format_export_value,
    format_percent_value as _format_percent_value,
    format_shandong_bucket as _format_shandong_bucket,
    format_shandong_confidence as _format_shandong_confidence,
    format_shandong_rank_basis as _format_shandong_rank_basis,
    format_shandong_rank_range as _format_shandong_rank_range,
    format_shandong_risk_preference as _format_shandong_risk_preference,
    format_shandong_source_mode as _format_shandong_source_mode,
    format_shandong_student_type as _format_shandong_student_type,
    format_source_document_ids as _format_source_document_ids,
    format_years as _format_years,
    read_nested_number as _read_nested_number,
    read_snapshot_value as _read_snapshot_value,
)
from app.exporters._recommendation_export_risks import (
    _build_recommendation_risk_group_label,
    _build_recommendation_risk_rows,
)
from app.exporters._recommendation_export_volunteer import (
    _build_candidate_boundary_notes,
    _build_candidate_reference_copy,
    _build_candidate_rule_copy,
    _build_volunteer_draft_boundary_rows,
    _build_volunteer_draft_overview_rows,
    _build_volunteer_draft_rule_rows,
)
from app.utils.parsers import make_timestamped_filename, relative_to_project


def export_recommendation_summary(settings: Settings, meta: dict[str, object], rows: list[dict[str, object]]) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "推荐概况"
    summary.append(["推荐方案", meta.get("scheme_name")])
    summary.append(["学生", meta.get("student_name")])
    summary.append(["考试", meta.get("exam_name")])
    summary.append(["省份", meta.get("province")])
    summary.append(["目标年份", meta.get("target_year")])
    summary.append(["分数模式", meta.get("score_input_label")])
    summary.append(["模拟说明", meta.get("simulation_note")])
    summary.append(["目标方向", meta.get("target_direction_summary") or "未维护"])
    summary.append(["可接受路径", meta.get("accepted_path_summary") or "未维护"])
    summary.append(["结果数量", len(rows)])

    risk_rows = _build_recommendation_risk_rows(meta, rows)

    risk_overview = workbook.create_sheet("风险概览")
    risk_overview.append(["分组", "类型", "摘要", "说明"])
    for item in risk_rows:
        risk_overview.append([
            _build_recommendation_risk_group_label(item["title"]),
            item["title"],
            item["summary"],
            item["detail"],
        ])

    overview_sheet = workbook.create_sheet("导出前摘要")
    overview_sheet.append(["分组", "标题", "摘要", "说明"])
    for item in risk_rows:
        overview_sheet.append([
            _build_recommendation_risk_group_label(item["title"]),
            item["title"],
            item["summary"],
            item["detail"],
        ])

    detail = workbook.create_sheet("推荐结果")
    detail.append(
        [
            "分组",
            "院校",
            "专业",
            "参考位次",
            "学生位次",
            "比值",
            "依据",
            "录取结果年份",
            "计划年份",
            "批次",
            "选科要求",
            "职业匹配",
            "对应目标方向",
            "路径提示",
            "职业说明",
            "理由",
            "风险提示",
            "章程链接",
            "章程状态",
            "章程限制状态",
            "校区备注",
            "章程备注",
        ]
    )
    for row in rows:
        detail.append(
            [
                row.get("result_type"),
                row.get("college_name"),
                row.get("major_name"),
                row.get("reference_rank"),
                row.get("student_rank"),
                row.get("ratio"),
                row.get("score_basis"),
                _format_years(row.get("reference_years_json") or _read_snapshot_value(row, "reference_years")),
                _read_snapshot_value(row, "plan_year") or _read_snapshot_value(row, "target_year"),
                _read_snapshot_value(row, "batch") or _read_snapshot_value(row, "matched_rule_batch"),
                _read_snapshot_value(row, "subject_requirement") or _read_snapshot_value(row, "required_subjects"),
                row.get("career_match_strength"),
                " / ".join(row.get("matched_direction_names_json") or []),
                _build_path_hint(row),
                row.get("career_match_summary"),
                row.get("reason_text"),
                _format_risk_flags(row.get("risk_flags_json")),
                _read_snapshot_value(row, "chapter_url"),
                _read_snapshot_value(row, "chapter_review_status"),
                _build_chapter_restriction_export_copy(row),
                _read_snapshot_value(row, "chapter_campus_note"),
                _read_snapshot_value(row, "chapter_other_risk_note"),
            ]
        )

    filename = make_timestamped_filename("recommendation_summary_report", ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)


def export_volunteer_draft_summary(settings: Settings, meta: dict[str, object], rows: list[dict[str, object]]) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "志愿草稿概况"
    summary.append(["草稿名称", meta.get("draft_name")])
    summary.append(["学生", meta.get("student_name")])
    summary.append(["考试", meta.get("exam_name")])
    summary.append(["省份", meta.get("province")])
    summary.append(["目标年份", meta.get("target_year")])
    summary.append(["批次", meta.get("batch")])
    summary.append(["科类/模式", meta.get("exam_mode")])
    summary.append(["分数模式", meta.get("score_input_label")])
    summary.append(["模拟说明", meta.get("simulation_note")])
    summary.append(["目标方向", meta.get("target_direction_summary") or "未维护"])
    summary.append(["可接受路径", meta.get("accepted_path_summary") or "未维护"])
    summary.append(["规则", meta.get("rule_label")])
    summary.append(["志愿数量", len(rows)])
    summary.append(["备注", meta.get("note")])

    boundary = workbook.create_sheet("边界概览")
    boundary.append(["类型", "摘要", "说明"])
    for item in _build_volunteer_draft_boundary_rows(rows, meta.get("rule_alerts") or []):
        boundary.append([item["title"], item["summary"], item["detail"]])

    rule_sheet = workbook.create_sheet("规则差异摘要")
    rule_sheet.append(["规则", "定位", "总分口径", "志愿上限", "志愿单位", "平行方式", "选科口径", "每单位专业数", "征集志愿", "调剂", "说明"])
    for item in _build_volunteer_draft_rule_rows(meta):
        rule_sheet.append(
            [
                item["title"],
                item["position"],
                item["total_score"],
                item["volunteer_limit"],
                item["volunteer_unit_type"],
                item["parallel_rule_mode"],
                item["subject_requirement_mode"],
                item["max_major_per_unit"],
                item["support_collect_round"],
                item["allow_adjustment"],
                item["notes"],
            ]
        )

    overview_sheet = workbook.create_sheet("导出前摘要")
    overview_sheet.append(["分组", "标题", "摘要", "说明"])
    for item in _build_volunteer_draft_overview_rows(meta, rows):
        overview_sheet.append([
            item["group"],
            item["title"],
            item["summary"],
            item["detail"],
        ])

    detail = workbook.create_sheet("志愿草稿")
    detail.append(
        [
            "顺序",
            "分层",
            "院校",
            "院校代码",
            "专业",
            "专业代码",
            "专业组",
            "计划数",
            "批次",
            "科类/模式",
            "参考位次",
            "最近最低位次",
            "最近最低分",
            "依据",
            "命中规则",
            "录取口径",
            "边界说明",
            "职业匹配",
            "对应目标方向",
            "路径提示",
            "职业说明",
            "理由",
            "风险提示",
            "章程链接",
            "章程状态",
            "校区备注",
            "章程备注",
        ]
    )
    for row in rows:
        detail.append(
            [
                row.get("order"),
                row.get("result_type"),
                row.get("college_name"),
                row.get("college_code_snapshot"),
                row.get("major_name"),
                row.get("major_code_snapshot"),
                row.get("major_group_code"),
                row.get("plan_count"),
                row.get("batch"),
                row.get("exam_mode"),
                row.get("reference_rank"),
                row.get("latest_min_rank"),
                row.get("latest_min_score"),
                row.get("score_basis"),
                _build_candidate_rule_copy(row),
                _build_candidate_reference_copy(row),
                "；".join(_build_candidate_boundary_notes(row, meta.get("rule_alerts") or [])),
                row.get("career_match_strength"),
                " / ".join(row.get("matched_direction_names_json") or []),
                _build_path_hint(row),
                row.get("career_match_summary"),
                row.get("reason_text"),
                _format_risk_flags(row.get("risk_flags_json")),
                row.get("chapter_url"),
                row.get("chapter_review_status"),
                row.get("chapter_campus_note"),
                row.get("chapter_other_risk_note"),
            ]
        )

    filename = make_timestamped_filename("volunteer_draft_summary_report", ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)


def export_shandong_recommendation_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总页"

    grouped = _group_shandong_candidates(payload)
    all_candidates = [item for key in ("rush", "stable", "safe", "watch") for item in grouped[key]]
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}

    summary_sheet.append(["报告", "山东普通类冲稳保推荐报告"])
    summary_sheet.append(["学生", payload.get("student_name") or "未命名学生"])
    summary_sheet.append(["生源地", payload.get("province") or "山东"])
    summary_sheet.append(["目标年份", payload.get("target_year")])
    summary_sheet.append(["考生类型", _format_shandong_student_type(payload.get("student_type"))])
    summary_sheet.append(["输入来源", _format_shandong_source_mode(payload.get("source_mode"))])
    summary_sheet.append(["预估分数", _format_export_value(payload.get("predicted_score"))])
    summary_sheet.append(["预估位次", _format_export_value(payload.get("predicted_rank"))])
    summary_sheet.append(["位次区间", _format_shandong_rank_range(payload)])
    summary_sheet.append(["位次依据", _format_shandong_rank_basis(payload.get("rank_projection_basis"))])
    summary_sheet.append(["风险偏好", _format_shandong_risk_preference(payload.get("risk_preference"))])
    summary_sheet.append(["数据年份", _format_years(payload.get("data_years"))])
    summary_sheet.append(["冲 / 稳 / 保 / 仅关注", f"{summary.get('rush_count', 0)} / {summary.get('stable_count', 0)} / {summary.get('safe_count', 0)} / {summary.get('watch_count', 0)}"])
    summary_sheet.append(["选科不符已排除", summary.get("excluded_subject_mismatch_count", 0)])
    summary_sheet.append(["输入提示", "；".join(str(item) for item in payload.get("input_notes") or [] if item) or "-"])
    summary_sheet.append(["2026 数据提示", SHANDONG_2026_DATA_NOTICE])

    risk_sheet = workbook.create_sheet("风险说明")
    risk_sheet.append(["风险", "影响范围", "说明"])
    for row in _build_shandong_risk_overview_rows(payload, all_candidates):
        risk_sheet.append([row["title"], row["summary"], row["detail"]])

    _append_shandong_candidate_sheet(workbook, "冲列表", grouped["rush"])
    _append_shandong_candidate_sheet(workbook, "稳列表", grouped["stable"])
    _append_shandong_candidate_sheet(workbook, "保列表", grouped["safe"])
    _append_shandong_candidate_sheet(workbook, "数据不足与风险列表", grouped["watch"])
    _append_shandong_source_sheet(workbook, payload, all_candidates)

    filename = make_timestamped_filename("shandong_recommendation_report", ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)


def export_gaokao_pathway_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总页"

    cards = _as_dict_list(payload.get("cards"))
    profile_summary = _as_dict_list(payload.get("profile_summary"))
    material_gaps = _as_dict_list(payload.get("material_gaps"))
    next_actions = _as_dict_list(payload.get("next_actions"))
    publication_status = _as_dict_list(payload.get("publication_status"))
    p0_gaps = [str(item) for item in payload.get("p0_gaps") or [] if item]

    summary_sheet.append(["报告", "山东升学路径规划报告"])
    summary_sheet.append(["学生", payload.get("student_name") or "未命名学生"])
    summary_sheet.append(["目标年份", payload.get("target_year")])
    summary_sheet.append(["生成时间", payload.get("generated_at")])
    summary_sheet.append(["路径卡数量", len(cards)])
    summary_sheet.append(["材料缺口数量", len(material_gaps)])
    summary_sheet.append(["P0 数据缺口数量", len(p0_gaps)])
    summary_sheet.append(["数据健康状态", payload.get("data_health_summary") or "未读取数据健康状态"])
    summary_sheet.append(["边界说明", "普通类常规批可进入位次型推荐；其他路径只做资格初筛、政策提醒和人工复核，不输出录取概率。"])

    profile_sheet = workbook.create_sheet("学生画像")
    profile_sheet.append(["字段", "当前值", "是否已维护"])
    for item in profile_summary:
        profile_sheet.append([
            item.get("label"),
            item.get("value"),
            "是" if item.get("filled") else "否",
        ])

    pathway_sheet = workbook.create_sheet("路径建议")
    pathway_sheet.append([
        "路径",
        "分组",
        "状态",
        "推荐深度",
        "置信程度",
        "评估分",
        "适用对象",
        "志愿方式",
        "摘要",
        "关键要求",
        "缺失材料",
        "风险提示",
        "下一步",
        "来源编号",
    ])
    for row in cards:
        pathway_sheet.append([
            row.get("name"),
            row.get("group"),
            row.get("status_label"),
            row.get("depth_label"),
            row.get("confidence_label"),
            _format_export_value(row.get("score")),
            row.get("applicable_object"),
            row.get("volunteer_mode"),
            row.get("summary"),
            _join_export_list(row.get("key_requirements")),
            _join_export_list(row.get("missing_materials")),
            _join_export_list(row.get("risk_messages")),
            _join_export_list(row.get("next_actions")),
            _format_export_value(row.get("source_document_id")),
        ])

    gap_sheet = workbook.create_sheet("材料缺口")
    gap_sheet.append(["材料或画像字段", "影响路径数", "影响路径", "下一步"])
    for row in material_gaps:
        gap_sheet.append([
            row.get("label"),
            row.get("count"),
            _join_export_list(row.get("pathways")),
            row.get("nextAction") or row.get("next_action"),
        ])

    action_sheet = workbook.create_sheet("下一步行动")
    action_sheet.append(["标题", "说明", "优先级"])
    for row in next_actions:
        action_sheet.append([row.get("title"), row.get("detail"), row.get("tone")])

    data_sheet = workbook.create_sheet("数据风险")
    data_sheet.append(["数据项", "状态", "下一步", "说明", "是否阻断推荐"])
    for row in publication_status:
        data_sheet.append([
            row.get("label"),
            row.get("status_label"),
            row.get("action_label"),
            row.get("explanation"),
            "是" if row.get("blocks_recommendation") else "否",
        ])
    for gap in p0_gaps:
        data_sheet.append(["P0 数据缺口", "需关注", "正式填报前补齐或人工复核", gap, "否"])

    filename = make_timestamped_filename("gaokao_pathway_report", ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)


def _build_path_hint(row: dict[str, object]) -> str:
    hints: list[str] = []
    if row.get("requires_postgraduate_path") is True:
        hints.append("读研")
    if row.get("requires_certificate_path") is True:
        hints.append("资格证")
    if row.get("requires_long_training_path") is True:
        hints.append("长培养周期")
    return " / ".join(hints)


def _as_dict_list(value: object) -> list[dict[str, object]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _join_export_list(value: object) -> str:
    if not isinstance(value, list):
        return "-"
    items = [str(item) for item in value if str(item or "").strip()]
    return "；".join(items) if items else "-"


def _format_risk_flags(flags: object) -> str:
    if not isinstance(flags, list) or not flags:
        return "-"
    labels = [
        RECOMMENDATION_RISK_FLAG_LABELS.get(str(flag), str(flag))
        for flag in flags
        if str(flag or "").strip()
    ]
    return " / ".join(labels) or "-"


def _group_shandong_candidates(payload: dict[str, object]) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = {}
    for key in ("rush", "stable", "safe", "watch"):
        value = payload.get(key)
        grouped[key] = [item for item in value if isinstance(item, dict)] if isinstance(value, list) else []
    return grouped


def _append_shandong_candidate_sheet(workbook: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(
        [
            "分层",
            "院校",
            "院校代码",
            "专业（专业类）",
            "专业代码",
            "计划数",
            "参考位次",
            "最近最低分",
            "最近最低位次",
            "位次差",
            "边际比例",
            "历史年份",
            "历年最低分/位次",
            "计划变化",
            "选科要求",
            "数据置信度",
            "风险说明",
            "来源编号",
            "推荐理由",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row.get("bucket_label") or _format_shandong_bucket(row.get("bucket")),
                row.get("college_name"),
                row.get("college_code_snapshot"),
                row.get("major_name") or "院校级计划",
                row.get("major_code_snapshot"),
                row.get("plan_count"),
                _read_nested_number(row.get("score_summary"), "reference_rank"),
                _read_nested_number(row.get("score_summary"), "latest_min_score"),
                _read_nested_number(row.get("score_summary"), "latest_min_rank"),
                row.get("rank_margin"),
                _format_percent_value(row.get("rank_margin_ratio")),
                _format_years(row.get("years_used")),
                _format_shandong_rank_rows(row),
                _format_shandong_plan_change(row),
                row.get("subject_requirement") or "不限或待补",
                _format_shandong_confidence(row.get("data_confidence")),
                _format_risk_flags(row.get("risk_flags")),
                _format_source_document_ids(row.get("source_document_ids")),
                row.get("explanation_text"),
            ]
        )


def _append_shandong_source_sheet(
    workbook: Workbook,
    payload: dict[str, object],
    candidates: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("数据来源页")
    sheet.append(["类型", "来源编号/年份", "说明"])
    sheet.append(["推荐依据", "2023 / 2024 / 2025", "普通类历史投档数据用于位次参考，分数只作为展示和换算辅助。"])
    sheet.append(["2026 数据提示", "待官方最终发布", SHANDONG_2026_DATA_NOTICE])
    for note in payload.get("input_notes") or []:
        if note:
            sheet.append(["输入提示", "-", str(note)])

    seen: set[tuple[str, str, str]] = set()
    for row in candidates:
        for source_id in row.get("source_document_ids") or []:
            item = ("来源编号", str(source_id), "候选明细关联的官方来源登记编号")
            if item not in seen:
                seen.add(item)
                sheet.append(list(item))
        for rank_row in _get_shandong_rank_rows(row):
            note = rank_row.get("source_note")
            if not note:
                continue
            item = ("历年投档表", str(rank_row.get("year") or "-"), str(note))
            if item not in seen:
                seen.add(item)
                sheet.append(list(item))


def _build_shandong_risk_overview_rows(
    payload: dict[str, object],
    candidates: list[dict[str, object]],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_mode = str(payload.get("source_mode") or "")
    rank_basis = str(payload.get("rank_projection_basis") or "")
    if source_mode in {"projection", "exam_projection"} or "exam" in rank_basis:
        rows.append(
            {
                "title": "校内考试估算",
                "summary": "本报告使用校内考试生成的预估位次",
                "detail": "校内名次不能直接等同于山东全省位次，正式填报前建议用省内位次或最终高考成绩复核。",
            }
        )
    if rank_basis == "previous_year_score_rank_segment":
        rows.append(
            {
                "title": "上一年一分一段估算",
                "summary": "当前分数换位次使用最近已导入的一分一段表",
                "detail": "2026 一分一段未发布时可临时估算，但正式填报前应改用当年官方一分一段。",
            }
        )
    if int(payload.get("target_year") or 0) == 2026:
        rows.append(
            {
                "title": "2026 数据未完全公开",
                "summary": "正式招生计划、投档线和一分一段需等待官方最终发布",
                "detail": SHANDONG_2026_DATA_NOTICE,
            }
        )

    risk_counts: dict[str, int] = {}
    for row in candidates:
        for flag in row.get("risk_flags") or []:
            key = str(flag)
            if key:
                risk_counts[key] = risk_counts.get(key, 0) + 1
    for flag, count in sorted(risk_counts.items(), key=lambda item: item[0]):
        rows.append(
            {
                "title": RECOMMENDATION_RISK_FLAG_LABELS.get(flag, flag),
                "summary": f"{count} 条结果带有该风险",
                "detail": _shandong_risk_detail(flag),
            }
        )

    if not rows:
        rows.append(
            {
                "title": "暂无明显系统风险",
                "summary": "当前结果未发现需要单独列出的系统性风险标签",
                "detail": "仍建议在正式填报前核对山东省教育招生考试院最终计划、选科要求和高校招生章程。",
            }
        )
    return rows


def _shandong_risk_detail(flag: str) -> str:
    mapping = {
        "rank_projection_from_previous_year": "分数换位次使用上一年一分一段，适合临时估算，不适合直接定稿。",
        "rank_projection_from_school_exam": "位次来自校内考试估算，不能直接等同山东全省位次。",
        "three_year_data_incomplete": "该专业近三年历史样本不完整，分层置信度会下降。",
        "historical_data_missing": "只有单年或更少历史样本，建议先列为关注对象。",
        "plan_missing": "目标年份招生计划暂缺，正式计划发布后可能影响分层。",
        "plan_decreased": "计划数减少可能抬高录取风险，不能按保底处理。",
        "subject_requirement_check": "存在选科要求，最终填报前仍需逐条复核。",
    }
    return mapping.get(flag, "请结合推荐理由、数据来源和高校招生章程人工复核。")



def _get_shandong_rank_rows(candidate: dict[str, object]) -> list[dict[str, object]]:
    historical_summary = candidate.get("historical_summary")
    if not isinstance(historical_summary, dict):
        return []
    rank_rows = historical_summary.get("rank_rows")
    return [item for item in rank_rows if isinstance(item, dict)] if isinstance(rank_rows, list) else []


def _format_shandong_rank_rows(candidate: dict[str, object]) -> str:
    rows: list[str] = []
    for row in _get_shandong_rank_rows(candidate):
        year = row.get("year") or "-"
        min_score = row.get("min_score") if row.get("min_score") is not None else "-"
        min_rank = row.get("min_rank") if row.get("min_rank") is not None else "-"
        rows.append(f"{year}: {min_score} 分 / {min_rank} 位")
    return "；".join(rows) or "-"


def _format_shandong_plan_change(candidate: dict[str, object]) -> str:
    historical_summary = candidate.get("historical_summary")
    if not isinstance(historical_summary, dict):
        return "目标年份计划暂缺，需要等正式计划或人工导入后复核。"
    plan_change = historical_summary.get("plan_change")
    if not isinstance(plan_change, dict):
        return "目标年份计划暂缺，需要等正式计划或人工导入后复核。"
    target_count = plan_change.get("target_year_plan_count")
    latest_count = plan_change.get("latest_historical_plan_count")
    factor = plan_change.get("plan_change_factor")
    if target_count is None and latest_count is None:
        return "目标年份计划暂缺，需要等正式计划或人工导入后复核。"
    parts = [
        "目标年份计划暂缺" if target_count is None else f"目标年份计划 {target_count} 人",
        "近年计划待补" if latest_count is None else f"近年参考计划 {latest_count} 人",
    ]
    if isinstance(factor, (int, float)):
        parts.append(f"变化系数 {float(factor):.2f}")
    return "；".join(parts)


def _build_chapter_restriction_export_copy(row: dict[str, object]) -> str:
    labels = [
        ("语种", _read_snapshot_value(row, "chapter_language_requirement")),
        ("单科", _read_snapshot_value(row, "chapter_single_subject_requirement")),
        ("性别", _read_snapshot_value(row, "chapter_gender_requirement")),
        ("身高", _read_snapshot_value(row, "chapter_height_requirement")),
        ("视力", _read_snapshot_value(row, "chapter_vision_requirement")),
        ("色觉", _read_snapshot_value(row, "chapter_color_vision_requirement")),
        ("体检", _read_snapshot_value(row, "chapter_physical_exam_requirement")),
    ]
    parts = [f"{label}：{value}" for label, value in labels if value]
    return "；".join(parts) or "-"
