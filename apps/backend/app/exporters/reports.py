from __future__ import annotations

from openpyxl import Workbook

from app.core.config import Settings
from app.utils.parsers import make_timestamped_filename, relative_to_project


def _save_workbook(settings: Settings, workbook: Workbook, prefix: str) -> str:
    filename = make_timestamped_filename(prefix, ".xlsx")
    path = settings.exports_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)


def _append_analysis_insight_sheet(
    workbook: Workbook,
    rows: list[dict[str, object]],
) -> None:
    insight_sheet = workbook.create_sheet("摘要概览")
    insight_sheet.append(["标题", "摘要", "说明", "提示级别"])
    for row in rows:
        insight_sheet.append([
            row.get("title"),
            row.get("summary"),
            row.get("detail"),
            row.get("tone"),
        ])


def _format_number(value: object) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value or "-")


def _format_percent(value: object) -> str:
    if isinstance(value, (int, float)):
        return f"{float(value) * 100:.1f}%"
    return "-"


def _format_gap(value: object, unit: str) -> str:
    if not isinstance(value, (int, float)):
        return "-"
    prefix = "+" if float(value) > 0 else ""
    return f"{prefix}{_format_number(value)}{unit}"


def _join_values(values: object) -> str:
    if not isinstance(values, list):
        return "-"
    return "；".join(str(item) for item in values if item) or "-"


def _trend_tone(delta: object) -> str:
    if not isinstance(delta, (int, float)):
        return "info"
    if delta > 0:
        return "success"
    if delta < 0:
        return "warning"
    return "info"


def _pick_max_by(items: list[dict[str, object]], key: str) -> dict[str, object] | None:
    valid_items = [item for item in items if isinstance(item.get(key), (int, float))]
    if not valid_items:
        return None
    return max(valid_items, key=lambda item: float(item.get(key) or 0))


def _pick_min_by(items: list[dict[str, object]], key: str) -> dict[str, object] | None:
    valid_items = [item for item in items if isinstance(item.get(key), (int, float))]
    if not valid_items:
        return None
    return min(valid_items, key=lambda item: float(item.get(key) or 0))


def _get_student_standing_score(subject: dict[str, object]) -> float | None:
    if isinstance(subject.get("grade_percentile"), (int, float)):
        return float(subject["grade_percentile"])
    if isinstance(subject.get("class_percentile"), (int, float)):
        return float(subject["class_percentile"])
    if isinstance(subject.get("score"), (int, float)):
        return float(subject["score"])
    return None


def _build_student_subject_detail(subject: dict[str, object], trailing_copy: str) -> str:
    parts: list[str] = []
    if isinstance(subject.get("score"), (int, float)):
        parts.append(f"分数 {_format_number(subject['score'])}")
    if subject.get("class_rank") is not None:
        parts.append(f"班级第 {subject.get('class_rank')}")
    if subject.get("grade_rank") is not None:
        parts.append(f"年级第 {subject.get('grade_rank')}")
    if isinstance(subject.get("grade_percentile"), (int, float)):
        parts.append(f"年百分位 {_format_percent(subject['grade_percentile'])}")
    elif isinstance(subject.get("class_percentile"), (int, float)):
        parts.append(f"班百分位 {_format_percent(subject['class_percentile'])}")
    if isinstance(subject.get("score_delta"), (int, float)):
        delta = float(subject["score_delta"])
        parts.append(f"较上次 {'+' if delta > 0 else '-' if delta < 0 else ''}{_format_number(abs(delta))} 分")
    if trailing_copy:
        parts.append(trailing_copy)
    return "；".join(parts)


def _build_student_analysis_insight_rows(payload: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    rank_segments: list[str] = []
    if payload.get("class_rank") is not None:
        rank_segments.append(f"班级第 {payload.get('class_rank')}")
    if payload.get("grade_rank") is not None:
        rank_segments.append(f"年级第 {payload.get('grade_rank')}")
    rows.append(
        {
            "title": "本次成绩摘要",
            "summary": f"{payload.get('student_name')} 在 {payload.get('exam_name')} 取得总分 {_format_number(payload.get('total_score'))}",
            "detail": payload.get("overview_sentence") or " / ".join(rank_segments) or "当前摘要可用于导出前快速复核学生本次成绩定位，避免把错误考试或学生参数带入报表。",
            "tone": "info",
        }
    )

    if isinstance(payload.get("total_score_delta"), (int, float)):
        delta = float(payload["total_score_delta"])
        trend_label = "与上次考试基本持平"
        if delta > 0:
            trend_label = "较上次考试提升"
        elif delta < 0:
            trend_label = "较上次考试回落"
        rows.append(
            {
                "title": "阶段趋势",
                "summary": f"{trend_label} {_format_number(abs(delta))} 分",
                "detail": (
                    f"对比考试：{payload.get('previous_exam_name')}。导出前建议确认本次分析是否使用了正确的对比考试。"
                    if payload.get("previous_exam_name")
                    else "当前暂无上一场考试可对比，导出结果会以单次考试分析为主。"
                ),
                "tone": _trend_tone(delta),
            }
        )

    subjects = [item for item in payload.get("subjects") or [] if isinstance(item, dict)]
    strongest_subject = None
    valid_standing_subjects = [item for item in subjects if _get_student_standing_score(item) is not None]
    if valid_standing_subjects:
        strongest_subject = max(valid_standing_subjects, key=lambda item: _get_student_standing_score(item) or 0)
        rows.append(
            {
                "title": "优势学科",
                "summary": f"{strongest_subject.get('subject_name')} 当前表现最强",
                "detail": _build_student_subject_detail(strongest_subject, "可在汇报时作为稳定得分点优先展示。"),
                "tone": "success",
            }
        )

    focus_subject = min(valid_standing_subjects, key=lambda item: _get_student_standing_score(item) or 0) if valid_standing_subjects else None
    if focus_subject and focus_subject.get("subject_id") != (strongest_subject or {}).get("subject_id"):
        rows.append(
            {
                "title": "重点关注学科",
                "summary": f"{focus_subject.get('subject_name')} 建议继续重点复核",
                "detail": _build_student_subject_detail(focus_subject, "如需导出给班主任或任课教师，这一科更适合作为后续跟进重点。"),
                "tone": "warning",
            }
        )
    for suggestion in payload.get("action_suggestions") or []:
        if not isinstance(suggestion, dict):
            continue
        rows.append(
            {
                "title": suggestion.get("title") or "行动建议",
                "summary": suggestion.get("summary") or "-",
                "detail": f"涉及科目：{_join_values(suggestion.get('subject_names'))}",
                "tone": "warning" if suggestion.get("category") in {"fix_weakness", "target_warning"} else "success",
            }
        )
    return rows


def _build_class_analysis_insight_rows(payload: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = [
        {
            "title": "班级整体状态",
            "summary": f"{payload.get('class_name')} 共 {payload.get('student_count')} 人，总分均分 {_format_number(payload.get('total_average'))}，中位数 {_format_number(payload.get('total_median'))}",
            "detail": "当前摘要可用于导出前确认班级规模和整体分布是否符合预期，避免误选班级或考试。",
            "tone": "info",
        }
    ]
    if isinstance(payload.get("grade_average"), (int, float)) and isinstance(payload.get("total_average"), (int, float)):
        delta = float(payload["total_average"]) - float(payload["grade_average"])
        rows.append(
            {
                "title": "与年级对比",
                "summary": (
                    f"班级均分高于年级 {_format_number(delta)} 分"
                    if delta >= 0
                    else f"班级均分低于年级 {_format_number(abs(delta))} 分"
                ),
                "detail": f"班级均分 {_format_number(payload.get('total_average'))}，年级均分 {_format_number(payload.get('grade_average'))}。",
                "tone": _trend_tone(delta),
            }
        )

    subjects = [item for item in payload.get("subject_breakdown") or [] if isinstance(item, dict)]
    strongest_subject = _pick_max_by(subjects, "average_score")
    if strongest_subject:
        rows.append(
            {
                "title": "优势学科",
                "summary": f"{strongest_subject.get('subject_name')} 当前均分最高",
                "detail": f"均分 {_format_number(strongest_subject.get('average_score'))}，优秀率 {_format_percent(strongest_subject.get('excellent_rate'))}，及格率 {_format_percent(strongest_subject.get('pass_rate'))}。",
                "tone": "success",
            }
        )
    focus_subject = _pick_min_by(subjects, "pass_rate")
    if focus_subject and focus_subject.get("subject_id") != (strongest_subject or {}).get("subject_id"):
        rows.append(
            {
                "title": "重点攻坚学科",
                "summary": f"{focus_subject.get('subject_name')} 当前及格率最低",
                "detail": f"均分 {_format_number(focus_subject.get('average_score'))}，及格率 {_format_percent(focus_subject.get('pass_rate'))}，优秀率 {_format_percent(focus_subject.get('excellent_rate'))}。",
                "tone": "warning",
            }
        )
    return rows


def _build_grade_analysis_insight_rows(
    meta: dict[str, object],
    class_rows: list[dict[str, object]],
    subject_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = [
        {
            "title": "年级整体状态",
            "summary": f"{meta.get('grade_name')} 共 {meta.get('student_count')} 人，均分 {_format_number(meta.get('total_average'))}，中位数 {_format_number(meta.get('total_median'))}",
            "detail": (
                f"当前优秀率 {_format_percent(meta.get('excellent_rate'))}，适合作为导出前的整体判断基线。"
                if isinstance(meta.get("excellent_rate"), (int, float))
                else "当前摘要可用于导出前确认年级整体成绩面貌和人数口径。"
            ),
            "tone": "info",
        }
    ]
    best_class = _pick_max_by(class_rows, "average_score")
    if best_class:
        detail = f"班级均分 {_format_number(best_class.get('average_score'))}，人数 {best_class.get('student_count')}"
        if isinstance(best_class.get("excellent_rate"), (int, float)):
            detail += f"，优秀率 {_format_percent(best_class.get('excellent_rate'))}"
        rows.append(
            {
                "title": "领先班级",
                "summary": f"{best_class.get('class_name')} 当前均分领先",
                "detail": f"{detail}。",
                "tone": "success",
            }
        )
    risk_class = _pick_min_by(class_rows, "average_score")
    if risk_class and risk_class.get("class_name") != (best_class or {}).get("class_name"):
        detail = f"班级均分 {_format_number(risk_class.get('average_score'))}，人数 {risk_class.get('student_count')}"
        if isinstance(risk_class.get("excellent_rate"), (int, float)):
            detail += f"，优秀率 {_format_percent(risk_class.get('excellent_rate'))}"
        rows.append(
            {
                "title": "需重点跟进班级",
                "summary": f"{risk_class.get('class_name')} 当前均分相对靠后",
                "detail": f"{detail}。",
                "tone": "warning",
            }
        )
    focus_subject = _pick_min_by(subject_rows, "pass_rate")
    if focus_subject:
        rows.append(
            {
                "title": "学科攻坚重点",
                "summary": f"{focus_subject.get('subject_name')} 当前及格率最低",
                "detail": f"均分 {_format_number(focus_subject.get('average_score'))}，及格率 {_format_percent(focus_subject.get('pass_rate'))}，优秀率 {_format_percent(focus_subject.get('excellent_rate'))}。",
                "tone": "warning",
            }
        )
    return rows


def export_adviser_weekly_summary_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "班主任周报"
    overview = payload.get("overview") if isinstance(payload.get("overview"), dict) else {}
    score_summary = payload.get("score_summary") if isinstance(payload.get("score_summary"), dict) else {}
    attendance_summary = payload.get("attendance_summary") if isinstance(payload.get("attendance_summary"), dict) else {}
    behavior_summary = payload.get("behavior_summary") if isinstance(payload.get("behavior_summary"), dict) else {}

    sheet.append(["班主任周报"])
    sheet.append(["年级", payload.get("grade_name") or "全部"])
    sheet.append(["班级", payload.get("class_name") or "全部"])
    sheet.append(["考试", payload.get("exam_name") or "未选择"])
    sheet.append(["数据范围", f"{payload.get('start_date')} 至 {payload.get('end_date')}"])
    sheet.append(["生成时间", payload.get("generated_at") or "-"])
    sheet.append([])
    sheet.append(["概览项", "数值"])
    sheet.append(["学生数", overview.get("student_count", 0)])
    sheet.append(["成绩样本", overview.get("score_sample_count", 0)])
    sheet.append(["缺勤风险", overview.get("absence_risk_count", 0)])
    sheet.append(["行为风险", overview.get("behavior_risk_count", 0)])
    sheet.append(["需跟进人数", overview.get("follow_up_count", 0)])
    sheet.append([])
    sheet.append(["成绩概况", "说明"])
    sheet.append(["样本数", score_summary.get("sample_count", 0)])
    sheet.append(["平均总分", _format_number(score_summary.get("average_total_score"))])
    sheet.append(["低位人数", score_summary.get("low_score_count", 0)])
    sheet.append(["下滑人数", score_summary.get("decline_count", 0)])
    sheet.append([])
    sheet.append(["考勤概况", "说明"])
    sheet.append(["导入状态", "已导入" if attendance_summary.get("imported") else "未导入"])
    sheet.append(["总记录", attendance_summary.get("total_records", 0)])
    sheet.append(["迟到", attendance_summary.get("late_count", 0)])
    sheet.append(["旷课", attendance_summary.get("truancy_count", 0)])
    sheet.append([])
    sheet.append(["行为概况", "说明"])
    sheet.append(["导入状态", "已导入" if behavior_summary.get("imported") else "未导入"])
    sheet.append(["总记录", behavior_summary.get("total_records", 0)])
    sheet.append(["高关注行为", behavior_summary.get("severe_count", 0)])

    risk_sheet = workbook.create_sheet("重点学生")
    risk_sheet.append(["姓名", "班级", "风险等级", "主要原因", "建议动作"])
    for row in payload.get("risk_students") or []:
        if not isinstance(row, dict):
            continue
        risk_sheet.append([
            row.get("student_name"),
            row.get("class_name"),
            row.get("risk_label"),
            row.get("primary_reason"),
            row.get("suggested_action"),
        ])

    action_sheet = workbook.create_sheet("行动清单")
    action_sheet.append(["类型", "任务", "数量", "学生 ID"])
    for row in payload.get("action_items") or []:
        if not isinstance(row, dict):
            continue
        action_sheet.append([
            row.get("action_type"),
            row.get("title"),
            row.get("count"),
            _join_values(row.get("student_ids")),
        ])

    flag_sheet = workbook.create_sheet("缺失数据")
    flag_sheet.append(["提示"])
    data_flags = payload.get("data_flags") if isinstance(payload.get("data_flags"), list) else []
    if data_flags:
        for item in data_flags:
            flag_sheet.append([item])
    else:
        flag_sheet.append(["当前范围未发现明显缺失提示"])
    return _save_workbook(settings, workbook, "adviser_weekly_summary")


def export_student_followup_package(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "学生跟进包"
    score_summary = payload.get("score_summary") if isinstance(payload.get("score_summary"), dict) else {}
    attendance_summary = payload.get("attendance_summary") if isinstance(payload.get("attendance_summary"), dict) else {}
    behavior_summary = payload.get("behavior_summary") if isinstance(payload.get("behavior_summary"), dict) else {}
    growth_summary = payload.get("growth_summary") if isinstance(payload.get("growth_summary"), dict) else {}

    sheet.append(["学生跟进包"])
    sheet.append(["学生", payload.get("student_name")])
    sheet.append(["学号", payload.get("student_no")])
    sheet.append(["年级", payload.get("grade_name")])
    sheet.append(["班级", payload.get("class_name")])
    sheet.append(["风险等级", payload.get("risk_label")])
    sheet.append(["生成时间", payload.get("generated_at") or "-"])
    sheet.append([])
    sheet.append(["主要原因", _join_values(payload.get("reasons"))])
    sheet.append(["建议行动", _join_values(payload.get("suggested_actions"))])
    sheet.append(["缺失数据", _join_values(payload.get("data_flags"))])
    sheet.append([])
    sheet.append(["成绩趋势", "值"])
    sheet.append(["考试", score_summary.get("exam_name") or "-"])
    sheet.append(["总分", _format_number(score_summary.get("total_score"))])
    sheet.append(["上次总分", _format_number(score_summary.get("previous_total_score"))])
    sheet.append(["总分变化", _format_number(score_summary.get("total_score_delta"))])
    sheet.append(["班级排名", _format_number(score_summary.get("class_rank"))])
    sheet.append(["年级排名", _format_number(score_summary.get("grade_rank"))])
    sheet.append([])
    sheet.append(["考勤摘要", "值"])
    sheet.append(["导入状态", "已导入" if attendance_summary.get("imported") else "未导入"])
    sheet.append(["总记录", attendance_summary.get("total_records", 0)])
    sheet.append(["迟到", attendance_summary.get("late_count", 0)])
    sheet.append(["早退", attendance_summary.get("early_leave_count", 0)])
    sheet.append(["病假", attendance_summary.get("sick_leave_count", 0)])
    sheet.append(["事假", attendance_summary.get("personal_leave_count", 0)])
    sheet.append(["旷课", attendance_summary.get("truancy_count", 0)])
    sheet.append([])
    sheet.append(["行为摘要", "值"])
    sheet.append(["导入状态", "已导入" if behavior_summary.get("imported") else "未导入"])
    sheet.append(["总记录", behavior_summary.get("total_records", 0)])
    sheet.append(["表扬", behavior_summary.get("positive_count", 0)])
    sheet.append(["违纪/奖惩", behavior_summary.get("discipline_count", 0)])
    sheet.append(["高关注行为", behavior_summary.get("severe_count", 0)])
    sheet.append([])
    sheet.append(["成长档案", "值"])
    sheet.append(["记录数", growth_summary.get("record_count", 0)])
    sheet.append(["最近记录日期", growth_summary.get("latest_record_date") or "-"])
    return _save_workbook(settings, workbook, "student_followup_package")


def export_planning_followup_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "升学规划跟进表"
    student = payload.get("student") if isinstance(payload.get("student"), dict) else {}
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    risk = payload.get("risk") if isinstance(payload.get("risk"), dict) else {}
    profile = payload.get("pathway_profile") if isinstance(payload.get("pathway_profile"), dict) else {}

    sheet.append(["学生升学规划跟进表"])
    sheet.append(["学生", student.get("name")])
    sheet.append(["学号", student.get("student_no")])
    sheet.append(["年级", student.get("grade_name")])
    sheet.append(["班级", student.get("class_name")])
    sheet.append(["生成时间", payload.get("generated_at") or "-"])
    sheet.append([])
    sheet.append(["规划状态", "值"])
    sheet.append(["进行中任务", summary.get("open_task_count", 0)])
    sheet.append(["已完成任务", summary.get("completed_task_count", 0)])
    sheet.append(["逾期任务", summary.get("overdue_task_count", 0)])
    sheet.append(["材料任务", summary.get("material_gap_task_count", 0)])
    sheet.append(["风险等级", risk.get("risk_label") or "-"])
    sheet.append(["风险原因", _join_values(risk.get("reasons"))])
    sheet.append([])
    sheet.append(["升学画像", "值"])
    sheet.append(["生源地", profile.get("province") or "-"])
    sheet.append(["考生类型", profile.get("candidate_type") or "-"])
    sheet.append(["考试类型", profile.get("exam_type") or "-"])
    sheet.append(["选科组合", profile.get("subject_combination") or "-"])

    goals_sheet = workbook.create_sheet("目标路径")
    goals_sheet.append(["目标年份", "路径", "目标院校", "目标专业", "目标分数", "目标位次", "状态", "优先级", "备选路径", "备注"])
    for row in payload.get("goals") or []:
        if not isinstance(row, dict):
            continue
        goals_sheet.append([
            row.get("target_year"),
            row.get("pathway_name"),
            row.get("target_college"),
            row.get("target_major"),
            row.get("target_score"),
            row.get("target_rank"),
            row.get("status_label"),
            row.get("priority_label"),
            row.get("backup_pathways"),
            row.get("note"),
        ])

    tasks_sheet = workbook.create_sheet("任务清单")
    tasks_sheet.append(["任务", "类型", "状态", "优先级", "截止日期", "是否逾期", "来源", "说明"])
    for row in payload.get("tasks") or []:
        if not isinstance(row, dict):
            continue
        tasks_sheet.append([
            row.get("title"),
            row.get("task_type_label"),
            row.get("status_label"),
            row.get("priority_label"),
            row.get("due_date"),
            "是" if row.get("is_overdue") else "否",
            row.get("source_type"),
            row.get("description"),
        ])

    evaluations_sheet = workbook.create_sheet("路径初筛")
    evaluations_sheet.append(["路径", "状态", "置信程度", "缺失材料", "下一步"])
    for row in payload.get("pathway_evaluations") or []:
        if not isinstance(row, dict):
            continue
        gaps = row.get("missing_materials_json") if isinstance(row.get("missing_materials_json"), list) else []
        evaluations_sheet.append([
            row.get("pathway_name"),
            row.get("status_label"),
            row.get("confidence_level"),
            "；".join(str(item.get("material_label") or item.get("material_key")) for item in gaps if isinstance(item, dict)) or "-",
            _join_values(row.get("next_actions_json")),
        ])

    notes_sheet = workbook.create_sheet("复盘记录")
    notes_sheet.append(["时间", "类型", "内容"])
    for row in payload.get("notes") or []:
        if not isinstance(row, dict):
            continue
        notes_sheet.append([row.get("created_at"), row.get("note_type_label"), row.get("content")])

    suggested_sheet = workbook.create_sheet("建议补充")
    suggested_sheet.append(["建议任务", "类型", "优先级", "说明"])
    for row in payload.get("suggested_tasks") or []:
        if not isinstance(row, dict):
            continue
        suggested_sheet.append([
            row.get("title"),
            row.get("task_type_label"),
            row.get("priority_label"),
            row.get("description"),
        ])
    return _save_workbook(settings, workbook, "planning_followup")


def _format_teacher_assignment(row: dict[str, object]) -> str:
    segments = [str(item).strip() for item in [row.get("class_name"), row.get("subject_name")] if str(item or "").strip()]
    return " / ".join(segments) or "未命名任教拆分"


def _format_adviser_item_type(value: object) -> str:
    mapping = {
        "daily_management": "班级常规管理",
        "discipline": "卫生纪律",
        "home_school": "家校沟通",
        "activity": "活动组织",
        "guidance": "学生发展指导",
        "bonus": "特殊加分项",
        "penalty": "扣分项",
        "positive": "加分项",
        "negative": "扣分项",
    }
    current = str(value or "").strip()
    return mapping.get(current, current or "-")


def _build_teacher_analysis_insight_rows(payload: dict[str, object]) -> list[dict[str, object]]:
    assignments = [item for item in payload.get("assignment_breakdown") or [] if isinstance(item, dict)]
    rows: list[dict[str, object]] = [
        {
            "title": "任教整体状态",
            "summary": (
                f"{payload.get('teacher_name')} 当前任教均分 {_format_number(payload.get('overall_average'))}"
                if isinstance(payload.get("overall_average"), (int, float))
                else f"{payload.get('teacher_name')} 当前共有 {len(assignments)} 条任教分析"
            ),
            "detail": f"当前分析基于 {len(assignments)} 条任教拆分结果，可用于导出前确认教师与考试参数是否正确。",
            "tone": "info",
        }
    ]
    best_assignment = _pick_max_by(assignments, "average_score")
    if best_assignment:
        rows.append(
            {
                "title": "表现最佳任教拆分",
                "summary": f"{_format_teacher_assignment(best_assignment)} 当前均分最高",
                "detail": f"均分 {_format_number(best_assignment.get('average_score'))}，优秀率 {_format_percent(best_assignment.get('excellent_rate'))}，及格率 {_format_percent(best_assignment.get('pass_rate'))}。",
                "tone": "success",
            }
        )
    focus_assignment = _pick_min_by(assignments, "pass_rate")
    if focus_assignment and focus_assignment.get("assignment_id") != (best_assignment or {}).get("assignment_id"):
        rows.append(
            {
                "title": "需重点跟进任教拆分",
                "summary": f"{_format_teacher_assignment(focus_assignment)} 当前及格率最低",
                "detail": f"均分 {_format_number(focus_assignment.get('average_score'))}，及格率 {_format_percent(focus_assignment.get('pass_rate'))}，优秀率 {_format_percent(focus_assignment.get('excellent_rate'))}。",
                "tone": "warning",
            }
        )
    return rows


def _pick_top_dimension(items: list[dict[str, object]]) -> dict[str, object] | None:
    scores: dict[str, dict[str, float]] = {}
    for item in items:
        for name, score in (item.get("dimension_scores_json") or {}).items():
            if not isinstance(name, str) or not isinstance(score, (int, float)):
                continue
            current = scores.get(name) or {"total": 0.0, "count": 0.0}
            current["total"] += float(score)
            current["count"] += 1.0
            scores[name] = current
    rows = [
        {"name": name, "score": value["total"] / value["count"]}
        for name, value in scores.items()
        if value["count"] > 0
    ]
    return _pick_max_by(rows, "score")


def _pick_top_category(items: list[dict[str, object]]) -> dict[str, object] | None:
    scores: dict[str, float] = {}
    for item in items:
        for name, score in (item.get("category_scores_json") or {}).items():
            if not isinstance(name, str) or not isinstance(score, (int, float)):
                continue
            scores[name] = scores.get(name, 0.0) + float(score)
    rows = [{"name": name, "score": score} for name, score in scores.items()]
    return _pick_max_by(rows, "score")


def _format_category_scores(value: object) -> str:
    if not isinstance(value, dict) or not value:
        return "-"
    return " / ".join(f"{_format_adviser_item_type(key)}:{score}" for key, score in value.items())


def _build_evaluation_insight_rows(
    meta: dict[str, object],
    teacher_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = [
        {
            "title": "评教整体状态",
            "summary": f"{meta.get('template_name')} 当前覆盖 {len(teacher_rows)} 位教师",
            "detail": f"{meta.get('semester_name') or '未标注学期'}。导出前建议先确认模板和学期是否对应当前批次。",
            "tone": "info",
        }
    ]

    best_teacher = _pick_max_by(teacher_rows, "overall_avg_score")
    if best_teacher:
        rows.append(
            {
                "title": "当前领先教师",
                "summary": f"{best_teacher.get('teacher_name')} 当前综合均分最高",
                "detail": f"综合均分 {_format_number(best_teacher.get('overall_avg_score'))}，响应数 {best_teacher.get('response_count')}{f'，排名 {best_teacher.get('rank')}' if best_teacher.get('rank') is not None else ''}。",
                "tone": "success",
            }
        )

    focus_teacher = _pick_min_by(teacher_rows, "overall_avg_score")
    if focus_teacher and focus_teacher.get("teacher_id") != (best_teacher or {}).get("teacher_id"):
        rows.append(
            {
                "title": "需重点复核教师",
                "summary": f"{focus_teacher.get('teacher_name')} 当前综合均分相对靠后",
                "detail": f"综合均分 {_format_number(focus_teacher.get('overall_avg_score'))}，响应数 {focus_teacher.get('response_count')}{f'，排名 {focus_teacher.get('rank')}' if focus_teacher.get('rank') is not None else ''}。",
                "tone": "warning",
            }
        )

    top_dimension = _pick_top_dimension(teacher_rows)
    if top_dimension:
        rows.append(
            {
                "title": "当前高频优势维度",
                "summary": f"{top_dimension.get('name')} 在本批次表现相对更强",
                "detail": f"平均维度得分 {_format_number(top_dimension.get('score'))}。导出前可确认这是否符合本次模板重点。",
                "tone": "info",
            }
        )

    return rows


def _build_adviser_quant_insight_rows(summary_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    if not summary_rows:
        return []

    total_score = sum(float(item.get("total_score") or 0) for item in summary_rows if isinstance(item.get("total_score"), (int, float)))
    total_record_count = sum(int(item.get("record_count") or 0) for item in summary_rows if item.get("record_count") is not None)
    rows: list[dict[str, object]] = [
        {
            "title": "量化整体状态",
            "summary": f"{summary_rows[0].get('semester_name') or '当前学期'}共 {len(summary_rows)} 位教师进入量化汇总",
            "detail": f"总分合计 {_format_number(total_score)}，记录数 {total_record_count}，规则版本 {summary_rows[0].get('rule_version_name') or '未指定'}。",
            "tone": "info",
        }
    ]

    top_teacher = _pick_max_by(summary_rows, "total_score")
    if top_teacher:
        rows.append(
            {
                "title": "总分最高教师",
                "summary": f"{top_teacher.get('teacher_name')} 当前总分最高",
                "detail": f"总分 {_format_number(top_teacher.get('total_score'))}，加分 {_format_number(top_teacher.get('positive_score'))}，扣分 {_format_number(top_teacher.get('negative_score'))}。",
                "tone": "success",
            }
        )

    focus_teacher = _pick_max_by(summary_rows, "negative_score")
    if (
        focus_teacher
        and focus_teacher.get("teacher_id") != (top_teacher or {}).get("teacher_id")
        and isinstance(focus_teacher.get("negative_score"), (int, float))
        and float(focus_teacher.get("negative_score") or 0) > 0
    ):
        rows.append(
            {
                "title": "扣分较多教师",
                "summary": f"{focus_teacher.get('teacher_name')} 当前扣分最多",
                "detail": f"扣分 {_format_number(focus_teacher.get('negative_score'))}，记录数 {focus_teacher.get('record_count')}，班级 {' / '.join(focus_teacher.get('class_names') or []) or '-'}。",
                "tone": "warning",
            }
        )

    top_category = _pick_top_category(summary_rows)
    if top_category:
        rows.append(
            {
                "title": "高频得分类别",
                "summary": f"{_format_adviser_item_type(top_category.get('name'))} 当前累计分值最高",
                "detail": f"累计分值 {_format_number(top_category.get('score'))}。导出前可结合规则版本确认该类别是否为本学期重点。",
                "tone": "info",
            }
        )

    return rows


def export_student_analysis_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "学生分析"
    summary.append(["考试", payload.get("exam_name")])
    summary.append(["学生", payload.get("student_name")])
    summary.append(["总分", payload.get("total_score")])
    summary.append(["班级名次", payload.get("class_rank")])
    summary.append(["年级名次", payload.get("grade_rank")])
    summary.append(["校内PR", _format_percent(payload.get("grade_percentile"))])
    summary.append(["上次考试", payload.get("previous_exam_name")])
    summary.append(["总分变化", payload.get("total_score_delta")])
    summary.append(["本次画像", payload.get("overview_sentence") or "-"])

    overview_sheet = workbook.create_sheet("核心概况")
    overview_sheet.append(["项目", "值"])
    overview_sheet.append(["考试", payload.get("exam_name")])
    overview_sheet.append(["学生", payload.get("student_name")])
    overview_sheet.append(["总分", payload.get("total_score")])
    overview_sheet.append(["总分口径", payload.get("score_value_label")])
    overview_sheet.append(["班级名次", payload.get("class_rank")])
    overview_sheet.append(["校内名次", payload.get("grade_rank")])
    overview_sheet.append(["校内PR", _format_percent(payload.get("grade_percentile"))])
    overview_sheet.append(["较上次总分", _format_gap(payload.get("total_score_delta"), "分")])
    overview_sheet.append(["较上次校内名次", _format_gap(payload.get("grade_rank_delta"), "名")])
    overview_sheet.append(["本次画像", payload.get("overview_sentence") or "-"])
    for row in payload.get("target_line_gaps") or []:
        if not isinstance(row, dict):
            continue
        overview_sheet.append([
            f"目标线：{row.get('line_name')}",
            f"{row.get('threshold_label') or '-'} / 差距 {_format_gap(row.get('gap_score'), '分')}",
        ])

    detail = workbook.create_sheet("学科明细")
    detail.append(["科目", "分数", "班级名次", "年级名次", "班百分位", "年百分位", "分数变化", "名次变化", "T分", "排名离差", "诊断"])
    for row in payload.get("subjects", []):
        detail.append(
            [
                row.get("subject_name"),
                row.get("score"),
                row.get("class_rank"),
                row.get("grade_rank"),
                row.get("class_percentile"),
                row.get("grade_percentile"),
                row.get("score_delta"),
                row.get("rank_delta"),
                row.get("t_score"),
                row.get("rank_deviation"),
                row.get("diagnosis"),
            ]
        )

    diagnosis_sheet = workbook.create_sheet("学科诊断")
    diagnosis_sheet.append([
        "科目",
        "分数",
        "校内名次",
        "PR",
        "Z分",
        "T分",
        "排名离差",
        "同档均分",
        "同档差距",
        "波动标准差",
        "有效分目标",
        "有效分差距",
        "诊断",
        "标签",
    ])
    for row in payload.get("subjects", []):
        if not isinstance(row, dict):
            continue
        diagnosis_sheet.append([
            row.get("subject_name"),
            row.get("score"),
            row.get("grade_rank"),
            _format_percent(row.get("grade_percentile")),
            row.get("z_score"),
            row.get("t_score"),
            row.get("rank_deviation"),
            row.get("peer_average_score"),
            row.get("peer_average_delta"),
            row.get("trend_rank_stddev"),
            row.get("primary_effective_score"),
            row.get("primary_effective_score_gap"),
            row.get("diagnosis"),
            _join_values(row.get("diagnosis_tags")),
        ])

    trend_sheet = workbook.create_sheet("趋势轨迹")
    trend_sheet.append(["考试", "日期", "总分", "班级名次", "校内名次", "PR"])
    for row in payload.get("trend_points") or []:
        if not isinstance(row, dict):
            continue
        trend_sheet.append([
            row.get("exam_name"),
            row.get("exam_date"),
            row.get("total_score"),
            row.get("class_rank"),
            row.get("grade_rank"),
            _format_percent(row.get("grade_percentile")),
        ])
    trend_sheet.append([])
    trend_sheet.append(["科目", "考试", "日期", "分数", "校内名次", "PR"])
    for series in payload.get("subject_trends") or []:
        if not isinstance(series, dict):
            continue
        for point in series.get("points") or []:
            if not isinstance(point, dict):
                continue
            trend_sheet.append([
                series.get("subject_name"),
                point.get("exam_name"),
                point.get("exam_date"),
                point.get("score"),
                point.get("grade_rank"),
                _format_percent(point.get("grade_percentile")),
            ])

    action_sheet = workbook.create_sheet("行动建议")
    action_sheet.append(["类别", "标题", "建议", "涉及科目", "优先级"])
    for row in payload.get("action_suggestions") or []:
        if not isinstance(row, dict):
            continue
        action_sheet.append([
            row.get("category"),
            row.get("title"),
            row.get("summary"),
            _join_values(row.get("subject_names")),
            row.get("priority"),
        ])
    _append_analysis_insight_sheet(workbook, _build_student_analysis_insight_rows(payload))
    return _save_workbook(settings, workbook, "student_analysis_report")


def export_class_analysis_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "班级分析"
    summary.append(["考试", payload.get("exam_name")])
    summary.append(["班级", payload.get("class_name")])
    summary.append(["学生数", payload.get("student_count")])
    summary.append(["总分均分", payload.get("total_average")])
    summary.append(["总分中位数", payload.get("total_median")])
    summary.append(["年级均分", payload.get("grade_average")])

    detail = workbook.create_sheet("学科统计")
    detail.append(["科目", "均分", "中位数", "最高分", "最低分", "标准差", "优秀率", "及格率", "有效人数"])
    for row in payload.get("subject_breakdown", []):
        detail.append(
            [
                row.get("subject_name"),
                row.get("average_score"),
                row.get("median_score"),
                row.get("max_score"),
                row.get("min_score"),
                row.get("standard_deviation"),
                row.get("excellent_rate"),
                row.get("pass_rate"),
                row.get("valid_count"),
            ]
        )
    _append_analysis_insight_sheet(workbook, _build_class_analysis_insight_rows(payload))
    return _save_workbook(settings, workbook, "class_analysis_report")


def export_grade_summary_report(
    settings: Settings,
    meta: dict[str, object],
    summary_rows: list[dict[str, object]],
    subject_rows: list[dict[str, object]],
    title: str,
) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "年级概况"
    summary.append(["考试", meta.get("exam_name")])
    summary.append(["年级", meta.get("grade_name")])
    summary.append(["学生数", meta.get("student_count")])
    summary.append(["年级均分", meta.get("total_average")])
    summary.append(["年级中位数", meta.get("total_median")])
    summary.append(["优秀率", meta.get("excellent_rate")])

    class_sheet = workbook.create_sheet("班级汇总")
    class_sheet.append(["班级", "人数", "总分均分", "总分中位数", "最高分", "最低分"])
    for row in summary_rows:
        class_sheet.append(
            [
                row.get("class_name"),
                row.get("student_count"),
                row.get("average_score"),
                row.get("median_score"),
                row.get("max_score"),
                row.get("min_score"),
            ]
        )

    detail = workbook.create_sheet("学科汇总")
    detail.append(["科目", "有效人数", "均分", "优秀率", "及格率"])
    for row in subject_rows:
        detail.append(
            [
                row.get("subject_name"),
                row.get("valid_count"),
                row.get("average_score"),
                row.get("excellent_rate"),
                row.get("pass_rate"),
            ]
        )
    _append_analysis_insight_sheet(workbook, _build_grade_analysis_insight_rows(meta, summary_rows, subject_rows))
    workbook.properties.title = title
    return _save_workbook(settings, workbook, "grade_summary_report")


def export_teacher_analysis_report(settings: Settings, payload: dict[str, object]) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "教师分析"
    summary.append(["考试", payload.get("exam_name")])
    summary.append(["教师", payload.get("teacher_name")])
    summary.append(["整体均分", payload.get("overall_average")])
    summary.append(["任教拆分", len(payload.get("assignment_breakdown") or [])])

    detail = workbook.create_sheet("任教明细")
    detail.append(["班级", "学科", "均分", "优秀率", "及格率", "有效人数"])
    for row in payload.get("assignment_breakdown", []):
        detail.append(
            [
                row.get("class_name"),
                row.get("subject_name"),
                row.get("average_score"),
                row.get("excellent_rate"),
                row.get("pass_rate"),
                row.get("valid_count"),
            ]
        )
    _append_analysis_insight_sheet(workbook, _build_teacher_analysis_insight_rows(payload))
    return _save_workbook(settings, workbook, "teacher_analysis_report")


def export_evaluation_summary_report(
    settings: Settings,
    meta: dict[str, object],
    teacher_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "评教汇总"
    summary.append(["模板", meta.get("template_name")])
    summary.append(["学期", meta.get("semester_name")])
    summary.append(["教师数", len(teacher_rows)])

    teacher_sheet = workbook.create_sheet("教师总览")
    teacher_sheet.append(["教师", "综合得分", "有效条数", "名次", "维度得分"])
    for row in teacher_rows:
        dimension_scores = row.get("dimension_scores_json") or {}
        teacher_sheet.append(
            [
                row.get("teacher_name"),
                row.get("overall_avg_score"),
                row.get("response_count"),
                row.get("rank"),
                " / ".join(f"{key}:{value}" for key, value in dimension_scores.items()),
            ]
        )

    detail_sheet = workbook.create_sheet("维度明细")
    detail_sheet.append(["教师", "维度", "平均分", "样本数"])
    for row in detail_rows:
        detail_sheet.append(
            [
                row.get("teacher_name"),
                row.get("dimension_name"),
                row.get("avg_score"),
                row.get("response_count"),
            ]
        )
    _append_analysis_insight_sheet(workbook, _build_evaluation_insight_rows(meta, teacher_rows))
    return _save_workbook(settings, workbook, "evaluation_summary_report")


def export_adviser_quant_summary_report(
    settings: Settings,
    meta: dict[str, object],
    summary_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "量化汇总"
    summary.append(["学期", meta.get("semester_name")])
    summary.append(["规则版本", meta.get("rule_version_name")])
    summary.append(["教师数", len(summary_rows)])

    teacher_sheet = workbook.create_sheet("教师汇总")
    teacher_sheet.append(["教师", "总分", "加分", "扣分", "记录数", "班级", "分类汇总"])
    for row in summary_rows:
        category_scores = row.get("category_scores_json") or {}
        teacher_sheet.append(
            [
                row.get("teacher_name"),
                row.get("total_score"),
                row.get("positive_score"),
                row.get("negative_score"),
                row.get("record_count"),
                " / ".join(row.get("class_names") or []),
                _format_category_scores(category_scores),
            ]
        )

    detail_sheet = workbook.create_sheet("量化明细")
    detail_sheet.append(["教师", "班级", "月份", "量化项", "类型", "得分", "说明"])
    for row in detail_rows:
        detail_sheet.append(
            [
                row.get("teacher_name"),
                row.get("class_name"),
                row.get("record_month"),
                row.get("item_name"),
                _format_adviser_item_type(row.get("item_type")),
                row.get("score"),
                row.get("description"),
            ]
        )
    _append_analysis_insight_sheet(workbook, _build_adviser_quant_insight_rows(summary_rows))
    return _save_workbook(settings, workbook, "adviser_quant_summary_report")
