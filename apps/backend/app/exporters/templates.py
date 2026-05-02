from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

from app.core.config import Settings


@dataclass(frozen=True)
class TemplateSpec:
    filename: str
    title: str
    headers: list[str]
    sample_rows: list[list[str]]
    instructions: list[str]


TEMPLATE_SPECS = [
    TemplateSpec(
        filename="students_import_template.xlsx",
        title="学生导入模板",
        headers=[
            "学号",
            "姓名",
            "性别",
            "出生日期",
            "入学年份",
            "年级",
            "班级",
            "学生状态",
            "学生类别",
            "艺体方向",
            "生源地省份",
            "联系电话",
            "家庭住址",
            "备注",
        ],
        sample_rows=[
            ["2026001", "张三", "男", "2009-03-12", "2024", "高一", "1班", "在读", "普通生", "", "广东", "13800000001", "示例地址", ""]
        ],
        instructions=[
            "学号为唯一键。",
            "年级、班级、学生状态、学生类别、艺体方向需要先在系统基础数据中维护。",
            "生源地省份建议按考生高考报名所在地填写，用于高考志愿推荐默认带入。",
            "出生日期统一使用 YYYY-MM-DD 格式。",
        ],
    ),
    TemplateSpec(
        filename="teachers_import_template.xlsx",
        title="教师导入模板",
        headers=[
            "工号",
            "姓名",
            "性别",
            "学科",
            "联系方式",
            "职称",
            "岗位",
            "是否班主任",
            "任教状态",
            "入职日期",
            "备注",
        ],
        sample_rows=[
            ["T001", "李老师", "女", "语文", "13900000001", "一级教师", "语文教师", "是", "在岗", "2021-08-01", ""]
        ],
        instructions=[
            "工号为唯一键。",
            "学科、职称、岗位、任教状态需先维护。",
            "是否班主任支持 是/否。",
        ],
    ),
    TemplateSpec(
        filename="exam_scores_import_template.xlsx",
        title="成绩导入模板",
        headers=["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注", "原始分", "赋分", "成绩口径"],
        sample_rows=[["2026届高一3月月考", "2026001", "张三", "1班", "物理", "86", "", "", "72", "86", "赋分"]],
        instructions=[
            "考试需先创建。",
            "同一考试 + 学生 + 科目必须唯一。",
            "有赋分时填写“赋分”和“成绩口径=赋分”，系统用赋分参与总分和名次；没有赋分时只填分数或原始分。",
        ],
    ),
    TemplateSpec(
        filename="score_question_details_import_template.xlsx",
        title="题分明细导入模板",
        headers=["考试名称", "学号", "姓名", "科目", "题号", "题目满分", "学生得分", "知识点", "题型", "能力层级", "错因标签", "错因备注", "备注"],
        sample_rows=[
            ["2026届高一3月月考", "2026001", "张三", "数学", "12", "5", "3", "函数>函数单调性", "选择题", "理解", "概念不清", "", ""],
            ["2026届高一3月月考", "2026001", "张三", "数学", "18", "12", "8", "函数单调性、导数应用", "解答题", "综合应用", "方法不会、步骤不全", "", ""],
        ],
        instructions=[
            "考试和科目需先创建并配置。",
            "同一考试 + 学生 + 科目 + 题号会覆盖更新题分。",
            "一题多个知识点可用顿号、逗号或分号分隔；知识点支持别名和“模块>子模块>知识点”路径写法。",
            "错因标签可用顿号、逗号、分号或竖线分隔；未知错因会自动创建为自定义标签。",
            "学生得分留空视为缺答或无有效得分，会按 0 分参与知识点诊断。",
        ],
    ),
    TemplateSpec(
        filename="knowledge_base_import_template.xlsx",
        title="知识库导入模板",
        headers=["科目", "知识点路径", "说明"],
        sample_rows=[
            ["数学", "函数>基本初等函数>函数单调性", "按同科单父级知识树维护。"],
            ["数学", "函数>导数应用", ""],
        ],
        instructions=[
            "本模板包含“知识点 / 别名 / 错因标签”三个工作表。",
            "知识点路径使用“模块>子模块>知识点”写法；别名精确匹配到标准知识点；错因标签可维护导入题分时的错因。",
        ],
    ),
    TemplateSpec(
        filename="timetable_import_template.xlsx",
        title="课表导入模板",
        headers=["学期", "星期", "节次", "教师", "班级", "学科", "课程类型", "周次规则", "备注"],
        sample_rows=[["2025-2026 下学期", "1", "1", "李老师", "1班", "语文", "正课", "全周", ""]],
        instructions=["教师、班级、学科、课程类型需已存在。"],
    ),
    TemplateSpec(
        filename="admission_records_import_template.xlsx",
        title="录取数据导入模板",
        headers=["年份", "省份", "批次", "院校", "专业", "最低分", "最低位次", "学生类别", "数据来源说明"],
        sample_rows=[["2025", "广东", "本科批", "示例大学", "汉语言文学", "580", "32000", "普通", "示例"]],
        instructions=["用于录取数据维护与升学推荐。"],
    ),
    TemplateSpec(
        filename="enrollment_plans_import_template.xlsx",
        title="招生计划导入模板",
        headers=[
            "年份",
            "省份",
            "批次",
            "科类/模式",
            "院校",
            "院校代码",
            "专业组编号",
            "专业",
            "专业代码",
            "计划人数",
            "选科要求",
            "学费",
            "学制",
            "培养地点",
            "学生类别",
            "数据来源",
            "导入批次",
        ],
        sample_rows=[
            [
                "2026",
                "广东",
                "本科批",
                "物理类",
                "示例大学",
                "10561",
                "201",
                "软件工程",
                "080902",
                "120",
                "物理+化学",
                "6850 元/年",
                "4年",
                "广州校区",
                "普通生",
                "招生计划简章",
                "2026-广东-本科",
            ]
        ],
        instructions=["用于维护高考志愿招生计划库，专业或专业组编号至少填写一项。"],
    ),
    TemplateSpec(
        filename="evaluation_import_template.xlsx",
        title="评教导入模板",
        headers=["模板名称", "教师", "班级", "题目", "分值", "评价对象类型"],
        sample_rows=[["通用评教", "李老师", "1班", "教学目标清晰", "5", "学生"]],
        instructions=["用于教师评教数据导入。"],
    ),
    TemplateSpec(
        filename="adviser_quant_import_template.xlsx",
        title="班主任量化导入模板",
        headers=["教师", "班级", "学期", "月份", "量化项", "得分", "说明"],
        sample_rows=[["李老师", "1班", "2025-2026 下学期", "2026-03", "班级常规管理", "5", "示例"]],
        instructions=["用于班主任量化记录导入。"],
    ),
]


def _build_workbook(spec: TemplateSpec) -> Workbook:
    workbook = Workbook()
    intro_sheet = workbook.active
    intro_sheet.title = "说明"
    intro_sheet.append([spec.title])
    for instruction in spec.instructions:
        intro_sheet.append([instruction])

    data_sheet = workbook.create_sheet("数据")
    data_sheet.append(spec.headers)
    for row in spec.sample_rows:
        data_sheet.append(row)
    if spec.filename == "knowledge_base_import_template.xlsx":
        data_sheet.title = "知识点"
        alias_sheet = workbook.create_sheet("别名")
        alias_sheet.append(["科目", "标准知识点", "别名", "说明"])
        alias_sheet.append(["数学", "函数>基本初等函数>函数单调性", "单调性", "平台简称"])
        error_sheet = workbook.create_sheet("错因标签")
        error_sheet.append(["错因标签", "说明", "排序"])
        error_sheet.append(["概念不清", "基础概念理解不到位", "1"])
    return workbook


def generate_import_templates(settings: Settings) -> list[Path]:
    generated_paths: list[Path] = []
    settings.templates_dir.mkdir(parents=True, exist_ok=True)

    for spec in TEMPLATE_SPECS:
        path = settings.templates_dir / spec.filename
        workbook = _build_workbook(spec)
        workbook.save(path)
        generated_paths.append(path)

    return generated_paths
