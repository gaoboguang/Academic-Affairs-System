from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.config import Settings
from app.utils.parsers import make_timestamped_filename, relative_to_project


@dataclass
class RowError:
    row_number: int
    values: dict[str, Any]
    message: str
    column_name: str | None = None
    field_name: str | None = None
    raw_value: Any | None = None
    suggestion: str | None = None


IMPORT_STATUS_PENDING = "pending"
IMPORT_STATUS_RUNNING = "running"
IMPORT_STATUS_SUCCESS = "success"
IMPORT_STATUS_FAILED = "failed"
IMPORT_STATUS_PARTIALLY_FAILED = "partially_failed"
IMPORT_STATUS_ROLLED_BACK = "rolled_back"

IMPORT_STATUS_ALIASES = {
    "processing": IMPORT_STATUS_RUNNING,
    "completed": IMPORT_STATUS_SUCCESS,
    "completed_with_unresolved": IMPORT_STATUS_PARTIALLY_FAILED,
    "partial_success": IMPORT_STATUS_PARTIALLY_FAILED,
}


def normalize_import_status(value: str | None) -> str:
    if not value:
        return IMPORT_STATUS_SUCCESS
    return IMPORT_STATUS_ALIASES.get(value, value)


def resolve_import_status(
    *,
    total_rows: int,
    success_rows: int,
    failed_rows: int,
    unresolved_rows: int = 0,
) -> str:
    if failed_rows <= 0 and unresolved_rows <= 0:
        return IMPORT_STATUS_SUCCESS
    if success_rows <= 0 and failed_rows >= total_rows:
        return IMPORT_STATUS_FAILED
    return IMPORT_STATUS_PARTIALLY_FAILED


def build_row_error(row_number: int, values: dict[str, Any], message: str) -> RowError:
    field_name, raw_value, suggestion = _infer_error_detail(values, message)
    return RowError(
        row_number=row_number,
        values=values,
        message=message,
        column_name=field_name,
        field_name=field_name,
        raw_value=raw_value,
        suggestion=suggestion,
    )


def build_error_preview(errors: list[RowError], limit: int = 3) -> list[str]:
    return [f"第 {item.row_number} 行：{item.message}" for item in errors[:limit]]


def _infer_error_detail(values: dict[str, Any], message: str) -> tuple[str | None, Any | None, str | None]:
    if message.endswith("不能为空"):
        field_name = message.removesuffix("不能为空").strip()
        return field_name, values.get(field_name), f"请补齐“{field_name}”后重新导入。"

    for marker in ("无法识别: ", "不存在: "):
        if marker in message:
            field_name, raw_value = message.split(marker, 1)
            field_name = field_name.strip()
            raw_value = raw_value.strip()
            if marker == "无法识别: ":
                suggestion = f"请确认“{raw_value}”是否已在基础数据或字典中维护。"
            else:
                suggestion = f"请先维护“{raw_value}”，或按模板填写系统中已有的“{field_name}”。"
            return field_name, raw_value, suggestion

    if "重复" in message:
        return None, None, "请检查导入文件中是否有重复记录，或调整导入策略后重试。"
    if "不匹配" in message:
        return None, None, "请核对同一行内的身份、姓名、班级、考试或学期是否一致。"
    if "格式错误" in message or "无法识别" in message:
        return None, None, "请按模板说明调整字段格式后重新导入。"
    return None, None, None


def read_template_rows(content: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    worksheet = workbook["数据"] if "数据" in workbook.sheetnames else workbook.active

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(cell).strip() if cell is not None else "" for cell in (header_row or [])]

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row_map = {headers[index]: value for index, value in enumerate(values)}
        if not any(value not in (None, "") for value in row_map.values()):
            continue
        rows.append((row_number, row_map))
    return headers, rows


def save_error_report(
    *,
    settings: Settings,
    prefix: str,
    headers: list[str],
    errors: list[RowError],
) -> str | None:
    if not errors:
        return None

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "错误报告"
    sheet.append(["行号", "列名", "字段名", "原始值", "错误原因", "建议修复", *headers])
    for error in errors:
        row = [
            error.row_number,
            error.column_name or "",
            error.field_name or "",
            error.raw_value,
            error.message,
            error.suggestion or "",
        ]
        for header in headers:
            row.append(error.values.get(header))
        sheet.append(row)

    filename = make_timestamped_filename(prefix, ".xlsx")
    path = settings.logs_dir / filename
    workbook.save(path)
    return relative_to_project(path, settings.project_root)
