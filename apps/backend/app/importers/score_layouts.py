from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session
from xlrd import open_workbook

from app.models import Exam, ExamSubject, Subject
from app.utils.parsers import clean_text


STANDARD_SCORE_HEADERS = [
    "考试名称",
    "学号",
    "姓名",
    "班级",
    "科目",
    "分数",
    "缺考标记",
    "备注",
    "原始分",
    "赋分",
    "成绩口径",
]

FIELD_LABELS = {
    "exam_name": "考试名称",
    "student_no": "学号",
    "student_name": "姓名",
    "class_name": "班级",
    "subject": "科目",
    "score": "分数",
    "original_score": "原始分",
    "converted_score": "赋分",
    "score_value_type": "成绩口径",
    "absent_flag": "缺考标记",
    "note": "备注",
}

METADATA_FIELD_LABELS = {
    "source_total_score": "源总分",
    "source_class_rank": "源班级名次",
    "source_school_rank": "源学校名次",
    "source_student_no": "源学籍号",
    "source_exam_no": "考生号",
}

REQUIRED_ROLES = {"student_no"}
LONG_REQUIRED_ROLES = {"student_no", "subject", "score"}

FIELD_ALIASES = {
    "exam_name": {"考试名称", "考试", "测验名称", "考试场次", "考试批次"},
    "student_no": {"学号", "考号", "准考证号", "学籍号", "学生编号", "学生代码", "账号", "报名号"},
    "student_name": {"姓名", "学生姓名", "考生姓名", "名字"},
    "class_name": {"班级", "班级名称", "行政班", "原班级", "所在班级"},
    "subject": {"科目", "学科", "科目名称", "课程", "课程名称"},
    "original_score": {"原始分", "原始成绩", "卷面分"},
    "converted_score": {"赋分", "等级分", "等级成绩", "赋分成绩"},
    "score_value_type": {"成绩口径", "分数口径", "成绩类型", "分数类型", "赋分类型"},
    "score": {"分数", "成绩", "得分", "原始分", "卷面分", "赋分", "等级分", "等级成绩", "总得分"},
    "absent_flag": {"缺考标记", "缺考", "状态", "考试状态", "成绩状态"},
    "note": {"备注", "说明", "评语", "异常说明"},
}

METADATA_ALIASES = {
    "source_total_score": {"总分", "总成绩", "合计", "总计", "源总分"},
    "source_class_rank": {"班级名次", "班名次", "班排", "班级排名", "源班级名次"},
    "source_school_rank": {"年级名次", "级名次", "年排", "校排", "学校名次", "校名次", "源学校名次"},
    "source_student_no": {"源学籍号", "原学籍号", "新教育号", "平台学籍号"},
    "source_exam_no": {"考生号", "准考证号", "源考号", "平台考号"},
}

SUMMARY_ALIASES = {
    "总分",
    "总成绩",
    "合计",
    "总计",
    "排名",
    "名次",
    "班级名次",
    "班名次",
    "班排",
    "年级名次",
    "级名次",
    "年排",
    "校排",
    "百分位",
    "平均分",
    "标准分",
}

ABSENT_VALUES = {"缺考", "缺", "未考", "旷考", "缓考", "免考", "absent", "missing"}
SUBJECT_SUFFIXES = ("等级成绩", "等级分", "赋分成绩", "成绩", "分数", "得分", "原始成绩", "原始分", "卷面分", "等级", "赋分")
SCORE_VALUE_TYPES = {"original", "converted", "display"}


@dataclass
class ScoreSourceSheet:
    name: str
    rows: list[list[Any]]


@dataclass
class ScoreImportMapping:
    layout_type: str
    sheet_name: str | None = None
    header_row: int = 1
    field_mapping: dict[str, str] = field(default_factory=dict)
    subject_mapping: dict[str, str] = field(default_factory=dict)
    subject_score_types: dict[str, str] = field(default_factory=dict)
    ignored_columns: list[str] = field(default_factory=list)
    metadata_mapping: dict[str, str] = field(default_factory=dict)

    @classmethod
    def from_dict(cls, value: dict[str, Any] | str | None) -> "ScoreImportMapping | None":
        if value is None:
            return None
        if isinstance(value, str):
            value = json.loads(value)
        return cls(
            layout_type=str(value.get("layout_type") or "wide"),
            sheet_name=clean_text(value.get("sheet_name")),
            header_row=max(int(value.get("header_row") or 1), 1),
            field_mapping={str(k): str(v) for k, v in (value.get("field_mapping") or {}).items() if v},
            subject_mapping={str(k): str(v) for k, v in (value.get("subject_mapping") or {}).items() if v},
            subject_score_types={
                str(k): str(v)
                for k, v in (value.get("subject_score_types") or value.get("subject_score_type_json") or {}).items()
                if v in SCORE_VALUE_TYPES
            },
            ignored_columns=[str(item) for item in (value.get("ignored_columns") or []) if item],
            metadata_mapping={str(k): str(v) for k, v in (value.get("metadata_mapping") or {}).items() if v},
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "layout_type": self.layout_type,
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "field_mapping": self.field_mapping,
            "subject_mapping": self.subject_mapping,
            "subject_score_types": self.subject_score_types,
            "ignored_columns": self.ignored_columns,
            "metadata_mapping": self.metadata_mapping,
        }


@dataclass
class ScoreImportPreview:
    source_filename: str | None
    sheet_name: str | None
    header_row: int
    layout_type: str
    confidence: float
    messages: list[str]
    columns: list[dict[str, Any]]
    sample_rows: list[dict[str, Any]]
    normalized_preview: list[dict[str, Any]]
    mapping: ScoreImportMapping
    import_ready: bool
    source_row_count: int
    detected_record_count: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_filename": self.source_filename,
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "layout_type": self.layout_type,
            "confidence": self.confidence,
            "messages": self.messages,
            "columns": self.columns,
            "sample_rows": self.sample_rows,
            "normalized_preview": self.normalized_preview,
            "mapping": self.mapping.to_dict(),
            "import_ready": self.import_ready,
            "source_row_count": self.source_row_count,
            "detected_record_count": self.detected_record_count,
        }


class ScoreLayoutAdapter:
    def __init__(self, session: Session, exam: Exam) -> None:
        self.session = session
        self.exam = exam
        self.exam_subjects = list(
            session.scalars(select(ExamSubject).where(ExamSubject.exam_id == exam.id, ExamSubject.is_active.is_(True)))
        )
        subject_ids = [item.subject_id for item in self.exam_subjects]
        self.subjects = list(session.scalars(select(Subject).where(Subject.id.in_(subject_ids))).all()) if subject_ids else []
        self.subject_alias_map = self._build_subject_alias_map()

    def preview(
        self,
        *,
        filename: str | None,
        content: bytes,
        mapping: ScoreImportMapping | None = None,
    ) -> ScoreImportPreview:
        sheets = load_score_sheets(filename, content)
        if not sheets:
            raise ValueError("成绩文件为空，未读取到可识别的工作表。")

        if mapping is None:
            sheet, mapping, confidence, messages = self._detect_mapping(sheets)
        else:
            sheet = self._select_sheet(sheets, mapping.sheet_name)
            confidence = 0.95
            messages = ["已按所选平台模板生成预览。"]

        source_rows = self._read_mapped_rows(sheet, mapping)
        normalized_rows: list[tuple[int, dict[str, Any]]] = []
        import_ready = True
        try:
            normalized_rows = self.normalize_rows(sheet, mapping)
            readiness_error = self._validate_mapping_ready(mapping)
            if readiness_error:
                import_ready = False
                messages.append(readiness_error)
        except ValueError as exc:
            import_ready = False
            messages.append(str(exc))

        return ScoreImportPreview(
            source_filename=filename,
            sheet_name=sheet.name,
            header_row=mapping.header_row,
            layout_type=mapping.layout_type,
            confidence=round(confidence, 2),
            messages=messages,
            columns=self._build_column_preview(sheet, mapping),
            sample_rows=[_stringify_row(row) for _, row in source_rows[:8]],
            normalized_preview=[_stringify_row(row) for _, row in normalized_rows[:12]],
            mapping=mapping,
            import_ready=import_ready,
            source_row_count=len(source_rows),
            detected_record_count=len(normalized_rows),
        )

    def normalize_content(
        self,
        *,
        filename: str | None,
        content: bytes,
        mapping: ScoreImportMapping,
    ) -> tuple[list[str], list[tuple[int, dict[str, Any]]], dict[str, Any]]:
        sheets = load_score_sheets(filename, content)
        sheet = self._select_sheet(sheets, mapping.sheet_name)
        error = self._validate_mapping_ready(mapping)
        if error:
            raise ValueError(error)
        rows = self.normalize_rows(sheet, mapping)
        summary = {
            "layout_type": mapping.layout_type,
            "sheet_name": sheet.name,
            "header_row": mapping.header_row,
            "source_row_count": len(self._read_mapped_rows(sheet, mapping)),
            "normalized_record_count": len(rows),
            "field_mapping": mapping.field_mapping,
            "subject_mapping": mapping.subject_mapping,
            "subject_score_types": mapping.subject_score_types,
            "ignored_columns": mapping.ignored_columns,
            "metadata_mapping": mapping.metadata_mapping,
        }
        return STANDARD_SCORE_HEADERS, rows, summary

    def normalize_rows(
        self,
        sheet: ScoreSourceSheet,
        mapping: ScoreImportMapping,
    ) -> list[tuple[int, dict[str, Any]]]:
        rows = self._read_mapped_rows(sheet, mapping)
        if mapping.layout_type == "long":
            return self._normalize_long_rows(rows, mapping)
        if mapping.layout_type == "wide":
            return self._normalize_wide_rows(rows, mapping)
        raise ValueError(f"不支持的成绩单布局: {mapping.layout_type}")

    def _detect_mapping(
        self,
        sheets: list[ScoreSourceSheet],
    ) -> tuple[ScoreSourceSheet, ScoreImportMapping, float, list[str]]:
        best: tuple[float, ScoreSourceSheet, int, list[str]] | None = None
        for sheet in sheets:
            for row_index, row in enumerate(sheet.rows[:20], start=1):
                headers = _make_headers(row)
                if not any(header for header in headers):
                    continue
                roles, subject_columns, subject_score_types, ignored_columns, metadata_columns = self._infer_columns(headers)
                score = self._score_header_row(roles, subject_columns, ignored_columns)
                score += len(metadata_columns) * 0.5
                if best is None or score > best[0]:
                    best = (score, sheet, row_index, headers)

        if best is None:
            raise ValueError("未找到可识别的成绩表头。")

        _, sheet, header_row, headers = best
        roles, subject_columns, subject_score_types, ignored_columns, metadata_columns = self._infer_columns(headers)
        if roles.get("subject") and roles.get("score"):
            layout_type = "long"
        else:
            layout_type = "wide"

        if layout_type == "long":
            subject_mapping: dict[str, str] = {}
        else:
            subject_mapping = {header: subject for header, subject in subject_columns.items()}

        mapping = ScoreImportMapping(
            layout_type=layout_type,
            sheet_name=sheet.name,
            header_row=header_row,
            field_mapping=roles,
            subject_mapping=subject_mapping,
            subject_score_types=subject_score_types,
            ignored_columns=ignored_columns,
            metadata_mapping=metadata_columns,
        )
        confidence = self._confidence(mapping)
        messages = self._detection_messages(mapping, confidence)
        return sheet, mapping, confidence, messages

    def _infer_columns(
        self,
        headers: list[str],
    ) -> tuple[dict[str, str], dict[str, str], dict[str, str], list[str], dict[str, str]]:
        roles: dict[str, str] = {}
        subject_columns: dict[str, str] = {}
        subject_score_types: dict[str, str] = {}
        ignored_columns: list[str] = []
        metadata_columns: dict[str, str] = {}

        for header in headers:
            if not header:
                continue
            normalized = normalize_header(header)
            role = self._infer_role(normalized)
            if role and role not in roles:
                roles[role] = header
                continue
            metadata_role = self._infer_metadata_role(normalized)
            if metadata_role and metadata_role not in metadata_columns:
                metadata_columns[metadata_role] = header
                continue
            subject_name = self._infer_subject_header(normalized)
            if subject_name:
                subject_columns[header] = subject_name
                subject_score_types[header] = _infer_score_value_type(header)
                continue
            if normalized in {normalize_header(item) for item in SUMMARY_ALIASES}:
                ignored_columns.append(header)

        if roles.get("score") in subject_columns:
            subject_columns.pop(roles["score"], None)
            subject_score_types.pop(roles["score"], None)
        if "score" not in roles:
            if "converted_score" in roles:
                roles["score"] = roles["converted_score"]
            elif "original_score" in roles:
                roles["score"] = roles["original_score"]
        return roles, subject_columns, subject_score_types, ignored_columns, metadata_columns

    def _infer_role(self, normalized_header: str) -> str | None:
        if normalized_header in {normalize_header(item) for item in SUMMARY_ALIASES}:
            return None
        for role, aliases in FIELD_ALIASES.items():
            normalized_aliases = {normalize_header(alias) for alias in aliases}
            if normalized_header in normalized_aliases:
                return role
        return None

    def _infer_metadata_role(self, normalized_header: str) -> str | None:
        for role, aliases in METADATA_ALIASES.items():
            if normalized_header in {normalize_header(alias) for alias in aliases}:
                return role
        return None

    def _infer_subject_header(self, normalized_header: str) -> str | None:
        for alias, subject_name in sorted(self.subject_alias_map.items(), key=lambda item: len(item[0]), reverse=True):
            if normalized_header == alias:
                return subject_name
            if normalized_header.startswith(alias):
                tail = normalized_header[len(alias) :]
                if tail in {"等级成绩", "等级分", "赋分成绩", "赋分", "成绩", "分数", "得分", "原始成绩", "原始分", "卷面分"}:
                    return subject_name
                if tail in {
                    "等级",
                    "校内排名",
                    "县区排名",
                    "全市排名",
                    "班级名次",
                    "年级名次",
                }:
                    return None
        candidate = normalized_header
        for suffix in SUBJECT_SUFFIXES:
            normalized_suffix = normalize_header(suffix)
            if candidate.endswith(normalized_suffix):
                candidate = candidate[: -len(normalized_suffix)]
        return self.subject_alias_map.get(candidate)

    def _build_subject_alias_map(self) -> dict[str, str]:
        result: dict[str, str] = {}
        extra_aliases = {
            "思想政治": "政治",
            "道法": "政治",
            "道德与法治": "政治",
            "英语听说": "英语",
            "外语": "英语",
            "生物学": "生物",
        }
        for subject in self.subjects:
            result[normalize_header(subject.name)] = subject.name
            result[normalize_header(subject.code)] = subject.name
        for alias, target in extra_aliases.items():
            if normalize_header(target) in result:
                result[normalize_header(alias)] = result[normalize_header(target)]
        return result

    def _score_header_row(
        self,
        roles: dict[str, str],
        subject_columns: dict[str, str],
        ignored_columns: list[str],
    ) -> float:
        score = len(roles) * 3 + len(subject_columns) * 2 + len(ignored_columns) * 0.2
        if "student_no" in roles:
            score += 5
        if {"subject", "score"} <= set(roles):
            score += 5
        if subject_columns:
            score += 4
        return score

    def _confidence(self, mapping: ScoreImportMapping) -> float:
        score = 0.35
        if "student_no" in mapping.field_mapping:
            score += 0.25
        if "student_name" in mapping.field_mapping:
            score += 0.08
        if mapping.layout_type == "long" and {"subject", "score"} <= set(mapping.field_mapping):
            score += 0.25
        if mapping.layout_type == "wide" and mapping.subject_mapping:
            score += 0.25
        if "class_name" in mapping.field_mapping:
            score += 0.04
        if mapping.ignored_columns:
            score += 0.03
        if "student_no" not in mapping.field_mapping:
            score = min(score, 0.45)
        return min(score, 0.98)

    def _detection_messages(self, mapping: ScoreImportMapping, confidence: float) -> list[str]:
        messages = [
            f"已识别为{'宽表' if mapping.layout_type == 'wide' else '长表'}成绩单，表头在第 {mapping.header_row} 行。"
        ]
        if confidence < 0.7:
            messages.append("识别置信度偏低，请先确认字段映射后再导入。")
        readiness_error = self._validate_mapping_ready(mapping)
        if readiness_error:
            messages.append(readiness_error)
        if mapping.metadata_mapping:
            messages.append("已识别平台总分或名次列，将作为原始核对信息保留，系统仍会重新计算本地快照。")
        if any(value == "converted" for value in mapping.subject_score_types.values()):
            messages.append("已识别赋分列；同一科目同时存在原始分和赋分时，将优先用赋分参与展示、总分和名次。")
        if mapping.layout_type == "wide" and mapping.ignored_columns:
            messages.append("未参与成绩导入的统计列已默认忽略。")
        return messages

    def _validate_mapping_ready(self, mapping: ScoreImportMapping) -> str | None:
        missing = [FIELD_LABELS[item] for item in REQUIRED_ROLES if item not in mapping.field_mapping]
        if missing:
            return f"必须指定唯一编号列：{', '.join(missing)}。"
        if mapping.layout_type == "long":
            missing_long = [FIELD_LABELS[item] for item in LONG_REQUIRED_ROLES if item not in mapping.field_mapping]
            if missing_long:
                return f"长表导入必须指定：{', '.join(missing_long)}。"
        if mapping.layout_type == "wide" and not mapping.subject_mapping:
            return "宽表导入必须至少映射一个科目成绩列。"
        return None

    def _build_column_preview(self, sheet: ScoreSourceSheet, mapping: ScoreImportMapping) -> list[dict[str, Any]]:
        reverse_roles = {header: role for role, header in mapping.field_mapping.items()}
        reverse_metadata_roles = {header: role for role, header in mapping.metadata_mapping.items()}
        if 1 <= mapping.header_row <= len(sheet.rows):
            ordered_headers = _make_headers(sheet.rows[mapping.header_row - 1])
        else:
            ordered_headers = list(dict.fromkeys(list(mapping.field_mapping.values()) + list(mapping.subject_mapping.keys())))
        return [
            {
                "header": header,
                "role": reverse_roles.get(header),
                "metadata_role": reverse_metadata_roles.get(header),
                "subject_name": mapping.subject_mapping.get(header),
                "score_value_type": mapping.subject_score_types.get(header) or _infer_score_value_type(header),
                "ignored": header in mapping.ignored_columns,
            }
            for header in ordered_headers
        ]

    def _read_mapped_rows(self, sheet: ScoreSourceSheet, mapping: ScoreImportMapping) -> list[tuple[int, dict[str, Any]]]:
        if mapping.header_row < 1 or mapping.header_row > len(sheet.rows):
            raise ValueError("表头行超出文件范围，请重新选择表头行。")
        headers = _make_headers(sheet.rows[mapping.header_row - 1])
        rows: list[tuple[int, dict[str, Any]]] = []
        for source_row_number, values in enumerate(sheet.rows[mapping.header_row :], start=mapping.header_row + 1):
            row_map = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
            if not any(value not in (None, "") for value in row_map.values()):
                continue
            rows.append((source_row_number, row_map))
        return rows

    def _normalize_long_rows(
        self,
        rows: list[tuple[int, dict[str, Any]]],
        mapping: ScoreImportMapping,
    ) -> list[tuple[int, dict[str, Any]]]:
        normalized: list[tuple[int, dict[str, Any]]] = []
        for row_number, row in rows:
            score_value = row.get(mapping.field_mapping["score"])
            score_header = mapping.field_mapping["score"]
            score_value_type = _normalize_score_value_type(
                _mapped_value(row, mapping, "score_value_type")
                or _infer_score_value_type(score_header)
            )
            original_score = _mapped_value(row, mapping, "original_score")
            converted_score = _mapped_value(row, mapping, "converted_score")
            if score_value_type == "converted" and converted_score in (None, ""):
                converted_score = score_value
            if score_value_type != "converted" and original_score in (None, ""):
                original_score = score_value
            absent_value = _mapped_value(row, mapping, "absent_flag")
            if _looks_absent(score_value) and not absent_value:
                absent_value = score_value
                score_value = None
            normalized.append(
                (
                    row_number,
                    {
                        "考试名称": _mapped_value(row, mapping, "exam_name") or self.exam.name,
                        "学号": _mapped_value(row, mapping, "student_no"),
                        "姓名": _mapped_value(row, mapping, "student_name"),
                        "班级": _mapped_value(row, mapping, "class_name"),
                        "科目": _mapped_value(row, mapping, "subject"),
                        "分数": score_value,
                        "缺考标记": absent_value,
                        "备注": _mapped_value(row, mapping, "note"),
                        "原始分": original_score,
                        "赋分": converted_score,
                        "成绩口径": score_value_type,
                        **_metadata_values(row, mapping),
                    },
                )
            )
        return normalized

    def _normalize_wide_rows(
        self,
        rows: list[tuple[int, dict[str, Any]]],
        mapping: ScoreImportMapping,
    ) -> list[tuple[int, dict[str, Any]]]:
        normalized: list[tuple[int, dict[str, Any]]] = []
        for row_number, row in rows:
            subject_columns: dict[str, list[tuple[str, str]]] = {}
            for column_name, subject_name in mapping.subject_mapping.items():
                if not subject_name:
                    continue
                score_value_type = _normalize_score_value_type(
                    mapping.subject_score_types.get(column_name) or _infer_score_value_type(column_name)
                )
                subject_columns.setdefault(subject_name, []).append((column_name, score_value_type))

            for subject_name, columns in subject_columns.items():
                original_value = _first_value(
                    row.get(column_name)
                    for column_name, score_value_type in columns
                    if score_value_type in {"original", "display"}
                )
                converted_value = _first_value(
                    row.get(column_name)
                    for column_name, score_value_type in columns
                    if score_value_type == "converted"
                )
                score_value = converted_value if converted_value not in (None, "") else original_value
                if score_value in (None, "") and original_value in (None, "") and converted_value in (None, ""):
                    continue
                score_value_type = "converted" if converted_value not in (None, "") else "original"
                absent_value = _mapped_value(row, mapping, "absent_flag")
                if _looks_absent(score_value):
                    absent_value = score_value
                    score_value = None
                normalized.append(
                    (
                        row_number,
                        {
                            "考试名称": _mapped_value(row, mapping, "exam_name") or self.exam.name,
                            "学号": _mapped_value(row, mapping, "student_no"),
                            "姓名": _mapped_value(row, mapping, "student_name"),
                            "班级": _mapped_value(row, mapping, "class_name"),
                            "科目": subject_name,
                            "分数": score_value,
                            "缺考标记": absent_value,
                            "备注": _mapped_value(row, mapping, "note"),
                            "原始分": original_value,
                            "赋分": converted_value,
                            "成绩口径": score_value_type,
                            **_metadata_values(row, mapping),
                        },
                    )
                )
        return normalized

    def _select_sheet(self, sheets: list[ScoreSourceSheet], sheet_name: str | None) -> ScoreSourceSheet:
        if sheet_name:
            for sheet in sheets:
                if sheet.name == sheet_name:
                    return sheet
        return sheets[0]


def load_score_sheets(filename: str | None, content: bytes) -> list[ScoreSourceSheet]:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return [ScoreSourceSheet(name="CSV", rows=_load_csv_rows(content))]
    if lowered.endswith(".xls"):
        return _load_xls_sheets(content)
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        try:
            return [ScoreSourceSheet(name="CSV", rows=_load_csv_rows(content))]
        except Exception:
            raise ValueError("仅支持 Excel 或 CSV 成绩文件。") from exc
    return [
        ScoreSourceSheet(name=worksheet.title, rows=[list(row) for row in worksheet.iter_rows(values_only=True)])
        for worksheet in workbook.worksheets
    ]


def normalize_header(value: Any) -> str:
    text = clean_text(value) or ""
    return re.sub(r"[\s_\-—:：/\\（）()【】\\[\\].。,，]+", "", text).lower()


def _make_headers(row: list[Any]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        base = clean_text(value) or f"空列{index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _mapped_value(row: dict[str, Any], mapping: ScoreImportMapping, role: str) -> Any:
    header = mapping.field_mapping.get(role)
    if not header:
        return None
    return row.get(header)


def _metadata_values(row: dict[str, Any], mapping: ScoreImportMapping) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for role, header in mapping.metadata_mapping.items():
        label = METADATA_FIELD_LABELS.get(role)
        if label:
            result[label] = row.get(header)
    return result


def _infer_score_value_type(header: Any) -> str:
    normalized = normalize_header(header)
    if any(marker in normalized for marker in ("赋分", "等级分", "等级成绩")):
        return "converted"
    if any(marker in normalized for marker in ("原始", "卷面")):
        return "original"
    return "original"


def _normalize_score_value_type(value: Any) -> str:
    normalized = normalize_header(value)
    if normalized in {"converted", "afterconversion", "赋分", "赋分后", "等级分", "等级成绩"}:
        return "converted"
    if normalized in {"display", "最终分", "展示分"}:
        return "display"
    return "original"


def _first_value(values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _looks_absent(value: Any) -> bool:
    text = clean_text(value)
    return bool(text and normalize_header(text) in {normalize_header(item) for item in ABSENT_VALUES})


def _stringify_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: (value.isoformat() if hasattr(value, "isoformat") else value) for key, value in row.items()}


def _load_csv_rows(content: bytes) -> list[list[str]]:
    text: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 文件编码无法识别，请另存为 UTF-8 后重试。")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(StringIO(text), dialect)
    return [list(row) for row in reader]


def _load_xls_sheets(content: bytes) -> list[ScoreSourceSheet]:
    try:
        workbook = open_workbook(file_contents=content)
    except Exception as exc:
        raise ValueError("旧版 xls 成绩文件无法读取，请另存为 Excel 工作簿后重试。") from exc
    sheets: list[ScoreSourceSheet] = []
    for sheet in workbook.sheets():
        rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            rows.append([sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)])
        sheets.append(ScoreSourceSheet(name=sheet.name, rows=rows))
    return sheets
