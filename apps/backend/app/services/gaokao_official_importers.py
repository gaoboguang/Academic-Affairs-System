from __future__ import annotations

import hashlib
import html
import re
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.importers.base import RowError, build_row_error, resolve_import_status, save_error_report
from app.models import AdmissionRecord, College, GaokaoImportRun, GaokaoSourceDocument, Major
from app.repositories.recommendations import ensure_college_major, get_admission_record_by_key
from app.services.gaokao_imports import (
    DEFAULT_PARSER_VERSION,
    GAOKAO_IMPORT_STATUS_PENDING,
    GAOKAO_SOURCE_STATUS_FILE_READY,
    get_gaokao_source_document,
    register_gaokao_local_file,
    seed_default_gaokao_sources,
)
from app.utils.parsers import relative_to_project


B1_PARSER_VERSION = "e6-official-core-20260426"
B1_DATA_VERSION_LABEL = "shandong_official_core_20260426"
B1_USER_AGENT = "local-edu-tool-b1-official-import/2026-04-25"
B1_COVERAGE_DOC = "docs/gaokao-shandong-2020-2025-coverage.md"
B1_IMPORT_YEARS = (2023, 2024, 2025)
B1_COVERAGE_YEARS = (2020, 2021, 2022, 2023, 2024, 2025)
B1_SOURCE_TYPES = ("admission_result", "score_rank_segment", "score_line")
_COLUMN_CACHE: dict[tuple[str, str], set[str]] = {}


DEFAULT_OFFICIAL_FILE_URLS: dict[tuple[int, str], str] = {
    (2023, "admission_result"): "https://www.sdzk.cn/Floadup/file/20230719/6382538122655052185031609.xls",
    (2024, "admission_result"): "https://www.sdzk.cn/Floadup/file/20240719/6385700532268895241675882.xls",
    (2025, "admission_result"): "https://www.sdzk.cn/Floadup/file/20250719/6388855130412530367357143.xls",
    (2023, "score_rank_segment"): "https://www.sdzk.cn/Floadup/file/20230625/6382330466178390033845468.xls",
    (2024, "score_rank_segment"): "https://www.sdzk.cn/Floadup/file/20240625/6385492724297110442689837.xls",
    (2025, "score_rank_segment"): "https://www.sdzk.cn/Floadup/file/20250625/6388646133710894671069456.xls",
    (2020, "score_rank_segment"): "https://www.sdzk.cn/Floadup/file/20200726/6373137724431062332329137.xls",
    (2021, "score_rank_segment"): "https://www.sdzk.cn/Floadup/file/20210625/6376023552787296505320142.xls",
    (2022, "score_rank_segment"): "https://www.sdzk.cn/Floadup/file/20220626/6379183893483079468961719.xls",
    (2020, "score_line"): "https://www.sdzk.cn/Floadup/image/20250618/6388585391998227954417320.png",
    (2021, "score_line"): "https://www.sdzk.cn/Floadup/image/20250618/6388585383987448818844597.png",
    (2022, "score_line"): "https://www.sdzk.cn/Floadup/image/20250618/6388583976098931994629755.png",
    (2023, "score_line"): "https://www.sdzk.cn/Floadup/image/20250618/6388585429679669883269908.png",
    (2024, "score_line"): "https://www.sdzk.cn/Floadup/image/20250618/6388585458310919561839075.png",
    (2025, "score_line"): "https://www.sdzk.cn/Floadup/file/20250625/6388646601596985266621362.pdf",
}


PUBLISHED_DATES: dict[tuple[int, str], str] = {
    (2023, "admission_result"): "2023-07-19",
    (2024, "admission_result"): "2024-07-19",
    (2025, "admission_result"): "2025-07-19",
    (2023, "score_rank_segment"): "2023-06-25",
    (2024, "score_rank_segment"): "2024-06-25",
    (2025, "score_rank_segment"): "2025-06-25",
    (2020, "score_rank_segment"): "2020-07-26",
    (2021, "score_rank_segment"): "2021-06-25",
    (2022, "score_rank_segment"): "2022-06-26",
    (2020, "score_line"): "2020-07-26",
    (2021, "score_line"): "2021-06-25",
    (2022, "score_line"): "2022-06-25",
    (2023, "score_line"): "2023-06-25",
    (2024, "score_line"): "2024-06-25",
    (2025, "score_line"): "2025-06-25",
}


SCORE_LINE_ROWS: dict[int, tuple[tuple[str, str, str, int, str], ...]] = {
    2020: (
        ("普通类", "常规批", "special_type_control", 532, "特殊类型招生控制线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "first_section", 449, "普通类一段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "second_section", 150, "普通类二段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "3plus2_vocational_qualification", 399, "3+2 对口贯通分段培养高职志愿填报资格线；根据官方分数线图片校对。"),
        ("普通类", "高水平运动队", "high_level_sports_preferential_65", 291, "享受 65% 优惠政策的高水平运动员文化录取控制线。"),
        ("艺术类", "本科", "undergrad_culture_literature_directing_broadcast_photo", 381, "文学编导、播音主持、摄影类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_art_music_calligraphy", 314, "美术、音乐、书法类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_dance_performance_fashion", 291, "舞蹈、影视戏剧表演、服装表演类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "专科", "junior_culture_control", 150, "艺术类专科文化控制线；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "first_section_composite", 561, "体育类一段线，划线成绩为综合分；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "second_section_composite", 457, "体育类二段线，划线成绩为综合分；根据官方分数线图片校对。"),
    ),
    2021: (
        ("普通类", "常规批", "special_type_control", 518, "特殊类型招生控制线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "first_section", 444, "普通类一段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "second_section", 150, "普通类二段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "3plus2_vocational_qualification", 394, "3+2 对口贯通分段培养高职志愿填报资格线；根据官方分数线图片校对。"),
        ("普通类", "高水平运动队", "high_level_sports_preferential_65", 288, "享受 65% 优惠政策的高水平运动员文化录取控制线。"),
        ("艺术类", "本科", "undergrad_culture_literature_directing_broadcast_photo", 444, "文学编导、播音主持、摄影类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_art_music_calligraphy_aviation", 333, "美术、音乐、书法、航空服务艺术类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_dance_performance_opera_fashion", 288, "舞蹈、影视戏剧表演、服装表演类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "专科", "junior_culture_control", 150, "艺术类专科文化控制线；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "first_section_composite", 569, "体育类一段线，划线成绩为综合分；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "second_section_composite", 470, "体育类二段线，划线成绩为综合分；根据官方分数线图片校对。"),
    ),
    2022: (
        ("普通类", "常规批", "special_type_control", 513, "特殊类型招生控制线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "first_section", 437, "普通类一段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "second_section", 150, "普通类二段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "3plus2_vocational_qualification", 387, "3+2 对口贯通分段培养高职志愿填报资格线；根据官方分数线图片校对。"),
        ("普通类", "高水平运动队", "high_level_sports_preferential_65", 284, "享受 65% 优惠政策的高水平运动员文化录取控制线。"),
        ("艺术类", "本科", "undergrad_culture_literature_directing_broadcast_photo", 437, "文学编导、播音主持、摄影类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_art_music_calligraphy_aviation", 327, "美术、音乐、书法、航空服务艺术类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_dance_performance_opera_fashion", 284, "舞蹈、影视戏剧表演、服装表演类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "专科", "junior_culture_control", 150, "艺术类专科文化控制线；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "first_section_composite", 583, "体育类一段线，划线成绩为综合分；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "second_section_composite", 474, "体育类二段线，划线成绩为综合分；根据官方分数线图片校对。"),
    ),
    2023: (
        ("普通类", "常规批", "special_type_control", 520, "特殊类型招生控制线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "first_section", 443, "普通类一段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "second_section", 150, "普通类二段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "3plus2_vocational_qualification", 393, "3+2 对口贯通分段培养高职志愿填报资格线；根据官方分数线图片校对。"),
        ("普通类", "高水平运动队", "high_level_sports_world_class", 443, "世界一流大学建设高校高水平运动队本科文化控制线。"),
        ("普通类", "高水平运动队", "high_level_sports_other", 287, "其他高校高水平运动队本科文化控制线。"),
        ("艺术类", "本科", "undergrad_culture_literature_directing_broadcast_photo", 443, "文学编导类、播音主持类、摄影类本科文化控制线。"),
        ("艺术类", "本科", "undergrad_culture_art_music_calligraphy", 332, "美术类、音乐类、书法类、航空服务艺术类本科文化控制线。"),
        ("艺术类", "本科", "undergrad_culture_dance_performance_opera", 287, "舞蹈类、影视戏剧表演类、服装表演类本科文化控制线。"),
        ("艺术类", "专科", "junior_culture_control", 150, "艺术类专科文化控制线。"),
        ("体育类", "本科/专科", "first_section_composite", 587, "体育类一段线，划线成绩为综合分。"),
        ("体育类", "本科/专科", "second_section_composite", 480, "体育类二段线，划线成绩为综合分。"),
        ("春季高考", "技能拔尖人才", "spring_exam_skill_top", 150, "春季高考技能拔尖人才本科录取控制线。"),
        ("春季高考", "专科", "spring_exam_junior", 150, "春季高考各专业类别专科录取控制线。"),
        ("春季高考", "3+4对口贯通培养转段", "spring_exam_3plus4_transfer", 160, "3+4 对口贯通培养转段文化课总分合格标准。"),
    ),
    2024: (
        ("普通类", "常规批", "special_type_control", 521, "特殊类型招生控制线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "first_section", 444, "普通类一段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "second_section", 150, "普通类二段线；根据官方分数线图片校对。"),
        ("普通类", "常规批", "3plus2_vocational_qualification", 394, "3+2 对口贯通分段培养高职志愿填报资格线；根据官方分数线图片校对。"),
        ("普通类", "高水平运动队", "high_level_sports_world_class", 444, "世界一流大学建设高校高水平运动队本科文化控制线。"),
        ("普通类", "高水平运动队", "high_level_sports_other", 355, "其他高校高水平运动队本科文化控制线。"),
        ("艺术类", "本科", "undergrad_culture_broadcast_hosting", 444, "播音与主持类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_art_music_calligraphy", 333, "美术与设计类、音乐类、书法类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "本科", "undergrad_culture_dance_performance_opera", 288, "舞蹈类、表（导）演类、戏曲类本科文化控制线；根据官方分数线图片校对。"),
        ("艺术类", "专科", "junior_culture_control", 150, "艺术类专科文化控制线；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "first_section_composite", 594, "体育类一段线，划线成绩为综合分；根据官方分数线图片校对。"),
        ("体育类", "本科/专科", "second_section_composite", 470, "体育类二段线，划线成绩为综合分；根据官方分数线图片校对。"),
    ),
    2025: (
        ("普通类", "常规批", "special_type_control", 521, "特殊类型招生控制线；根据官方 PDF 校对。"),
        ("普通类", "常规批", "first_section", 441, "普通类一段线；根据官方 PDF 校对。"),
        ("普通类", "常规批", "second_section", 150, "普通类二段线；根据官方 PDF 校对。"),
        ("普通类", "常规批", "3plus2_vocational_qualification", 391, "3+2 对口贯通分段培养高职志愿填报资格线；根据官方 PDF 校对。"),
        ("普通类", "高水平运动队", "high_level_sports_world_class", 441, "世界一流大学建设高校高水平运动队本科文化控制线。"),
        ("普通类", "高水平运动队", "high_level_sports_other", 352, "其他高校高水平运动队本科文化控制线。"),
        ("艺术类", "本科", "undergrad_culture_broadcast_hosting", 441, "播音与主持类本科文化控制线；根据官方 PDF 校对。"),
        ("艺术类", "本科", "undergrad_culture_art_music_calligraphy", 330, "美术与设计类、音乐类、书法类本科文化控制线；根据官方 PDF 校对。"),
        ("艺术类", "本科", "undergrad_culture_dance_performance_opera", 286, "舞蹈类、表（导）演类、戏曲类本科文化控制线；根据官方 PDF 校对。"),
        ("艺术类", "专科", "junior_culture_control", 150, "艺术类专科文化控制线；根据官方 PDF 校对。"),
        ("体育类", "本科", "undergrad_culture_control", 286, "体育类本科文化录取控制线。"),
        ("体育类", "专科", "junior_culture_control", 150, "体育类专科文化录取控制线。"),
        ("体育类", "本科/专科", "first_section_composite", 566, "体育类一段线，划线成绩为综合分；根据官方 PDF 校对。"),
        ("体育类", "本科/专科", "second_section_composite", 428, "体育类二段线，划线成绩为综合分；根据官方 PDF 校对。"),
    ),
}


@dataclass
class ParsedAdmissionRow:
    row_number: int
    college_code: str | None
    college_name: str
    major_code: str | None
    major_name: str
    plan_count: int | None
    min_score: float | None
    min_rank: int | None


@dataclass
class OfficialImportStats:
    source_document_id: int
    import_run_id: int
    source_type: str
    year: int
    total_rows: int = 0
    success_rows: int = 0
    failed_rows: int = 0
    skipped_rows: int = 0
    created_rows: int = 0
    updated_rows: int = 0
    error_report_path: str | None = None
    local_file_path: str | None = None
    notes: list[str] = field(default_factory=list)


def download_gaokao_source_document_file(
    session: Session,
    settings: Settings,
    *,
    source_document_id: int,
) -> Path:
    document = get_gaokao_source_document(session, source_document_id)
    if document is None:
        raise ValueError(f"source document not found: {source_document_id}")
    file_url = DEFAULT_OFFICIAL_FILE_URLS.get((document.year, document.source_type))
    if not file_url:
        file_url = _discover_file_url(document)
    suffix = _suffix_from_url(file_url)
    filename = _safe_filename(f"{document.year}_{document.source_type}_{document.title}") + suffix
    target_dir = settings.data_dir / "imports" / "gaokao" / "official" / str(document.year)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / filename
    if target_path.exists():
        return target_path
    _download_file(file_url, target_path)
    return target_path


def import_b1_shandong_core(
    session: Session,
    settings: Settings,
    *,
    download_missing: bool,
    years: tuple[int, ...] = B1_IMPORT_YEARS,
    source_types: tuple[str, ...] = B1_SOURCE_TYPES,
    coverage_doc: Path | None = None,
) -> dict[str, Any]:
    seed_default_gaokao_sources(session, settings)
    stats: list[OfficialImportStats] = []
    for year in years:
        for source_type in source_types:
            document = _get_b1_source_document(session, year=year, source_type=source_type)
            file_path = _resolve_document_file(settings, document)
            if file_path is None:
                if not download_missing:
                    raise ValueError(
                        f"{year} {source_type} has no local file under "
                        "data/imports/gaokao/official/{year}/ or manual/{year}; "
                        "rerun without --no-download or place the official file there."
                    )
                file_path = download_gaokao_source_document_file(
                    session,
                    settings,
                    source_document_id=document.id,
                )
            _apply_published_at(document)
            document, run = register_gaokao_local_file(
                session,
                settings,
                source_document_id=document.id,
                file_path=file_path,
                importer_name=document.parser_name or _default_importer_name(source_type),
            )
            stats.append(import_registered_gaokao_file(session, settings, import_run_id=run.id))
    doc_path = coverage_doc or (settings.project_root / B1_COVERAGE_DOC)
    coverage = write_b1_coverage_doc(session, settings, doc_path)
    return {
        "imports": [stats_item_to_dict(item) for item in stats],
        "coverage_doc": relative_to_project(doc_path, settings.project_root),
        "coverage": coverage,
    }


def import_registered_gaokao_file(
    session: Session,
    settings: Settings,
    *,
    import_run_id: int,
) -> OfficialImportStats:
    run = session.get(GaokaoImportRun, import_run_id)
    if run is None:
        raise ValueError(f"import run not found: {import_run_id}")
    document = session.get(GaokaoSourceDocument, run.source_document_id)
    if document is None:
        raise ValueError(f"source document not found: {run.source_document_id}")
    file_path = _resolve_document_file(settings, document)
    if file_path is None:
        raise ValueError(f"source document has no local file: {document.id}")

    run.status = "running"
    run.started_at = datetime.now()
    run.finished_at = None
    session.flush()

    row_errors: list[RowError] = []
    try:
        if document.source_type == "admission_result":
            stats = _import_admission_result(session, settings, document, run, file_path, row_errors)
        elif document.source_type == "score_rank_segment":
            stats = _import_score_rank_segment(session, settings, document, run, file_path, row_errors)
        elif document.source_type == "score_line":
            stats = _import_score_line(session, settings, document, run, file_path, row_errors)
        else:
            raise ValueError(f"unsupported gaokao source type: {document.source_type}")
    except Exception as exc:
        run.status = "failed"
        run.finished_at = datetime.now()
        run.note = str(exc)
        document.status = "failed"
        session.flush()
        raise

    stats.error_report_path = _save_b1_error_report(settings, document, row_errors)
    run.total_rows = stats.total_rows
    run.success_rows = stats.success_rows
    run.failed_rows = stats.failed_rows
    run.skipped_rows = stats.skipped_rows
    run.created_rows = stats.created_rows
    run.updated_rows = stats.updated_rows
    run.error_report_path = stats.error_report_path
    run.raw_snapshot_path = document.local_file_path
    run.status = resolve_import_status(
        total_rows=stats.total_rows,
        success_rows=stats.success_rows,
        failed_rows=stats.failed_rows,
    )
    run.finished_at = datetime.now()
    run.note = "B1 官方核心数据导入完成。" if run.status == "success" else "B1 官方核心数据导入完成，但存在需人工修复的行。"

    document.status = "imported" if run.status == "success" else "partially_imported"
    document.parser_version = B1_PARSER_VERSION
    document.fetched_at = document.fetched_at or datetime.now()
    document.local_file_path = stats.local_file_path or document.local_file_path
    session.flush()
    return stats


def write_b1_coverage_doc(session: Session, settings: Settings, doc_path: Path) -> dict[str, Any]:
    coverage = build_b1_coverage(session)
    doc_path.parent.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# 山东高考 2020-2025 普通类核心数据覆盖报告",
        "",
        f"- 生成时间：{generated_at}",
        "- 执行窗口：B1/E6，2020-2025 山东普通类核心数据导入",
        "- 数据范围：山东普通类常规批第 1 次志愿投档情况表、一分一段表、各类别分数线 / 省控线；其中 E6 重点补 2020-2022 一分一段和分数线。",
        "",
        "## 1. 覆盖矩阵",
        "",
        "| 年份 | 应用侧录取结果 | raw 投档结果 | 一分一段 | 省控线 / 分数线 | source_document 追溯结论 |",
        "| --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for year in B1_COVERAGE_YEARS:
        item = coverage["years"][year]
        lines.append(
            "| {year} | {admission_record_total} / {admission_record_with_source} | "
            "{raw_admission_total} / {raw_admission_with_source} | "
            "{score_rank_total} / {score_rank_with_source} | "
            "{score_line_total} / {score_line_with_source} | {trace_status} |".format(
                year=year,
                **item,
            )
        )
    lines.extend(
        [
            "",
            "说明：每个单元格格式为“当前记录数 / 已关联来源记录数”。",
            "",
            "## 2. 官方来源文件",
            "",
            "| 年份 | 类型 | source_document_id | 状态 | 本地文件 | SHA256 | 最近导入批次 |",
            "| --- | --- | ---: | --- | --- | --- | --- |",
        ]
    )
    for source in coverage["sources"]:
        lines.append(
            "| {year} | {source_type_label} | {id} | {status} | {local_file_path} | {file_sha256} | {run_summary} |".format(
                **source
            )
        )
    lines.extend(
        [
            "",
            "## 3. B1 判断",
            "",
            f"- 普通类推荐最低数据条件：{coverage['ordinary_recommendation_readiness_label']}。",
            "- 2020-2025 普通类投档表已保留最低位次；后续推荐算法应继续优先使用位次，不直接用裸分跨年比较。",
            "- 2026 普通类正式招生计划、投档结果、一分一段和省控线未在本窗口伪造；仍保持待官方发布 / 待导入语义。",
            "- 分数线来源中 2023、2024 为官方页面图片，2025 为官方 PDF；当前导入保留原始文件和 SHA256，并把结构化分数线作为 B1 校对数据写入。",
            "",
            "## 4. 尚需后续处理",
            "",
        ]
    )
    if coverage["missing_files"]:
        for item in coverage["missing_files"]:
            lines.append(f"- {item}")
    else:
        lines.append("- 2020-2025 核心范围内投档表、一分一段、分数线官方文件均已登记本地文件。")
    lines.extend(
        [
            "- 2023 招生计划缺失、2024 招生计划偏少不是本窗口投档 / 一分一段 / 省控线导入能完全解决的问题，后续应由招生计划专项继续核验。",
            "- 特殊类型专门录取结果仍缺失，只能作为初筛和人工核对，不满足完整录取把握。",
            "",
        ]
    )
    doc_path.write_text("\n".join(lines), encoding="utf-8")
    return coverage


def build_b1_coverage(session: Session) -> dict[str, Any]:
    years: dict[int, dict[str, Any]] = {}
    for year in B1_COVERAGE_YEARS:
        admission_total, admission_with_source = _count_table(
            session,
            "admission_record",
            year=year,
            province_values=("山东",),
            filters={"student_type": "general", "batch": "常规批"},
        )
        raw_total, raw_with_source = _count_table(
            session,
            "gaokao_admission_result",
            year=year,
            province_values=("sd", "山东"),
            filters={"candidate_type": "普通类", "batch_name": "常规批"},
        )
        rank_total, rank_with_source = _count_table(
            session,
            "score_rank_segment",
            year=year,
            province_values=("sd", "山东"),
        )
        line_total, line_with_source = _count_table(
            session,
            "gaokao_score_line",
            year=year,
            province_values=("sd", "山东"),
        )
        trace_ready = (
            admission_with_source > 0
            and raw_with_source > 0
            and rank_with_source > 0
            and line_with_source > 0
        )
        years[year] = {
            "admission_record_total": admission_total,
            "admission_record_with_source": admission_with_source,
            "raw_admission_total": raw_total,
            "raw_admission_with_source": raw_with_source,
            "score_rank_total": rank_total,
            "score_rank_with_source": rank_with_source,
            "score_line_total": line_total,
            "score_line_with_source": line_with_source,
            "trace_status": "已关联来源" if trace_ready else "仍需补来源",
        }

    sources = []
    missing_files = []
    for document in _list_b1_source_documents(session):
        latest_run = _latest_import_run(session, document.id)
        local_file = document.local_file_path or ""
        if not local_file:
            missing_files.append(f"{document.year} {document.source_type} 缺少本地官方文件。")
        sources.append(
            {
                "year": document.year,
                "source_type_label": _source_type_label(document.source_type),
                "id": document.id,
                "status": document.status,
                "local_file_path": local_file or "未登记",
                "file_sha256": (document.file_sha256 or "")[:12] + ("..." if document.file_sha256 else ""),
                "run_summary": _format_run_summary(latest_run),
            }
        )

    readiness = all(
        years[year]["raw_admission_with_source"] > 0
        and years[year]["score_rank_with_source"] > 0
        and years[year]["score_line_with_source"] > 0
        for year in B1_IMPORT_YEARS
    )
    return {
        "years": years,
        "sources": sources,
        "missing_files": missing_files,
        "ordinary_recommendation_readiness": readiness,
        "ordinary_recommendation_readiness_label": "满足最近三年普通类推荐最低数据条件" if readiness else "未满足，仍缺核心来源或数据",
    }


def stats_item_to_dict(item: OfficialImportStats) -> dict[str, Any]:
    return {
        "source_document_id": item.source_document_id,
        "import_run_id": item.import_run_id,
        "source_type": item.source_type,
        "year": item.year,
        "total_rows": item.total_rows,
        "success_rows": item.success_rows,
        "failed_rows": item.failed_rows,
        "skipped_rows": item.skipped_rows,
        "created_rows": item.created_rows,
        "updated_rows": item.updated_rows,
        "error_report_path": item.error_report_path,
        "local_file_path": item.local_file_path,
        "notes": item.notes,
    }


def _import_admission_result(
    session: Session,
    settings: Settings,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    file_path: Path,
    row_errors: list[RowError],
) -> OfficialImportStats:
    rows = _parse_admission_rows(file_path)
    stats = OfficialImportStats(
        source_document_id=document.id,
        import_run_id=run.id,
        source_type=document.source_type,
        year=document.year,
        total_rows=len(rows),
        local_file_path=document.local_file_path,
    )
    if len(rows) > 1000:
        return _bulk_import_admission_result(session, document, run, rows, stats)
    for row in rows:
        try:
            raw_created = _upsert_raw_admission_result(session, document, run, row)
            app_created = _upsert_application_admission_record(session, document, run, row)
            stats.success_rows += 1
            stats.created_rows += int(raw_created) + int(app_created)
            stats.updated_rows += int(not raw_created) + int(not app_created)
        except Exception as exc:
            stats.failed_rows += 1
            row_errors.append(
                build_row_error(
                    row_number=row.row_number,
                    values={
                        "院校": row.college_name,
                        "专业": row.major_name,
                        "投档计划数": row.plan_count,
                        "最低位次": row.min_rank,
                    },
                    message=str(exc),
                )
            )
    return stats


def _bulk_import_admission_result(
    session: Session,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    rows: list[ParsedAdmissionRow],
    stats: OfficialImportStats,
) -> OfficialImportStats:
    if not _table_exists(session, "gaokao_admission_result"):
        raise ValueError("gaokao_admission_result table does not exist; run migrations or gaokao materialization first.")
    temp_table = f"tmp_b1_admission_{run.id}"
    session.execute(text(f"DROP TABLE IF EXISTS {temp_table}"))
    session.execute(
        text(
            f"""
            CREATE TEMP TABLE {temp_table} (
                college_code TEXT,
                college_name TEXT NOT NULL,
                major_code TEXT,
                major_name TEXT NOT NULL,
                plan_count INTEGER,
                min_score NUMERIC,
                min_rank INTEGER,
                source_record_hash TEXT
            )
            """
        )
    )
    session.execute(
        text(
            f"""
            INSERT INTO {temp_table} (
                college_code, college_name, major_code, major_name,
                plan_count, min_score, min_rank, source_record_hash
            ) VALUES (
                :college_code, :college_name, :major_code, :major_name,
                :plan_count, :min_score, :min_rank, :source_record_hash
            )
            """
        ),
        [
            {
                "college_code": row.college_code,
                "college_name": row.college_name,
                "major_code": row.major_code,
                "major_name": row.major_name,
                "plan_count": row.plan_count,
                "min_score": row.min_score,
                "min_rank": row.min_rank,
                "source_record_hash": _hash_text(
                    f"admission|{document.year}|{row.college_code}|{row.major_code}|{row.major_name}|{row.min_rank}|{row.plan_count}"
                ),
            }
            for row in rows
        ],
    )
    session.execute(text(f"CREATE INDEX idx_{temp_table}_code ON {temp_table} (college_code, major_code, major_name)"))
    session.execute(text(f"CREATE INDEX idx_{temp_table}_name ON {temp_table} (college_name, major_name)"))
    now = _now_text()
    source_payload = {
        "year": document.year,
        "source_level": "official",
        "source_title": document.title,
        "source_url": document.url,
        "local_source_path": document.local_file_path,
        "parser_script_name": run.importer_name,
        "published_at": _published_at(document),
        "review_status": "confirmed_official_public",
        "data_version_label": B1_DATA_VERSION_LABEL,
        "source_document_id": document.id,
        "import_run_id": run.id,
        "updated_at": now,
    }
    raw_updated = session.execute(
        text(
            f"""
            UPDATE gaokao_admission_result AS r
            SET
                plan_count = t.plan_count,
                min_rank = t.min_rank,
                min_score = COALESCE(t.min_score, r.min_score),
                source_level = :source_level,
                source_title = :source_title,
                source_url = :source_url,
                local_source_path = :local_source_path,
                parser_script_name = :parser_script_name,
                published_at = :published_at,
                review_status = :review_status,
                source_record_hash = t.source_record_hash,
                data_version_label = :data_version_label,
                source_document_id = :source_document_id,
                import_run_id = :import_run_id,
                updated_at = :updated_at
            FROM {temp_table} AS t
            WHERE r.province IN ('sd', '山东')
              AND r.year = :year
              AND r.candidate_type = '普通类'
              AND r.batch_name = '常规批'
              AND r.round_no = '1'
              AND t.college_code = r.college_code_snapshot
              AND t.major_code = r.major_code_snapshot
              AND t.major_name = r.major_name_snapshot
            """
        ),
        source_payload,
    ).rowcount or 0

    raw_created = 0
    unmatched_raw_rows = max(len(rows) - int(raw_updated), 0)
    app_updated = session.execute(
        text(
            f"""
            UPDATE admission_record AS ar
            SET
                plan_count = matched.plan_count,
                min_rank = matched.min_rank,
                min_score = COALESCE(matched.min_score, ar.min_score),
                source_note = :source_title,
                source_document_id = :source_document_id,
                import_run_id = :import_run_id,
                updated_at = :updated_at
            FROM (
                SELECT
                  ar2.id AS admission_record_id,
                  MAX(t.plan_count) AS plan_count,
                  MAX(t.min_score) AS min_score,
                  MAX(t.min_rank) AS min_rank
                FROM {temp_table} AS t
                JOIN college c ON (
                  (t.college_code IS NOT NULL AND c.college_code = t.college_code)
                  OR c.name = t.college_name
                )
                JOIN major m ON (
                  (t.major_code IS NOT NULL AND m.major_code = t.major_code AND m.name = t.major_name)
                  OR m.name = t.major_name
                )
                JOIN admission_record ar2 ON ar2.college_id = c.id AND ar2.major_id = m.id
                WHERE ar2.year = :year
                  AND ar2.province = '山东'
                  AND ar2.batch = '常规批'
                  AND ar2.student_type = 'general'
                GROUP BY ar2.id
            ) AS matched
            WHERE ar.year = :year
              AND ar.id = matched.admission_record_id
            """
        ),
        {
            "year": document.year,
            "source_title": document.title,
            "source_document_id": document.id,
            "import_run_id": run.id,
            "updated_at": now,
        },
    ).rowcount or 0
    session.execute(text(f"DROP TABLE IF EXISTS {temp_table}"))
    stats.success_rows = len(rows) - unmatched_raw_rows
    stats.created_rows = int(raw_created)
    stats.updated_rows = int(raw_updated) + int(app_updated)
    stats.skipped_rows = unmatched_raw_rows
    if unmatched_raw_rows:
        stats.notes.append(f"有 {unmatched_raw_rows} 条官方投档行未在既有 raw 投档表中匹配到，已保留本地官方文件，后续可专项补插。")
    if app_updated < len(rows):
        stats.notes.append(f"应用侧录取结果按现有院校/专业匹配回填 {app_updated} 条；未匹配行保留在 raw 投档结果中。")
    return stats


def _insert_missing_raw_admission_rows(
    session: Session,
    temp_table: str,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    now: str,
) -> int:
    columns = _table_columns(session, "gaokao_admission_result")
    wanted_columns = [
        "province",
        "year",
        "candidate_type",
        "batch_name",
        "round_no",
        "college_code_snapshot",
        "college_name_snapshot",
        "major_code_snapshot",
        "major_name_snapshot",
        "min_score",
        "min_rank",
        "plan_count",
        "original_min_rank_text",
        "remark",
        "created_at",
        "updated_at",
        "source_level",
        "source_title",
        "source_url",
        "local_source_path",
        "parser_script_name",
        "published_at",
        "review_status",
        "source_record_hash",
        "data_version_label",
        "source_document_id",
        "import_run_id",
    ]
    insert_columns = [column for column in wanted_columns if column in columns]
    expressions = {
        "province": "'sd'",
        "year": ":year",
        "candidate_type": "'普通类'",
        "batch_name": "'常规批'",
        "round_no": "'1'",
        "college_code_snapshot": "t.college_code",
        "college_name_snapshot": "t.college_name",
        "major_code_snapshot": "t.major_code",
        "major_name_snapshot": "t.major_name",
        "min_score": "t.min_score",
        "min_rank": "t.min_rank",
        "plan_count": "t.plan_count",
        "original_min_rank_text": "CAST(t.min_rank AS TEXT)",
        "remark": "'山东普通类常规批第1次志愿投档情况表'",
        "created_at": ":now",
        "updated_at": ":now",
        "source_level": "'official'",
        "source_title": ":source_title",
        "source_url": ":source_url",
        "local_source_path": ":local_source_path",
        "parser_script_name": ":parser_script_name",
        "published_at": ":published_at",
        "review_status": "'confirmed_official_public'",
        "source_record_hash": "t.source_record_hash",
        "data_version_label": ":data_version_label",
        "source_document_id": ":source_document_id",
        "import_run_id": ":import_run_id",
    }
    return int(
        session.execute(
            text(
                f"""
                INSERT INTO gaokao_admission_result ({', '.join(insert_columns)})
                SELECT {', '.join(expressions[column] for column in insert_columns)}
                FROM {temp_table} t
                WHERE NOT EXISTS (
                  SELECT 1
                  FROM gaokao_admission_result r
                  WHERE r.province IN ('sd', '山东')
                    AND r.year = :year
                    AND r.candidate_type = '普通类'
                    AND r.batch_name = '常规批'
                    AND r.round_no = '1'
                    AND r.college_code_snapshot = t.college_code
                    AND r.major_code_snapshot = t.major_code
                    AND r.major_name_snapshot = t.major_name
                )
                """
            ),
            {
                "year": document.year,
                "now": now,
                "source_title": document.title,
                "source_url": document.url,
                "local_source_path": document.local_file_path,
                "parser_script_name": run.importer_name,
                "published_at": _published_at(document),
                "data_version_label": B1_DATA_VERSION_LABEL,
                "source_document_id": document.id,
                "import_run_id": run.id,
            },
        ).rowcount
        or 0
    )


def _import_score_rank_segment(
    session: Session,
    settings: Settings,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    file_path: Path,
    row_errors: list[RowError],
) -> OfficialImportStats:
    rows = _parse_score_rank_rows(file_path, document)
    stats = OfficialImportStats(
        source_document_id=document.id,
        import_run_id=run.id,
        source_type=document.source_type,
        year=document.year,
        total_rows=len(rows),
        local_file_path=document.local_file_path,
    )
    if not _table_exists(session, "score_rank_segment"):
        raise ValueError("score_rank_segment table does not exist; run migrations or gaokao materialization first.")
    deleted = session.execute(
        text("DELETE FROM score_rank_segment WHERE year = :year AND province IN ('sd', '山东')"),
        {"year": document.year},
    ).rowcount or 0
    for row in rows:
        row["source_document_id"] = document.id
        row["import_run_id"] = run.id
        _insert_dynamic(session, "score_rank_segment", row)
    stats.success_rows = len(rows)
    stats.created_rows = len(rows)
    stats.updated_rows = int(deleted)
    return stats


def _import_score_line(
    session: Session,
    settings: Settings,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    file_path: Path,
    row_errors: list[RowError],
) -> OfficialImportStats:
    if not _table_exists(session, "gaokao_score_line"):
        raise ValueError("gaokao_score_line table does not exist; run migrations or gaokao materialization first.")
    source_rows = SCORE_LINE_ROWS.get(document.year)
    if not source_rows:
        raise ValueError(f"unsupported score line year: {document.year}")
    deleted = session.execute(
        text("DELETE FROM gaokao_score_line WHERE year = :year AND province IN ('sd', '山东')"),
        {"year": document.year},
    ).rowcount or 0
    now = _now_text()
    for candidate_type, batch_name, line_type, score, remark in source_rows:
        row = {
            "province": "sd",
            "year": document.year,
            "candidate_type": candidate_type,
            "batch_name": batch_name,
            "line_type": line_type,
            "score": score,
            "remark": remark,
            "created_at": now,
            "updated_at": now,
            "source_level": "official",
            "source_title": document.title,
            "source_url": document.url,
            "local_source_path": document.local_file_path,
            "parser_script_name": run.importer_name,
            "published_at": _published_at(document),
            "review_status": "confirmed_official_public",
            "source_record_hash": _hash_text(f"score_line|{document.year}|{candidate_type}|{batch_name}|{line_type}|{score}"),
            "data_version_label": B1_DATA_VERSION_LABEL,
            "source_document_id": document.id,
            "import_run_id": run.id,
        }
        _insert_dynamic(session, "gaokao_score_line", row)
    return OfficialImportStats(
        source_document_id=document.id,
        import_run_id=run.id,
        source_type=document.source_type,
        year=document.year,
        total_rows=len(source_rows),
        success_rows=len(source_rows),
        created_rows=len(source_rows),
        updated_rows=int(deleted),
        local_file_path=document.local_file_path,
        notes=["分数线来自官方页面图片/PDF，B1 以结构化校对数据写入并保留原始文件。"],
    )


def _parse_admission_rows(path: Path) -> list[ParsedAdmissionRow]:
    table = _read_table_rows(path)
    header_index = _find_header_index(table, required=("院校", "位次"))
    headers = table[header_index]
    col_major = _find_column(headers, ("专业",))
    col_college = _find_column(headers, ("院校",))
    col_plan = _find_column(headers, ("计划",))
    col_rank = _find_column(headers, ("位次",))
    col_score = _find_optional_column(headers, ("最低分", "投档最低分"))
    parsed = []
    for row_index, row in enumerate(table[header_index + 1 :], start=header_index + 2):
        if not any(_cell_text(cell) for cell in row):
            continue
        major_text = _cell(row, col_major)
        college_text = _cell(row, col_college)
        if not major_text or not college_text:
            continue
        college_code, college_name = _split_code_name(college_text, min_code_length=4)
        major_code, major_name = _split_code_name(major_text, min_code_length=2)
        min_rank = _to_int(_cell(row, col_rank))
        plan_count = _to_int(_cell(row, col_plan))
        min_score = _to_float(_cell(row, col_score)) if col_score is not None else None
        if not college_name or not major_name:
            continue
        if min_rank is None:
            raise ValueError(f"第 {row_index} 行缺少最低位次")
        parsed.append(
            ParsedAdmissionRow(
                row_number=row_index,
                college_code=college_code,
                college_name=college_name,
                major_code=major_code,
                major_name=major_name,
                plan_count=plan_count,
                min_score=min_score,
                min_rank=min_rank,
            )
        )
    return parsed


def _parse_score_rank_rows(path: Path, document: GaokaoSourceDocument) -> list[dict[str, Any]]:
    table = _read_table_rows(path)
    header_index = _find_header_index(table, required=("分数段", "全体"))
    group_row = table[header_index]
    subheader_row = table[header_index + 1]
    group_starts = _detect_score_rank_group_starts(group_row, subheader_row)

    now = _now_text()
    parsed: list[dict[str, Any]] = []
    for row in table[header_index + 2 :]:
        score = _to_int(row[0] if row else None)
        if score is None:
            continue
        for column, subject_group in group_starts:
            segment_count = _to_int(row[column] if column < len(row) else None)
            cumulative_count = _to_int(row[column + 1] if column + 1 < len(row) else None)
            if segment_count is None and cumulative_count is None:
                continue
            parsed.append(
                {
                    "province": "sd",
                    "year": document.year,
                    "score_type": "summer_total",
                    "subject_group": subject_group,
                    "score": score,
                    "segment_count": segment_count,
                    "cumulative_count": cumulative_count,
                    "rank_value": cumulative_count,
                    "created_at": now,
                    "updated_at": now,
                    "source_level": "official",
                    "source_title": document.title,
                    "source_url": document.url,
                    "local_source_path": document.local_file_path,
                    "parser_script_name": document.parser_name or "shandong_score_rank_segment",
                    "published_at": _published_at(document),
                    "review_status": "confirmed_official_public",
                    "source_record_hash": _hash_text(f"score_rank|{document.year}|{subject_group}|{score}|{segment_count}|{cumulative_count}"),
                    "data_version_label": B1_DATA_VERSION_LABEL,
                }
            )
    return parsed


def _upsert_raw_admission_result(
    session: Session,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    row: ParsedAdmissionRow,
) -> bool:
    if not _table_exists(session, "gaokao_admission_result"):
        raise ValueError("gaokao_admission_result table does not exist; run migrations or gaokao materialization first.")
    existing_id = session.execute(
        text(
            """
            SELECT id
            FROM gaokao_admission_result
            WHERE province IN ('sd', '山东')
              AND year = :year
              AND candidate_type = '普通类'
              AND batch_name = '常规批'
              AND round_no = '1'
              AND college_code_snapshot = :college_code
              AND major_code_snapshot = :major_code
              AND major_name_snapshot = :major_name
            ORDER BY id
            LIMIT 1
            """
        ),
        {
            "year": document.year,
            "college_code": row.college_code,
            "major_code": row.major_code,
            "major_name": row.major_name,
        },
    ).scalar_one_or_none()
    payload = _raw_admission_payload(document, run, row)
    if existing_id is None:
        _insert_dynamic(session, "gaokao_admission_result", payload)
        return True
    _update_dynamic(session, "gaokao_admission_result", existing_id, payload)
    return False


def _raw_admission_payload(
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    row: ParsedAdmissionRow,
) -> dict[str, Any]:
    now = _now_text()
    return {
        "province": "sd",
        "year": document.year,
        "candidate_type": "普通类",
        "batch_name": "常规批",
        "round_no": "1",
        "college_code_snapshot": row.college_code,
        "college_name_snapshot": row.college_name,
        "major_code_snapshot": row.major_code,
        "major_name_snapshot": row.major_name,
        "min_score": row.min_score,
        "min_rank": row.min_rank,
        "plan_count": row.plan_count,
        "original_min_rank_text": str(row.min_rank) if row.min_rank is not None else None,
        "remark": "山东普通类常规批第1次志愿投档情况表",
        "created_at": now,
        "updated_at": now,
        "source_level": "official",
        "source_title": document.title,
        "source_url": document.url,
        "local_source_path": document.local_file_path,
        "parser_script_name": run.importer_name,
        "published_at": _published_at(document),
        "review_status": "confirmed_official_public",
        "source_record_hash": _hash_text(
            f"admission|{document.year}|{row.college_code}|{row.major_code}|{row.major_name}|{row.min_rank}|{row.plan_count}"
        ),
        "data_version_label": B1_DATA_VERSION_LABEL,
        "source_document_id": document.id,
        "import_run_id": run.id,
    }


def _upsert_application_admission_record(
    session: Session,
    document: GaokaoSourceDocument,
    run: GaokaoImportRun,
    row: ParsedAdmissionRow,
) -> bool:
    college = _get_or_create_college(session, row.college_name, row.college_code)
    major = _get_or_create_major(session, row.major_name, row.major_code)
    ensure_college_major(session, college.id, major.id)
    current = get_admission_record_by_key(
        session,
        year=document.year,
        province="山东",
        batch="常规批",
        college_id=college.id,
        major_id=major.id,
        student_type="general",
        art_track=None,
    )
    created = current is None
    if current is None:
        current = AdmissionRecord(
            year=document.year,
            province="山东",
            batch="常规批",
            college_id=college.id,
            major_id=major.id,
            student_type="general",
            art_track=None,
        )
        session.add(current)
    if row.min_score is not None:
        current.min_score = row.min_score
    current.min_rank = row.min_rank
    current.plan_count = row.plan_count
    current.source_note = document.title
    current.source_document_id = document.id
    current.import_run_id = run.id
    current.is_active = True
    session.flush()
    return created


def _get_or_create_college(session: Session, name: str, code: str | None) -> College:
    college = None
    if code:
        college = session.scalar(select(College).where(College.college_code == code))
    if college is None:
        college = session.scalar(select(College).where(College.name == name))
    if college is None:
        college = College(name=name, college_code=code, province="山东")
        session.add(college)
        session.flush()
    else:
        if code and not college.college_code:
            college.college_code = code
        if not college.province:
            college.province = "山东"
    return college


def _get_or_create_major(session: Session, name: str, code: str | None) -> Major:
    major = None
    if code:
        major = session.scalar(select(Major).where(Major.major_code == code, Major.name == name))
    if major is None:
        major = session.scalar(select(Major).where(Major.name == name))
    if major is None:
        major = Major(name=name, major_code=code)
        session.add(major)
        session.flush()
    elif code and not major.major_code:
        major.major_code = code
    return major


def _read_table_rows(path: Path) -> list[list[Any]]:
    data = path.read_bytes()
    if data[:256].lstrip().lower().startswith((b"<html", b"<!doctype", b"<table")):
        return _read_html_table_rows(data.decode("utf-8", errors="ignore"))
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        return [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    sheet = xlrd.open_workbook(str(path)).sheet_by_index(0)
    return [[_cell_text(value) for value in sheet.row_values(row_index)] for row_index in range(sheet.nrows)]


def _read_html_table_rows(content: str) -> list[list[str]]:
    rows = []
    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", content, flags=re.S | re.I):
        row = []
        for cell_html in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_html, flags=re.S | re.I):
            cell_text = re.sub(r"<[^>]+>", "", cell_html)
            row.append(_cell_text(html.unescape(cell_text)))
        if row:
            rows.append(row)
    return rows


def _find_header_index(rows: list[list[Any]], *, required: tuple[str, ...]) -> int:
    for index, row in enumerate(rows):
        text_value = "|".join(_cell_text(item) for item in row)
        if all(item in text_value for item in required):
            return index
    raise ValueError(f"无法识别表头：需要包含 {'、'.join(required)}")


def _find_column(headers: list[Any], keywords: tuple[str, ...]) -> int:
    index = _find_optional_column(headers, keywords)
    if index is None:
        raise ValueError(f"无法识别列：{'/'.join(keywords)}")
    return index


def _find_optional_column(headers: list[Any], keywords: tuple[str, ...]) -> int | None:
    for index, value in enumerate(headers):
        text_value = _cell_text(value)
        if all(keyword in text_value for keyword in keywords):
            return index
    return None


def _cell(row: list[Any], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _cell_text(row[index])


def _split_code_name(value: str, *, min_code_length: int) -> tuple[str | None, str]:
    text_value = _cell_text(value)
    pattern = r"^([A-Z0-9]{" + str(min_code_length) + r"})(.+)$"
    match = re.match(pattern, text_value, flags=re.I)
    if match:
        return match.group(1).upper(), _cell_text(match.group(2))
    return None, text_value


def _normalize_subject_group(value: str) -> str:
    mapping = {
        "全体": "all",
        "选考物理": "physics",
        "选考化学": "chemistry",
        "选考生物": "biology",
        "选考政治": "politics",
        "选考思想政治": "politics",
        "选考历史": "history",
        "选考地理": "geography",
    }
    return mapping.get(_cell_text(value), _cell_text(value))


def _detect_score_rank_group_starts(group_row: list[Any], subheader_row: list[Any]) -> list[tuple[int, str]]:
    group_starts: list[tuple[int, str]] = []
    # Some older official Excel-HTML files omit the blank cells created by colspan
    # in the group header row. The subheader still has paired 本段人数/累计人数
    # columns, so reconstruct starts from the compact group labels before trying
    # the normal merged-cell shape.
    if _cell_text(subheader_row[0] if subheader_row else "") == "本段人数":
        compact_groups = [_normalize_subject_group(value) for value in group_row[1:] if _cell_text(value)]
        for group_index, subject_group in enumerate(compact_groups):
            column = 1 + group_index * 2
            subheader_value = _cell_text(subheader_row[column - 1] if column - 1 < len(subheader_row) else "")
            next_subheader_value = _cell_text(subheader_row[column] if column < len(subheader_row) else "")
            if subheader_value == "本段人数" and next_subheader_value == "累计人数":
                group_starts.append((column, subject_group))
        return group_starts

    current_group: str | None = None
    max_columns = max(len(group_row), len(subheader_row))
    for index in range(1, max_columns):
        group_value = _cell_text(group_row[index] if index < len(group_row) else "")
        if group_value:
            current_group = _normalize_subject_group(group_value)
        subheader_value = _cell_text(subheader_row[index] if index < len(subheader_row) else "")
        next_subheader_value = _cell_text(subheader_row[index + 1] if index + 1 < len(subheader_row) else "")
        if current_group and subheader_value == "本段人数" and next_subheader_value == "累计人数":
            group_starts.append((index, current_group))
    return group_starts


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def _to_int(value: Any) -> int | None:
    text_value = _cell_text(value)
    if not text_value or text_value in {"-", "—"}:
        return None
    match = re.search(r"-?\d+(?:\.0+)?", text_value.replace(",", ""))
    if not match:
        return None
    return int(float(match.group(0)))


def _to_float(value: Any) -> float | None:
    text_value = _cell_text(value)
    if not text_value or text_value in {"-", "—"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text_value.replace(",", ""))
    if not match:
        return None
    return float(match.group(0))


def _discover_file_url(document: GaokaoSourceDocument) -> str:
    if not document.url:
        raise ValueError(f"source document has no url: {document.id}")
    if _is_downloadable_url(document.url):
        return document.url
    content = _download_text(document.url)
    allowed_suffixes = (".xls", ".xlsx") if document.source_type in {"admission_result", "score_rank_segment"} else (".pdf", ".png", ".jpg", ".jpeg")
    links = re.findall(r"""(?:href|src)=["']([^"']+)["']""", content, flags=re.I)
    for link in links:
        file_url = urllib.parse.urljoin(document.url, html.unescape(link))
        lowered = urllib.parse.urlparse(file_url).path.lower()
        if "/floadup/" in lowered and lowered.endswith(allowed_suffixes):
            return file_url
    raise ValueError(f"无法从官方页面识别附件链接：{document.url}")


def _download_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": B1_USER_AGENT})
    with urllib.request.urlopen(request, timeout=30) as response:
        content_type = response.headers.get_content_charset() or "utf-8"
        return response.read().decode(content_type, errors="ignore")


def _download_file(url: str, target: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": B1_USER_AGENT})
    with urllib.request.urlopen(request, timeout=60) as response:
        target.write_bytes(response.read())


def _suffix_from_url(url: str) -> str:
    suffix = Path(urllib.parse.urlparse(url).path).suffix.lower()
    if suffix:
        return suffix
    return ".bin"


def _is_downloadable_url(url: str) -> bool:
    return Path(urllib.parse.urlparse(url).path).suffix.lower() in {".xls", ".xlsx", ".pdf", ".png", ".jpg", ".jpeg"}


def _safe_filename(value: str) -> str:
    text_value = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", value.strip(), flags=re.U)
    return text_value.strip("._")[:160] or "gaokao_source"


def _get_b1_source_document(session: Session, *, year: int, source_type: str) -> GaokaoSourceDocument:
    document = session.scalar(
        select(GaokaoSourceDocument).where(
            GaokaoSourceDocument.province == "山东",
            GaokaoSourceDocument.year == year,
            GaokaoSourceDocument.source_type == source_type,
        )
    )
    if document is None:
        raise ValueError(f"missing B1 source document: {year} {source_type}")
    return document


def _list_b1_source_documents(session: Session) -> list[GaokaoSourceDocument]:
    return list(
        session.scalars(
            select(GaokaoSourceDocument)
            .where(
                GaokaoSourceDocument.province == "山东",
                GaokaoSourceDocument.year.in_(B1_COVERAGE_YEARS),
                GaokaoSourceDocument.source_type.in_(B1_SOURCE_TYPES),
            )
            .order_by(GaokaoSourceDocument.year, GaokaoSourceDocument.source_type)
        )
    )


def _resolve_document_file(settings: Settings, document: GaokaoSourceDocument) -> Path | None:
    if not document.local_file_path:
        return _find_existing_local_file(settings, document)
    path = Path(document.local_file_path)
    if not path.is_absolute():
        path = settings.project_root / path
    return path if path.exists() else _find_existing_local_file(settings, document)


def _find_existing_local_file(settings: Settings, document: GaokaoSourceDocument) -> Path | None:
    allowed_suffixes = (
        (".xls", ".xlsx")
        if document.source_type in {"admission_result", "score_rank_segment"}
        else (".pdf", ".png", ".jpg", ".jpeg")
    )
    search_dirs = [
        settings.data_dir / "imports" / "gaokao" / "official" / str(document.year),
        settings.data_dir / "imports" / "gaokao" / "manual" / str(document.year),
        settings.data_dir / "imports" / "gaokao" / "official",
        settings.data_dir / "imports" / "gaokao" / "manual",
    ]
    for directory in search_dirs:
        if not directory.exists():
            continue
        candidates = [
            path
            for path in sorted(directory.iterdir())
            if path.is_file()
            and path.suffix.lower() in allowed_suffixes
            and str(document.year) in path.name
            and document.source_type in path.name
        ]
        if candidates:
            return candidates[0]
    return None


def _apply_published_at(document: GaokaoSourceDocument) -> None:
    if document.published_at:
        return
    date_text = PUBLISHED_DATES.get((document.year, document.source_type))
    if date_text:
        document.published_at = datetime.strptime(date_text, "%Y-%m-%d").date()


def _published_at(document: GaokaoSourceDocument) -> str | None:
    _apply_published_at(document)
    return document.published_at.isoformat() if document.published_at else None


def _default_importer_name(source_type: str) -> str:
    return {
        "admission_result": "shandong_admission_result",
        "score_rank_segment": "shandong_score_rank_segment",
        "score_line": "shandong_score_line",
    }.get(source_type, "shandong_official_import")


def _save_b1_error_report(
    settings: Settings,
    document: GaokaoSourceDocument,
    row_errors: list[RowError],
) -> str | None:
    return save_error_report(
        settings=settings,
        prefix=f"gaokao_b1_{document.source_type}_{document.year}_errors",
        headers=["院校", "专业", "投档计划数", "最低位次"],
        errors=row_errors,
    )


def _table_exists(session: Session, table_name: str) -> bool:
    return bool(
        session.execute(
            text("SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = :table_name"),
            {"table_name": table_name},
        ).first()
    )


def _table_columns(session: Session, table_name: str) -> set[str]:
    bind = session.get_bind()
    cache_key = (str(getattr(bind, "url", "")), table_name)
    if cache_key in _COLUMN_CACHE:
        return _COLUMN_CACHE[cache_key]
    rows = session.execute(text(f"PRAGMA table_info({table_name})")).mappings().all()
    columns = {str(row["name"]) for row in rows}
    _COLUMN_CACHE[cache_key] = columns
    return columns


def _insert_dynamic(session: Session, table_name: str, payload: dict[str, Any]) -> None:
    columns = [column for column in payload if column in _table_columns(session, table_name)]
    sql = text(
        f"INSERT INTO {table_name} ({', '.join(columns)}) "
        f"VALUES ({', '.join(':' + column for column in columns)})"
    )
    session.execute(sql, {column: payload[column] for column in columns})


def _update_dynamic(session: Session, table_name: str, row_id: int, payload: dict[str, Any]) -> None:
    columns = [
        column
        for column in payload
        if column in _table_columns(session, table_name)
        and column not in {"id", "created_at"}
        and not (payload[column] is None and column in {"min_score"})
    ]
    assignments = ", ".join(f"{column} = :{column}" for column in columns)
    values = {column: payload[column] for column in columns}
    values["id"] = row_id
    session.execute(text(f"UPDATE {table_name} SET {assignments} WHERE id = :id"), values)


def _count_table(
    session: Session,
    table_name: str,
    *,
    year: int,
    province_values: tuple[str, ...],
    filters: dict[str, Any] | None = None,
) -> tuple[int, int]:
    if not _table_exists(session, table_name):
        return 0, 0
    filters = filters or {}
    conditions = ["year = :year"]
    values: dict[str, Any] = {"year": year}
    placeholders = []
    for index, province in enumerate(province_values):
        key = f"province_{index}"
        placeholders.append(f":{key}")
        values[key] = province
    conditions.append(f"province IN ({', '.join(placeholders)})")
    columns = _table_columns(session, table_name)
    for key, value in filters.items():
        if key in columns:
            conditions.append(f"{key} = :{key}")
            values[key] = value
    where_clause = " AND ".join(conditions)
    row = session.execute(
        text(
            f"""
            SELECT
              COUNT(*) AS total,
              SUM(CASE WHEN source_document_id IS NOT NULL THEN 1 ELSE 0 END) AS with_source
            FROM {table_name}
            WHERE {where_clause}
            """
        ),
        values,
    ).mappings().one()
    return int(row["total"] or 0), int(row["with_source"] or 0)


def _latest_import_run(session: Session, source_document_id: int) -> GaokaoImportRun | None:
    return session.scalar(
        select(GaokaoImportRun)
        .where(GaokaoImportRun.source_document_id == source_document_id)
        .order_by(GaokaoImportRun.id.desc())
        .limit(1)
    )


def _format_run_summary(run: GaokaoImportRun | None) -> str:
    if run is None:
        return "未运行"
    return f"#{run.id} {run.status}，成功 {run.success_rows}/{run.total_rows}"


def _source_type_label(source_type: str) -> str:
    return {
        "admission_result": "普通类投档表",
        "score_rank_segment": "一分一段",
        "score_line": "分数线 / 省控线",
    }.get(source_type, source_type)


def _now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _hash_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()
