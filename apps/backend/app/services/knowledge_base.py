from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.models import ErrorReasonTag, KnowledgePoint, KnowledgePointAlias, Subject
from app.repositories.system import write_audit_log
from app.schemas.common import ImportResult
from app.schemas.knowledge import (
    ErrorReasonTagPayload,
    ErrorReasonTagRead,
    KnowledgeBaseImportResponse,
    KnowledgePointAliasPayload,
    KnowledgePointAliasRead,
    KnowledgePointNodeRead,
    KnowledgePointPayload,
    KnowledgePointRead,
    KnowledgeTreeResponse,
)
from app.utils.parsers import clean_text


UNCATEGORIZED_NAME = "未归类"
DEFAULT_ERROR_TAGS = [
    "概念不清",
    "审题失误",
    "计算错误",
    "方法不会",
    "步骤不全",
    "表达不规范",
    "时间分配",
    "知识遗忘",
    "粗心漏答",
    "其他",
]


@dataclass
class KnowledgeResolveResult:
    point: KnowledgePoint
    match_source: str
    raw_text: str


def normalize_lookup_text(value: str | None) -> str:
    text = clean_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).lower()


def split_error_tags(value: str | None) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"[、,，;；|]+", text)
    return list(dict.fromkeys(part.strip() for part in parts if part.strip()))


def ensure_default_error_tags(session: Session) -> None:
    existing = {
        normalize_lookup_text(item.name): item
        for item in session.scalars(select(ErrorReasonTag)).all()
    }
    changed = False
    for index, name in enumerate(DEFAULT_ERROR_TAGS):
        normalized = normalize_lookup_text(name)
        if normalized in existing:
            existing[normalized].is_active = True
            continue
        session.add(
            ErrorReasonTag(
                name=name,
                normalized_name=normalized,
                sort_order=index,
                is_builtin=True,
                is_active=True,
            )
        )
        changed = True
    if changed:
        session.flush()


def list_knowledge_tree(session: Session, subject_id: int | None = None) -> KnowledgeTreeResponse:
    stmt = (
        select(KnowledgePoint)
        .options(selectinload(KnowledgePoint.subject))
        .where(KnowledgePoint.is_active.is_(True))
        .order_by(KnowledgePoint.subject_id.asc(), KnowledgePoint.sort_order.asc(), KnowledgePoint.id.asc())
    )
    if subject_id is not None:
        stmt = stmt.where(KnowledgePoint.subject_id == subject_id)
    items = list(session.scalars(stmt).all())
    paths = build_knowledge_paths(items)
    children_by_parent: dict[int | None, list[KnowledgePoint]] = {}
    for item in items:
        children_by_parent.setdefault(item.parent_id, []).append(item)

    def build_node(item: KnowledgePoint) -> KnowledgePointNodeRead:
        children = [
            build_node(child)
            for child in sorted(
                children_by_parent.get(item.id, []),
                key=lambda current: (current.sort_order, current.id),
            )
        ]
        return KnowledgePointNodeRead(
            id=item.id,
            subject_id=item.subject_id,
            subject_name=item.subject.name if item.subject else None,
            parent_id=item.parent_id,
            name=item.name,
            code=item.code,
            description=item.description,
            sort_order=item.sort_order,
            source_type=item.source_type,
            path=paths.get(item.id, item.name),
            is_active=item.is_active,
            children=children,
        )

    roots = [
        item
        for item in items
        if item.parent_id is None or item.parent_id not in {current.id for current in items}
    ]
    nodes = [build_node(item) for item in sorted(roots, key=lambda current: (current.subject_id, current.sort_order, current.id))]
    return KnowledgeTreeResponse(subject_id=subject_id, items=nodes)


def list_knowledge_points(session: Session, subject_id: int | None = None) -> list[KnowledgePointRead]:
    stmt = select(KnowledgePoint).where(KnowledgePoint.is_active.is_(True)).order_by(
        KnowledgePoint.subject_id.asc(),
        KnowledgePoint.sort_order.asc(),
        KnowledgePoint.id.asc(),
    )
    if subject_id is not None:
        stmt = stmt.where(KnowledgePoint.subject_id == subject_id)
    items = list(session.scalars(stmt).all())
    paths = build_knowledge_paths(items)
    return [_serialize_point(item, paths) for item in items]


def create_knowledge_point(session: Session, payload: KnowledgePointPayload) -> KnowledgePointRead:
    _validate_point_payload(session, payload)
    point = KnowledgePoint(**payload.model_dump())
    session.add(point)
    _flush_or_400(session)
    write_audit_log(session, module="knowledge", action="create_point", target_type="knowledge_point", target_id=str(point.id))
    return _serialize_point(point, build_knowledge_paths(list(session.scalars(select(KnowledgePoint)).all())))


def update_knowledge_point(session: Session, point_id: int, payload: KnowledgePointPayload) -> KnowledgePointRead:
    point = session.get(KnowledgePoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="知识点不存在")
    _validate_point_payload(session, payload, current_id=point_id)
    for key, value in payload.model_dump().items():
        setattr(point, key, clean_text(value) if isinstance(value, str) else value)
    _flush_or_400(session)
    write_audit_log(session, module="knowledge", action="update_point", target_type="knowledge_point", target_id=str(point.id))
    return _serialize_point(point, build_knowledge_paths(list(session.scalars(select(KnowledgePoint)).all())))


def delete_knowledge_point(session: Session, point_id: int) -> dict[str, str]:
    point = session.get(KnowledgePoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="知识点不存在")
    point.is_active = False
    for child in session.scalars(select(KnowledgePoint).where(KnowledgePoint.parent_id == point_id)).all():
        child.parent_id = point.parent_id
    session.flush()
    write_audit_log(session, module="knowledge", action="delete_point", target_type="knowledge_point", target_id=str(point.id))
    return {"message": "知识点已停用"}


def list_aliases(session: Session, subject_id: int | None = None) -> list[KnowledgePointAliasRead]:
    stmt = (
        select(KnowledgePointAlias)
        .options(selectinload(KnowledgePointAlias.subject), selectinload(KnowledgePointAlias.knowledge_point))
        .where(KnowledgePointAlias.is_active.is_(True))
        .order_by(KnowledgePointAlias.subject_id.asc(), KnowledgePointAlias.alias_name.asc())
    )
    if subject_id is not None:
        stmt = stmt.where(KnowledgePointAlias.subject_id == subject_id)
    aliases = list(session.scalars(stmt).all())
    subject_ids = list({item.subject_id for item in aliases})
    points = (
        list(session.scalars(select(KnowledgePoint).where(KnowledgePoint.subject_id.in_(subject_ids))).all())
        if subject_ids
        else []
    )
    paths = build_knowledge_paths(points)
    return [_serialize_alias(item, paths) for item in aliases]


def create_alias(session: Session, payload: KnowledgePointAliasPayload) -> KnowledgePointAliasRead:
    alias = KnowledgePointAlias()
    _apply_alias_payload(session, alias, payload)
    session.add(alias)
    _flush_or_400(session, detail="同一科目下别名不能重复。")
    write_audit_log(session, module="knowledge", action="create_alias", target_type="knowledge_point_alias", target_id=str(alias.id))
    return _serialize_alias(alias, _paths_for_alias(session, alias))


def update_alias(session: Session, alias_id: int, payload: KnowledgePointAliasPayload) -> KnowledgePointAliasRead:
    alias = session.get(KnowledgePointAlias, alias_id)
    if not alias:
        raise HTTPException(status_code=404, detail="知识点别名不存在")
    _apply_alias_payload(session, alias, payload)
    _flush_or_400(session, detail="同一科目下别名不能重复。")
    write_audit_log(session, module="knowledge", action="update_alias", target_type="knowledge_point_alias", target_id=str(alias.id))
    return _serialize_alias(alias, _paths_for_alias(session, alias))


def delete_alias(session: Session, alias_id: int) -> dict[str, str]:
    alias = session.get(KnowledgePointAlias, alias_id)
    if not alias:
        raise HTTPException(status_code=404, detail="知识点别名不存在")
    alias.is_active = False
    session.flush()
    return {"message": "知识点别名已停用"}


def list_error_tags(session: Session) -> list[ErrorReasonTagRead]:
    ensure_default_error_tags(session)
    rows = session.scalars(
        select(ErrorReasonTag)
        .where(ErrorReasonTag.is_active.is_(True))
        .order_by(ErrorReasonTag.sort_order.asc(), ErrorReasonTag.id.asc())
    ).all()
    return [ErrorReasonTagRead.model_validate(item) for item in rows]


def create_error_tag(session: Session, payload: ErrorReasonTagPayload) -> ErrorReasonTagRead:
    tag = ErrorReasonTag()
    _apply_error_tag_payload(tag, payload)
    session.add(tag)
    _flush_or_400(session, detail="错因标签名称不能重复。")
    return ErrorReasonTagRead.model_validate(tag)


def update_error_tag(session: Session, tag_id: int, payload: ErrorReasonTagPayload) -> ErrorReasonTagRead:
    tag = session.get(ErrorReasonTag, tag_id)
    if not tag:
        raise HTTPException(status_code=404, detail="错因标签不存在")
    _apply_error_tag_payload(tag, payload)
    _flush_or_400(session, detail="错因标签名称不能重复。")
    return ErrorReasonTagRead.model_validate(tag)


def delete_error_tag(session: Session, tag_id: int) -> dict[str, str]:
    tag = session.get(ErrorReasonTag, tag_id)
    if not tag:
        raise HTTPException(status_code=404, detail="错因标签不存在")
    tag.is_active = False
    session.flush()
    return {"message": "错因标签已停用"}


def resolve_knowledge_point(session: Session, subject_id: int, raw_name: str) -> KnowledgeResolveResult:
    text = clean_text(raw_name)
    if not text:
        raise ValueError("知识点不能为空")
    normalized = normalize_lookup_text(text)
    standard = session.scalar(
        select(KnowledgePoint).where(
            KnowledgePoint.subject_id == subject_id,
            KnowledgePoint.name == text,
        )
    )
    if standard:
        standard.is_active = True
        return KnowledgeResolveResult(point=standard, match_source="standard", raw_text=text)

    alias = session.scalar(
        select(KnowledgePointAlias)
        .where(
            KnowledgePointAlias.subject_id == subject_id,
            KnowledgePointAlias.normalized_alias == normalized,
            KnowledgePointAlias.is_active.is_(True),
        )
        .order_by(KnowledgePointAlias.id.asc())
    )
    if alias and alias.knowledge_point:
        alias.knowledge_point.is_active = True
        return KnowledgeResolveResult(point=alias.knowledge_point, match_source="alias", raw_text=text)

    if ">" in text or "＞" in text:
        return KnowledgeResolveResult(
            point=ensure_knowledge_path(session, subject_id, text, source_type="import_path"),
            match_source="path",
            raw_text=text,
        )

    parent = ensure_uncategorized_point(session, subject_id)
    point = upsert_point_by_name(session, subject_id, text, parent_id=parent.id, source_type="import_flat")
    return KnowledgeResolveResult(point=point, match_source="uncategorized", raw_text=text)


def ensure_error_tags_by_names(session: Session, names: list[str]) -> list[dict[str, object]]:
    ensure_default_error_tags(session)
    rows: list[dict[str, object]] = []
    for name in names:
        clean_name = clean_text(name)
        if not clean_name:
            continue
        normalized = normalize_lookup_text(clean_name)
        tag = session.scalar(select(ErrorReasonTag).where(ErrorReasonTag.normalized_name == normalized))
        if tag is None:
            tag = ErrorReasonTag(
                name=clean_name,
                normalized_name=normalized,
                sort_order=1000,
                is_builtin=False,
                is_active=True,
            )
            session.add(tag)
            session.flush()
        else:
            tag.is_active = True
        rows.append({"id": tag.id, "name": tag.name})
    return rows


def import_knowledge_base(session: Session, content: bytes) -> KnowledgeBaseImportResponse:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    point_count = 0
    alias_count = 0
    error_tag_count = 0
    errors: list[str] = []

    if "知识点" in workbook.sheetnames:
        sheet = workbook["知识点"]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {headers[index]: value for index, value in enumerate(values)}
            if not any(value not in (None, "") for value in row.values()):
                continue
            try:
                subject = _resolve_subject(session, clean_text(row.get("科目")))
                path_text = clean_text(row.get("知识点路径")) or clean_text(row.get("知识点"))
                if not path_text:
                    raise ValueError("知识点路径不能为空")
                point = ensure_knowledge_path(session, subject.id, path_text, source_type="knowledge_import")
                point.description = clean_text(row.get("说明")) or point.description
                point_count += 1
            except Exception as exc:
                errors.append(f"知识点 sheet 第 {row_number} 行：{exc}")

    if "别名" in workbook.sheetnames:
        sheet = workbook["别名"]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {headers[index]: value for index, value in enumerate(values)}
            if not any(value not in (None, "") for value in row.values()):
                continue
            try:
                subject = _resolve_subject(session, clean_text(row.get("科目")))
                point = ensure_knowledge_path(session, subject.id, clean_text(row.get("标准知识点")) or "", source_type="knowledge_import")
                alias_name = clean_text(row.get("别名"))
                if not alias_name:
                    raise ValueError("别名不能为空")
                normalized_alias = normalize_lookup_text(alias_name)
                alias = session.scalar(
                    select(KnowledgePointAlias).where(
                        KnowledgePointAlias.subject_id == subject.id,
                        KnowledgePointAlias.normalized_alias == normalized_alias,
                    )
                )
                if alias is None:
                    alias = KnowledgePointAlias(subject_id=subject.id, alias_name=alias_name)
                    session.add(alias)
                alias.knowledge_point_id = point.id
                alias.alias_name = alias_name
                alias.normalized_alias = normalized_alias
                alias.source_type = "knowledge_import"
                alias.is_active = True
                alias_count += 1
            except Exception as exc:
                errors.append(f"别名 sheet 第 {row_number} 行：{exc}")

    if "错因标签" in workbook.sheetnames:
        sheet = workbook["错因标签"]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {headers[index]: value for index, value in enumerate(values)}
            name = clean_text(row.get("错因标签"))
            if not name:
                continue
            tag = session.scalar(select(ErrorReasonTag).where(ErrorReasonTag.name == name))
            if tag is None:
                tag = ErrorReasonTag(name=name, normalized_name=normalize_lookup_text(name))
                session.add(tag)
            tag.description = clean_text(row.get("说明"))
            tag.sort_order = _to_int(row.get("排序"), default=1000)
            tag.is_builtin = False
            tag.is_active = True
            error_tag_count += 1

    session.flush()
    status = "success" if not errors else ("failed" if not (point_count or alias_count or error_tag_count) else "partially_failed")
    return KnowledgeBaseImportResponse(
        status=status,
        total_rows=point_count + alias_count + error_tag_count + len(errors),
        success_rows=point_count + alias_count + error_tag_count,
        failed_rows=len(errors),
        error_preview=errors[:3],
        notice_preview=["知识库导入完成，路径会自动创建为同科单父级知识树。"],
        message=f"知识库导入完成，知识点 {point_count} 条，别名 {alias_count} 条，错因标签 {error_tag_count} 条。",
        point_count=point_count,
        alias_count=alias_count,
        error_tag_count=error_tag_count,
    )


def ensure_knowledge_path(session: Session, subject_id: int, path_text: str, *, source_type: str = "manual") -> KnowledgePoint:
    parts = [part.strip() for part in re.split(r"[>＞]", clean_text(path_text) or "") if part.strip()]
    if not parts:
        raise ValueError("知识点路径不能为空")
    parent_id: int | None = None
    point: KnowledgePoint | None = None
    for index, part in enumerate(parts):
        point = upsert_point_by_name(
            session,
            subject_id,
            part,
            parent_id=parent_id,
            source_type=source_type,
            sort_order=index,
        )
        parent_id = point.id
    if point is None:
        raise ValueError("知识点路径不能为空")
    return point


def ensure_uncategorized_point(session: Session, subject_id: int) -> KnowledgePoint:
    return upsert_point_by_name(session, subject_id, UNCATEGORIZED_NAME, parent_id=None, source_type="system")


def upsert_point_by_name(
    session: Session,
    subject_id: int,
    name: str,
    *,
    parent_id: int | None,
    source_type: str,
    sort_order: int = 0,
) -> KnowledgePoint:
    clean_name = clean_text(name)
    if not clean_name:
        raise ValueError("知识点名称不能为空")
    point = session.scalar(
        select(KnowledgePoint).where(
            KnowledgePoint.subject_id == subject_id,
            KnowledgePoint.name == clean_name,
        )
    )
    if point is None:
        point = KnowledgePoint(
            subject_id=subject_id,
            parent_id=parent_id,
            name=clean_name,
            sort_order=sort_order,
            source_type=source_type,
            is_active=True,
        )
        session.add(point)
        session.flush()
        return point
    if point.parent_id is None and parent_id is not None and point.name != UNCATEGORIZED_NAME:
        point.parent_id = parent_id
    point.sort_order = point.sort_order or sort_order
    point.source_type = point.source_type or source_type
    point.is_active = True
    return point


def build_knowledge_paths(points: list[KnowledgePoint]) -> dict[int, str]:
    point_map = {item.id: item for item in points}
    paths: dict[int, str] = {}

    def resolve(point: KnowledgePoint, seen: set[int] | None = None) -> str:
        if point.id in paths:
            return paths[point.id]
        seen = seen or set()
        if point.id in seen or point.parent_id is None or point.parent_id not in point_map:
            paths[point.id] = point.name
            return paths[point.id]
        seen.add(point.id)
        parent_path = resolve(point_map[point.parent_id], seen)
        paths[point.id] = f"{parent_path}>{point.name}"
        return paths[point.id]

    for point in points:
        resolve(point)
    return paths


def _serialize_point(item: KnowledgePoint, paths: dict[int, str]) -> KnowledgePointRead:
    return KnowledgePointRead(
        id=item.id,
        subject_id=item.subject_id,
        parent_id=item.parent_id,
        name=item.name,
        code=item.code,
        description=item.description,
        sort_order=item.sort_order,
        source_type=item.source_type,
        is_active=item.is_active,
        path=paths.get(item.id, item.name),
    )


def _serialize_alias(item: KnowledgePointAlias, paths: dict[int, str]) -> KnowledgePointAliasRead:
    return KnowledgePointAliasRead(
        id=item.id,
        subject_id=item.subject_id,
        subject_name=item.subject.name if item.subject else None,
        knowledge_point_id=item.knowledge_point_id,
        knowledge_point_name=item.knowledge_point.name if item.knowledge_point else None,
        knowledge_point_path=paths.get(item.knowledge_point_id),
        alias_name=item.alias_name,
        normalized_alias=item.normalized_alias,
        source_type=item.source_type,
        note=item.note,
        is_active=item.is_active,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _paths_for_alias(session: Session, alias: KnowledgePointAlias) -> dict[int, str]:
    points = list(session.scalars(select(KnowledgePoint).where(KnowledgePoint.subject_id == alias.subject_id)).all())
    return build_knowledge_paths(points)


def _validate_point_payload(session: Session, payload: KnowledgePointPayload, current_id: int | None = None) -> None:
    if not session.get(Subject, payload.subject_id):
        raise HTTPException(status_code=404, detail="学科不存在")
    if not clean_text(payload.name):
        raise HTTPException(status_code=400, detail="知识点名称不能为空")
    if payload.parent_id is not None:
        parent = session.get(KnowledgePoint, payload.parent_id)
        if not parent or parent.subject_id != payload.subject_id:
            raise HTTPException(status_code=400, detail="父级知识点必须属于同一科目")
        if current_id is not None and payload.parent_id == current_id:
            raise HTTPException(status_code=400, detail="知识点不能把自己作为父级")


def _apply_alias_payload(session: Session, alias: KnowledgePointAlias, payload: KnowledgePointAliasPayload) -> None:
    subject = session.get(Subject, payload.subject_id)
    if not subject:
        raise HTTPException(status_code=404, detail="学科不存在")
    point = session.get(KnowledgePoint, payload.knowledge_point_id)
    if not point or point.subject_id != payload.subject_id:
        raise HTTPException(status_code=400, detail="标准知识点必须属于同一科目")
    alias_name = clean_text(payload.alias_name)
    if not alias_name:
        raise HTTPException(status_code=400, detail="别名不能为空")
    alias.subject_id = payload.subject_id
    alias.knowledge_point_id = payload.knowledge_point_id
    alias.alias_name = alias_name
    alias.normalized_alias = normalize_lookup_text(alias_name)
    alias.source_type = clean_text(payload.source_type) or "manual"
    alias.note = clean_text(payload.note)
    alias.is_active = payload.is_active
    _ensure_alias_normalized_unique(session, alias)


def _apply_error_tag_payload(tag: ErrorReasonTag, payload: ErrorReasonTagPayload) -> None:
    name = clean_text(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="错因标签名称不能为空")
    tag.name = name
    tag.normalized_name = normalize_lookup_text(name)
    tag.description = clean_text(payload.description)
    tag.sort_order = payload.sort_order
    tag.is_builtin = payload.is_builtin
    tag.is_active = payload.is_active


def _ensure_alias_normalized_unique(session: Session, alias: KnowledgePointAlias) -> None:
    conditions = [
        KnowledgePointAlias.subject_id == alias.subject_id,
        KnowledgePointAlias.normalized_alias == alias.normalized_alias,
        KnowledgePointAlias.is_active.is_(True),
    ]
    if alias.id is not None:
        conditions.append(KnowledgePointAlias.id != alias.id)
    with session.no_autoflush:
        existing = session.scalar(select(KnowledgePointAlias).where(*conditions))
    if existing:
        raise HTTPException(status_code=400, detail="同一科目下别名不能重复。")


def _resolve_subject(session: Session, subject_text: str | None) -> Subject:
    name = clean_text(subject_text)
    if not name:
        raise ValueError("科目不能为空")
    subject = session.scalar(select(Subject).where((Subject.name == name) | (Subject.code == name)))
    if subject is None:
        raise ValueError(f"科目不存在: {name}")
    return subject


def _flush_or_400(session: Session, detail: str = "数据存在重复或约束冲突。") -> None:
    try:
        session.flush()
    except IntegrityError as exc:
        raise HTTPException(status_code=400, detail=detail) from exc


def _to_int(value: object, *, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default
