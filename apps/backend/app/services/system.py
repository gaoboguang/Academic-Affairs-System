from __future__ import annotations

import json
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import HTTPException, Request
from openpyxl import load_workbook
from sqlalchemy import desc, func, or_, select
from sqlalchemy.orm import Session

from app.core.bootstrap import ensure_runtime_directories
from app.db.session import DatabaseManager
from app.importers.base import normalize_import_status
from app.models import (
    AuditLog,
    BackupRecord,
    ConfigItem,
    EvaluationBatch,
    EvaluationResponse,
    ImportJob,
    ReportExportRecord,
    ScoreImportBatch,
    StoredFile,
    TimetableBatch,
    TimetableEntry,
)
from app.repositories.system import (
    get_backup_record,
    get_stored_file,
    list_audit_logs as repo_list_audit_logs,
    list_backups as repo_list_backups,
    write_audit_log,
)
from app.schemas.archive import StoredFileRead
from app.schemas.system import (
    AuditLogRead,
    BackupCreateResponse,
    BackupRecordRead,
    BackupRestorePayload,
    BackupVerificationRead,
    DataRepairExecutePayload,
    DataRepairExecuteResponse,
    DataRepairScanRead,
    ImportCenterBatchDetailRead,
    ImportCenterBatchRead,
    ImportCenterErrorItemRead,
    ImportCenterResponse,
    ImportCenterSummaryRead,
    ImportCenterTemplateRead,
    SystemConfigGroupRead,
    SystemConfigItemRead,
    SystemConfigItemUpdatePayload,
    SystemSafetyStatusRead,
    SystemTemplateRead,
)
from app.services.data_quality import build_data_repair_scan, execute_repair_action
from app.utils.files import resolve_allowed_file_path, resolve_named_file_in_directory, validate_upload_category
from app.utils.parsers import make_timestamped_filename, relative_to_project

CONFIG_GROUP_TITLES = {
    "analytics": "成绩分析参数",
    "recommendation": "推荐策略参数",
    "system": "系统运行参数",
}

IMPORT_CENTER_TYPES = {
    "students": {
        "label": "学生信息导入",
        "business_path": "/students",
        "template": "students_import_template.xlsx",
        "guidance": "先下载学生模板，按学号或学籍号维护唯一身份，再回到学生中心上传。",
        "rollback": "学生导入当前没有逐行回滚记录；需要撤销时优先恢复导入前备份，或按学号重新导入正确台账覆盖修正。",
    },
    "teachers": {
        "label": "教师信息导入",
        "business_path": "/teachers",
        "template": "teachers_import_template.xlsx",
        "guidance": "先下载教师模板，使用工号作为稳定身份，再回到教师中心上传。",
        "rollback": "教师导入当前没有逐行回滚记录；需要撤销时优先恢复导入前备份，或按工号重新导入正确台账覆盖修正。",
    },
    "scores": {
        "label": "成绩导入",
        "business_path": "/exams",
        "template": "exam_scores_import_template.xlsx",
        "guidance": "先在考试成绩页创建考试并配置科目，再下载成绩模板上传。",
        "rollback": "成绩导入可用同一考试的覆盖/跳过策略重新导入修正；当前不自动删除批次内成绩，避免误删后续分析快照。",
    },
    "timetable": {
        "label": "课表导入",
        "business_path": "/workload",
        "template": "timetable_import_template.xlsx",
        "guidance": "先下载课表模板，确认学期、教师、班级和学科名称，再回到课表工作量页上传。",
        "rollback": "课表工作量默认读取最新有效课表批次；需要撤销时重新导入正确批次并复核未匹配项，当前不自动删除历史批次。",
    },
    "admissions": {
        "label": "录取数据导入",
        "business_path": "/recommendations",
        "template": "admission_records_import_template.xlsx",
        "guidance": "先下载录取数据模板，确认省份、年份、批次、院校和专业口径，再回到高考志愿页上传。",
        "rollback": "录取数据会影响推荐参考口径；当前没有逐行回滚记录，导入前应先创建备份。",
    },
    "enrollment_plans": {
        "label": "招生计划导入",
        "business_path": "/recommendations",
        "template": "enrollment_plans_import_template.xlsx",
        "guidance": "先下载招生计划模板，确认省份、年份、批次、选科和计划数，再回到高考志愿页上传。",
        "rollback": "招生计划会影响候选池；当前没有逐行回滚记录，导入前应先创建备份。",
    },
    "evaluation": {
        "label": "评教数据导入",
        "business_path": "/evaluation-quant",
        "template": "evaluation_import_template.xlsx",
        "guidance": "先在评教量化页确认模板和学期，再下载评教模板上传。",
        "rollback": "评教批次当前保留历史记录；需要撤销时恢复备份或重新导入正确批次并重新分析。",
    },
}

IMPORT_JOB_TYPE_ALIASES = {
    "student_import": "students",
    "teacher_import": "teachers",
    "exam_score_import": "scores",
    "score_import": "scores",
    "timetable_import": "timetable",
    "admission_import": "admissions",
    "evaluation_import": "evaluation",
}

IMPORT_SOURCE_LABELS = {
    "import_job": "通用导入任务",
    "score_import_batch": "成绩批次",
    "timetable_batch": "课表批次",
    "evaluation_batch": "评教批次",
}


def _serialize_stored_file(item: StoredFile) -> StoredFileRead:
    return StoredFileRead(
        id=item.id,
        original_filename=item.original_filename,
        file_path=item.file_path,
        content_type=item.content_type,
        file_size=item.file_size,
        category=item.category,
        created_at=item.created_at,
        download_url=f"/api/files/{item.id}",
    )


def _serialize_backup(item: BackupRecord) -> BackupRecordRead:
    return BackupRecordRead(
        id=item.id,
        backup_name=item.backup_name,
        file_path=item.file_path,
        file_size=item.file_size,
        created_at=item.created_at,
        status=item.status,
        download_url=f"/api/system/backups/{item.id}/download",
    )


def list_backups(session: Session) -> list[BackupRecordRead]:
    return [_serialize_backup(item) for item in repo_list_backups(session)]


def get_system_safety_status(session: Session, settings) -> SystemSafetyStatusRead:
    backups = repo_list_backups(session)
    latest_backup = _serialize_backup(backups[0]) if backups else None
    latest_restore = session.scalar(
        select(AuditLog)
        .where(AuditLog.module == "system", AuditLog.action == "restore")
        .order_by(desc(AuditLog.created_at), desc(AuditLog.id))
        .limit(1)
    )
    sqlite_integrity = _check_sqlite_integrity(settings.db_path)
    warnings: list[str] = []
    data_dir_size = _safe_directory_size(settings.data_dir)
    backup_dir_size = _safe_directory_size(settings.backups_dir)
    if latest_backup is None:
        warnings.append("当前还没有备份记录，导入真实数据或恢复前建议先创建备份。")
    if sqlite_integrity != "ok":
        warnings.append(f"SQLite 完整性检查结果为 {sqlite_integrity}，建议先停止写入并备份现场。")
    if data_dir_size >= 5 * 1024 * 1024 * 1024:
        warnings.append("data 目录已超过 5GB，建议检查附件、导出和备份占用。")
    if len(backups) >= 20:
        warnings.append("备份包数量较多，建议人工确认哪些需要长期保留；系统不会自动删除。")

    return SystemSafetyStatusRead(
        main_db_path=str(settings.db_path),
        main_db_size=_safe_file_size(settings.db_path),
        data_dir_path=str(settings.data_dir),
        data_dir_size=data_dir_size,
        backup_dir_path=str(settings.backups_dir),
        backup_dir_size=backup_dir_size,
        backup_count=len(backups),
        latest_backup=latest_backup,
        latest_restore_at=latest_restore.created_at if latest_restore else None,
        sqlite_integrity=sqlite_integrity,
        alembic_version=_read_alembic_version(settings.db_path),
        warnings=warnings,
    )


def verify_backup(
    session: Session,
    settings,
    backup_id: int,
    *,
    restore_dry_run: bool = False,
) -> BackupVerificationRead:
    record = get_backup_record(session, backup_id)
    if not record:
        raise HTTPException(status_code=404, detail="备份记录不存在")
    backup_path = resolve_allowed_file_path(
        record.file_path,
        allowed_roots=[settings.backups_dir],
        project_root=settings.project_root,
        not_found_detail="备份文件不存在",
    )
    return _verify_backup_file(
        record,
        backup_path,
        settings.db_path,
        restore_dry_run=restore_dry_run,
    )


def list_audit_logs(session: Session, limit: int = 100) -> list[AuditLogRead]:
    return [AuditLogRead.model_validate(item) for item in repo_list_audit_logs(session, limit=limit)]


def list_config_groups(session: Session) -> list[SystemConfigGroupRead]:
    rows = session.scalars(
        select(ConfigItem)
        .where(ConfigItem.is_active.is_(True))
        .order_by(ConfigItem.config_group, ConfigItem.config_key)
    ).all()
    grouped: dict[str, list[SystemConfigItemRead]] = {}
    for row in rows:
        grouped.setdefault(row.config_group, []).append(_serialize_config_item(row))
    return [
        SystemConfigGroupRead(
            config_group=config_group,
            title=CONFIG_GROUP_TITLES.get(config_group, config_group.replace("_", " ").title()),
            items=items,
        )
        for config_group, items in grouped.items()
    ]


def update_config_items(
    session: Session,
    payloads: list[SystemConfigItemUpdatePayload],
) -> list[SystemConfigGroupRead]:
    existing = {
        (row.config_group, row.config_key): row
        for row in session.scalars(select(ConfigItem)).all()
    }
    for payload in payloads:
        key = (payload.config_group, payload.config_key)
        config_value = _stringify_config_value(payload.config_value, payload.value_type)
        row = existing.get(key)
        if row:
            row.config_value = config_value
            row.value_type = payload.value_type
            if payload.description is not None:
                row.description = payload.description
            row.is_active = True
            continue
        session.add(
            ConfigItem(
                config_group=payload.config_group,
                config_key=payload.config_key,
                config_value=config_value,
                value_type=payload.value_type,
                description=payload.description,
            )
        )
    session.flush()
    write_audit_log(
        session,
        module="system",
        action="update_config",
        target_type="config_group",
        detail_json={"item_count": len(payloads)},
    )
    return list_config_groups(session)


def list_templates(settings) -> list[SystemTemplateRead]:
    if not settings.templates_dir.exists():
        return []
    rows: list[SystemTemplateRead] = []
    for path in sorted(settings.templates_dir.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True):
        if not path.is_file():
            continue
        stat = path.stat()
        rows.append(
            SystemTemplateRead(
                name=path.stem.replace("_", " "),
                file_name=path.name,
                file_size=stat.st_size,
                updated_at=datetime.fromtimestamp(stat.st_mtime),
                download_url=f"/api/system/templates/{path.name}/download",
            )
        )
    return rows


def get_template_path(settings, template_name: str) -> Path:
    return resolve_named_file_in_directory(
        settings.templates_dir,
        template_name,
        not_found_detail="模板文件不存在",
    )


def list_import_center_batches(
    session: Session,
    settings,
    *,
    limit: int = 100,
    job_type: str | None = None,
    status: str | None = None,
) -> ImportCenterResponse:
    batches = _build_import_center_batches(session, settings, limit=max(limit, 20))
    if job_type:
        normalized_job_type = _normalize_import_job_type(job_type)
        batches = [item for item in batches if item.job_type == normalized_job_type]
    if status:
        normalized_status = normalize_import_status(status)
        batches = [item for item in batches if item.status == normalized_status]
    batches = batches[:limit]
    return ImportCenterResponse(
        generated_at=datetime.now(),
        summary=_build_import_center_summary(batches),
        latest_backup=_get_latest_backup(session),
        templates=_build_import_center_templates(),
        batches=batches,
    )


def get_import_center_batch_detail(
    session: Session,
    settings,
    *,
    source_type: str,
    batch_id: int,
) -> ImportCenterBatchDetailRead:
    batch = _find_import_center_batch(session, settings, source_type=source_type, batch_id=batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="导入批次不存在")
    result_json = _read_batch_result_json(session, batch)
    return ImportCenterBatchDetailRead(
        batch=batch,
        result_json=result_json,
        audit_logs=_list_import_batch_audit_logs(session, batch),
        error_preview=_string_list(result_json.get("error_preview") if result_json else None),
        error_items=_build_import_error_items(settings, batch, result_json),
        notice_preview=_string_list(result_json.get("notice_preview") if result_json else None),
        rollback_steps=_build_rollback_steps(batch),
    )


def _get_latest_backup(session: Session) -> BackupRecordRead | None:
    item = session.scalar(select(BackupRecord).order_by(desc(BackupRecord.created_at), desc(BackupRecord.id)).limit(1))
    return _serialize_backup(item) if item else None


def _build_import_error_items(
    settings,
    batch: ImportCenterBatchRead,
    result_json: dict | None,
) -> list[ImportCenterErrorItemRead]:
    explicit_items = result_json.get("error_items") if result_json else None
    if isinstance(explicit_items, list):
        rows: list[ImportCenterErrorItemRead] = []
        for item in explicit_items[:10]:
            if not isinstance(item, dict):
                continue
            message = str(item.get("message") or item.get("error") or "").strip()
            if not message:
                continue
            rows.append(
                ImportCenterErrorItemRead(
                    row_number=_optional_int(item.get("row_number") or item.get("row")),
                    field_name=_optional_str(item.get("field_name") or item.get("field") or item.get("column_name")),
                    raw_value=item.get("raw_value"),
                    message=message,
                    suggestion=_optional_str(item.get("suggestion")),
                )
            )
        if rows:
            return rows

    if batch.error_report_path:
        rows = _read_error_items_from_report(settings, batch.error_report_path)
        if rows:
            return rows

    return [
        ImportCenterErrorItemRead(row_number=_extract_error_row_number(item), message=item)
        for item in _string_list(result_json.get("error_preview") if result_json else None)[:10]
    ]


def _read_error_items_from_report(settings, error_report_path: str) -> list[ImportCenterErrorItemRead]:
    try:
        path = resolve_allowed_file_path(
            error_report_path,
            allowed_roots=[settings.logs_dir],
            project_root=settings.project_root,
        )
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["错误报告"] if "错误报告" in workbook.sheetnames else workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(value).strip() if value is not None else "" for value in (header_row or [])]
        index = {header: offset for offset, header in enumerate(headers)}
        rows: list[ImportCenterErrorItemRead] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            message = _cell_value(values, index, "错误原因")
            if not message:
                continue
            rows.append(
                ImportCenterErrorItemRead(
                    row_number=_optional_int(_cell_value(values, index, "行号")),
                    field_name=_optional_str(_cell_value(values, index, "字段名") or _cell_value(values, index, "列名")),
                    raw_value=_cell_value(values, index, "原始值"),
                    message=str(message),
                    suggestion=_optional_str(_cell_value(values, index, "建议修复")),
                )
            )
            if len(rows) >= 10:
                break
        return rows
    except Exception:
        return []


def _cell_value(values: tuple, index: dict[str, int], header: str):
    offset = index.get(header)
    if offset is None or offset >= len(values):
        return None
    return values[offset]


def _extract_error_row_number(value: str) -> int | None:
    if not value.startswith("第 ") or " 行" not in value:
        return None
    raw = value.removeprefix("第 ").split(" 行", 1)[0]
    return _optional_int(raw)


def _build_import_center_templates() -> list[ImportCenterTemplateRead]:
    rows: list[ImportCenterTemplateRead] = []
    for job_type, meta in IMPORT_CENTER_TYPES.items():
        template_file = str(meta["template"])
        rows.append(
            ImportCenterTemplateRead(
                job_type=job_type,
                job_type_label=str(meta["label"]),
                template_name=str(meta["label"]),
                file_name=template_file,
                download_url=_runtime_file_url(f"data/templates/{template_file}"),
                business_path=str(meta["business_path"]),
                guidance=str(meta["guidance"]),
            )
        )
    return rows


def _build_import_center_summary(batches: list[ImportCenterBatchRead]) -> ImportCenterSummaryRead:
    return ImportCenterSummaryRead(
        total_batches=len(batches),
        failed_batches=sum(1 for item in batches if item.status == "failed"),
        partial_batches=sum(1 for item in batches if item.status == "partially_failed"),
        error_report_count=sum(1 for item in batches if item.error_report_path),
    )


def _build_import_center_batches(session: Session, settings, *, limit: int) -> list[ImportCenterBatchRead]:
    rows: list[ImportCenterBatchRead] = []
    related_keys: set[tuple[str, int]] = set()
    jobs = session.scalars(
        select(ImportJob).order_by(desc(ImportJob.started_at), desc(ImportJob.created_at), desc(ImportJob.id)).limit(limit)
    ).all()
    for job in jobs:
        row = _serialize_import_job_batch(job)
        rows.append(row)
        related_key = _related_batch_key(row, job.result_json)
        if related_key:
            related_keys.add(related_key)

    for score_batch in session.scalars(
        select(ScoreImportBatch).order_by(desc(ScoreImportBatch.import_time), desc(ScoreImportBatch.id)).limit(limit)
    ).all():
        key = ("score_import_batch", score_batch.id)
        if key not in related_keys:
            rows.append(_serialize_score_batch(score_batch))

    for timetable_batch in session.scalars(
        select(TimetableBatch).order_by(desc(TimetableBatch.import_time), desc(TimetableBatch.id)).limit(limit)
    ).all():
        key = ("timetable_batch", timetable_batch.id)
        if key not in related_keys:
            rows.append(_serialize_timetable_batch(session, timetable_batch))

    for evaluation_batch in session.scalars(
        select(EvaluationBatch).order_by(desc(EvaluationBatch.import_time), desc(EvaluationBatch.id)).limit(limit)
    ).all():
        key = ("evaluation_batch", evaluation_batch.id)
        if key not in related_keys:
            rows.append(_serialize_evaluation_batch(session, evaluation_batch))

    return sorted(rows, key=_import_batch_sort_key, reverse=True)[:limit]


def _find_import_center_batch(
    session: Session,
    settings,
    *,
    source_type: str,
    batch_id: int,
) -> ImportCenterBatchRead | None:
    if source_type == "import_job":
        item = session.get(ImportJob, batch_id)
        return _serialize_import_job_batch(item) if item else None
    if source_type == "score_import_batch":
        item = session.get(ScoreImportBatch, batch_id)
        return _serialize_score_batch(item) if item else None
    if source_type == "timetable_batch":
        item = session.get(TimetableBatch, batch_id)
        return _serialize_timetable_batch(session, item) if item else None
    if source_type == "evaluation_batch":
        item = session.get(EvaluationBatch, batch_id)
        return _serialize_evaluation_batch(session, item) if item else None
    return None


def _serialize_import_job_batch(job: ImportJob) -> ImportCenterBatchRead:
    result = job.result_json or {}
    job_type = _normalize_import_job_type(job.job_type)
    meta = _import_type_meta(job_type)
    status = normalize_import_status(str(result.get("status") or job.status))
    return ImportCenterBatchRead(
        id=f"import_job:{job.id}",
        numeric_id=job.id,
        source_type="import_job",
        source_type_label=IMPORT_SOURCE_LABELS["import_job"],
        job_type=job_type,
        job_type_label=str(meta["label"]),
        source_filename=job.source_filename,
        status=status,
        started_at=job.started_at or job.created_at,
        finished_at=job.finished_at,
        total_rows=_optional_int(result.get("total_rows")),
        success_rows=_optional_int(result.get("success_rows")),
        failed_rows=_optional_int(result.get("failed_rows")),
        skipped_rows=_int_value(result.get("skipped_rows")),
        created_rows=_int_value(result.get("created_rows")),
        updated_rows=_int_value(result.get("updated_rows")),
        error_report_path=_optional_str(result.get("error_report_path")),
        business_path=str(meta["business_path"]),
        template_download_url=_template_url(meta),
        rollback_supported=False,
        rollback_hint=str(meta["rollback"]),
        detail_summary=_build_detail_summary(result),
    )


def _serialize_score_batch(batch: ScoreImportBatch) -> ImportCenterBatchRead:
    meta = _import_type_meta("scores")
    return ImportCenterBatchRead(
        id=f"score_import_batch:{batch.id}",
        numeric_id=batch.id,
        source_type="score_import_batch",
        source_type_label=IMPORT_SOURCE_LABELS["score_import_batch"],
        job_type="scores",
        job_type_label=str(meta["label"]),
        source_filename=batch.source_filename,
        status=normalize_import_status(batch.status),
        started_at=batch.import_time,
        finished_at=None,
        total_rows=batch.total_rows,
        success_rows=batch.success_rows,
        failed_rows=batch.failed_rows,
        skipped_rows=0,
        error_report_path=batch.error_report_path,
        business_path=str(meta["business_path"]),
        template_download_url=_template_url(meta),
        rollback_supported=False,
        rollback_hint=str(meta["rollback"]),
        detail_summary=f"考试 ID {batch.exam_id}，成功 {batch.success_rows} 行，失败 {batch.failed_rows} 行。",
    )


def _serialize_timetable_batch(session: Session, batch: TimetableBatch) -> ImportCenterBatchRead:
    meta = _import_type_meta("timetable")
    total_rows = session.scalar(select(func.count()).select_from(TimetableEntry).where(TimetableEntry.batch_id == batch.id)) or 0
    unresolved_rows = (
        session.scalar(
            select(func.count())
            .select_from(TimetableEntry)
            .where(TimetableEntry.batch_id == batch.id, TimetableEntry.mapping_status != "matched")
        )
        or 0
    )
    return ImportCenterBatchRead(
        id=f"timetable_batch:{batch.id}",
        numeric_id=batch.id,
        source_type="timetable_batch",
        source_type_label=IMPORT_SOURCE_LABELS["timetable_batch"],
        job_type="timetable",
        job_type_label=str(meta["label"]),
        source_filename=batch.source_filename,
        status=normalize_import_status(batch.status),
        started_at=batch.import_time,
        finished_at=None,
        total_rows=total_rows,
        success_rows=total_rows - unresolved_rows,
        failed_rows=unresolved_rows,
        skipped_rows=0,
        error_report_path=None,
        business_path=str(meta["business_path"]),
        template_download_url=_template_url(meta),
        rollback_supported=False,
        rollback_hint=str(meta["rollback"]),
        detail_summary=f"学期 ID {batch.semester_id}，课表条目 {total_rows} 条，待修正 {unresolved_rows} 条。",
    )


def _serialize_evaluation_batch(session: Session, batch: EvaluationBatch) -> ImportCenterBatchRead:
    meta = _import_type_meta("evaluation")
    response_count = (
        session.scalar(select(func.count()).select_from(EvaluationResponse).where(EvaluationResponse.batch_id == batch.id))
        or 0
    )
    return ImportCenterBatchRead(
        id=f"evaluation_batch:{batch.id}",
        numeric_id=batch.id,
        source_type="evaluation_batch",
        source_type_label=IMPORT_SOURCE_LABELS["evaluation_batch"],
        job_type="evaluation",
        job_type_label=str(meta["label"]),
        source_filename=batch.source_filename,
        status=normalize_import_status(batch.status),
        started_at=batch.import_time,
        finished_at=None,
        total_rows=response_count,
        success_rows=response_count,
        failed_rows=0,
        skipped_rows=0,
        error_report_path=None,
        business_path=str(meta["business_path"]),
        template_download_url=_template_url(meta),
        rollback_supported=False,
        rollback_hint=str(meta["rollback"]),
        detail_summary=f"模板 ID {batch.template_id}，评教响应 {response_count} 条。",
    )


def _read_batch_result_json(session: Session, batch: ImportCenterBatchRead) -> dict | None:
    if batch.source_type == "import_job":
        item = session.get(ImportJob, batch.numeric_id)
        return item.result_json if item else None
    related_job = _find_related_import_job(session, batch)
    return related_job.result_json if related_job else None


def _find_related_import_job(session: Session, batch: ImportCenterBatchRead) -> ImportJob | None:
    jobs = session.scalars(
        select(ImportJob)
        .where(ImportJob.job_type.in_(_job_type_aliases_for(batch.job_type)))
        .order_by(desc(ImportJob.started_at), desc(ImportJob.created_at), desc(ImportJob.id))
        .limit(200)
    ).all()
    for job in jobs:
        if _related_batch_key(_serialize_import_job_batch(job), job.result_json) == (batch.source_type, batch.numeric_id):
            return job
    return None


def _list_import_batch_audit_logs(session: Session, batch: ImportCenterBatchRead) -> list[AuditLogRead]:
    targets = [(batch.source_type, str(batch.numeric_id))]
    if batch.source_type != "import_job":
        related_job = _find_related_import_job(session, batch)
        if related_job:
            targets.append(("import_job", str(related_job.id)))
    conditions = [
        (AuditLog.target_type == target_type) & (AuditLog.target_id == target_id)
        for target_type, target_id in targets
    ]
    if not conditions:
        return []
    logs = session.scalars(
        select(AuditLog)
        .where(or_(*conditions))
        .order_by(desc(AuditLog.created_at), desc(AuditLog.id))
        .limit(20)
    ).all()
    return [AuditLogRead.model_validate(item) for item in logs]


def _related_batch_key(batch: ImportCenterBatchRead, result_json: dict | None) -> tuple[str, int] | None:
    if not result_json:
        return None
    batch_id = _optional_int(result_json.get("batch_id"))
    if batch_id is None:
        return None
    if batch.job_type == "scores":
        return ("score_import_batch", batch_id)
    if batch.job_type == "timetable":
        return ("timetable_batch", batch_id)
    if batch.job_type == "evaluation":
        return ("evaluation_batch", batch_id)
    return None


def _build_rollback_steps(batch: ImportCenterBatchRead) -> list[str]:
    if batch.rollback_supported:
        return ["确认影响范围", "创建当前备份", "执行批次回滚", "重新查看业务页面和审计日志"]
    return [
        "当前批次不支持自动逐行回滚，页面只给出撤销说明，不直接删除业务数据。",
        "如必须恢复导入前状态，先到系统设置创建当前备份，再选择导入前备份恢复。",
        "如果只是少量字段错误，优先回到对应业务页面下载模板，按稳定标识重新导入正确文件覆盖修正。",
    ]


def _build_detail_summary(result: dict) -> str:
    total_rows = _optional_int(result.get("total_rows"))
    success_rows = _optional_int(result.get("success_rows"))
    failed_rows = _optional_int(result.get("failed_rows"))
    if total_rows is None:
        return "该批次未记录行级摘要，请查看业务页或审计日志。"
    return f"共 {total_rows} 行，成功 {success_rows or 0} 行，失败 {failed_rows or 0} 行。"


def _normalize_import_job_type(job_type: str) -> str:
    return IMPORT_JOB_TYPE_ALIASES.get(job_type, job_type)


def _job_type_aliases_for(job_type: str) -> list[str]:
    aliases = [job_type]
    aliases.extend(alias for alias, target in IMPORT_JOB_TYPE_ALIASES.items() if target == job_type)
    return aliases


def _import_type_meta(job_type: str) -> dict[str, str]:
    return IMPORT_CENTER_TYPES.get(
        job_type,
        {
            "label": job_type.replace("_", " "),
            "business_path": "/system-tools",
            "template": "",
            "guidance": "该导入类型暂无专用模板说明，请查看对应业务页面。",
            "rollback": "该导入类型暂无自动回滚能力，建议通过备份恢复或业务页手工修正。",
        },
    )


def _template_url(meta: dict[str, str]) -> str | None:
    template_file = str(meta.get("template") or "")
    if not template_file:
        return None
    return _runtime_file_url(f"data/templates/{template_file}")


def _runtime_file_url(path: str) -> str:
    from urllib.parse import quote

    return f"/api/system/files?path={quote(path, safe='')}"


def _import_batch_sort_key(item: ImportCenterBatchRead) -> tuple[datetime, int]:
    return (item.started_at or item.finished_at or datetime.min, item.numeric_id)


def _optional_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _int_value(value: object) -> int:
    return _optional_int(value) or 0


def _optional_str(value: object) -> str | None:
    if value in (None, ""):
        return None
    return str(value)


def _string_list(value: object) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value]


def get_data_repair_scan(session: Session) -> DataRepairScanRead:
    return build_data_repair_scan(session)


def execute_data_repair(
    session: Session,
    payload: DataRepairExecutePayload,
) -> DataRepairExecuteResponse:
    repaired_count, message = execute_repair_action(session, payload.action_code)
    return DataRepairExecuteResponse(
        action_code=payload.action_code,
        repaired_count=repaired_count,
        message=message,
        scan=build_data_repair_scan(session),
    )


def upload_file(
    session: Session,
    settings,
    *,
    filename: str | None,
    content: bytes,
    content_type: str | None,
    category: str = "general",
) -> StoredFileRead:
    if not filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")
    safe_category = validate_upload_category(category)
    safe_name = Path(filename).name
    suffix = Path(safe_name).suffix
    stored_name = f"{uuid4().hex}{suffix}"
    target_dir = settings.uploads_dir / safe_category
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / stored_name
    target_path.write_bytes(content)
    record = StoredFile(
        original_filename=safe_name,
        file_path=relative_to_project(target_path, settings.project_root),
        content_type=content_type,
        file_size=len(content),
        category=safe_category,
    )
    session.add(record)
    session.flush()
    write_audit_log(
        session,
        module="files",
        action="upload",
        target_type="stored_file",
        target_id=str(record.id),
        detail_json={"category": safe_category, "original_filename": safe_name},
    )
    return _serialize_stored_file(record)


def get_file_path(session: Session, settings, file_id: int) -> Path:
    item = get_stored_file(session, file_id)
    if not item or not item.is_active:
        raise HTTPException(status_code=404, detail="文件不存在")
    return resolve_allowed_file_path(
        item.file_path,
        allowed_roots=[settings.uploads_dir],
        project_root=settings.project_root,
        not_found_detail="文件不存在",
    )


def create_backup(request: Request) -> BackupCreateResponse:
    settings = request.app.state.settings
    ensure_runtime_directories(settings)
    with TemporaryDirectory() as tmp_dir:
        tmp_root = Path(tmp_dir)
        snapshot_db = tmp_root / "app.db"
        _snapshot_sqlite_db(settings.db_path, snapshot_db)
        manifest = {
            "app_name": settings.app_name,
            "version": "0.1.0",
            "database_file": "db/app.db",
            "created_from": str(settings.project_root),
        }
        backup_name = make_timestamped_filename("local_edu_backup", ".zip")
        backup_path = settings.backups_dir / backup_name
        with ZipFile(backup_path, "w", compression=ZIP_DEFLATED) as archive:
            archive.write(snapshot_db, arcname="db/app.db")
            _add_directory_to_zip(archive, settings.uploads_dir, "uploads")
            _add_directory_to_zip(archive, settings.templates_dir, "templates")
            env_path = settings.project_root / ".env"
            if env_path.exists():
                archive.write(env_path, arcname="config/.env")
            archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    file_size = backup_path.stat().st_size if backup_path.exists() else 0
    with request.app.state.db.session_scope() as session:
        record = BackupRecord(
            backup_name=backup_name,
            file_path=relative_to_project(backup_path, settings.project_root),
            file_size=file_size,
            status="success",
        )
        session.add(record)
        session.flush()
        write_audit_log(
            session,
            module="system",
            action="backup",
            target_type="backup_record",
            target_id=str(record.id),
            detail_json={"backup_name": backup_name},
        )
        backup_id = record.id
    return BackupCreateResponse(message="备份完成", backup_id=backup_id)


def restore_backup(request: Request, payload: BackupRestorePayload) -> dict[str, str]:
    settings = request.app.state.settings
    if payload.auto_backup_current:
        create_backup(request)

    with request.app.state.db.session_scope() as session:
        record = get_backup_record(session, payload.backup_id)
        if not record:
            raise HTTPException(status_code=404, detail="备份记录不存在")
        backup_path = resolve_allowed_file_path(
            record.file_path,
            allowed_roots=[settings.backups_dir],
            project_root=settings.project_root,
            not_found_detail="备份文件不存在",
        )

    with TemporaryDirectory() as tmp_dir:
        tmp_root = Path(tmp_dir)
        with ZipFile(backup_path, "r") as archive:
            if "manifest.json" not in archive.namelist() or "db/app.db" not in archive.namelist():
                raise HTTPException(status_code=400, detail="备份包结构无效")
            archive.extractall(tmp_root)
        restored_db = tmp_root / "db" / "app.db"
        if not restored_db.exists():
            raise HTTPException(status_code=400, detail="备份包缺少数据库文件")

        request.app.state.db.dispose()
        _replace_file(restored_db, settings.db_path)
        _replace_directory(tmp_root / "uploads", settings.uploads_dir)
        _replace_directory(tmp_root / "templates", settings.templates_dir)
        env_path = tmp_root / "config" / ".env"
        if env_path.exists():
            _replace_file(env_path, settings.project_root / ".env")

    request.app.state.db = DatabaseManager(settings.database_url, settings.debug)
    request.app.state.db.initialize()
    ensure_runtime_directories(settings)
    with request.app.state.db.session_scope() as session:
        write_audit_log(
            session,
            module="system",
            action="restore",
            target_type="backup_record",
            target_id=str(payload.backup_id),
            detail_json={"backup_id": payload.backup_id},
        )
    return {"message": "恢复完成"}


def get_backup_path(session: Session, settings, backup_id: int) -> Path:
    record = get_backup_record(session, backup_id)
    if not record:
        raise HTTPException(status_code=404, detail="备份记录不存在")
    return resolve_allowed_file_path(
        record.file_path,
        allowed_roots=[settings.backups_dir],
        project_root=settings.project_root,
        not_found_detail="备份文件不存在",
    )


def _serialize_config_item(item: ConfigItem) -> SystemConfigItemRead:
    return SystemConfigItemRead(
        id=item.id,
        config_group=item.config_group,
        config_key=item.config_key,
        config_value=item.config_value,
        parsed_value=_parse_config_value(item.config_value, item.value_type),
        value_type=item.value_type,
        description=item.description,
    )


def _parse_config_value(value: str, value_type: str):
    if value_type == "bool":
        return str(value).strip().lower() in {"1", "true", "yes", "on"}
    if value_type == "int":
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return value
    if value_type == "float":
        try:
            return float(str(value).strip())
        except (TypeError, ValueError):
            return value
    if value_type == "json":
        try:
            return json.loads(value)
        except (TypeError, json.JSONDecodeError):
            return value
    return value


def _stringify_config_value(value, value_type: str) -> str:
    if value_type == "bool":
        if isinstance(value, str):
            normalized = value.strip().lower() in {"1", "true", "yes", "on"}
        else:
            normalized = bool(value)
        return "true" if normalized else "false"
    if value_type == "json":
        if isinstance(value, str):
            json.loads(value)
            return value
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


def _snapshot_sqlite_db(source_path: Path, target_path: Path) -> None:
    source = sqlite3.connect(source_path)
    target = sqlite3.connect(target_path)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()


def _verify_backup_file(
    record: BackupRecord,
    backup_path: Path,
    overwrite_path: Path,
    *,
    restore_dry_run: bool,
) -> BackupVerificationRead:
    try:
        with TemporaryDirectory() as tmp_dir:
            tmp_root = Path(tmp_dir)
            with ZipFile(backup_path, "r") as archive:
                names = set(archive.namelist())
                if "manifest.json" not in names or "db/app.db" not in names:
                    return BackupVerificationRead(
                        backup_id=record.id,
                        backup_name=record.backup_name,
                        valid=False,
                        message="备份包结构无效，缺少 manifest.json 或 db/app.db。",
                        file_size=_safe_file_size(backup_path),
                        will_overwrite_path=str(overwrite_path),
                    )
                archive.extract("manifest.json", tmp_root)
                archive.extract("db/app.db", tmp_root)
            manifest_path = tmp_root / "manifest.json"
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            restored_db = tmp_root / "db" / "app.db"
            integrity = _check_sqlite_integrity(restored_db)
            valid = integrity == "ok"
            message = "恢复演练通过，备份包可读取，临时库完整性为 ok。" if restore_dry_run else "备份包校验通过。"
            if not valid:
                message = f"备份包可读取，但临时库完整性检查结果为 {integrity}。"
            return BackupVerificationRead(
                backup_id=record.id,
                backup_name=record.backup_name,
                valid=valid,
                message=message,
                file_size=_safe_file_size(backup_path),
                will_overwrite_path=str(overwrite_path),
                sqlite_integrity=integrity,
                manifest=manifest if isinstance(manifest, dict) else None,
            )
    except (OSError, json.JSONDecodeError, sqlite3.Error, KeyError) as exc:
        return BackupVerificationRead(
            backup_id=record.id,
            backup_name=record.backup_name,
            valid=False,
            message=f"备份包校验失败：{exc}",
            file_size=_safe_file_size(backup_path),
            will_overwrite_path=str(overwrite_path),
        )


def _check_sqlite_integrity(db_path: Path) -> str:
    if not db_path.exists():
        return "missing"
    try:
        with sqlite3.connect(str(db_path)) as connection:
            row = connection.execute("PRAGMA integrity_check").fetchone()
            return str(row[0]) if row and row[0] is not None else "unknown"
    except sqlite3.Error as exc:
        return f"error: {exc}"


def _read_alembic_version(db_path: Path) -> str | None:
    if not db_path.exists():
        return None
    try:
        with sqlite3.connect(str(db_path)) as connection:
            exists = connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='alembic_version'"
            ).fetchone()
            if not exists:
                return None
            row = connection.execute("SELECT version_num FROM alembic_version LIMIT 1").fetchone()
            return str(row[0]) if row and row[0] is not None else None
    except sqlite3.Error:
        return None


def _safe_file_size(path: Path) -> int:
    try:
        return path.stat().st_size if path.exists() else 0
    except OSError:
        return 0


def _safe_directory_size(path: Path) -> int:
    if not path.exists():
        return 0
    total = 0
    try:
        for item in path.rglob("*"):
            if item.is_file():
                total += _safe_file_size(item)
    except OSError:
        return total
    return total


def _add_directory_to_zip(archive: ZipFile, directory: Path, prefix: str) -> None:
    if not directory.exists():
        return
    for path in directory.rglob("*"):
        if path.is_file():
            archive.write(path, arcname=f"{prefix}/{path.relative_to(directory).as_posix()}")


def _replace_directory(source: Path, target: Path) -> None:
    if target.exists():
        shutil.rmtree(target)
    if source.exists():
        shutil.copytree(source, target)
    else:
        target.mkdir(parents=True, exist_ok=True)


def _replace_file(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    wal_path = target.with_suffix(target.suffix + "-wal")
    shm_path = target.with_suffix(target.suffix + "-shm")
    if wal_path.exists():
        wal_path.unlink()
    if shm_path.exists():
        shm_path.unlink()
