from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import AuditLog, BackupRecord, Exam, ImportJob, ReportExportRecord, ScoreImportBatch, Semester, StoredFile


def test_growth_archive_upload_and_report_export(client) -> None:
    upload_response = client.post(
        "/api/files/upload",
        data={"category": "growth_archive"},
        files={"file": ("note.txt", b"growth attachment", "text/plain")},
    )
    assert upload_response.status_code == 200
    file_payload = upload_response.json()
    assert file_payload["original_filename"] == "note.txt"

    create_response = client.post(
        "/api/archives/students/1/records",
        json={
            "occurred_on": "2026-04-01",
            "record_type": "reward",
            "title": "市级作文竞赛二等奖",
            "content": "表现稳定，获得奖励。",
            "owner_name": "班主任",
            "note": "纳入成长档案",
            "attachment_file_ids": [file_payload["id"]],
            "is_active": True,
        },
    )
    assert create_response.status_code == 200
    record_payload = create_response.json()
    assert record_payload["student_id"] == 1
    assert len(record_payload["attachments"]) == 1

    list_response = client.get("/api/archives/students/1/records")
    assert list_response.status_code == 200
    list_payload = list_response.json()
    assert list_payload["total"] == 1
    assert list_payload["items"][0]["title"] == "市级作文竞赛二等奖"

    export_response = client.get("/api/archives/students/1/summary/export")
    assert export_response.status_code == 200
    assert "growth_summary_report" in export_response.headers["content-disposition"]

    report_response = client.post(
        "/api/reports/export",
        json={"report_type": "growth_summary", "student_id": 1},
    )
    assert report_response.status_code == 200
    report_payload = report_response.json()
    assert report_payload["report_name"] == "学生成长档案摘要"

    report_download = client.get(report_payload["download_url"])
    assert report_download.status_code == 200
    report_workbook = load_workbook(BytesIO(report_download.content))
    record_sheet = report_workbook["成长记录"]
    assert record_sheet.cell(row=2, column=2).value == "奖励记录"
    insight_sheet = report_workbook["摘要概览"]
    assert insight_sheet.cell(row=1, column=1).value == "标题"
    insight_titles = {
        insight_sheet.cell(row=index, column=1).value
        for index in range(2, insight_sheet.max_row + 1)
    }
    assert {"成长档案整体状态", "高频成长类型", "最近一条记录"} <= insight_titles


def test_backup_restore_roundtrip(client) -> None:
    initial_students = client.get("/api/students?page=1&page_size=200")
    assert initial_students.status_code == 200
    initial_total = initial_students.json()["total"]

    backup_response = client.post("/api/system/backup")
    assert backup_response.status_code == 200
    backup_id = backup_response.json()["backup_id"]

    create_response = client.post(
        "/api/students",
        json={
            "student_no": "2099001",
            "name": "测试恢复学生",
            "gender": "男",
            "birth_date": None,
            "id_number": None,
            "admission_year": 2024,
            "current_grade_id": 1,
            "current_class_id": 1,
            "status": "active",
            "student_type": "general",
            "art_track": None,
            "phone": None,
            "address": None,
            "note": None,
            "guardians": [],
            "is_active": True,
        },
    )
    assert create_response.status_code == 200

    after_create = client.get("/api/students?page=1&page_size=200")
    assert after_create.status_code == 200
    assert after_create.json()["total"] == initial_total + 1

    restore_response = client.post(
        "/api/system/restore",
        json={"backup_id": backup_id, "auto_backup_current": False},
    )
    assert restore_response.status_code == 200

    after_restore = client.get("/api/students?page=1&page_size=200")
    assert after_restore.status_code == 200
    restored_payload = after_restore.json()
    assert restored_payload["total"] == initial_total
    assert all(item["student_no"] != "2099001" for item in restored_payload["items"])


def test_system_safety_status_and_backup_dry_run(client) -> None:
    backup_response = client.post("/api/system/backup")
    assert backup_response.status_code == 200
    backup_id = backup_response.json()["backup_id"]

    status_response = client.get("/api/system/safety-status")
    assert status_response.status_code == 200
    status_payload = status_response.json()
    assert status_payload["sqlite_integrity"] == "ok"
    assert status_payload["main_db_path"].endswith(".db")
    assert status_payload["latest_backup"]["id"] == backup_id
    assert status_payload["backup_count"] >= 1

    verify_response = client.get(f"/api/system/backups/{backup_id}/verify")
    assert verify_response.status_code == 200
    verify_payload = verify_response.json()
    assert verify_payload["valid"] is True
    assert verify_payload["sqlite_integrity"] == "ok"
    assert verify_payload["will_overwrite_path"].endswith(".db")

    dry_run_response = client.post(f"/api/system/backups/{backup_id}/restore-dry-run")
    assert dry_run_response.status_code == 200
    assert dry_run_response.json()["message"].startswith("恢复演练通过")


def test_upload_category_rejects_path_traversal(client) -> None:
    response = client.post(
        "/api/files/upload",
        data={"category": "../escape"},
        files={"file": ("note.txt", b"bad", "text/plain")},
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "上传分类不合法"


def test_runtime_file_download_rejects_invalid_paths(client) -> None:
    traversal_response = client.get("/api/system/files", params={"path": "../README.md"})
    assert traversal_response.status_code == 400
    assert traversal_response.json()["detail"] == "非法文件路径"

    absolute_response = client.get("/api/system/files", params={"path": "/etc/hosts"})
    assert absolute_response.status_code == 400
    assert absolute_response.json()["detail"] == "非法文件路径"


def test_import_center_lists_batches_and_details(client, app) -> None:
    error_report_path = app.state.settings.logs_dir / "score_import_errors.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "错误报告"
    sheet.append(["行号", "列名", "字段名", "原始值", "错误原因", "建议修复"])
    sheet.append([3, "学号", "学号", "2026999", "学生不存在", "请先维护学生，或核对学号。"])
    workbook.save(error_report_path)
    error_report_relative_path = str(error_report_path)

    with app.state.db.session_scope() as session:
        session.add(
            BackupRecord(
                backup_name="trial_run_backup.zip",
                file_path="data/backups/trial_run_backup.zip",
                file_size=128,
                status="success",
            )
        )
        semester = session.scalar(select(Semester))
        assert semester is not None
        exam = Exam(
            name="导入中心测试考试",
            exam_type="monthly",
            exam_date=date(2026, 4, 24),
            semester_id=semester.id,
            status="published",
        )
        session.add(exam)
        session.flush()
        score_batch = ScoreImportBatch(
            exam_id=exam.id,
            source_filename="scores.xlsx",
            import_time=datetime(2026, 4, 24, 9, 30),
            total_rows=2,
            success_rows=1,
                failed_rows=1,
                status="partially_failed",
                error_report_path=error_report_relative_path,
            )
        session.add(score_batch)
        session.flush()
        job = ImportJob(
            job_type="scores",
            source_filename="scores.xlsx",
            started_at=datetime(2026, 4, 24, 9, 30),
            finished_at=datetime(2026, 4, 24, 9, 31),
            status="partially_failed",
            result_json={
                "batch_id": score_batch.id,
                "status": "partially_failed",
                "total_rows": 2,
                    "success_rows": 1,
                    "failed_rows": 1,
                    "skipped_rows": 0,
                    "error_report_path": error_report_relative_path,
                    "error_preview": ["第 3 行：学生不存在"],
                },
        )
        session.add(job)
        session.flush()
        session.add(
            AuditLog(
                module="exams",
                action="import_scores",
                target_type="score_import_batch",
                target_id=str(score_batch.id),
                detail_json={"batch_id": score_batch.id, "failed_rows": 1},
            )
        )
        import_job_id = job.id

    response = client.get("/api/import-center/batches")
    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["total_batches"] >= 1
    assert payload["summary"]["partial_batches"] >= 1
    assert payload["latest_backup"]["backup_name"] == "trial_run_backup.zip"
    score_rows = [item for item in payload["batches"] if item["job_type"] == "scores"]
    assert len([item for item in score_rows if item["source_filename"] == "scores.xlsx"]) == 1
    assert score_rows[0]["status"] == "partially_failed"
    assert score_rows[0]["template_download_url"].endswith("exam_scores_import_template.xlsx")
    assert score_rows[0]["rollback_supported"] is False

    detail_response = client.get(f"/api/import-center/batches/import_job/{import_job_id}")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["batch"]["job_type_label"] == "成绩导入"
    assert detail["error_preview"] == ["第 3 行：学生不存在"]
    assert detail["error_items"][0]["row_number"] == 3
    assert detail["error_items"][0]["field_name"] == "学号"
    assert detail["error_items"][0]["raw_value"] == "2026999"
    assert "核对学号" in detail["error_items"][0]["suggestion"]
    assert detail["rollback_steps"]


def test_report_download_rejects_absolute_db_path(client, app, tmp_path: Path) -> None:
    external_file = tmp_path / "outside-report.xlsx"
    external_file.write_bytes(b"report")

    with app.state.db.session_scope() as session:
        record = ReportExportRecord(
            report_type="growth_summary",
            report_name="测试报表",
            params_json={},
            file_path=str(external_file),
            status="success",
        )
        session.add(record)
        session.flush()
        export_id = record.id

    response = client.get(f"/api/reports/exports/{export_id}/download")
    assert response.status_code == 400
    assert response.json()["detail"] == "非法文件路径"


def test_stored_file_download_rejects_absolute_db_path(client, app, tmp_path: Path) -> None:
    external_file = tmp_path / "outside-upload.txt"
    external_file.write_bytes(b"upload")

    with app.state.db.session_scope() as session:
        record = StoredFile(
            original_filename="outside-upload.txt",
            file_path=str(external_file),
            content_type="text/plain",
            file_size=external_file.stat().st_size,
            category="general",
        )
        session.add(record)
        session.flush()
        file_id = record.id

    response = client.get(f"/api/files/{file_id}")
    assert response.status_code == 400
    assert response.json()["detail"] == "非法文件路径"


def test_restore_backup_rejects_absolute_db_path(client, app, tmp_path: Path) -> None:
    external_file = tmp_path / "outside-backup.zip"
    external_file.write_bytes(b"backup")

    with app.state.db.session_scope() as session:
        record = BackupRecord(
            backup_name="outside-backup.zip",
            file_path=str(external_file),
            file_size=external_file.stat().st_size,
            status="success",
        )
        session.add(record)
        session.flush()
        backup_id = record.id

    response = client.post(
        "/api/system/restore",
        json={"backup_id": backup_id, "auto_backup_current": False},
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "非法文件路径"


def test_system_templates_route_returns_structured_list(client) -> None:
    response = client.get("/api/system/templates")
    assert response.status_code == 200
    payload = response.json()
    assert isinstance(payload, list)
    assert payload
    assert "file_name" in payload[0]


def test_system_templates_route_is_unique(app) -> None:
    matching_routes = [
        route
        for route in app.routes
        if getattr(route, "path", None) == "/api/system/templates" and "GET" in getattr(route, "methods", set())
    ]
    assert len(matching_routes) == 1
