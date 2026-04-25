from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl import load_workbook


def build_score_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    sheet.append(["2026届高一4月月考", "2026001", "张三", "1班", "语文", "118", "", ""])
    sheet.append(["2026届高一4月月考", "2026001", "张三", "1班", "数学", "125", "", ""])
    sheet.append(["2026届高一4月月考", "2026002", "李四", "1班", "语文", "110", "", ""])
    sheet.append(["2026届高一4月月考", "2026002", "李四", "1班", "数学", "120", "", ""])
    sheet.append(["2026届高一4月月考", "2026003", "王五", "2班", "语文", "98", "", ""])
    sheet.append(["2026届高一4月月考", "2026003", "王五", "2班", "数学", "", "缺考", ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_invalid_score_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["错误表头", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    sheet.append(["坏模板", "2026001", "张三", "1班", "语文", "118", "", ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_score_workbook_with_rows(exam_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    for row in rows:
        sheet.append([exam_name, *row])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_published_exam_with_scores(client) -> tuple[int, dict]:
    exam_response = client.post(
        "/api/exams",
        json={
            "name": "2026届高一4月月考",
            "exam_type": "月考",
            "exam_date": "2026-04-10",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            },
            {
                "subject_id": 2,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 2,
                "is_active": True,
            },
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                "scores.xlsx",
                build_score_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    return exam_id, import_response.json()


def test_exam_score_import_and_analytics(client) -> None:
    exam_id, import_payload = create_published_exam_with_scores(client)

    subject_response = client.get(f"/api/exams/{exam_id}/subjects")
    assert subject_response.status_code == 200
    assert len(subject_response.json()) == 2

    assert import_payload["success_rows"] == 6
    assert import_payload["failed_rows"] == 0

    student_analytics = client.get(f"/api/analytics/students/1?exam_id={exam_id}")
    assert student_analytics.status_code == 200
    student_payload = student_analytics.json()
    assert student_payload["total_score"] == 243.0
    assert student_payload["class_rank"] == 1
    assert student_payload["grade_rank"] == 1

    class_analytics = client.get(f"/api/analytics/classes/1?exam_id={exam_id}")
    assert class_analytics.status_code == 200
    class_payload = class_analytics.json()
    assert class_payload["student_count"] == 2
    assert class_payload["total_average"] == 236.5
    assert len(class_payload["subject_breakdown"]) == 2

    teacher_analytics = client.get(f"/api/analytics/teachers/1?exam_id={exam_id}")
    assert teacher_analytics.status_code == 200
    teacher_payload = teacher_analytics.json()
    assert teacher_payload["teacher_name"] == "李语文"
    assert len(teacher_payload["assignment_breakdown"]) == 1
    assert teacher_payload["assignment_breakdown"][0]["average_score"] == 114.0


def test_teacher_analysis_export_includes_assignment_count(client) -> None:
    exam_id, _ = create_published_exam_with_scores(client)

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "teacher_analysis", "exam_id": exam_id, "teacher_id": 1},
    )
    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["report_name"] == "教师任教分析报表"

    download_response = client.get(export_payload["download_url"])
    assert download_response.status_code == 200

    workbook = load_workbook(BytesIO(download_response.content))
    summary_sheet = workbook["教师分析"]
    summary_rows = {
        summary_sheet.cell(row=index, column=1).value: summary_sheet.cell(row=index, column=2).value
        for index in range(1, summary_sheet.max_row + 1)
    }
    assert summary_rows["考试"] == "2026届高一4月月考"
    assert summary_rows["教师"] == "李语文"
    assert summary_rows["整体均分"] == 114
    assert summary_rows["任教拆分"] == 1

    detail_sheet = workbook["任教明细"]
    assert detail_sheet.cell(row=1, column=1).value == "班级"
    assert detail_sheet.max_row == 2

    insight_sheet = workbook["摘要概览"]
    assert insight_sheet.cell(row=1, column=1).value == "标题"
    insight_titles = {
        insight_sheet.cell(row=index, column=1).value
        for index in range(2, insight_sheet.max_row + 1)
    }
    assert {"任教整体状态", "表现最佳任教拆分"} <= insight_titles


def test_student_analysis_export_keeps_summary_structure(client) -> None:
    exam_id, _ = create_published_exam_with_scores(client)

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "student_analysis", "exam_id": exam_id, "student_id": 1},
    )
    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["report_name"] == "学生成绩分析单"

    download_response = client.get(export_payload["download_url"])
    assert download_response.status_code == 200

    workbook = load_workbook(BytesIO(download_response.content))
    summary_sheet = workbook["学生分析"]
    summary_rows = {
        summary_sheet.cell(row=index, column=1).value: summary_sheet.cell(row=index, column=2).value
        for index in range(1, summary_sheet.max_row + 1)
    }
    assert summary_rows["考试"] == "2026届高一4月月考"
    assert summary_rows["学生"] == "张三"
    assert summary_rows["总分"] == 243
    assert summary_rows["班级名次"] == 1
    assert summary_rows["年级名次"] == 1
    assert "总分变化" in summary_rows

    detail_sheet = workbook["学科明细"]
    assert detail_sheet.cell(row=1, column=1).value == "科目"
    assert detail_sheet.cell(row=1, column=3).value == "班级名次"
    assert detail_sheet.cell(row=1, column=4).value == "年级名次"
    assert detail_sheet.max_row == 3

    insight_sheet = workbook["摘要概览"]
    assert insight_sheet.cell(row=1, column=1).value == "标题"
    insight_rows = {
        insight_sheet.cell(row=index, column=1).value: insight_sheet.cell(row=index, column=2).value
        for index in range(2, insight_sheet.max_row + 1)
    }
    assert insight_rows["本次成绩摘要"] == "张三 在 2026届高一4月月考 取得总分 243"
    assert str(insight_rows["优势学科"]).endswith("当前表现最强")
    if "重点关注学科" in insight_rows:
        assert str(insight_rows["重点关注学科"]).endswith("建议继续重点复核")


def test_class_analysis_export_keeps_summary_structure(client) -> None:
    exam_id, _ = create_published_exam_with_scores(client)
    analytics_response = client.get(f"/api/analytics/classes/1?exam_id={exam_id}")
    assert analytics_response.status_code == 200
    analytics_payload = analytics_response.json()

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "class_analysis", "exam_id": exam_id, "class_id": 1},
    )
    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["report_name"] == "班级成绩分析报表"

    download_response = client.get(export_payload["download_url"])
    assert download_response.status_code == 200

    workbook = load_workbook(BytesIO(download_response.content))
    summary_sheet = workbook["班级分析"]
    summary_rows = {
        summary_sheet.cell(row=index, column=1).value: summary_sheet.cell(row=index, column=2).value
        for index in range(1, summary_sheet.max_row + 1)
    }
    assert summary_rows["考试"] == "2026届高一4月月考"
    assert summary_rows["班级"] == "1班"
    assert summary_rows["学生数"] == analytics_payload["student_count"]
    assert summary_rows["总分均分"] == analytics_payload["total_average"]
    assert summary_rows["总分中位数"] == analytics_payload["total_median"]
    assert summary_rows["年级均分"] == analytics_payload["grade_average"]

    detail_sheet = workbook["学科统计"]
    assert detail_sheet.cell(row=1, column=1).value == "科目"
    assert detail_sheet.max_row == 3

    insight_sheet = workbook["摘要概览"]
    assert insight_sheet.cell(row=1, column=1).value == "标题"
    insight_rows = {
        insight_sheet.cell(row=index, column=1).value: insight_sheet.cell(row=index, column=2).value
        for index in range(2, insight_sheet.max_row + 1)
    }
    assert str(insight_rows["班级整体状态"]).startswith("1班 共 2 人，总分均分 ")
    assert "中位数" in str(insight_rows["班级整体状态"])
    assert str(insight_rows["与年级对比"]).startswith("班级均分")
    assert insight_rows["优势学科"] == "数学 当前均分最高"


def test_exam_score_import_rejects_invalid_template(client) -> None:
    exam_response = client.post(
        "/api/exams",
        json={
            "name": "无效模板测试",
            "exam_type": "月考",
            "exam_date": "2026-04-11",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "draft",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            }
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                "invalid_scores.xlsx",
                build_invalid_score_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 400
    assert import_response.json()["detail"] == "成绩导入模板表头不匹配，请先下载系统模板。"


def test_exam_score_import_reports_duplicate_rows(client) -> None:
    exam_name = "重复成绩导入测试"
    exam_response = client.post(
        "/api/exams",
        json={
            "name": exam_name,
            "exam_type": "月考",
            "exam_date": "2026-04-12",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "draft",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            }
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "false"},
        files={
            "file": (
                "duplicate_scores.xlsx",
                build_score_workbook_with_rows(
                    exam_name,
                    [
                        ["2026001", "张三", "1班", "语文", "118", "", ""],
                        ["2026001", "张三", "1班", "语文", "119", "", ""],
                    ],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert import_response.status_code == 200
    payload = import_response.json()
    assert payload["status"] == "partially_failed"
    assert payload["success_rows"] == 1
    assert payload["failed_rows"] == 1
    assert payload["error_preview"] == ["第 3 行：同一学生同一科目在导入文件中重复出现"]
    assert payload["error_report_path"]


def test_exam_score_import_reports_identity_conflicts(client) -> None:
    exam_name = "成绩身份冲突测试"
    exam_response = client.post(
        "/api/exams",
        json={
            "name": exam_name,
            "exam_type": "月考",
            "exam_date": "2026-04-13",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "draft",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]
    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            }
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "false"},
        files={
            "file": (
                "conflict_scores.xlsx",
                build_score_workbook_with_rows(
                    exam_name,
                    [["2026001", "李四", "1班", "语文", "118", "", ""]],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert import_response.status_code == 200
    payload = import_response.json()
    assert payload["status"] == "failed"
    assert payload["success_rows"] == 0
    assert payload["failed_rows"] == 1
    assert payload["error_preview"] == ["第 2 行：学号与姓名不匹配: 2026001"]
