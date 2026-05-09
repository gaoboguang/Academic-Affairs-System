from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import StudentPathwayProfile


def _build_pathway_profile_workbook(rows: list[list[object]], headers: list[str] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        headers
        or [
            "学号",
            "姓名",
            "班级",
            "生源地",
            "考生类型",
            "考试类型",
            "选科组合",
            "春考专业类别",
            "艺术类别",
            "体育类别",
            "已完成高考报名",
            "普通高中应届",
            "中职学生",
            "社会人员",
            "高中阶段学历或同等学力",
            "接受专科",
            "接受民办院校",
            "接受中外合作",
            "接受省外院校",
            "接受提前批",
            "接受定向服务",
            "接受面试体检政审",
            "高考报名确认材料",
            "综合素质评价材料",
            "单招院校章程和分专业计划",
            "体检限制",
            "备注",
        ]
    )
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _first_student_id(client) -> int:
    response = client.get("/api/students", params={"page_size": 1})
    assert response.status_code == 200
    data = response.json()
    assert data["items"]
    return data["items"][0]["id"]


def _pathway_by_code(items: list[dict], code: str) -> dict:
    return next(item for item in items if item["pathway_code"] == code)


def _evaluation_by_code(items: list[dict], code: str) -> dict:
    return next(item for item in items if item["pathway_code"] == code)


def test_bootstrap_shandong_pathways_creates_core_paths(client):
    response = client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    assert response.status_code == 200
    payload = response.json()
    assert payload["total_count"] >= 10
    assert payload["created_count"] >= 10
    assert payload["rule_created_count"] >= 45

    list_response = client.get("/api/gaokao/pathways", params={"province": "山东"})
    assert list_response.status_code == 200
    pathways = list_response.json()
    codes = {item["pathway_code"] for item in pathways}
    assert {
        "summer_general_regular",
        "summer_general_early_a",
        "summer_special_type",
        "spring_exam_undergrad",
        "vocational_single_exam",
        "vocational_comprehensive",
        "art_undergrad",
        "sports_regular",
        "sports_single_exam",
        "high_level_sports",
    }.issubset(codes)

    general_pathway = _pathway_by_code(pathways, "summer_general_regular")
    assert general_pathway["source_document_id"]
    rule_response = client.get(f"/api/gaokao/pathways/{general_pathway['id']}/rules")
    assert rule_response.status_code == 200
    rules = rule_response.json()
    rule_codes = {item["rule_code"] for item in rules}
    assert {
        "d2_general_gaokao_registration",
        "d2_general_subject_combination",
        "d2_general_2026_plan_pending",
        "d2_general_chapter_review",
    }.issubset(rule_codes)
    assert all(item["source_document_id"] for item in rules if item["rule_code"].startswith("d2_"))

    single_pathway = _pathway_by_code(pathways, "vocational_single_exam")
    single_rules = client.get(f"/api/gaokao/pathways/{single_pathway['id']}/rules").json()
    assert {item["rule_code"] for item in single_rules}.issuperset(
        {
            "d2_single_registration",
            "d2_single_candidate_scope",
            "d2_single_school_chapter",
            "d6_single_major_category_match",
            "d6_single_chapter_plan_material",
            "d6_single_retired_soldier_review",
        }
    )
    assert any(item["manual_review_required"] for item in single_rules if item["rule_code"].startswith("d2_"))

    comprehensive_pathway = _pathway_by_code(pathways, "vocational_comprehensive")
    comprehensive_rules = client.get(f"/api/gaokao/pathways/{comprehensive_pathway['id']}/rules").json()
    assert {item["rule_code"] for item in comprehensive_rules}.issuperset(
        {
            "d6_comprehensive_registration",
            "d2_comprehensive_candidate_scope",
            "d2_comprehensive_quality_record",
            "d6_comprehensive_test_material",
            "d6_comprehensive_chapter_plan",
        }
    )

    spring_pathway = _pathway_by_code(pathways, "spring_exam_undergrad")
    spring_rules = client.get(f"/api/gaokao/pathways/{spring_pathway['id']}/rules").json()
    assert {item["rule_code"] for item in spring_rules}.issuperset(
        {
            "d6_spring_undergrad_registration",
            "d2_spring_undergrad_category",
            "d6_spring_undergrad_score_and_skill",
            "d6_spring_undergrad_score_line",
            "d6_spring_undergrad_plan_chapter",
        }
    )

    second_response = client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    assert second_response.status_code == 200
    assert second_response.json()["created_count"] == 0
    assert second_response.json()["rule_created_count"] == 0


def test_student_pathway_rule_engine_reports_passed_failed_and_unknown(client):
    client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    student_id = _first_student_id(client)
    pathways = client.get("/api/gaokao/pathways", params={"province": "山东"}).json()
    general_pathway = _pathway_by_code(pathways, "summer_general_regular")

    rule_response = client.post(
        f"/api/gaokao/pathways/{general_pathway['id']}/rules",
        json={
            "rule_code": "d1_test_gaokao_registration",
            "rule_name": "高考报名确认材料",
            "rule_type": "material_required",
            "severity": "required",
            "condition_json": {"type": "material_present", "key": "gaokao_registration"},
            "message_template": "补充高考报名确认材料后再评估。",
            "valid_from_year": 2026,
        },
    )
    assert rule_response.status_code == 200

    profile_response = client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "subject_combination": "物理,化学,生物",
            "has_gaokao_registration": True,
            "materials_json": {},
        },
    )
    assert profile_response.status_code == 200

    preview_response = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    )
    assert preview_response.status_code == 200
    general_eval = _evaluation_by_code(preview_response.json()["evaluations"], "summer_general_regular")
    assert general_eval["status"] == "insufficient_data"
    assert general_eval["matched_rules_json"][0]["result"] == "passed"
    assert general_eval["missing_materials_json"][0]["material_key"] == "gaokao_registration"
    assert general_eval["warning_rules_json"][0]["result"] == "unknown"

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "subject_combination": "物理,化学,生物",
            "has_gaokao_registration": True,
            "materials_json": {"gaokao_registration": True},
        },
    )
    passed_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    )
    passed_eval = _evaluation_by_code(passed_preview.json()["evaluations"], "summer_general_regular")
    assert passed_eval["status"] == "possible"
    assert {item["result"] for item in passed_eval["matched_rules_json"]} == {"passed"}
    assert {item["rule_code"] for item in passed_eval["warning_rules_json"]}.issuperset(
        {"d2_general_2026_plan_pending", "d2_general_chapter_review"}
    )

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "spring_exam",
            "subject_combination": "物理,化学,生物",
            "has_gaokao_registration": True,
            "materials_json": {"gaokao_registration": True},
        },
    )
    failed_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    )
    failed_eval = _evaluation_by_code(failed_preview.json()["evaluations"], "summer_general_regular")
    assert failed_eval["status"] == "not_recommended"
    assert failed_eval["failed_rules_json"][0]["result"] == "failed"


def test_student_pathway_profile_missing_fields_are_readable(client):
    client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    student_id = _first_student_id(client)
    profile_response = client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "materials_json": {},
        },
    )
    assert profile_response.status_code == 200

    preview_response = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    )
    assert preview_response.status_code == 200
    general_eval = _evaluation_by_code(preview_response.json()["evaluations"], "summer_general_regular")
    missing_items = general_eval["missing_materials_json"]
    missing_by_key = {item["material_key"]: item for item in missing_items}
    assert missing_by_key["has_gaokao_registration"]["material_label"] == "高考报名状态"
    assert missing_by_key["subject_combination"]["material_label"] == "选科组合"
    assert missing_by_key["subject_combination"]["gap_type"] == "profile_field"
    assert "补充选科组合" in missing_by_key["subject_combination"]["next_action"]


def test_pathway_profile_template_endpoint_contains_expected_headers(client) -> None:
    response = client.get("/api/gaokao/pathway-profiles/template")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook["数据"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 28)]
    assert headers[:7] == ["学号", "姓名", "班级", "生源地", "考生类型", "考试类型", "选科组合"]
    assert "高考报名确认材料" in headers
    assert "体检限制" in headers
    workbook.close()


def test_pathway_profile_import_updates_subjects_without_clearing_existing_values(client, app) -> None:
    student_id = _first_student_id(client)
    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "exam_type": "summer_gaokao",
            "subject_combination": "物理,化学,生物",
            "has_gaokao_registration": True,
            "materials_json": {"gaokao_registration": True},
            "known_body_limitations_json": {"note": "色觉正常"},
            "note": "原备注",
        },
    )

    response = client.post(
        "/api/gaokao/pathway-profiles/import",
        files={
            "file": (
                "pathway.xlsx",
                _build_pathway_profile_workbook(
                    headers=["学号", "姓名", "班级", "选科组合"],
                    rows=[
                        [
                            "2026001",
                            "张三",
                            "1班",
                            "物理,化学,政治",
                        ]
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["success_rows"] == 1
    assert payload["updated_rows"] == 1
    assert payload["failed_rows"] == 0

    with app.state.db.session_scope() as session:
        profile = session.scalar(select(StudentPathwayProfile).where(StudentPathwayProfile.student_id == student_id))
        assert profile is not None
        assert profile.subject_combination == "物理,化学,政治"
        assert profile.candidate_type == "general"
        assert profile.has_gaokao_registration is True
        assert profile.materials_json == {"gaokao_registration": True}
        assert profile.known_body_limitations_json == {"note": "色觉正常"}
        assert profile.note == "原备注"


def test_pathway_profile_import_creates_profile_and_reports_bad_rows(client, app, test_settings) -> None:
    response = client.post(
        "/api/gaokao/pathway-profiles/import",
        files={
            "file": (
                "pathway.xlsx",
                _build_pathway_profile_workbook(
                    rows=[
                        [
                            "2026002",
                            "李四",
                            "1班",
                            "山东",
                            "普通类",
                            "夏季高考",
                            "物理,化学,生物",
                            "",
                            "",
                            "",
                            "是",
                            "是",
                            "否",
                            "否",
                            "",
                            "是",
                            "否",
                            "否",
                            "是",
                            "否",
                            "否",
                            "否",
                            "是",
                            "否",
                            "是",
                            "无",
                            "批量导入",
                        ],
                        ["2026003", "错误姓名", "1班", "山东", "普通类", "夏季高考", "物理,化学,生物", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                        ["999999", "不存在", "1班", "山东", "普通类", "夏季高考", "物理,化学,生物", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                        ["2026001", "张三", "1班", "山东", "普通类", "夏季高考", "物理,化学,生物", "", "", "", "也许", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                        ["2026001", "张三", "1班", "山东", "火星类", "夏季高考", "物理,化学,生物", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "partially_failed"
    assert payload["success_rows"] == 1
    assert payload["failed_rows"] == 4
    assert payload["created_rows"] == 1
    assert payload["error_report_path"]
    assert any("姓名不一致" in item for item in payload["error_preview"])

    with app.state.db.session_scope() as session:
        profile = session.scalar(
            select(StudentPathwayProfile)
            .join(StudentPathwayProfile.student)
            .where(StudentPathwayProfile.student.has(student_no="2026002"))
        )
        assert profile is not None
        assert profile.candidate_type == "general"
        assert profile.exam_type == "summer_gaokao"
        assert profile.subject_combination == "物理,化学,生物"
        assert profile.has_gaokao_registration is True
        assert profile.is_vocational_student is False
        assert profile.materials_json == {
            "gaokao_registration": True,
            "comprehensive_quality_evaluation": False,
            "single_exam_college_chapter_plan": True,
        }
        assert profile.known_body_limitations_json == {"note": "无"}
        assert profile.note == "批量导入"

    error_workbook = load_workbook(test_settings.project_root / payload["error_report_path"])
    error_sheet = error_workbook["错误报告"]
    assert error_sheet.cell(row=2, column=5).value == "姓名不一致: 系统为王五，导入为错误姓名"
    assert error_sheet.cell(row=3, column=5).value == "学生不存在: 999999"
    assert "已完成高考报名 无法识别" in str(error_sheet.cell(row=4, column=5).value)
    assert "考生类型 无法识别" in str(error_sheet.cell(row=5, column=5).value)
    error_workbook.close()


def test_pathway_profile_export_includes_students_with_and_without_profiles(client) -> None:
    student_id = _first_student_id(client)
    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "exam_type": "summer_gaokao",
            "subject_combination": "物理,化学,生物",
            "has_gaokao_registration": True,
            "materials_json": {"gaokao_registration": True},
            "known_body_limitations_json": {"note": "无"},
            "note": "已确认",
        },
    )

    response = client.get("/api/gaokao/pathway-profiles/export")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook["数据"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 28)]
    assert headers[:7] == ["学号", "姓名", "班级", "生源地", "考生类型", "考试类型", "选科组合"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    row_by_no = {str(row[0]): row for row in rows}
    assert row_by_no["2026001"][2] == "1班"
    assert row_by_no["2026001"][6] == "物理,化学,生物"
    assert row_by_no["2026001"][10] == "是"
    assert row_by_no["2026001"][22] == "是"
    assert row_by_no["2026001"][25] == "无"
    assert row_by_no["2026002"][2] == "1班"
    assert row_by_no["2026002"][6] is None
    workbook.close()


def test_d6_vocational_and_spring_pathways_surface_screening_requirements(client):
    client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    student_id = _first_student_id(client)

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "independent_recruitment",
            "has_gaokao_registration": True,
            "is_vocational_student": True,
            "is_social_candidate": False,
            "materials_json": {},
        },
    )
    preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    single_eval = _evaluation_by_code(preview["evaluations"], "vocational_single_exam")
    single_missing = {item["material_key"]: item for item in single_eval["missing_materials_json"]}
    assert single_eval["status"] == "insufficient_data"
    assert "single_exam_major_category_match" in single_missing
    assert "single_exam_college_chapter_plan" in single_missing
    assert "high_school_equivalent" not in single_missing
    assert any("退役士兵" in item["message"] for item in single_eval["warning_rules_json"])
    assert "不能理解为录取概率" in single_eval["summary"]

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "independent_recruitment",
            "has_gaokao_registration": True,
            "is_vocational_student": False,
            "is_social_candidate": True,
            "materials_json": {},
        },
    )
    social_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    social_single_eval = _evaluation_by_code(social_preview["evaluations"], "vocational_single_exam")
    social_missing = {item["material_key"]: item for item in social_single_eval["missing_materials_json"]}
    assert social_missing["high_school_equivalent"]["material_label"] == "高中阶段毕业证书或同等学力材料"

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "has_gaokao_registration": True,
            "is_fresh_graduate": True,
            "materials_json": {"comprehensive_quality_evaluation": True},
        },
    )
    comprehensive_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    comprehensive_eval = _evaluation_by_code(comprehensive_preview["evaluations"], "vocational_comprehensive")
    comprehensive_missing = {item["material_key"]: item for item in comprehensive_eval["missing_materials_json"]}
    assert "comprehensive_test_or_interview" in comprehensive_missing
    assert "comprehensive_college_chapter_plan" in comprehensive_missing
    assert any("普通高中应届" in item["rule_name"] for item in comprehensive_eval["matched_rules_json"])

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "spring_exam",
            "has_gaokao_registration": True,
            "spring_exam_category": "软件与应用技术",
            "materials_json": {},
        },
    )
    spring_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    spring_eval = _evaluation_by_code(spring_preview["evaluations"], "spring_exam_undergrad")
    spring_missing = {item["material_key"]: item for item in spring_eval["missing_materials_json"]}
    assert any("专业类别一致" in item["rule_name"] for item in spring_eval["matched_rules_json"])
    assert "spring_exam_skill_score" in spring_missing
    assert "spring_exam_score_line" in spring_missing
    assert "spring_exam_college_plan_chapter" in spring_missing


def test_d7_special_early_art_and_sports_pathways_surface_manual_review_requirements(client):
    client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    student_id = _first_student_id(client)

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "art",
            "art_track": "美术与设计类",
            "materials_json": {},
        },
    )
    art_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    art_eval = _evaluation_by_code(art_preview["evaluations"], "art_undergrad")
    art_missing = {item["material_key"]: item for item in art_eval["missing_materials_json"]}
    assert art_eval["status"] == "insufficient_data"
    assert "art_exam_score" in art_missing
    assert "art_culture_composite_rule" in art_missing
    assert art_missing["art_chapter_restrictions"]["material_label"] == "艺术类院校章程限制"
    assert any("同批次兼报" in item["message"] for item in art_eval["warning_rules_json"])
    assert "不能理解为录取概率" in art_eval["summary"]

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "sports",
            "has_gaokao_registration": True,
            "sports_track": "田径",
            "materials_json": {},
        },
    )
    sports_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    sports_eval = _evaluation_by_code(sports_preview["evaluations"], "sports_regular")
    sports_missing = {item["material_key"]: item for item in sports_eval["missing_materials_json"]}
    assert "sports_test_score" in sports_missing
    assert "sports_culture_composite_rule" in sports_missing
    assert "sports_chapter_restrictions" in sports_missing
    assert any("体育类常规批、体育单招和高水平运动队" in item["message"] for item in sports_eval["warning_rules_json"])

    sports_single_eval = _evaluation_by_code(sports_preview["evaluations"], "sports_single_exam")
    sports_single_missing = {item["material_key"]: item for item in sports_single_eval["missing_materials_json"]}
    assert "athlete_level_certificate" in sports_single_missing
    assert "sports_single_exam_arrangement" in sports_single_missing
    assert "sports_single_college_chapter" in sports_single_missing

    high_level_eval = _evaluation_by_code(sports_preview["evaluations"], "high_level_sports")
    high_level_missing = {item["material_key"]: item for item in high_level_eval["missing_materials_json"]}
    assert "high_level_athlete_level" in high_level_missing
    assert "high_level_qualification_review" in high_level_missing
    assert "high_level_college_chapter" in high_level_missing

    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "has_gaokao_registration": True,
            "accept_early_batch": True,
            "accept_interview_or_physical_test": True,
            "accept_service_commitment": True,
            "materials_json": {},
        },
    )
    general_preview = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations/preview",
        params={"target_year": 2026, "province": "山东"},
    ).json()
    early_a_eval = _evaluation_by_code(general_preview["evaluations"], "summer_general_early_a")
    early_a_missing = {item["material_key"]: item for item in early_a_eval["missing_materials_json"]}
    assert "early_batch_physical_political_review" in early_a_missing
    assert "early_batch_chapter_limits" in early_a_missing

    early_b_eval = _evaluation_by_code(general_preview["evaluations"], "summer_general_early_b")
    early_b_missing = {item["material_key"]: item for item in early_b_eval["missing_materials_json"]}
    assert "early_b_service_contract" in early_b_missing
    assert "early_batch_chapter_limits" in early_b_missing

    special_eval = _evaluation_by_code(general_preview["evaluations"], "summer_special_type")
    special_missing = {item["material_key"]: item for item in special_eval["missing_materials_json"]}
    assert "special_type_score_line_ready" in special_missing
    assert "special_type_qualification" in special_missing
    assert "special_type_application_review" in special_missing
    assert "special_type_chapter_limits" in special_missing


def test_student_pathway_evaluations_can_be_persisted(client):
    client.post("/api/gaokao/pathways/bootstrap-shandong", params={"target_year": 2026})
    student_id = _first_student_id(client)
    client.put(
        f"/api/gaokao/students/{student_id}/pathway-profile",
        json={
            "province": "山东",
            "candidate_type": "general",
            "materials_json": {},
        },
    )

    response = client.post(
        f"/api/gaokao/students/{student_id}/pathway-evaluations",
        params={"target_year": 2026, "province": "山东"},
    )
    assert response.status_code == 200
    evaluations = response.json()["evaluations"]
    assert evaluations
    assert all(item["id"] for item in evaluations)
