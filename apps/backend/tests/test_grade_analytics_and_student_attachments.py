from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl import load_workbook


def build_score_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    sheet.append(["2026届高一5月月考", "2026001", "张三", "1班", "语文", "120", "", ""])
    sheet.append(["2026届高一5月月考", "2026001", "张三", "1班", "数学", "128", "", ""])
    sheet.append(["2026届高一5月月考", "2026002", "李四", "1班", "语文", "112", "", ""])
    sheet.append(["2026届高一5月月考", "2026002", "李四", "1班", "数学", "118", "", ""])
    sheet.append(["2026届高一5月月考", "2026003", "王五", "2班", "语文", "95", "", ""])
    sheet.append(["2026届高一5月月考", "2026003", "王五", "2班", "数学", "100", "", ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_exam_with_scores(client) -> int:
    exam_response = client.post(
        "/api/exams",
        json={
            "name": "2026届高一5月月考",
            "exam_type": "月考",
            "exam_date": "2026-05-10",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            },
            {
                "subject_id": 2,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 2,
                "is_active": True,
            },
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                "scores.xlsx",
                build_score_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    return exam_id


def create_previous_year_exam_with_scores(client) -> int:
    year_response = client.post(
        "/api/base/academic-years",
        json={
            "name": "2024-2025",
            "start_date": "2024-09-01",
            "end_date": "2025-08-31",
            "is_current": False,
            "is_active": True,
        },
    )
    assert year_response.status_code == 200
    year_id = year_response.json()["id"]

    semester_response = client.post(
        "/api/base/semesters",
        json={
            "academic_year_id": year_id,
            "name": "下学期",
            "start_date": "2025-02-15",
            "end_date": "2025-07-15",
            "week_count": 20,
            "is_current": False,
            "is_active": True,
        },
    )
    assert semester_response.status_code == 200
    semester_id = semester_response.json()["id"]

    previous_assignments_response = client.post(
        "/api/teachers/assignments",
        json={
            "teacher_id": 1,
            "semester_id": semester_id,
            "grade_id": 1,
            "class_id": 1,
            "subject_id": 1,
            "course_type": "regular",
            "weekly_periods_manual": 5,
            "is_active": True,
        },
    )
    assert previous_assignments_response.status_code == 200

    previous_math_assignment_response = client.post(
        "/api/teachers/assignments",
        json={
            "teacher_id": 2,
            "semester_id": semester_id,
            "grade_id": 1,
            "class_id": 1,
            "subject_id": 2,
            "course_type": "regular",
            "weekly_periods_manual": 6,
            "is_active": True,
        },
    )
    assert previous_math_assignment_response.status_code == 200

    exam_response = client.post(
        "/api/exams",
        json={
            "name": "2025届高一4月月考",
            "exam_type": "月考",
            "exam_date": "2025-04-10",
            "semester_id": semester_id,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            },
            {
                "subject_id": 2,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 2,
                "is_active": True,
            },
        ],
    )
    assert subject_response.status_code == 200

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    sheet.append(["2025届高一4月月考", "2026001", "张三", "1班", "语文", "108", "", ""])
    sheet.append(["2025届高一4月月考", "2026001", "张三", "1班", "数学", "115", "", ""])
    sheet.append(["2025届高一4月月考", "2026002", "李四", "1班", "语文", "102", "", ""])
    sheet.append(["2025届高一4月月考", "2026002", "李四", "1班", "数学", "110", "", ""])
    sheet.append(["2025届高一4月月考", "2026003", "王五", "2班", "语文", "88", "", ""])
    sheet.append(["2025届高一4月月考", "2026003", "王五", "2班", "数学", "92", "", ""])
    buffer = BytesIO()
    workbook.save(buffer)

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                "scores_prev.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    return exam_id


def test_grade_analytics_and_student_attachment_management(client) -> None:
    exam_id = create_exam_with_scores(client)
    create_previous_year_exam_with_scores(client)

    grade_response = client.get(f"/api/analytics/grades/1?exam_id={exam_id}")
    assert grade_response.status_code == 200
    grade_payload = grade_response.json()
    assert grade_payload["grade_name"] == "高一"
    assert grade_payload["student_count"] == 3
    assert len(grade_payload["class_breakdown"]) == 2
    assert len(grade_payload["subject_breakdown"]) == 2
    assert sum(item["count"] for item in grade_payload["score_bands"]) == 3

    panorama_response = client.get("/api/analytics/grades/1/panorama")
    assert panorama_response.status_code == 200
    panorama_payload = panorama_response.json()
    assert panorama_payload["grade_name"] == "高一"
    assert panorama_payload["academic_year_count"] == 2
    assert panorama_payload["exam_count"] == 2
    assert len(panorama_payload["year_summaries"]) == 2
    assert {item["academic_year_name"] for item in panorama_payload["year_summaries"]} == {"2024-2025", "2025-2026"}
    assert len(panorama_payload["subject_trends"]) == 2
    assert all(len(item["points"]) == 2 for item in panorama_payload["subject_trends"])

    class_panorama_response = client.get("/api/analytics/classes/1/panorama")
    assert class_panorama_response.status_code == 200
    class_panorama_payload = class_panorama_response.json()
    assert class_panorama_payload["class_name"] == "1班"
    assert class_panorama_payload["academic_year_count"] == 2
    assert class_panorama_payload["exam_count"] == 2
    assert len(class_panorama_payload["year_summaries"]) == 2
    assert len(class_panorama_payload["subject_trends"]) == 2
    assert all(len(item["points"]) == 2 for item in class_panorama_payload["subject_trends"])

    teacher_panorama_response = client.get("/api/analytics/teachers/1/panorama")
    assert teacher_panorama_response.status_code == 200
    teacher_panorama_payload = teacher_panorama_response.json()
    assert teacher_panorama_payload["teacher_name"] == "李语文"
    assert teacher_panorama_payload["academic_year_count"] == 2
    assert teacher_panorama_payload["exam_count"] == 2
    assert len(teacher_panorama_payload["year_summaries"]) == 2
    assert len(teacher_panorama_payload["subject_trends"]) == 1
    assert teacher_panorama_payload["subject_trends"][0]["subject_name"] == "语文"
    assert len(teacher_panorama_payload["subject_trends"][0]["points"]) == 2

    upload_response = client.post(
        "/api/files/upload",
        data={"category": "student_attachment"},
        files={"file": ("photo.jpg", b"binary-data", "image/jpeg")},
    )
    assert upload_response.status_code == 200
    file_id = upload_response.json()["id"]

    create_attachment_response = client.post(
        "/api/students/1/attachments",
        json={
            "stored_file_id": file_id,
            "attachment_type": "证件",
            "title": "学生照片",
            "note": "用于学籍核验",
            "is_active": True,
        },
    )
    assert create_attachment_response.status_code == 200
    attachment_payload = create_attachment_response.json()
    assert attachment_payload["title"] == "学生照片"
    assert attachment_payload["source_type"] == "student_attachment"

    list_attachment_response = client.get("/api/students/1/attachments")
    assert list_attachment_response.status_code == 200
    attachments = list_attachment_response.json()
    assert len(attachments) == 1
    assert attachments[0]["attachment_type"] == "证件"

    profile_response = client.get("/api/students/1/profile")
    assert profile_response.status_code == 200
    profile_payload = profile_response.json()
    assert any(item["title"] == "学生照片" for item in profile_payload["attachments"])

    delete_attachment_response = client.delete(f"/api/students/1/attachments/{attachment_payload['id']}")
    assert delete_attachment_response.status_code == 200
    assert delete_attachment_response.json()["message"] == "学生附件已删除"

    list_after_delete_response = client.get("/api/students/1/attachments")
    assert list_after_delete_response.status_code == 200
    assert list_after_delete_response.json() == []


def test_grade_summary_export_uses_grade_overview_sheet(client) -> None:
    exam_id = create_exam_with_scores(client)

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "grade_summary", "exam_id": exam_id, "grade_id": 1},
    )
    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["report_name"] == "年级成绩汇总表"

    download_response = client.get(export_payload["download_url"])
    assert download_response.status_code == 200

    workbook = load_workbook(BytesIO(download_response.content))
    assert workbook.sheetnames[:3] == ["年级概况", "班级汇总", "学科汇总"]

    overview_sheet = workbook["年级概况"]
    overview_rows = {
        overview_sheet.cell(row=index, column=1).value: overview_sheet.cell(row=index, column=2).value
        for index in range(1, overview_sheet.max_row + 1)
    }
    assert overview_rows["考试"] == "2026届高一5月月考"
    assert overview_rows["年级"] == "高一"
    assert overview_rows["学生数"] == 3
    assert overview_rows["年级均分"] is not None
    assert overview_rows["年级中位数"] is not None
    assert overview_rows["优秀率"] is not None

    class_sheet = workbook["班级汇总"]
    assert class_sheet.cell(row=1, column=1).value == "班级"
    assert class_sheet.max_row >= 3

    subject_sheet = workbook["学科汇总"]
    assert subject_sheet.cell(row=1, column=1).value == "科目"
    assert subject_sheet.max_row >= 3

    insight_sheet = workbook["摘要概览"]
    assert insight_sheet.cell(row=1, column=1).value == "标题"
    insight_rows = {
        insight_sheet.cell(row=index, column=1).value: insight_sheet.cell(row=index, column=2).value
        for index in range(2, insight_sheet.max_row + 1)
    }
    assert str(insight_rows["年级整体状态"]).startswith("高一 共 3 人，均分 ")
    assert "中位数" in str(insight_rows["年级整体状态"])
    assert insight_rows["领先班级"] == "1班 当前均分领先"
    assert insight_rows["需重点跟进班级"] == "2班 当前均分相对靠后"
