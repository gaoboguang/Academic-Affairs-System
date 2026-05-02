from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import (
    KnowledgePoint,
    KnowledgePointAlias,
    ScoreKnowledgeSnapshot,
    ScoreQuestionKnowledgePoint,
    ScoreQuestionRecord,
    StudentPlanningTask,
    Subject,
)


def _build_score_workbook(exam_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    for row in rows:
        sheet.append([exam_name, *row])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_question_workbook(exam_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "科目", "题号", "题目满分", "学生得分", "知识点", "题型", "能力层级", "错因标签", "错因备注", "备注"])
    for row in rows:
        normalized_row = list(row)
        if len(normalized_row) == 10:
            normalized_row = [*normalized_row[:9], "", "", normalized_row[9]]
        sheet.append([exam_name, *normalized_row])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_knowledge_base_workbook() -> bytes:
    workbook = Workbook()
    point_sheet = workbook.active
    point_sheet.title = "知识点"
    point_sheet.append(["科目", "知识点路径", "说明"])
    point_sheet.append(["数学", "函数>函数性质>单调性", "函数单调性标准知识点"])
    point_sheet.append(["数学", "函数>导数应用", "导数应用标准知识点"])

    alias_sheet = workbook.create_sheet("别名")
    alias_sheet.append(["科目", "标准知识点", "别名"])
    alias_sheet.append(["数学", "函数>函数性质>单调性", "增减性"])

    error_sheet = workbook.create_sheet("错因标签")
    error_sheet.append(["错因标签", "说明", "排序"])
    error_sheet.append(["模型迁移错误", "自定义错因", 30])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _math_subject_id(app) -> int:
    with app.state.db.session_scope() as session:
        subject = session.scalar(select(Subject).where(Subject.name == "数学"))
        assert subject is not None
        return subject.id


def _create_exam_with_scores(client, app, exam_name: str = "V22知识点闭环测试") -> int:
    math_subject_id = _math_subject_id(app)
    exam_response = client.post(
        "/api/exams",
        json={
            "name": exam_name,
            "exam_type": "阶段测试",
            "exam_date": "2026-05-12",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": math_subject_id,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            }
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                "scores.xlsx",
                _build_score_workbook(
                    exam_name,
                    [
                        ["2026001", "张三", "1班", "数学", 72, "", ""],
                        ["2026002", "李四", "1班", "数学", 126, "", ""],
                        ["2026003", "王五", "2班", "数学", 130, "", ""],
                    ],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    return exam_id


def _seed_knowledge_base(client) -> None:
    response = client.post(
        "/api/knowledge/import",
        files={
            "file": (
                "knowledge.xlsx",
                _build_knowledge_base_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["point_count"] == 2
    assert payload["alias_count"] == 1
    assert payload["error_tag_count"] == 1


def _import_question_details(client, exam_id: int, exam_name: str) -> None:
    rows = [
        ["2026001", "张三", "数学", "12", 10, 2, "增减性", "选择题", "理解", "概念不清、模型迁移错误", "别名命中"],
        ["2026001", "张三", "数学", "18", 10, 3, "函数>导数应用", "解答题", "综合应用", "方法不会|步骤不全", "路径创建/命中"],
        ["2026002", "李四", "数学", "12", 10, 8, "增减性", "选择题", "理解", "审题失误", ""],
        ["2026002", "李四", "数学", "18", 10, 8, "函数>导数应用", "解答题", "综合应用", "", ""],
        ["2026003", "王五", "数学", "12", 10, 9, "增减性", "选择题", "理解", "", ""],
        ["2026003", "王五", "数学", "18", 10, 10, "函数>导数应用", "解答题", "综合应用", "", ""],
    ]
    response = client.post(
        f"/api/exams/{exam_id}/score-questions/import",
        data={"strategy": "overwrite"},
        files={
            "file": (
                "questions.xlsx",
                _build_question_workbook(exam_name, rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    assert response.json()["failed_rows"] == 0


def _sheet_headers(sheet) -> dict[str, int]:
    return {cell.value: index for index, cell in enumerate(sheet[1], start=1)}


def test_knowledge_base_import_alias_normalization_and_tree(client, app) -> None:
    _seed_knowledge_base(client)
    math_subject_id = _math_subject_id(app)

    tree_response = client.get("/api/knowledge/tree", params={"subject_id": math_subject_id})
    assert tree_response.status_code == 200
    tree_payload = tree_response.json()
    root = next(item for item in tree_payload["items"] if item["name"] == "函数")
    child = next(item for item in root["children"] if item["name"] == "函数性质")
    assert child["children"][0]["path"] == "函数>函数性质>单调性"

    aliases_response = client.get("/api/knowledge/aliases", params={"subject_id": math_subject_id})
    assert aliases_response.status_code == 200
    alias_payload = aliases_response.json()[0]
    assert alias_payload["alias_name"] == "增减性"
    assert alias_payload["knowledge_point_path"] == "函数>函数性质>单调性"

    duplicate_response = client.post(
        "/api/knowledge/aliases",
        json={
            "subject_id": math_subject_id,
            "knowledge_point_id": alias_payload["knowledge_point_id"],
            "alias_name": " 增 减 性 ",
        },
    )
    assert duplicate_response.status_code == 400
    assert duplicate_response.json()["detail"] == "同一科目下别名不能重复。"

    with app.state.db.session_scope() as session:
        point = session.scalar(select(KnowledgePoint).where(KnowledgePoint.name == "单调性"))
        alias = session.scalar(select(KnowledgePointAlias).where(KnowledgePointAlias.alias_name == "增减性"))
        assert point is not None
        assert alias is not None
        assert alias.knowledge_point_id == point.id


def test_question_import_error_tags_student_analysis_and_class_briefing(client, app) -> None:
    _seed_knowledge_base(client)
    exam_name = "V22知识点闭环测试"
    exam_id = _create_exam_with_scores(client, app, exam_name)
    _import_question_details(client, exam_id, exam_name)

    with app.state.db.session_scope() as session:
        record = session.scalar(
            select(ScoreQuestionRecord)
            .where(ScoreQuestionRecord.student_id == 1)
            .order_by(ScoreQuestionRecord.id.asc())
        )
        assert record is not None
        assert [item["name"] for item in record.error_tags_json] == ["概念不清", "模型迁移错误"]

        alias_link = session.scalar(
            select(ScoreQuestionKnowledgePoint)
            .join(KnowledgePoint, KnowledgePoint.id == ScoreQuestionKnowledgePoint.knowledge_point_id)
            .where(KnowledgePoint.name == "单调性")
        )
        assert alias_link is not None
        assert alias_link.match_source == "alias"
        assert alias_link.raw_knowledge_text == "增减性"

        snapshot = session.scalar(
            select(ScoreKnowledgeSnapshot)
            .join(KnowledgePoint, KnowledgePoint.id == ScoreKnowledgeSnapshot.knowledge_point_id)
            .where(ScoreKnowledgeSnapshot.student_id == 1, KnowledgePoint.name == "单调性")
        )
        assert snapshot is not None
        assert snapshot.dominant_error_tag == "概念不清"
        assert {"tag": "模型迁移错误", "count": 1} in snapshot.error_tags_json

    student_response = client.get(f"/api/analytics/students/1?exam_id={exam_id}")
    assert student_response.status_code == 200
    student_payload = student_response.json()
    first_point = student_payload["knowledge_points"][0]
    assert first_point["knowledge_path"] in {"函数>函数性质>单调性", "函数>导数应用"}
    assert first_point["dominant_error_tag"]
    assert first_point["error_tag_stats"]

    briefing_response = client.get(f"/api/analytics/classes/1/knowledge-briefing?exam_id={exam_id}")
    assert briefing_response.status_code == 200
    briefing_payload = briefing_response.json()
    assert briefing_payload["items"]
    top_item = briefing_payload["items"][0]
    assert top_item["weak_student_count"] >= 1
    assert top_item["knowledge_path"] in {"函数>函数性质>单调性", "函数>导数应用"}
    assert top_item["error_tag_stats"]
    assert top_item["weak_students"][0]["student_name"] == "张三"

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "class_knowledge_briefing", "exam_id": exam_id, "class_id": 1},
    )
    assert export_response.status_code == 200
    download_response = client.get(export_response.json()["download_url"])
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    assert {"讲评摘要", "讲评清单", "弱项学生", "任务建议"} <= set(workbook.sheetnames)
    detail_headers = _sheet_headers(workbook["讲评清单"])
    assert workbook["讲评清单"].cell(row=2, column=detail_headers["知识路径"]).value in {
        "函数>函数性质>单调性",
        "函数>导数应用",
    }
    assert workbook["讲评清单"].cell(row=2, column=detail_headers["错因分布"]).value


def test_knowledge_remediation_task_preview_generate_is_idempotent(client, app) -> None:
    _seed_knowledge_base(client)
    exam_name = "V22任务生成测试"
    exam_id = _create_exam_with_scores(client, app, exam_name)
    _import_question_details(client, exam_id, exam_name)

    preview_response = client.get(f"/api/planning/students/1/knowledge-tasks/preview?exam_id={exam_id}")
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["create_count"] >= 1
    candidate = preview_payload["candidates"][0]
    assert candidate["source_ref_id"] == f"{exam_id}:1:{candidate['knowledge_point_id']}"
    assert candidate["will_create"] is True

    generate_response = client.post(
        f"/api/planning/students/1/knowledge-tasks/generate?exam_id={exam_id}",
        json={},
    )
    assert generate_response.status_code == 200
    generate_payload = generate_response.json()
    assert generate_payload["created_count"] >= 1
    assert generate_payload["tasks"][0]["source_type"] == "knowledge_remediation"
    assert generate_payload["tasks"][0]["task_type"] == "knowledge_remediation"

    second_generate = client.post(
        f"/api/planning/students/1/knowledge-tasks/generate?exam_id={exam_id}",
        json={},
    )
    assert second_generate.status_code == 200
    assert second_generate.json()["created_count"] == 0
    assert second_generate.json()["skipped_count"] >= generate_payload["created_count"]

    class_preview = client.get(
        "/api/planning/classes/1/knowledge-tasks/preview",
        params=[("exam_id", str(exam_id)), ("knowledge_point_ids", str(candidate["knowledge_point_id"]))],
    )
    assert class_preview.status_code == 200, class_preview.json()
    assert all(item["knowledge_point_id"] == candidate["knowledge_point_id"] for item in class_preview.json()["candidates"])

    class_generate = client.post(
        f"/api/planning/classes/1/knowledge-tasks/generate?exam_id={exam_id}",
        json={"knowledge_point_ids": [candidate["knowledge_point_id"]]},
    )
    assert class_generate.status_code == 200
    assert class_generate.json()["created_count"] >= 0

    with app.state.db.session_scope() as session:
        tasks = session.scalars(
            select(StudentPlanningTask).where(
                StudentPlanningTask.source_type == "knowledge_remediation",
                StudentPlanningTask.task_type == "knowledge_remediation",
                StudentPlanningTask.is_active.is_(True),
            )
        ).all()
        assert tasks
        assert all(task.source_ref_id and task.source_ref_id.startswith(f"{exam_id}:") for task in tasks)


def test_student_knowledge_plan_export_contains_v22_columns(client, app) -> None:
    _seed_knowledge_base(client)
    exam_name = "V22学生清单导出测试"
    exam_id = _create_exam_with_scores(client, app, exam_name)
    _import_question_details(client, exam_id, exam_name)

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "student_knowledge_plan", "exam_id": exam_id, "student_id": 1},
    )
    assert export_response.status_code == 200
    download_response = client.get(export_response.json()["download_url"])
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))

    diagnosis_headers = _sheet_headers(workbook["知识点诊断"])
    assert "知识路径" in diagnosis_headers
    assert "主错因" in diagnosis_headers
    assert "错因分布" in diagnosis_headers
    assert workbook["知识点诊断"].cell(row=2, column=diagnosis_headers["知识路径"]).value in {
        "函数>函数性质>单调性",
        "函数>导数应用",
    }
    assert workbook["知识点诊断"].cell(row=2, column=diagnosis_headers["主错因"]).value
    assert "任务建议" in workbook.sheetnames
