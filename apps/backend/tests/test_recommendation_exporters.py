from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.core.bootstrap import ensure_runtime_directories
from app.exporters.recommendations import (
    export_gaokao_pathway_report,
    export_recommendation_summary,
    export_shandong_recommendation_report,
    export_volunteer_draft_summary,
)


def build_candidate_row(**overrides):
    row = {
        "order": 1,
        "result_type": "steady",
        "college_name": "测试大学",
        "college_code_snapshot": "10561",
        "major_name": "软件工程",
        "major_code_snapshot": "080902",
        "major_group_code": "201",
        "plan_count": 80,
        "province": "广东",
        "year": 2026,
        "batch": "本科批",
        "exam_mode": "物理类",
        "major_id": 1,
        "reference_rank": 28000,
        "latest_min_rank": 28500,
        "latest_min_score": 585,
        "score_basis": "rank",
        "reference_scope": "major",
        "reference_years_json": [2025],
        "reference_record_count": 1,
        "reference_source_notes_json": ["近年数据"],
        "matched_rule_exam_mode": "物理类",
        "matched_rule_batch": "本科批",
        "matched_rule_candidate_type": "general",
        "matched_rule_is_baseline": False,
        "match_tags_json": ["专业线参考"],
        "risk_flags_json": [],
        "matched_direction_names_json": [],
        "reason_text": "测试结果",
    }
    row.update(overrides)
    return row


def load_exported_workbook(path: str, project_root: Path):
    export_path = project_root / path
    return load_workbook(BytesIO(export_path.read_bytes()))


def build_shandong_candidate(**overrides):
    row = {
        "college_id": 1,
        "college_name": "山东示例大学",
        "college_code_snapshot": "A001",
        "major_id": 2,
        "major_name": "计算机类",
        "major_code_snapshot": "0809",
        "bucket": "stable",
        "bucket_label": "稳",
        "rank_margin": 1200,
        "rank_margin_ratio": 0.08,
        "score_summary": {
            "reference_rank": 15000,
            "latest_min_score": 612,
            "latest_min_rank": 15200,
        },
        "years_used": [2025, 2024, 2023],
        "historical_summary": {
            "weighted_reference_rank": 15000,
            "rank_rows": [
                {"year": 2025, "min_rank": 15200, "min_score": 612, "plan_count": 12, "source_note": "2025 山东普通类投档表"},
                {"year": 2024, "min_rank": 15100, "min_score": 609, "plan_count": 10, "source_note": "2024 山东普通类投档表"},
            ],
            "plan_change": {
                "target_year_plan_count": None,
                "latest_historical_plan_count": 12,
                "plan_change_factor": None,
            },
        },
        "plan_count": None,
        "subject_requirement": "物理 化学",
        "data_confidence": "medium",
        "risk_flags": ["plan_missing", "three_year_data_incomplete"],
        "explanation_text": "按近三年山东普通类投档数据计算，归为稳。",
        "source_document_ids": [101, 102],
    }
    row.update(overrides)
    return row


def test_shandong_recommendation_export_splits_buckets_and_localizes_risks(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_shandong_recommendation_report(
        test_settings,
        {
            "student_id": 1,
            "student_name": "张三",
            "province": "山东",
            "target_year": 2026,
            "student_type": "general",
            "source_mode": "manual_rank",
            "predicted_score": 620,
            "predicted_rank": 13800,
            "rank_range_low": 13800,
            "rank_range_high": 13800,
            "rank_projection_basis": "manual_rank",
            "risk_preference": "balanced",
            "data_years": [2025, 2024, 2023],
            "input_notes": ["当前按手动填写的全省位次作为主排序依据。"],
            "summary": {
                "rush_count": 1,
                "stable_count": 1,
                "safe_count": 1,
                "watch_count": 1,
                "excluded_subject_mismatch_count": 2,
            },
            "rush": [build_shandong_candidate(bucket="rush", bucket_label="冲", major_name="人工智能")],
            "stable": [build_shandong_candidate()],
            "safe": [build_shandong_candidate(bucket="safe", bucket_label="保", major_name="软件工程", risk_flags=[])],
            "watch": [build_shandong_candidate(bucket="watch", bucket_label="仅关注", major_name="数据科学", risk_flags=["historical_data_missing"])],
        },
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    assert workbook.sheetnames == ["汇总页", "风险说明", "冲列表", "稳列表", "保列表", "数据不足与风险列表", "数据来源页"]
    summary_sheet = workbook["汇总页"]
    assert summary_sheet.cell(row=2, column=2).value == "张三"
    assert summary_sheet.cell(row=6, column=2).value == "手动全省位次"
    notice = str(summary_sheet.cell(row=16, column=2).value)
    assert "当前主要参考 2023-2025 历史投档数据" in notice
    assert "正式填报前需导入 2026 官方计划" in notice

    stable_sheet = workbook["稳列表"]
    assert stable_sheet.cell(row=2, column=1).value == "稳"
    assert "目标年份招生计划暂缺" in str(stable_sheet.cell(row=2, column=17).value)
    assert "plan_missing" not in str(stable_sheet.cell(row=2, column=17).value)

    source_sheet = workbook["数据来源页"]
    source_values = [source_sheet.cell(row=index, column=3).value for index in range(2, source_sheet.max_row + 1)]
    assert any("2025 山东普通类投档表" in str(item) for item in source_values)


def test_gaokao_pathway_export_includes_cards_gaps_and_data_risks(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_gaokao_pathway_report(
        test_settings,
        {
            "student_id": 3,
            "student_name": "张三",
            "target_year": 2026,
            "generated_at": "2026-04-25T22:00:00",
            "data_health_summary": "可验收但有数据警告",
            "profile_summary": [
                {"key": "province", "label": "生源地", "value": "山东", "filled": True},
                {"key": "subject_combination", "label": "选科组合", "value": "物理,化学,生物", "filled": True},
            ],
            "cards": [
                {
                    "code": "summer_general_regular",
                    "name": "普通类常规批",
                    "group": "夏季高考",
                    "status_label": "适合关注",
                    "depth_label": "可接完整位次推荐",
                    "confidence_label": "中等",
                    "score": 94,
                    "applicable_object": "普通类学生 · 夏季高考",
                    "volunteer_mode": "专业（专业类）+学校",
                    "summary": "可进入山东普通类推荐工作台。",
                    "key_requirements": ["报名：已完成山东高考报名"],
                    "missing_materials": [],
                    "risk_messages": ["正式填报前仍需导入 2026 官方计划"],
                    "next_actions": ["进入山东普通类推荐工作台查看冲稳保候选。"],
                    "source_document_id": 11,
                },
                {
                    "code": "vocational_comprehensive",
                    "name": "高职综合评价招生",
                    "group": "高职分类招生",
                    "status_label": "信息不足",
                    "depth_label": "资格初筛",
                    "confidence_label": "待补充",
                    "score": 45,
                    "applicable_object": "普通类学生",
                    "volunteer_mode": "院校报名与素质测试",
                    "summary": "只做资格初筛。",
                    "key_requirements": ["材料：综合素质评价、素质测试或面试安排"],
                    "missing_materials": ["综合素质评价材料"],
                    "risk_messages": ["当前只做资格初筛和人工复核清单，不输出录取概率。"],
                    "next_actions": ["补齐综合素质评价材料后重新评估该路径。"],
                    "source_document_id": 12,
                },
            ],
            "material_gaps": [
                {
                    "key": "comprehensive_quality_evaluation",
                    "label": "综合素质评价材料",
                    "count": 1,
                    "pathways": ["高职综合评价招生"],
                    "nextAction": "补齐综合素质评价材料后重新评估该路径。",
                }
            ],
            "next_actions": [
                {
                    "key": "recommendation-entry",
                    "title": "普通类常规批可以继续看冲稳保候选",
                    "detail": "进入山东普通类推荐工作台后，仍要以 2026 官方计划和高校章程复核结果。",
                    "tone": "primary",
                }
            ],
            "publication_status": [
                {
                    "key": "summer_general_plan",
                    "label": "2026 普通类正式招生计划",
                    "status_label": "待官方发布",
                    "action_label": "等待官方发布后导入",
                    "explanation": "不能把单招综评材料当作普通类正式计划。",
                    "blocks_recommendation": True,
                }
            ],
            "p0_gaps": ["2026 普通类正式计划未导入"],
        },
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    assert workbook.sheetnames == ["汇总页", "学生画像", "路径建议", "材料缺口", "下一步行动", "数据风险"]
    assert workbook["汇总页"].cell(row=2, column=2).value == "张三"
    assert workbook["汇总页"].cell(row=9, column=2).value == "普通类常规批可进入位次型推荐；其他路径只做资格初筛、政策提醒和人工复核，不输出录取概率。"
    pathway_sheet = workbook["路径建议"]
    assert pathway_sheet.cell(row=2, column=1).value == "普通类常规批"
    assert pathway_sheet.cell(row=3, column=4).value == "资格初筛"
    assert "综合素质评价材料" in str(pathway_sheet.cell(row=3, column=11).value)
    data_sheet = workbook["数据风险"]
    assert data_sheet.cell(row=2, column=1).value == "2026 普通类正式招生计划"
    assert data_sheet.cell(row=3, column=1).value == "P0 数据缺口"


def test_volunteer_draft_export_includes_missing_rule_and_general_rule_summaries(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_volunteer_draft_summary(
      test_settings,
      {
          "draft_name": "测试草稿",
          "student_name": "张三",
          "exam_name": "2026届一模",
          "province": "广东",
          "target_year": 2026,
          "batch": "本科批",
          "exam_mode": "物理类",
          "score_input_label": "正式位次",
          "simulation_note": "默认推荐链路",
          "target_direction_summary": "未维护",
          "accepted_path_summary": "未维护",
          "rule_label": "未命中明确规则",
          "rule_alerts": [
              {
                  "code": "missing_rule_year",
                  "level": "warning",
                  "title": "缺少目标年份规则",
                  "detail": "当前未找到 广东 2025 年省份规则；该省现有 2026 年规则，志愿上限与单位类型需按当年公告人工复核。",
              },
              {
                  "code": "fallback_general_candidate_rule",
                  "level": "info",
                  "title": "已回退到通用考生规则",
                  "detail": "当前未配置“general”专用规则，先按通用考生规则预览。",
              },
              {
                  "code": "fallback_general_reference_data",
                  "level": "info",
                  "title": "已回退到普通类录取参考",
                  "detail": "当前缺少“spring_exam”专门录取结果，已先回退参考普通类录取结果；正式填报前建议结合学校公告和类别专门批次再复核。",
              },
          ],
          "applicable_rules": [
              {
                  "id": 1,
                  "province": "广东",
                  "year": 2025,
                  "exam_mode": "物理类",
                  "batch": "本科批",
                  "candidate_type": "",
                  "batch_order": 1,
                  "total_score": 750,
                  "volunteer_limit": 45,
                  "volunteer_unit_type": "院校专业组",
                  "subject_requirement_mode": "first_choice_reselect",
                  "required_subjects_json": [],
                  "first_choice_subjects_json": ["物理", "历史"],
                  "reselect_subjects_json": ["化学", "生物"],
                  "score_rule_summary": "按等级赋分",
                  "parallel_rule_mode": "group_parallel",
                  "max_major_per_unit": 6,
                  "is_parallel": True,
                  "allow_adjustment": True,
                  "support_collect_round": True,
                  "special_rules_json": ["需核对选科要求"],
                  "note": "系统基线初始化生成",
                  "created_at": "2026-04-15T10:00:00",
                  "updated_at": "2026-04-15T10:00:00",
                  "is_active": True,
              },
              {
                  "id": 2,
                  "province": "广东",
                  "year": 2025,
                  "exam_mode": "历史类",
                  "batch": "本科批",
                  "candidate_type": "",
                  "batch_order": 1,
                  "total_score": 750,
                  "volunteer_limit": 30,
                  "volunteer_unit_type": "平行志愿",
                  "subject_requirement_mode": "unified_subject_requirement",
                  "required_subjects_json": ["历史"],
                  "first_choice_subjects_json": [],
                  "reselect_subjects_json": [],
                  "score_rule_summary": "按历史类顺序志愿说明执行",
                  "parallel_rule_mode": "ordered_sequential",
                  "max_major_per_unit": None,
                  "is_parallel": False,
                  "allow_adjustment": False,
                  "support_collect_round": False,
                  "special_rules_json": [],
                  "note": "兼容模式预览",
                  "created_at": "2026-04-15T10:00:00",
                  "updated_at": "2026-04-15T10:00:00",
                  "is_active": True,
              },
          ],
          "selected_rule": {
              "id": 1,
              "province": "广东",
              "year": 2025,
              "exam_mode": "物理类",
              "batch": "本科批",
              "candidate_type": "",
              "batch_order": 1,
              "total_score": 750,
              "volunteer_limit": 45,
              "volunteer_unit_type": "院校专业组",
              "subject_requirement_mode": "first_choice_reselect",
              "required_subjects_json": [],
              "first_choice_subjects_json": ["物理", "历史"],
              "reselect_subjects_json": ["化学", "生物"],
              "score_rule_summary": "按等级赋分",
              "parallel_rule_mode": "group_parallel",
              "max_major_per_unit": 6,
              "is_parallel": True,
              "allow_adjustment": True,
              "support_collect_round": True,
              "special_rules_json": ["需核对选科要求"],
              "note": "系统基线初始化生成",
              "created_at": "2026-04-15T10:00:00",
              "updated_at": "2026-04-15T10:00:00",
              "is_active": True,
          },
          "note": "",
      },
      [
          build_candidate_row(
              matched_rule_exam_mode=None,
              matched_rule_batch=None,
              matched_rule_candidate_type=None,
              matched_rule_is_baseline=False,
              match_tags_json=["专业线参考"],
              year=2026,
              reference_years_json=[2023],
              reference_record_count=1,
          ),
          build_candidate_row(
              order=2,
              matched_rule_exam_mode="物理类",
              matched_rule_batch="本科批",
              matched_rule_candidate_type="",
              matched_rule_is_baseline=False,
              match_tags_json=["专业线参考"],
              student_type="spring_exam",
              risk_flags_json=["general_reference_fallback"],
          ),
        build_candidate_row(
            order=3,
            matched_rule_exam_mode="物理类",
            matched_rule_batch="本科批",
            matched_rule_candidate_type="general",
            matched_rule_is_baseline=False,
            match_tags_json=["专业线参考"],
            reference_years_json=[2024],
            reference_record_count=1,
        ),
        build_candidate_row(
            order=4,
            province="四川",
            matched_rule_exam_mode="物理类",
            matched_rule_batch="本科批",
            matched_rule_candidate_type="general",
            matched_rule_is_baseline=False,
            match_tags_json=["专业线参考", "待核对选科"],
            risk_flags_json=["subject_requirement_check"],
            subject_requirement="物理 + 化学",
            reference_years_json=[2025],
            reference_record_count=1,
            year=2026,
          ),
      ],
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    boundary_sheet = workbook["边界概览"]
    rule_sheet = workbook["规则差异摘要"]
    overview_sheet = workbook["导出前摘要"]
    detail_sheet = workbook["志愿草稿"]

    boundary_rows = [
        (
            boundary_sheet.cell(row=index, column=1).value,
            boundary_sheet.cell(row=index, column=2).value,
            boundary_sheet.cell(row=index, column=3).value,
        )
        for index in range(2, boundary_sheet.max_row + 1)
    ]
    assert ("缺少目标年份规则", "1 条已选志愿当前缺少目标年份规则支撑", "当前未找到 广东 2025 年省份规则；该省现有 2026 年规则，志愿上限与单位类型需按当年公告人工复核。") in boundary_rows
    assert ("已回退到通用考生规则", "1 条已选志愿当前按通用考生规则解释", "当前未配置“general”专用规则，先按通用考生规则预览。") in boundary_rows
    assert ("已回退到普通类录取参考", "1 条已选志愿当前按普通类录取结果参考", "当前缺少“spring_exam”专门录取结果，已先回退参考普通类录取结果；正式填报前建议结合学校公告和类别专门批次再复核。") in boundary_rows
    assert ("类别专用规则口径", "2 条已选志愿当前按普通类专用规则解释", "当前命中的省份规则已细分到普通类；同省同年其他类别可能适用不同的志愿上限、单位结构和选科口径。") in boundary_rows
    assert ("跨年份参考样本", "4 条已选志愿最近录取样本分布在 2025 / 2024 / 2023 年", "这些志愿当前涉及 广东 / 四川 等多个口径，最近录取样本也并非同一年；跨省或跨年份比较时，录取位次、最低分和冲稳保分组变化属于正常现象。") in boundary_rows
    assert ("跨省口径差异", "2 个省份口径混合", "这些志愿当前涉及 广东 / 四川 等多个口径，跨省比较时，录取位次、最低分和冲稳保分组变化属于正常现象。") in boundary_rows
    assert ("选科待核对", "1 条已选志愿仍需逐条核对选科限制", "导出稿可用于讨论和汇报，但最终填报前仍应回到招生章程确认专业或专业组要求。") in boundary_rows
    assert ("参考年份偏旧", "2 条已选志愿最近录取样本与目标年份相差 2 年及以上", "这类志愿当前更适合作为方向性参考；若后续补齐近一年录取数据，排序、边界说明和最终取舍都可能变化。") in boundary_rows
    assert rule_sheet.cell(row=1, column=1).value == "规则"
    assert rule_sheet.cell(row=2, column=1).value == "广东 2025 物理类 · 本科批"
    assert rule_sheet.cell(row=2, column=2).value == "当前控制规则"
    assert "按等级赋分" in str(rule_sheet.cell(row=2, column=11).value)
    assert rule_sheet.cell(row=3, column=2).value == "兼容预览规则"

    overview_rows = [
        (
            overview_sheet.cell(row=index, column=1).value,
            overview_sheet.cell(row=index, column=2).value,
            overview_sheet.cell(row=index, column=3).value,
            overview_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, overview_sheet.max_row + 1)
    ]
    assert (
        "规则差异摘要",
        "缺少目标年份规则",
        "1 条已选志愿当前缺少目标年份规则支撑",
        "当前未找到 广东 2025 年省份规则；该省现有 2026 年规则，志愿上限与单位类型需按当年公告人工复核。",
    ) in overview_rows
    assert (
        "边界概览",
        "已回退到普通类录取参考",
        "1 条已选志愿当前按普通类录取结果参考",
        "当前缺少“spring_exam”专门录取结果，已先回退参考普通类录取结果；正式填报前建议结合学校公告和类别专门批次再复核。",
    ) in overview_rows
    assert (
        "边界概览",
        "跨省口径差异",
        "2 个省份口径混合",
        "这些志愿当前涉及 广东 / 四川 等多个口径，跨省比较时，录取位次、最低分和冲稳保分组变化属于正常现象。",
    ) in overview_rows
    assert (
        "风险概览",
        "选科待核对",
        "1 条已选志愿仍需逐条核对选科限制",
        "导出稿可用于讨论和汇报，但最终填报前仍应回到招生章程确认专业或专业组要求。",
    ) in overview_rows

    assert detail_sheet.cell(row=2, column=15).value is None
    assert "当前未找到 广东 2025 年省份规则" in str(detail_sheet.cell(row=2, column=17).value)
    assert "当前录取参考最近只到 2023 年，与 2026 目标年相差 3 年" in str(detail_sheet.cell(row=2, column=17).value)
    assert "通用规则" in str(detail_sheet.cell(row=3, column=15).value)
    assert "当前未配置“general”专用规则" in str(detail_sheet.cell(row=3, column=17).value)
    assert "普通类录取结果做方向性参考" in str(detail_sheet.cell(row=3, column=17).value)
    assert detail_sheet.cell(row=3, column=23).value == "缺少专门录取结果，按普通类参考"
    assert detail_sheet.cell(row=5, column=23).value == "需复核选科要求"


def test_recommendation_export_includes_stale_reference_year_summary(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_recommendation_summary(
        test_settings,
        {
            "scheme_name": "测试推荐",
            "student_name": "李四",
            "exam_name": "2026届一模",
            "province": "广东",
            "target_year": 2026,
            "score_input_label": "正式位次",
            "simulation_note": "默认推荐链路",
            "target_direction_summary": "未维护",
            "accepted_path_summary": "未维护",
        },
        [
            {
                "result_type": "steady",
                "college_name": "测试大学",
                "major_name": "软件工程",
                "reference_rank": 28000,
                "student_rank": 27500,
                "ratio": 0.98,
                "score_basis": "rank",
                "career_match_strength": "high",
                "matched_direction_names_json": ["软件平台工程"],
                "requires_postgraduate_path": False,
                "requires_certificate_path": False,
                "requires_long_training_path": False,
                "career_match_summary": "测试职业说明",
                "reason_text": "测试结果",
                "risk_flags_json": ["sample_insufficient", "rank_missing"],
                "snapshot_json": {
                    "reference_years": [2023],
                },
            }
        ],
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    risk_sheet = workbook["风险概览"]
    overview_sheet = workbook["导出前摘要"]
    detail_sheet = workbook["推荐结果"]
    risk_rows = [
        (
            risk_sheet.cell(row=index, column=1).value,
            risk_sheet.cell(row=index, column=2).value,
            risk_sheet.cell(row=index, column=3).value,
            risk_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, risk_sheet.max_row + 1)
    ]
    overview_rows = [
        (
            overview_sheet.cell(row=index, column=1).value,
            overview_sheet.cell(row=index, column=2).value,
            overview_sheet.cell(row=index, column=3).value,
            overview_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, overview_sheet.max_row + 1)
    ]

    assert (
        "边界概览",
        "统一风险口径",
        "正式输出保留普通类、特殊类型、2024/2026 数据和章程复核边界",
        "2026 正式招生计划、一分一段和省控线需等待官方发布，正式填报前必须替换为当年官方数据。",
    ) in risk_rows
    assert (
        "边界概览",
        "参考年份偏旧",
        "1 条结果最近录取样本与目标年份相差 2 年及以上",
        "这类推荐更适合作为方向性参考；若近一年录取数据尚未补齐，分层、排序和汇报口径都可能继续变化。",
    ) in risk_rows
    assert (
        "边界概览",
        "参考年份偏旧",
        "1 条结果最近录取样本与目标年份相差 2 年及以上",
        "这类推荐更适合作为方向性参考；若近一年录取数据尚未补齐，分层、排序和汇报口径都可能继续变化。",
    ) in overview_rows
    assert detail_sheet.cell(row=2, column=8).value == "2023"
    assert detail_sheet.cell(row=2, column=17).value == "样本不足 / 缺少位次，分数参考"


def test_recommendation_export_includes_history_comparison_rows(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_recommendation_summary(
        test_settings,
        {
            "scheme_name": "当前方案",
            "student_name": "李四",
            "exam_name": "2026届一模",
            "province": "广东",
            "target_year": 2026,
            "score_input_label": "正式位次",
            "simulation_note": "默认推荐链路",
            "target_direction_summary": "未维护",
            "accepted_path_summary": "未维护",
            "compare_scheme": {
                "scheme_id": 7,
                "scheme_name": "历史方案 A",
                "generated_at": "2026-04-10T10:00:00",
                "province": "河北",
                "target_year": 2025,
            },
            "compare_rows": [
                {
                    "result_type": "challenge",
                    "college_id": 11,
                    "college_name": "测试大学",
                    "major_id": 21,
                    "major_name": "软件工程",
                    "score_basis": "rank",
                    "snapshot_json": {
                        "reference_years": [2025],
                    },
                },
                {
                    "result_type": "safe",
                    "college_id": 15,
                    "college_name": "历史院校",
                    "major_id": 25,
                    "major_name": "历史专业",
                    "score_basis": "rank",
                    "snapshot_json": {
                        "reference_years": [2025],
                    },
                },
            ],
        },
        [
            {
                "result_type": "steady",
                "college_id": 11,
                "college_name": "测试大学",
                "major_id": 21,
                "major_name": "软件工程",
                "reference_rank": 28000,
                "student_rank": 27500,
                "ratio": 0.98,
                "score_basis": "rank",
                "career_match_strength": "high",
                "matched_direction_names_json": ["软件平台工程"],
                "requires_postgraduate_path": False,
                "requires_certificate_path": False,
                "requires_long_training_path": False,
                "career_match_summary": "测试职业说明",
                "reason_text": "测试结果",
                "risk_flags_json": [],
                "snapshot_json": {
                    "reference_years": [2023],
                },
            },
            {
                "result_type": "safe",
                "college_id": 12,
                "college_name": "新增院校",
                "major_id": 22,
                "major_name": "新增专业",
                "score_basis": "rank",
                "risk_flags_json": [],
                "snapshot_json": {
                    "reference_years": [2025],
                },
            },
        ],
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    risk_sheet = workbook["风险概览"]
    overview_sheet = workbook["导出前摘要"]
    risk_rows = [
        (
            risk_sheet.cell(row=index, column=1).value,
            risk_sheet.cell(row=index, column=2).value,
            risk_sheet.cell(row=index, column=3).value,
            risk_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, risk_sheet.max_row + 1)
    ]
    overview_rows = [
        (
            overview_sheet.cell(row=index, column=1).value,
            overview_sheet.cell(row=index, column=2).value,
            overview_sheet.cell(row=index, column=3).value,
            overview_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, overview_sheet.max_row + 1)
    ]

    assert (
        "历史对照摘要",
        "历史方案差异",
        "相对“历史方案 A”（2026-04-10），当前方案新增 1 条、移除 1 条、分组调整 1 条",
        "冲刺 0 对 1（-1）；稳妥 1 对 0（+1）；保底 1 对 1（0）；仅关注 0 对 0（0）；保留 0 条可比结果。",
    ) in risk_rows
    assert (
        "历史对照摘要",
        "历史方案参考变化",
        "1 条保留院校/专业的最近录取样本年份发生变化，其中 1 条同时伴随冲稳保分组变化，1 条结果的“年份偏旧”状态发生变化",
        "“历史方案 A”（2026-04-10）；当前方案目标年份 2026，对比方案目标年份 2025；在参考年份发生变化的保留结果里，有 1 条同时出现了冲稳保分组调整。；当前方案按 广东 口径，对比方案按 河北 口径；同校同专业在跨省比较时，录取位次和分组变化属于正常现象。；若同校同专业的参考年份被补新或回退，分组、排序和汇报口径出现变化属于正常现象。",
    ) in risk_rows
    assert (
        "历史对照摘要",
        "跨省口径差异",
        "1 条可比结果受跨省口径影响",
        "当前方案按 广东 口径，对比方案按 河北 口径；同校同专业在跨省比较时，录取位次和分组变化属于正常现象。",
    ) in risk_rows
    assert (
        "历史对照摘要",
        "历史方案差异",
        "相对“历史方案 A”（2026-04-10），当前方案新增 1 条、移除 1 条、分组调整 1 条",
        "冲刺 0 对 1（-1）；稳妥 1 对 0（+1）；保底 1 对 1（0）；仅关注 0 对 0（0）；保留 0 条可比结果。",
    ) in overview_rows


def test_recommendation_export_includes_same_province_year_change_copy(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_recommendation_summary(
        test_settings,
        {
            "scheme_name": "当前方案",
            "student_name": "李四",
            "exam_name": "2026届一模",
            "province": "广东",
            "target_year": 2026,
            "score_input_label": "正式位次",
            "simulation_note": "默认推荐链路",
            "target_direction_summary": "未维护",
            "accepted_path_summary": "未维护",
            "compare_scheme": {
                "scheme_id": 7,
                "scheme_name": "历史方案 B",
                "generated_at": "2026-04-10T10:00:00",
                "province": "广东",
                "target_year": 2025,
            },
            "compare_rows": [
                {
                    "result_type": "steady",
                    "college_id": 11,
                    "college_name": "测试大学",
                    "major_id": 21,
                    "major_name": "软件工程",
                    "reference_rank": 26000,
                    "score_basis": "rank",
                    "snapshot_json": {
                        "reference_years": [2025],
                        "latest_min_rank": 26800,
                        "latest_min_score": 592,
                    },
                },
            ],
        },
        [
            {
                "result_type": "steady",
                "college_id": 11,
                "college_name": "测试大学",
                "major_id": 21,
                "major_name": "软件工程",
                "reference_rank": 28000,
                "score_basis": "rank",
                "risk_flags_json": [],
                "snapshot_json": {
                    "reference_years": [2023],
                    "latest_min_rank": 28500,
                    "latest_min_score": 585,
                },
            },
        ],
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    risk_sheet = workbook["风险概览"]
    risk_rows = [
        (
            risk_sheet.cell(row=index, column=1).value,
            risk_sheet.cell(row=index, column=2).value,
            risk_sheet.cell(row=index, column=3).value,
            risk_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, risk_sheet.max_row + 1)
    ]
    matching_details = [
        str(risk_sheet.cell(row=index, column=4).value)
        for index in range(2, risk_sheet.max_row + 1)
        if risk_sheet.cell(row=index, column=2).value == "历史方案参考变化"
    ]
    assert any("两版方案都按 广东 口径，但目标年份不同" in detail for detail in matching_details)
    assert any("1 条参考位次变化" in detail for detail in matching_details)
    assert any("1 条最近最低位次变化" in detail for detail in matching_details)
    assert any("1 条最近最低分变化" in detail for detail in matching_details)
    assert any(
        row[1] == "同省跨年份差异"
        and row[2] == "1 条可比结果受同省跨年份口径变化影响"
        for row in risk_rows
    )


def test_recommendation_export_includes_stale_only_shift_copy(test_settings) -> None:
    ensure_runtime_directories(test_settings)
    export_path = export_recommendation_summary(
        test_settings,
        {
            "scheme_name": "当前方案",
            "student_name": "李四",
            "exam_name": "2027届一模",
            "province": "广东",
            "target_year": 2027,
            "score_input_label": "正式位次",
            "simulation_note": "默认推荐链路",
            "target_direction_summary": "未维护",
            "accepted_path_summary": "未维护",
            "compare_scheme": {
                "scheme_id": 8,
                "scheme_name": "历史方案 C",
                "generated_at": "2026-04-10T10:00:00",
                "province": "广东",
                "target_year": 2025,
            },
            "compare_rows": [
                {
                    "result_type": "steady",
                    "college_id": 11,
                    "college_name": "测试大学",
                    "major_id": 21,
                    "major_name": "软件工程",
                    "reference_rank": 28000,
                    "score_basis": "rank",
                    "snapshot_json": {
                        "reference_years": [2024],
                    },
                },
            ],
        },
        [
            {
                "result_type": "steady",
                "college_id": 11,
                "college_name": "测试大学",
                "major_id": 21,
                "major_name": "软件工程",
                "reference_rank": 28000,
                "score_basis": "rank",
                "risk_flags_json": [],
                "snapshot_json": {
                    "reference_years": [2024],
                },
            },
        ],
    )

    workbook = load_exported_workbook(export_path, test_settings.project_root)
    risk_sheet = workbook["风险概览"]
    risk_rows = [
        (
            risk_sheet.cell(row=index, column=1).value,
            risk_sheet.cell(row=index, column=2).value,
            risk_sheet.cell(row=index, column=3).value,
            risk_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, risk_sheet.max_row + 1)
    ]
    matching_details = [
        str(risk_sheet.cell(row=index, column=4).value)
        for index in range(2, risk_sheet.max_row + 1)
        if risk_sheet.cell(row=index, column=2).value == "历史方案参考变化"
    ]
    matching_summaries = [
        str(risk_sheet.cell(row=index, column=3).value)
        for index in range(2, risk_sheet.max_row + 1)
        if risk_sheet.cell(row=index, column=2).value == "历史方案参考变化"
    ]

    assert "1 条结果的“年份偏旧”状态发生变化" in matching_summaries
    assert any(
        row[1] == "同省跨年份差异"
        and row[2] == "1 条可比结果受同省跨年份口径变化影响"
        for row in risk_rows
    ) is True
    assert any("当前没有检测到明确的参考年份切换" in detail for detail in matching_details)
    assert any("其中 1 条只是“年份偏旧”状态切换" in detail for detail in matching_details)
    assert any("两版方案都按 广东 口径，但目标年份不同" in detail for detail in matching_details)
