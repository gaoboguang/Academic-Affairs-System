from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import text


def build_score_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    sheet.append(["2026届高一4月月考", "2026001", "张三", "1班", "语文", "118", "", ""])
    sheet.append(["2026届高一4月月考", "2026001", "张三", "1班", "数学", "125", "", ""])
    sheet.append(["2026届高一4月月考", "2026002", "李四", "1班", "语文", "110", "", ""])
    sheet.append(["2026届高一4月月考", "2026002", "李四", "1班", "数学", "120", "", ""])
    sheet.append(["2026届高一4月月考", "2026003", "王五", "2班", "语文", "98", "", ""])
    sheet.append(["2026届高一4月月考", "2026003", "王五", "2班", "数学", "96", "", ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_admission_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["年份", "省份", "批次", "院校", "专业", "最低分", "最低位次", "学生类别", "数据来源说明"])
    sheet.append([2025, "广东", "本科批", "岭南科技大学", "软件工程", 585, 28000, "普通生", "近年数据"])
    sheet.append([2024, "广东", "本科批", "湾区信息大学", "计算机科学与技术", 568, 32000, "普通生", "近年数据"])
    sheet.append([2025, "广东", "本科批", "南方应用大学", "信息管理", 545, 40000, "普通生", "近年数据"])
    sheet.append([2025, "广东", "艺术本科", "岭南艺术学院", "视觉传达设计", 510, 26000, "美术", "艺体数据"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_cross_province_admission_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["年份", "省份", "批次", "院校", "专业", "最低分", "最低位次", "学生类别", "数据来源说明"])
    sheet.append([2025, "山东", "本科批", "华北信息大学", "计算机科学与技术", 602, 12000, "普通生", "跨省对比样本"])
    sheet.append([2025, "河北", "本科批", "华北信息大学", "计算机科学与技术", 584, 22000, "普通生", "跨省对比样本"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_enrollment_plan_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        [
            "年份",
            "省份",
            "批次",
            "科类/模式",
            "院校",
            "院校代码",
            "专业组编号",
            "专业",
            "专业代码",
            "计划人数",
            "选科要求",
            "学费",
            "学制",
            "培养地点",
            "学生类别",
            "数据来源",
            "导入批次",
        ]
    )
    sheet.append(
        [
            2026,
            "广东",
            "本科批",
            "物理类",
            "岭南科技大学",
            "10561",
            "201",
            "软件工程",
            "080902",
            120,
            "物理+化学",
            "6850 元/年",
            "4年",
            "广州校区",
            "普通生",
            "招生计划简章",
            "2026-广东-本科",
        ]
    )
    sheet.append(
        [
            2026,
            "广东",
            "本科批",
            "物理类",
            "岭南科技大学",
            "10561",
            "202",
            "人工智能",
            "080717T",
            80,
            "物理+化学",
            "6850 元/年",
            "4年",
            "佛山校区",
            "普通生",
            "招生计划简章",
            "2026-广东-本科",
        ]
    )
    sheet.append(
        [
            2026,
            "广东",
            "本科批",
            "物理类",
            "粤东应用大学",
            "11888",
            "",
            "数据科学与大数据技术",
            "080910T",
            90,
            "物理",
            "6200 元/年",
            "4年",
            "汕头校区",
            "普通生",
            "招生计划表",
            "2026-广东-本科",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_large_enrollment_plan_workbook(row_count: int = 305) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        [
            "年份",
            "省份",
            "批次",
            "科类/模式",
            "院校",
            "院校代码",
            "专业组编号",
            "专业",
            "专业代码",
            "计划人数",
            "选科要求",
            "学费",
            "学制",
            "培养地点",
            "学生类别",
            "数据来源",
            "导入批次",
        ]
    )
    for index in range(row_count):
        sheet.append(
            [
                2026,
                "广东",
                "本科批",
                "物理类",
                "分页性能大学",
                "P001",
                f"P{index:03d}",
                f"分页测试专业{index:03d}",
                f"PX{index:03d}",
                1,
                "物理+化学",
                "5000 元/年",
                "4年",
                "广州校区",
                "普通生",
                "分页压力测试",
                "2026-广东-分页性能",
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_large_admission_workbook(row_count: int = 305) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["年份", "省份", "批次", "院校", "专业", "最低分", "最低位次", "学生类别", "数据来源说明"])
    for index in range(row_count):
        sheet.append(
            [
                2025,
                "广东",
                "本科批",
                "分页性能大学",
                f"分页测试专业{index:03d}",
                560,
                32000 + index,
                "普通生",
                "分页压力测试",
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_exam_with_scores(client) -> int:
    exam_response = client.post(
        "/api/exams",
        json={
            "name": "2026届高一4月月考",
            "exam_type": "月考",
            "exam_date": "2026-04-10",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]

    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            },
            {
                "subject_id": 2,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 2,
                "is_active": True,
            },
        ],
    )
    assert subject_response.status_code == 200

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                "scores.xlsx",
                build_score_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    return exam_id


def test_bootstrap_province_volunteer_rules(client) -> None:
    target_year = date.today().year + 1

    bootstrap_response = client.post(f"/api/province-volunteer-rules/bootstrap?year={target_year}")
    assert bootstrap_response.status_code == 200
    bootstrap_payload = bootstrap_response.json()
    assert bootstrap_payload["year"] == target_year
    assert bootstrap_payload["total_count"] == 81
    assert bootstrap_payload["created_count"] == 81
    assert bootstrap_payload["skipped_count"] == 0

    list_response = client.get(f"/api/province-volunteer-rules?year={target_year}")
    assert list_response.status_code == 200
    rules_payload = list_response.json()
    assert len(rules_payload) == 81
    assert any(
        item["province"] == "北京"
        and item["exam_mode"] == "3+3"
        and item["subject_requirement_mode"] == "unified_subject_requirement"
        for item in rules_payload
    )
    assert any(
        item["province"] == "上海"
        and item["total_score"] == 660
        for item in rules_payload
    )
    assert any(
        item["province"] == "海南"
        and item["total_score"] == 900
        for item in rules_payload
    )
    assert any(
        item["province"] == "广东"
        and item["exam_mode"] == "物理类"
        and item["volunteer_unit_type"] == "院校专业组"
        and item["max_major_per_unit"] == 6
        for item in rules_payload
    )
    assert any(
        item["province"] == "西藏"
        and item["exam_mode"] == "文科"
        and item["parallel_rule_mode"] == "ordered_sequential"
        and item["is_parallel"] is False
        for item in rules_payload
    )
    assert any(
        item["province"] == "山东"
        and item["candidate_type"] == "spring_exam"
        and item["exam_mode"] == "春季高考"
        and item["batch"] == "春季高考本科批"
        and item["volunteer_limit"] == 96
        for item in rules_payload
    )
    assert any(
        item["province"] == "山东"
        and item["candidate_type"] == "independent_recruitment"
        and item["batch"] == "专科批"
        and item["is_parallel"] is False
        for item in rules_payload
    )
    assert any(
        item["province"] == "山东"
        and item["candidate_type"] == "comprehensive_evaluation"
        and item["batch"] == "专科批"
        and item["is_parallel"] is False
        for item in rules_payload
    )

    repeat_response = client.post(f"/api/province-volunteer-rules/bootstrap?year={target_year}")
    assert repeat_response.status_code == 200
    repeat_payload = repeat_response.json()
    assert repeat_payload["created_count"] == 0
    assert repeat_payload["skipped_count"] == 81


def test_bootstrap_province_score_transform_rules(client) -> None:
    target_year = date.today().year + 1

    bootstrap_response = client.post(f"/api/province-score-transform-rules/bootstrap?year={target_year}")
    assert bootstrap_response.status_code == 200
    bootstrap_payload = bootstrap_response.json()
    assert bootstrap_payload["year"] == target_year
    assert bootstrap_payload["total_count"] == 438
    assert bootstrap_payload["created_count"] == 438
    assert bootstrap_payload["skipped_count"] == 0

    list_response = client.get(f"/api/province-score-transform-rules?year={target_year}")
    assert list_response.status_code == 200
    rules_payload = list_response.json()
    assert len(rules_payload) == 438
    assert any(
        item["province"] == "北京"
        and item["exam_mode"] == "3+3"
        and item["subject_name"] == "物理"
        and item["score_mode"] == "grade_assigned"
        for item in rules_payload
    )
    assert any(
        item["province"] == "广东"
        and item["exam_mode"] == "物理类"
        and item["subject_name"] == "物理"
        and item["score_mode"] == "raw"
        for item in rules_payload
    )
    assert any(
        item["province"] == "新疆"
        and item["exam_mode"] == "理科"
        and item["subject_name"] == "理综"
        and item["score_mode"] == "raw"
        for item in rules_payload
    )

    repeat_response = client.post(f"/api/province-score-transform-rules/bootstrap?year={target_year}")
    assert repeat_response.status_code == 200
    repeat_payload = repeat_response.json()
    assert repeat_payload["created_count"] == 0
    assert repeat_payload["skipped_count"] == 438


def test_bootstrap_subject_requirement_dicts(client) -> None:
    target_year = date.today().year + 1

    bootstrap_response = client.post(f"/api/subject-requirement-dicts/bootstrap?year={target_year}")
    assert bootstrap_response.status_code == 200
    bootstrap_payload = bootstrap_response.json()
    assert bootstrap_payload["year"] == target_year
    assert bootstrap_payload["total_count"] == 292
    assert bootstrap_payload["created_count"] == 292
    assert bootstrap_payload["skipped_count"] == 0

    list_response = client.get(f"/api/subject-requirement-dicts?year={target_year}")
    assert list_response.status_code == 200
    rules_payload = list_response.json()
    assert len(rules_payload) == 292
    assert any(
        item["province"] == "北京"
        and item["exam_mode"] == "3+3"
        and item["requirement_code"] == "UNLIMITED"
        and item["match_mode"] == "no_requirement"
        for item in rules_payload
    )
    assert any(
        item["province"] == "广东"
        and item["exam_mode"] == "物理类"
        and item["requirement_code"] == "PHYSICS_CHEMISTRY"
        and item["normalized_subjects_json"] == ["化学", "物理"]
        for item in rules_payload
    )
    assert any(
        item["province"] == "西藏"
        and item["exam_mode"] == "文科"
        and item["requirement_code"] == "LIBERAL_ARTS"
        and item["match_mode"] == "track_only"
        for item in rules_payload
    )

    repeat_response = client.post(f"/api/subject-requirement-dicts/bootstrap?year={target_year}")
    assert repeat_response.status_code == 200
    repeat_payload = repeat_response.json()
    assert repeat_payload["created_count"] == 0
    assert repeat_payload["skipped_count"] == 292


def test_bootstrap_employment_directions_and_major_mappings(client) -> None:
    majors_payload = client.get("/api/majors?keyword=软件工程").json()
    if not any(item["name"] == "软件工程" for item in majors_payload):
        client.post(
            "/api/majors",
            json={
                "name": "软件工程",
                "major_code": "080902",
                "category": "工学",
                "direction": "学士",
                "career_path": None,
                "is_art_related": False,
                "note": None,
                "is_active": True,
            },
        )
    majors_payload = client.get("/api/majors?keyword=人工智能").json()
    if not any(item["name"] == "人工智能" for item in majors_payload):
        client.post(
            "/api/majors",
            json={
                "name": "人工智能",
                "major_code": "080717T",
                "category": "工学",
                "direction": "学士",
                "career_path": None,
                "is_art_related": False,
                "note": None,
                "is_active": True,
            },
        )
    majors_payload = client.get("/api/majors?keyword=视觉传达设计").json()
    if not any(item["name"] == "视觉传达设计" for item in majors_payload):
        client.post(
            "/api/majors",
            json={
                "name": "视觉传达设计",
                "major_code": "130502",
                "category": "艺术学",
                "direction": "学士",
                "career_path": None,
                "is_art_related": True,
                "note": None,
                "is_active": True,
            },
        )

    bootstrap_direction_response = client.post("/api/employment-directions/bootstrap")
    assert bootstrap_direction_response.status_code == 200
    bootstrap_direction_payload = bootstrap_direction_response.json()
    assert bootstrap_direction_payload["total_count"] >= 10
    assert bootstrap_direction_payload["created_count"] >= 10

    direction_response = client.get("/api/employment-directions?keyword=软件平台工程")
    assert direction_response.status_code == 200
    direction_payload = direction_response.json()
    assert len(direction_payload) == 1
    assert direction_payload[0]["category"] == "技术研发类"

    bootstrap_mapping_response = client.post("/api/major-employment-maps/bootstrap")
    assert bootstrap_mapping_response.status_code == 200
    bootstrap_mapping_payload = bootstrap_mapping_response.json()
    assert bootstrap_mapping_payload["major_total_count"] >= 3
    assert bootstrap_mapping_payload["matched_major_count"] >= 3
    assert bootstrap_mapping_payload["created_count"] >= 3

    software_mapping_response = client.get("/api/major-employment-maps?keyword=软件工程")
    assert software_mapping_response.status_code == 200
    software_mapping_payload = software_mapping_response.json()
    assert any(item["direction_name"] == "软件平台工程" for item in software_mapping_payload)

    ai_mapping_response = client.get("/api/major-employment-maps?keyword=人工智能")
    assert ai_mapping_response.status_code == 200
    ai_mapping_payload = ai_mapping_response.json()
    assert any(item["direction_name"] == "数据智能分析" for item in ai_mapping_payload)

    art_mapping_response = client.get("/api/major-employment-maps?keyword=视觉传达设计")
    assert art_mapping_response.status_code == 200
    art_mapping_payload = art_mapping_response.json()
    assert any(item["direction_name"] == "传媒内容与视觉设计" for item in art_mapping_payload)


def test_candidate_filters_match_exact_special_student_types(app) -> None:
    from app.models import AdmissionRecord, College, EnrollmentPlan, Major, Student
    from app.repositories.recommendations import list_admission_candidates, list_enrollment_plan_candidates
    from app.services._recommendations_candidates import detect_student_type, get_admission_reference_candidates
    from app.services._recommendations_shared import RecommendationSettingsState

    with app.state.db.session_scope() as session:
        college = College(name="特殊类型测试大学", province="山东", supports_art=True, is_active=True)
        major = Major(name="特殊类型测试专业", is_art_related=True, is_active=True)
        session.add_all([college, major])
        session.flush()

        session.add_all(
            [
                AdmissionRecord(
                    year=2025,
                    province="山东",
                    batch="常规批",
                    college_id=college.id,
                    major_id=major.id,
                    student_type="general",
                    art_track="",
                    min_rank=1000,
                    is_active=True,
                ),
                AdmissionRecord(
                    year=2025,
                    province="山东",
                    batch="艺术类本科批统考",
                    college_id=college.id,
                    major_id=major.id,
                    student_type="art",
                    art_track="",
                    min_rank=2000,
                    is_active=True,
                ),
            ]
        )
        session.add_all(
            [
                EnrollmentPlan(
                    year=2026,
                    province="山东",
                    batch="常规批",
                    exam_mode="3+3",
                    college_id=college.id,
                    major_id=major.id,
                    major_group_code="A",
                    major_name_snapshot="特殊类型测试专业",
                    student_type="general",
                    plan_count=10,
                    is_active=True,
                ),
                EnrollmentPlan(
                    year=2026,
                    province="山东",
                    batch="艺术类本科批统考",
                    exam_mode="3+3",
                    college_id=college.id,
                    major_id=major.id,
                    major_group_code="ART",
                    major_name_snapshot="特殊类型测试专业",
                    student_type="art",
                    plan_count=8,
                    is_active=True,
                ),
                EnrollmentPlan(
                    year=2026,
                    province="山东",
                    batch="春季高考本科批",
                    exam_mode="春季高考",
                    college_id=college.id,
                    major_id=major.id,
                    major_group_code="SPRING",
                    major_name_snapshot="特殊类型测试专业",
                    student_type="spring_exam",
                    plan_count=12,
                    is_active=True,
                ),
            ]
        )
        sports_student = Student(name="体育生张三", student_no="TS001", gender="男", student_type="art", art_track="sports", is_active=True)
        spring_student = Student(name="春考李四", student_no="TS002", gender="女", student_type="spring_exam", art_track=None, is_active=True)
        repeat_student = Student(name="复读王五", student_no="TS003", gender="男", student_type="repeat", art_track=None, is_active=True)
        session.add_all([sports_student, spring_student, repeat_student])
        session.flush()

        assert detect_student_type(sports_student) == "sports"
        assert detect_student_type(spring_student) == "spring_exam"
        assert detect_student_type(repeat_student) == "repeat"

        general_admissions = list_admission_candidates(session, province="山东", student_type="general")
        art_admissions = list_admission_candidates(session, province="山东", student_type="art")
        spring_plans = list_enrollment_plan_candidates(session, year=2026, province="山东", student_type="spring_exam")
        art_plans = list_enrollment_plan_candidates(session, year=2026, province="山东", student_type="art")
        repeat_plans = list_enrollment_plan_candidates(session, year=2026, province="山东", student_type="repeat")
        spring_admission_reference, used_general_reference_fallback = get_admission_reference_candidates(
            session,
            province="山东",
            student=spring_student,
            student_type="spring_exam",
            target_regions=[],
            school_levels=[],
            major_keyword=None,
            subject_combination=None,
            settings=RecommendationSettingsState(
                safe_ratio_max=0.85,
                steady_ratio_max=1.0,
                rush_ratio_max=1.15,
                whitelist_college_ids=[],
                blacklist_college_ids=[],
            ),
        )

    assert {item.student_type for item in general_admissions} == {"general"}
    assert {item.student_type for item in art_admissions} == {"art"}
    assert {item.student_type for item in spring_plans} == {"spring_exam"}
    assert {item.student_type for item in art_plans} == {"art"}
    assert {item.student_type for item in repeat_plans} == {"general"}
    assert used_general_reference_fallback is True
    assert {item.student_type for item in spring_admission_reference} == {"general"}


def test_art_candidate_filter_uses_normalized_request_track(app) -> None:
    from app.models import AdmissionRecord, College, Major, Student
    from app.services._recommendations_candidates import get_admission_reference_candidates
    from app.services._recommendations_shared import RecommendationSettingsState

    with app.state.db.session_scope() as session:
        college = College(name="艺术类别请求测试大学", province="山东", supports_art=True, is_active=True)
        music_major = Major(name="音乐类请求测试专业", is_art_related=True, is_active=True)
        fine_art_major = Major(name="美术类请求测试专业", is_art_related=True, is_active=True)
        session.add_all([college, music_major, fine_art_major])
        session.flush()
        session.add_all(
            [
                AdmissionRecord(
                    year=2025,
                    province="山东",
                    batch="艺术类本科批统考",
                    college_id=college.id,
                    major_id=music_major.id,
                    student_type="art",
                    art_track="music",
                    min_score=500,
                    is_active=True,
                ),
                AdmissionRecord(
                    year=2025,
                    province="山东",
                    batch="艺术类本科批统考",
                    college_id=college.id,
                    major_id=fine_art_major.id,
                    student_type="art",
                    art_track="fine_art_design",
                    min_score=500,
                    is_active=True,
                ),
            ]
        )
        student = Student(
            name="艺术类别请求测试生",
            student_no="ARTREQ001",
            gender="女",
            student_type="art",
            art_track="fine_art_design",
            is_active=True,
        )
        session.add(student)
        session.flush()

        items, _ = get_admission_reference_candidates(
            session,
            province="山东",
            student=student,
            student_type="art",
            art_track="music",
            target_regions=[],
            school_levels=[],
            major_keyword=None,
            subject_combination=None,
            settings=RecommendationSettingsState(
                safe_ratio_max=0.85,
                steady_ratio_max=1.0,
                rush_ratio_max=1.15,
                whitelist_college_ids=[],
                blacklist_college_ids=[],
            ),
        )

    assert {item.major.name for item in items if item.major} == {"音乐类请求测试专业"}


def update_major_profile(client, major_name: str, *, direction: str, career_path: str, note: str) -> None:
    major_response = client.get(f"/api/majors?keyword={major_name}")
    assert major_response.status_code == 200
    majors_payload = major_response.json()
    target_major = next(item for item in majors_payload if item["name"] == major_name)

    update_response = client.put(
        f"/api/majors/{target_major['id']}",
        json={
            "name": target_major["name"],
            "major_code": target_major["major_code"],
            "category": target_major["category"],
            "direction": direction,
            "career_path": career_path,
            "is_art_related": target_major["is_art_related"],
            "note": note,
            "is_active": target_major["is_active"],
        },
    )
    assert update_response.status_code == 200


def test_art_score_line_fallback_supports_preview_and_generation(client, app) -> None:
    from app.models import College, EnrollmentPlan, Major, Student

    exam_id = create_exam_with_scores(client)

    with app.state.db.session_scope() as session:
        student = session.get(Student, 1)
        assert student is not None
        student.student_type = "art"
        student.art_track = "music"

        college = College(
            name="山东艺术测试大学",
            college_code="SD001",
            province="山东",
            city="济南",
            school_type="本科",
            school_level_tags_json=["省属"],
            supports_art=True,
            is_active=True,
        )
        major = Major(
            name="音乐表演测试专业",
            major_code="130201",
            category="艺术学",
            is_art_related=True,
            is_active=True,
        )
        session.add_all([college, major])
        session.flush()

        session.add(
            EnrollmentPlan(
                year=2026,
                province="山东",
                batch="本科批",
                exam_mode="3+3",
                college_id=college.id,
                major_id=major.id,
                college_code_snapshot="SD001",
                major_group_code="A01",
                major_name_snapshot=major.name,
                major_code_snapshot="130201",
                plan_count=12,
                subject_requirement=None,
                tuition_fee="9800 元/年",
                schooling_years="4年",
                training_location="主校区",
                student_type="art",
                source_note="山东艺术计划",
                import_batch_name="sd-art-test",
                is_active=True,
            )
        )
        session.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS gaokao_score_line (
                    id INTEGER PRIMARY KEY,
                    province TEXT NOT NULL,
                    year INTEGER NOT NULL,
                    candidate_type TEXT,
                    batch_name TEXT,
                    line_type TEXT NOT NULL,
                    score NUMERIC(8, 2) NOT NULL,
                    remark TEXT,
                    created_at TEXT,
                    updated_at TEXT,
                    source_title TEXT,
                    source_url TEXT
                )
                """
            )
        )
        session.execute(
            text(
                """
                INSERT INTO gaokao_score_line (
                    province, year, candidate_type, batch_name, line_type, score, remark, created_at, updated_at, source_title, source_url
                ) VALUES (
                    'sd', 2025, '艺术类', '本科', 'undergrad_culture_art_music_calligraphy', 330,
                    '艺术类本科文化控制线。', '2026-04-23 10:00:00', '2026-04-23 10:00:00', '山东省教育招生考试院', 'https://example.com/sd-art-line'
                )
                """
            )
        )
        session.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS gaokao_batch_dict (
                    province_scope TEXT NOT NULL,
                    batch_code TEXT NOT NULL,
                    batch_name TEXT NOT NULL,
                    pathway_code TEXT,
                    year_from INTEGER,
                    year_to INTEGER,
                    sort_order INTEGER,
                    note TEXT,
                    id INTEGER PRIMARY KEY,
                    created_at TEXT,
                    updated_at TEXT,
                    status TEXT
                )
                """
            )
        )
        session.execute(
            text(
                """
                INSERT INTO gaokao_batch_dict (
                    province_scope, batch_code, batch_name, pathway_code, sort_order, note, created_at, updated_at, status
                ) VALUES (
                    'sd', 'sd:艺术类:本科批', '本科批', 'art', 1, '艺术类本科批按艺术专业类别分别投档。', '2026-04-24 10:00:00', '2026-04-24 10:00:00', 'active'
                )
                """
            )
        )
        session.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS gaokao_policy_reference (
                    province TEXT,
                    year INTEGER,
                    policy_type TEXT NOT NULL,
                    title TEXT NOT NULL,
                    url TEXT,
                    local_path TEXT,
                    summary TEXT,
                    source_level TEXT,
                    version_id INTEGER,
                    import_batch_id INTEGER,
                    published_at TEXT,
                    id INTEGER PRIMARY KEY,
                    created_at TEXT,
                    updated_at TEXT,
                    status TEXT NOT NULL
                )
                """
            )
        )
        session.execute(
            text(
                """
                INSERT INTO gaokao_policy_reference (
                    province, year, policy_type, title, url, summary, created_at, updated_at, status
                ) VALUES (
                    'sd', 2025, 'province_rule', '山东省2025年普通高校考试招生工作实施办法', 'https://example.com/sd-policy-2025',
                    '艺术类本科文化线按专业类别分别执行，正式填报前需核对当年公告。', '2026-04-24 10:00:00', '2026-04-24 10:00:00', 'active'
                )
                """
            )
        )

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "山东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "3+3",
            "candidate_type": "art",
            "score_input_mode": "estimated_score",
            "culture_score": 360,
            "professional_score": 240,
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["candidate_count"] == 1
    candidate = preview_payload["candidates"][0]
    assert candidate["result_type"] == "watch"
    assert candidate["reference_scope"] == "score_line"
    assert candidate["reference_years_json"] == [2025]
    assert candidate["reference_record_count"] == 0
    assert candidate["latest_min_score"] == 330
    assert candidate["score_basis"] == "culture_score"
    assert preview_payload["culture_score"] == 360
    assert preview_payload["professional_score"] == 240
    assert preview_payload["art_comprehensive_score"] == 480
    assert candidate["fallback_priority_label"] in {"重点比较", "优先核看"}
    assert any("省级控制线" in note for note in candidate["fallback_priority_notes_json"])
    assert candidate["fallback_category_label"] == "艺术音乐类"
    assert any("核对艺术统考类别" in note for note in candidate["fallback_review_notes_json"])
    assert any("高出省控线 30" in note for note in candidate["fallback_priority_notes_json"])
    assert "省控线参考" in candidate["match_tags_json"]
    assert "score_line_reference_only" in (candidate["risk_flags_json"] or [])
    assert "cross_year_score_line_reference" in (candidate["risk_flags_json"] or [])
    assert any("本科文化控制线" in note for note in candidate["match_notes_json"])
    assert any("批次词典：" in note for note in candidate["match_notes_json"])
    assert any("政策摘要：" in note for note in candidate["match_notes_json"])

    generate_response = client.post(
        "/api/recommendations/generate",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "山东",
            "target_year": 2026,
            "score_input_mode": "estimated_score",
            "culture_score": 360,
            "professional_score": 240,
        },
    )
    assert generate_response.status_code == 200
    generate_payload = generate_response.json()
    assert generate_payload["result_count"] == 1
    assert len(generate_payload["challenge"]) == 0
    assert len(generate_payload["steady"]) == 0
    assert len(generate_payload["safe"]) == 0
    assert len(generate_payload["watch"]) == 1
    result = generate_payload["watch"][0]
    assert result["score_basis"] == "culture_score"
    assert result["fallback_priority_label"] in {"重点比较", "优先核看"}
    assert any("省级控制线" in note for note in result["fallback_priority_notes_json"])
    assert result["fallback_category_label"] == "艺术音乐类"
    assert any("核对艺术统考类别" in note for note in result["fallback_review_notes_json"])
    assert "score_line_reference_only" in (result["risk_flags_json"] or [])
    assert "cross_year_score_line_reference" in (result["risk_flags_json"] or [])
    assert result["snapshot_json"]["score_line_reference"] is True
    assert result["snapshot_json"]["score_line_year"] == 2025
    assert result["snapshot_json"]["score_line_line_type"] == "undergrad_culture_art_music_calligraphy"
    assert result["snapshot_json"]["reference_scope"] == "score_line"
    assert result["snapshot_json"]["batch_dict_note"] == "艺术类本科批按艺术专业类别分别投档。"
    assert result["snapshot_json"]["province_policy_title"] == "山东省2025年普通高校考试招生工作实施办法"


def test_plan_only_fallback_supports_special_workbench_preview_and_generation(client, app) -> None:
    from app.models import College, EnrollmentPlan, Major, Student

    exam_id = create_exam_with_scores(client)

    with app.state.db.session_scope() as session:
        student = session.get(Student, 1)
        assert student is not None
        student.student_type = "comprehensive_evaluation"
        student.art_track = None

        college = College(
            name="山东综评测试大学",
            college_code="SD002",
            province="山东",
            city="青岛",
            school_type="专科",
            school_level_tags_json=["公办"],
            supports_art=False,
            is_active=True,
        )
        second_college = College(
            name="山东综评补充大学",
            college_code="SD003",
            province="山东",
            city="烟台",
            school_type="专科",
            school_level_tags_json=["民办"],
            supports_art=False,
            is_active=True,
        )
        major = Major(
            name="智能制造测试专业",
            major_code="460101",
            category="装备制造",
            is_art_related=False,
            is_active=True,
        )
        second_major = Major(
            name="市场营销综评测试专业",
            major_code="530605",
            category="财经商贸",
            is_art_related=False,
            is_active=True,
        )
        session.add_all([college, second_college, major, second_major])
        session.flush()

        session.add_all([
            EnrollmentPlan(
                year=2026,
                province="山东",
                batch="专科批",
                exam_mode="3+3",
                college_id=college.id,
                major_id=major.id,
                college_code_snapshot="SD002",
                major_group_code="Z01",
                major_name_snapshot=major.name,
                major_code_snapshot="460101",
                plan_count=20,
                subject_requirement=None,
                tuition_fee="6800 元/年",
                schooling_years="3年",
                training_location="青岛校区",
                student_type="comprehensive_evaluation",
                source_note="山东综评计划",
                import_batch_name="sd-ce-test",
                is_active=True,
            ),
            EnrollmentPlan(
                year=2026,
                province="山东",
                batch="专科批",
                exam_mode="3+3",
                college_id=second_college.id,
                major_id=second_major.id,
                college_code_snapshot="SD003",
                major_group_code="Z02",
                major_name_snapshot=second_major.name,
                major_code_snapshot="530605",
                plan_count=5,
                subject_requirement=None,
                tuition_fee="8800 元/年",
                schooling_years="3年",
                training_location="烟台校区",
                student_type="comprehensive_evaluation",
                source_note="山东综评补充计划",
                import_batch_name="sd-ce-test",
                is_active=True,
            ),
        ])
        session.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS gaokao_batch_dict (
                    province_scope TEXT NOT NULL,
                    batch_code TEXT NOT NULL,
                    batch_name TEXT NOT NULL,
                    pathway_code TEXT,
                    year_from INTEGER,
                    year_to INTEGER,
                    sort_order INTEGER,
                    note TEXT,
                    id INTEGER PRIMARY KEY,
                    created_at TEXT,
                    updated_at TEXT,
                    status TEXT
                )
                """
            )
        )
        session.execute(
            text(
                """
                INSERT INTO gaokao_batch_dict (
                    province_scope, batch_code, batch_name, pathway_code, sort_order, note, created_at, updated_at, status
                ) VALUES (
                    'sd', 'sd:综合评价招生:专科批', '专科批', '', 1, '综合评价招生专科批需结合学校综合测试与高考成绩。', '2026-04-24 10:00:00', '2026-04-24 10:00:00', 'active'
                )
                """
            )
        )
        session.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS gaokao_policy_reference (
                    province TEXT,
                    year INTEGER,
                    policy_type TEXT NOT NULL,
                    title TEXT NOT NULL,
                    url TEXT,
                    local_path TEXT,
                    summary TEXT,
                    source_level TEXT,
                    version_id INTEGER,
                    import_batch_id INTEGER,
                    published_at TEXT,
                    id INTEGER PRIMARY KEY,
                    created_at TEXT,
                    updated_at TEXT,
                    status TEXT NOT NULL
                )
                """
            )
        )
        session.execute(
            text(
                """
                INSERT INTO gaokao_policy_reference (
                    province, year, policy_type, title, url, summary, created_at, updated_at, status
                ) VALUES (
                    'sd', 2025, 'province_rule', '山东省2025年普通高校考试招生工作实施办法', 'https://example.com/sd-policy-2025',
                    '综合评价招生需结合高校测试和高考成绩综合认定，正式填报前需核对学校章程。', '2026-04-24 10:00:00', '2026-04-24 10:00:00', 'active'
                )
                """
            )
        )

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "山东",
            "target_year": 2026,
            "batch": "专科批",
            "exam_mode": "3+3",
            "candidate_type": "comprehensive_evaluation",
            "score_input_mode": "actual_rank",
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["candidate_count"] == 0
    assert preview_payload["candidates"] == []

    generate_response = client.post(
        "/api/recommendations/generate",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "山东",
            "target_year": 2026,
            "score_input_mode": "actual_rank",
        },
    )
    assert generate_response.status_code == 404


def test_special_type_rule_dictionary_bootstrap_and_list(client) -> None:
    bootstrap_response = client.post("/api/special-type-rules/bootstrap?year=2026")
    assert bootstrap_response.status_code == 200
    bootstrap_payload = bootstrap_response.json()
    assert bootstrap_payload["total_count"] >= 30

    list_response = client.get("/api/special-type-rules?province=山东&year=2026&student_type=spring_exam")
    assert list_response.status_code == 200
    payload = list_response.json()
    labels = {item["category_label"] for item in payload}
    assert "春考信息技术类" in labels
    assert "春考设备制造类" in labels
    info_rule = next(item for item in payload if item["category_code"] == "spring_it")
    assert "计算机" in info_rule["match_keywords_json"]
    assert any("技能考试类别" in note for note in info_rule["review_notes_json"])


def test_recommendation_large_tables_support_paginated_queries(client) -> None:
    for index, province in enumerate(["广东", "广东", "山东"], start=1):
        response = client.post(
            "/api/colleges",
            json={
                "name": f"分页测试大学{index}",
                "college_code": None,
                "province": province,
                "city": None,
                "school_type": None,
                "school_level_tags_json": [],
                "intro": None,
                "website": None,
                "supports_art": index == 3,
                "note": None,
                "alias_names": [],
                "is_active": True,
            },
        )
        assert response.status_code == 200

    college_page = client.get("/api/colleges/page?keyword=分页测试大学&province=广东&page=1&page_size=1")
    assert college_page.status_code == 200
    college_payload = college_page.json()
    assert college_payload["total"] == 2
    assert college_payload["page"] == 1
    assert college_payload["page_size"] == 1
    assert len(college_payload["items"]) == 1
    assert college_payload["items"][0]["province"] == "广东"

    for major_name in ["分页测试专业A", "分页测试专业B", "分页测试专业C"]:
        response = client.post(
            "/api/majors",
            json={
                "name": major_name,
                "major_code": None,
                "category": "测试门类",
                "direction": None,
                "career_path": None,
                "is_art_related": False,
                "note": None,
                "is_active": True,
            },
        )
        assert response.status_code == 200

    major_page = client.get("/api/majors/page?keyword=分页测试专业&page=2&page_size=2")
    assert major_page.status_code == 200
    major_payload = major_page.json()
    assert major_payload["total"] == 3
    assert major_payload["page"] == 2
    assert major_payload["page_size"] == 2
    assert len(major_payload["items"]) == 1

    direction_response = client.post(
        "/api/employment-directions",
        json={
            "name": "分页测试就业方向",
            "category": "技术研发类",
            "alias_names_json": [],
            "description": None,
            "common_job_types_json": [],
            "common_industries_json": [],
            "prefers_postgraduate": False,
            "requires_certificate": False,
            "requires_long_cycle": False,
            "supports_art": False,
            "risk_note": None,
            "source_note": None,
            "is_active": True,
        },
    )
    assert direction_response.status_code == 200
    direction_id = direction_response.json()["id"]
    major_list = client.get("/api/majors?keyword=分页测试专业")
    assert major_list.status_code == 200
    for major_item in major_list.json():
        mapping_response = client.post(
            "/api/major-employment-maps",
            json={
                "major_id": major_item["id"],
                "direction_id": direction_id,
                "strength": "high",
                "recommendation_note": "分页查询测试",
                "requires_postgraduate": False,
                "requires_certificate": False,
                "supported_student_types_json": [],
                "supports_art": False,
                "note": None,
                "is_active": True,
            },
        )
        assert mapping_response.status_code == 200

    mapping_page = client.get("/api/major-employment-maps/page?keyword=分页测试&page=2&page_size=2")
    assert mapping_page.status_code == 200
    mapping_payload = mapping_page.json()
    assert mapping_payload["total"] == 3
    assert mapping_payload["page"] == 2
    assert len(mapping_payload["items"]) == 1
    assert mapping_payload["items"][0]["direction_name"] == "分页测试就业方向"

    import_response = client.post(
        "/api/admissions/import",
        files={
            "file": (
                "admissions.xlsx",
                build_admission_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200

    admissions_page = client.get("/api/admissions/page?province=广东&page=1&page_size=2")
    assert admissions_page.status_code == 200
    admissions_payload = admissions_page.json()
    assert admissions_payload["total"] == 4
    assert len(admissions_payload["items"]) == 2
    assert {item["province"] for item in admissions_payload["items"]} == {"广东"}

    plan_import_response = client.post(
        "/api/enrollment-plans/import",
        files={
            "file": (
                "enrollment-plans.xlsx",
                build_enrollment_plan_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert plan_import_response.status_code == 200

    plans_page = client.get("/api/enrollment-plans/page?province=广东&year=2026&page=1&page_size=2")
    assert plans_page.status_code == 200
    plans_payload = plans_page.json()
    assert plans_payload["total"] == 3
    assert len(plans_payload["items"]) == 2
    assert {item["province"] for item in plans_payload["items"]} == {"广东"}

    too_large_page_size = client.get("/api/enrollment-plans/page?page_size=500")
    assert too_large_page_size.status_code == 422
    too_large_college_page_size = client.get("/api/colleges/page?page_size=500")
    assert too_large_college_page_size.status_code == 422
    too_large_mapping_page_size = client.get("/api/major-employment-maps/page?page_size=500")
    assert too_large_mapping_page_size.status_code == 422


def test_admission_import_and_recommendation_generation(client, app) -> None:
    exam_id = create_exam_with_scores(client)

    import_response = client.post(
        "/api/admissions/import",
        files={
            "file": (
                "admissions.xlsx",
                build_admission_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["success_rows"] == 4
    assert import_payload["created_college_count"] == 4
    assert import_payload["created_major_count"] == 4

    admissions_response = client.get("/api/admissions?province=广东")
    assert admissions_response.status_code == 200
    assert len(admissions_response.json()) == 4

    plan_import_response = client.post(
        "/api/enrollment-plans/import",
        files={
            "file": (
                "enrollment-plans.xlsx",
                build_enrollment_plan_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert plan_import_response.status_code == 200
    plan_import_payload = plan_import_response.json()
    assert plan_import_payload["success_rows"] == 3
    assert plan_import_payload["created_college_count"] == 1
    assert plan_import_payload["created_major_count"] == 2

    plans_response = client.get("/api/enrollment-plans?province=广东&year=2026")
    assert plans_response.status_code == 200
    plans_payload = plans_response.json()
    assert len(plans_payload) == 3
    assert plans_payload[0]["batch"] == "本科批"
    assert {item["college_name"] for item in plans_payload} == {"岭南科技大学", "粤东应用大学"}
    lingnan_plan = next(item for item in plans_payload if item["college_name"] == "岭南科技大学")

    with app.state.db.session_scope() as session:
        session.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS gaokao_college_chapter_rule (
                    id INTEGER PRIMARY KEY,
                    province TEXT NOT NULL,
                    year INTEGER NOT NULL,
                    college_id INTEGER,
                    chapter_url TEXT,
                    retrieval_status TEXT,
                    review_status TEXT,
                    campus_note TEXT,
                    other_risk_note TEXT,
                    language_requirement TEXT,
                    single_subject_requirement TEXT,
                    gender_requirement TEXT,
                    height_requirement TEXT,
                    vision_requirement TEXT,
                    color_vision_requirement TEXT,
                    physical_exam_requirement TEXT,
                    updated_at TEXT
                )
                """
            )
        )
        session.execute(
            text(
                """
                INSERT INTO gaokao_college_chapter_rule (
                    province, year, college_id, chapter_url, retrieval_status, review_status, campus_note, other_risk_note,
                    language_requirement, single_subject_requirement, gender_requirement, height_requirement,
                    vision_requirement, color_vision_requirement, physical_exam_requirement, updated_at
                ) VALUES (
                    '广东', 2026, :college_id, 'https://chapter.example/lingnan', 'chapter_url_filled_official_result',
                    'confirmed_manual_review', '广州校区', '部分专业需关注校区安排。',
                    '英语语种考生优先。', '数学单科不低于 100 分。', '建议男生。', '身高不低于 168cm。',
                    '裸眼视力不低于 4.8。', '色弱慎报。', '需满足体检结论要求。', '2026-04-23 15:30:00'
                )
                """
            ),
            {"college_id": lingnan_plan["college_id"]},
        )

    create_direction_response = client.post(
        "/api/employment-directions",
        json={
            "name": "企业软件开发",
            "category": "技术研发类",
            "alias_names_json": ["软件研发", "平台开发"],
            "description": "面向企业级软件、平台和信息系统研发岗位。",
            "common_job_types_json": ["后端开发", "平台研发"],
            "common_industries_json": ["互联网", "企业信息化"],
            "prefers_postgraduate": False,
            "requires_certificate": True,
            "requires_long_cycle": False,
            "supports_art": False,
            "risk_note": "岗位变化快，需要持续学习。",
            "source_note": "测试方向",
            "is_active": True,
        },
    )
    assert create_direction_response.status_code == 200
    direction_payload = create_direction_response.json()
    assert direction_payload["name"] == "企业软件开发"
    assert direction_payload["category"] == "技术研发类"
    assert direction_payload["requires_certificate"] is True

    list_direction_response = client.get("/api/employment-directions?keyword=软件&category=技术研发类")
    assert list_direction_response.status_code == 200
    list_direction_payload = list_direction_response.json()
    assert len(list_direction_payload) == 1
    assert list_direction_payload[0]["alias_names_json"] == ["软件研发", "平台开发"]

    update_direction_response = client.put(
        f"/api/employment-directions/{direction_payload['id']}",
        json={
            "name": "企业软件开发",
            "category": "技术研发类",
            "alias_names_json": ["软件研发", "企业平台开发"],
            "description": "面向企业级软件、平台和信息系统研发岗位。",
            "common_job_types_json": ["后端开发", "平台研发", "实施顾问"],
            "common_industries_json": ["互联网", "企业信息化"],
            "prefers_postgraduate": True,
            "requires_certificate": True,
            "requires_long_cycle": False,
            "supports_art": False,
            "risk_note": "岗位变化快，需要持续学习。",
            "source_note": "更新后方向",
            "is_active": True,
        },
    )
    assert update_direction_response.status_code == 200
    update_direction_payload = update_direction_response.json()
    assert update_direction_payload["prefers_postgraduate"] is True
    assert update_direction_payload["common_job_types_json"][-1] == "实施顾问"

    majors_response = client.get("/api/majors?keyword=软件工程")
    assert majors_response.status_code == 200
    majors_payload = majors_response.json()
    software_major = next(item for item in majors_payload if item["name"] == "软件工程")

    create_mapping_response = client.post(
        "/api/major-employment-maps",
        json={
            "major_id": software_major["id"],
            "direction_id": direction_payload["id"],
            "strength": "high",
            "recommendation_note": "软件工程是企业软件开发的优先入口专业。",
            "requires_postgraduate": False,
            "requires_certificate": True,
            "supported_student_types_json": ["general"],
            "supports_art": False,
            "note": "测试映射",
            "is_active": True,
        },
    )
    assert create_mapping_response.status_code == 200
    mapping_payload = create_mapping_response.json()
    assert mapping_payload["major_name"] == "软件工程"
    assert mapping_payload["direction_name"] == "企业软件开发"
    assert mapping_payload["strength"] == "high"

    list_mapping_response = client.get(f"/api/major-employment-maps?major_id={software_major['id']}&strength=high")
    assert list_mapping_response.status_code == 200
    list_mapping_payload = list_mapping_response.json()
    assert len(list_mapping_payload) == 1
    assert list_mapping_payload[0]["supported_student_types_json"] == ["general"]

    update_mapping_response = client.put(
        f"/api/major-employment-maps/{mapping_payload['id']}",
        json={
            "major_id": software_major["id"],
            "direction_id": direction_payload["id"],
            "strength": "transferable",
            "recommendation_note": "软件工程可转向更多企业数字化岗位。",
            "requires_postgraduate": False,
            "requires_certificate": False,
            "supported_student_types_json": ["general", "repeat"],
            "supports_art": False,
            "note": "更新后映射",
            "is_active": True,
        },
    )
    assert update_mapping_response.status_code == 200
    update_mapping_payload = update_mapping_response.json()
    assert update_mapping_payload["strength"] == "transferable"
    assert update_mapping_payload["supported_student_types_json"] == ["general", "repeat"]

    update_major_profile(
        client,
        "软件工程",
        direction="企业软件与平台开发",
        career_path="企业应用开发、系统实施，部分岗位需要读研或资格证",
        note="建议关注软件架构与项目管理路径",
    )
    update_major_profile(
        client,
        "人工智能",
        direction="智能系统与算法应用",
        career_path="算法工程、模型训练，可先就业也可继续深造",
        note="可延展到机器人与计算机视觉岗位",
    )

    create_rule_response = client.post(
        "/api/province-volunteer-rules",
        json={
            "province": "广东",
            "year": 2026,
            "exam_mode": "3+1+2",
            "batch": "本科批",
            "candidate_type": "general",
            "batch_order": 1,
            "total_score": 750,
            "volunteer_limit": 45,
            "volunteer_unit_type": "院校专业组",
            "subject_requirement_mode": "first_choice_reselect",
            "required_subjects_json": [],
            "first_choice_subjects_json": ["物理", "历史"],
            "reselect_subjects_json": ["化学", "生物", "政治", "地理"],
            "score_rule_summary": "再选科目按等级赋分",
            "parallel_rule_mode": "group_parallel",
            "max_major_per_unit": 6,
            "is_parallel": True,
            "allow_adjustment": True,
            "support_collect_round": True,
            "special_rules_json": ["专业组内最多 6 个专业", "需核对选科要求"],
            "note": "测试规则",
            "is_active": True,
        },
    )
    assert create_rule_response.status_code == 200
    create_rule_payload = create_rule_response.json()
    assert create_rule_payload["candidate_type"] == "general"
    assert create_rule_payload["total_score"] == 750
    assert create_rule_payload["volunteer_limit"] == 45
    assert create_rule_payload["max_major_per_unit"] == 6
    assert create_rule_payload["special_rules_json"] == ["专业组内最多 6 个专业", "需核对选科要求"]

    update_rule_response = client.put(
        f"/api/province-volunteer-rules/{create_rule_payload['id']}",
        json={
            "province": "广东",
            "year": 2026,
            "exam_mode": "3+1+2",
            "batch": "本科批",
            "candidate_type": "general",
            "batch_order": 1,
            "total_score": 750,
            "volunteer_limit": 48,
            "volunteer_unit_type": "院校专业组",
            "subject_requirement_mode": "first_choice_reselect",
            "required_subjects_json": [],
            "first_choice_subjects_json": ["物理", "历史"],
            "reselect_subjects_json": ["化学", "生物", "政治", "地理"],
            "score_rule_summary": "更新后的赋分摘要",
            "parallel_rule_mode": "group_parallel",
            "max_major_per_unit": 6,
            "is_parallel": True,
            "allow_adjustment": False,
            "support_collect_round": True,
            "special_rules_json": ["需核对选科要求"],
            "note": "更新后规则",
            "is_active": True,
        },
    )
    assert update_rule_response.status_code == 200
    update_rule_payload = update_rule_response.json()
    assert update_rule_payload["volunteer_limit"] == 48
    assert update_rule_payload["allow_adjustment"] is False
    assert update_rule_payload["score_rule_summary"] == "更新后的赋分摘要"

    create_art_rule_response = client.post(
        "/api/province-volunteer-rules",
        json={
            "province": "广东",
            "year": 2026,
            "exam_mode": "3+1+2",
            "batch": "本科批",
            "candidate_type": "art",
            "batch_order": 1,
            "total_score": 750,
            "volunteer_limit": 30,
            "volunteer_unit_type": "院校专业组",
            "subject_requirement_mode": "first_choice_reselect",
            "required_subjects_json": [],
            "first_choice_subjects_json": ["历史"],
            "reselect_subjects_json": ["政治", "地理"],
            "score_rule_summary": "艺体类按专业组补充说明执行",
            "parallel_rule_mode": "group_parallel",
            "max_major_per_unit": 4,
            "is_parallel": True,
            "allow_adjustment": True,
            "support_collect_round": False,
            "special_rules_json": ["艺体类需同时核对专业成绩"],
            "note": "艺体规则",
            "is_active": True,
        },
    )
    assert create_art_rule_response.status_code == 200

    list_rule_response = client.get("/api/province-volunteer-rules?province=广东&year=2026&candidate_type=general")
    assert list_rule_response.status_code == 200
    rules_payload = list_rule_response.json()
    assert len(rules_payload) == 1
    assert rules_payload[0]["note"] == "更新后规则"

    create_primary_direction_response = client.post(
        "/api/employment-directions",
        json={
            "name": "人工智能研发",
            "category": "技术研发类",
            "alias_names_json": ["AI研发"],
            "description": "偏算法、模型训练和智能系统",
            "common_job_types_json": ["算法工程师"],
            "common_industries_json": ["人工智能"],
            "prefers_postgraduate": True,
            "requires_certificate": False,
            "requires_long_cycle": False,
            "supports_art": False,
            "risk_note": "需关注学历要求",
            "source_note": "测试",
            "is_active": True,
        },
    )
    assert create_primary_direction_response.status_code == 200
    primary_direction_payload = create_primary_direction_response.json()

    create_secondary_direction_response = client.post(
        "/api/employment-directions",
        json={
            "name": "软件平台工程",
            "category": "技术研发类",
            "alias_names_json": ["平台研发"],
            "description": "偏企业软件与平台建设",
            "common_job_types_json": ["平台开发工程师"],
            "common_industries_json": ["企业数字化"],
            "prefers_postgraduate": False,
            "requires_certificate": False,
            "requires_long_cycle": False,
            "supports_art": False,
            "risk_note": None,
            "source_note": "测试",
            "is_active": True,
        },
    )
    assert create_secondary_direction_response.status_code == 200
    secondary_direction_payload = create_secondary_direction_response.json()

    create_alternative_direction_response = client.post(
        "/api/employment-directions",
        json={
            "name": "数字化产品运营",
            "category": "服务运营类",
            "alias_names_json": [],
            "description": "偏产品与运营协同岗位",
            "common_job_types_json": ["产品运营"],
            "common_industries_json": ["互联网服务"],
            "prefers_postgraduate": False,
            "requires_certificate": False,
            "requires_long_cycle": False,
            "supports_art": False,
            "risk_note": None,
            "source_note": "测试",
            "is_active": True,
        },
    )
    assert create_alternative_direction_response.status_code == 200
    alternative_direction_payload = create_alternative_direction_response.json()

    ai_major_response = client.get("/api/majors?keyword=人工智能")
    assert ai_major_response.status_code == 200
    ai_major_payload = next(item for item in ai_major_response.json() if item["name"] == "人工智能")

    create_ai_mapping_response = client.post(
        "/api/major-employment-maps",
        json={
            "major_id": ai_major_payload["id"],
            "direction_id": primary_direction_payload["id"],
            "strength": "core",
            "recommendation_note": "人工智能是算法研发的核心入口专业。",
            "requires_postgraduate": True,
            "requires_certificate": False,
            "supported_student_types_json": ["general"],
            "supports_art": False,
            "note": "AI 方向映射",
            "is_active": True,
        },
    )
    assert create_ai_mapping_response.status_code == 200

    create_software_direction_mapping_response = client.post(
        "/api/major-employment-maps",
        json={
            "major_id": software_major["id"],
            "direction_id": secondary_direction_payload["id"],
            "strength": "high",
            "recommendation_note": "软件工程与软件平台工程直接相关。",
            "requires_postgraduate": False,
            "requires_certificate": True,
            "supported_student_types_json": ["general"],
            "supports_art": False,
            "note": "软件平台方向映射",
            "is_active": True,
        },
    )
    assert create_software_direction_mapping_response.status_code == 200

    create_career_preference_response = client.post(
        "/api/students/1/career-preference",
        json={
            "primary_direction_id": primary_direction_payload["id"],
            "secondary_direction_id": secondary_direction_payload["id"],
            "alternative_direction_id": alternative_direction_payload["id"],
            "priority_focuses_json": ["stability", "salary"],
            "preferred_industries_json": ["人工智能", "企业数字化"],
            "preferred_job_types_json": ["算法工程师", "平台开发工程师"],
            "target_employment_cities_json": ["深圳", "广州"],
            "accepts_postgraduate": True,
            "accepts_public_service": False,
            "accepts_certificate": True,
            "accepts_long_training": False,
        },
    )
    assert create_career_preference_response.status_code == 200
    career_preference_payload = create_career_preference_response.json()
    assert career_preference_payload["primary_direction_name"] == "人工智能研发"

    get_career_preference_response = client.get("/api/students/1/career-preference")
    assert get_career_preference_response.status_code == 200
    assert get_career_preference_response.json()["preferred_job_types_json"] == ["算法工程师", "平台开发工程师"]

    update_career_preference_response = client.put(
        "/api/students/1/career-preference",
        json={
            "primary_direction_id": secondary_direction_payload["id"],
            "secondary_direction_id": primary_direction_payload["id"],
            "alternative_direction_id": None,
            "priority_focuses_json": ["interest"],
            "preferred_industries_json": ["互联网服务"],
            "preferred_job_types_json": ["产品运营"],
            "target_employment_cities_json": ["珠海"],
            "accepts_postgraduate": False,
            "accepts_public_service": True,
            "accepts_certificate": False,
            "accepts_long_training": True,
        },
    )
    assert update_career_preference_response.status_code == 200
    updated_career_preference_payload = update_career_preference_response.json()
    assert updated_career_preference_payload["primary_direction_name"] == "软件平台工程"
    assert updated_career_preference_payload["accepts_public_service"] is True

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "estimated_score_and_rank",
            "reference_exam_name": "2026届一模",
            "use_historical_mapping": True,
            "risk_preference": "balanced",
            "subject_combination": "物理+化学",
            "primary_direction_id": primary_direction_payload["id"],
            "secondary_direction_id": secondary_direction_payload["id"],
            "alternative_direction_id": alternative_direction_payload["id"],
            "priority_focuses_json": ["stability", "salary"],
            "preferred_industries_json": ["人工智能", "企业数字化"],
            "preferred_job_types_json": ["算法工程师", "平台开发工程师"],
            "target_employment_cities_json": ["深圳", "广州"],
            "accepts_postgraduate": True,
            "accepts_public_service": False,
            "accepts_certificate": True,
            "accepts_long_training": False,
            "student_rank_override": 31000,
            "comprehensive_score": 580,
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["student_name"] == "张三"
    assert preview_payload["candidate_type"] == "general"
    assert preview_payload["score_input_label"] == "预估分数 + 预估位次"
    assert preview_payload["score_confidence"] == "estimated"
    assert any("2026届一模" in item for item in preview_payload["input_notes"])
    assert preview_payload["applicable_rule_count"] == 1
    assert preview_payload["candidate_count"] == 2
    assert preview_payload["candidates"][0]["major_name"] == "人工智能"
    assert preview_payload["candidates"][0]["career_match_strength"] == "core"
    assert preview_payload["candidates"][0]["matched_direction_names_json"] == ["人工智能研发"]
    assert preview_payload["candidates"][0]["requires_postgraduate_path"] is True
    assert "读研" in (preview_payload["candidates"][0]["career_match_summary"] or "")
    assert {item["major_name"] for item in preview_payload["candidates"]} == {"软件工程", "人工智能"}
    assert {item["college_code_snapshot"] for item in preview_payload["candidates"]} == {"10561"}
    assert {item["major_code_snapshot"] for item in preview_payload["candidates"]} == {"080717T", "080902"}
    assert {item["major_direction"] for item in preview_payload["candidates"]} == {
        "企业软件与平台开发",
        "智能系统与算法应用",
    }
    assert preview_payload["candidates"][0]["chapter_url"] == "https://chapter.example/lingnan"
    assert preview_payload["candidates"][0]["chapter_campus_note"] == "广州校区"
    assert preview_payload["candidates"][0]["chapter_review_status"] == "confirmed_manual_review"
    assert preview_payload["candidates"][0]["chapter_language_requirement"] == "英语语种考生优先。"
    assert preview_payload["candidates"][0]["chapter_single_subject_requirement"] == "数学单科不低于 100 分。"
    assert preview_payload["candidates"][0]["chapter_gender_requirement"] == "建议男生。"
    assert preview_payload["candidates"][0]["chapter_height_requirement"] == "身高不低于 168cm。"
    assert preview_payload["candidates"][0]["chapter_vision_requirement"] == "裸眼视力不低于 4.8。"
    assert preview_payload["candidates"][0]["chapter_color_vision_requirement"] == "色弱慎报。"
    assert preview_payload["candidates"][0]["chapter_physical_exam_requirement"] == "需满足体检结论要求。"
    assert any("校区备注：广州校区" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert any("章程备注：部分专业需关注校区安排。" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert any("语言要求：英语语种考生优先。" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert any("单科要求：数学单科不低于 100 分。" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert any("体检要求：需满足体检结论要求。" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert "chapter_special_requirement" in (preview_payload["candidates"][0]["risk_flags_json"] or [])
    assert any("资格证" in (item["career_path"] or "") for item in preview_payload["candidates"])
    assert any("项目管理路径" in (item["major_note"] or "") for item in preview_payload["candidates"])
    assert any("major_baseline_missing" in (item["risk_flags_json"] or []) for item in preview_payload["candidates"])

    create_draft_response = client.post(
        "/api/recommendations/volunteer-drafts",
        json={
            "name": "张三-本科批草稿",
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "estimated_score_and_rank",
            "reference_exam_name": "2026届一模",
            "use_historical_mapping": True,
            "risk_preference": "balanced",
            "target_regions_json": ["广东"],
            "school_level_tags_json": ["双一流"],
            "major_keyword": "人工",
            "subject_combination": "物理+化学",
            "primary_direction_id": primary_direction_payload["id"],
            "secondary_direction_id": secondary_direction_payload["id"],
            "alternative_direction_id": alternative_direction_payload["id"],
            "priority_focuses_json": ["stability", "salary"],
            "preferred_industries_json": ["人工智能", "企业数字化"],
            "preferred_job_types_json": ["算法工程师", "平台开发工程师"],
            "target_employment_cities_json": ["深圳", "广州"],
            "accepts_postgraduate": True,
            "accepts_public_service": False,
            "accepts_certificate": True,
            "accepts_long_training": False,
            "student_rank_override": 31000,
            "note": "第一次保存",
            "selected_rule": preview_payload["applicable_rules"][0],
            "items": [
                {
                    "order": index + 1,
                    "plan_id": item["plan_id"],
                    "candidate": item,
                }
                for index, item in enumerate(preview_payload["candidates"])
            ],
        },
    )
    assert create_draft_response.status_code == 200
    draft_payload = create_draft_response.json()
    assert draft_payload["name"] == "张三-本科批草稿"
    assert len(draft_payload["items"]) == 2
    assert draft_payload["primary_direction_id"] == primary_direction_payload["id"]
    assert draft_payload["preferred_industries_json"] == ["人工智能", "企业数字化"]

    list_draft_response = client.get(f"/api/recommendations/volunteer-drafts?student_id=1&exam_id={exam_id}")
    assert list_draft_response.status_code == 200
    list_draft_payload = list_draft_response.json()
    assert len(list_draft_payload) == 1
    assert list_draft_payload[0]["item_count"] == 2

    detail_draft_response = client.get(f"/api/recommendations/volunteer-drafts/{draft_payload['id']}")
    assert detail_draft_response.status_code == 200
    assert detail_draft_response.json()["selected_rule"]["volunteer_limit"] == 48
    assert detail_draft_response.json()["score_input_mode"] == "estimated_score_and_rank"
    assert detail_draft_response.json()["reference_exam_name"] == "2026届一模"

    update_draft_response = client.put(
        f"/api/recommendations/volunteer-drafts/{draft_payload['id']}",
        json={
            "name": "张三-本科批二次调整",
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "score_range",
            "score_range_min": 570,
            "score_range_max": 585,
            "reference_exam_name": "2026届二模",
            "use_historical_mapping": True,
            "risk_preference": "conservative",
            "target_regions_json": ["广东", "深圳"],
            "school_level_tags_json": [],
            "major_keyword": "",
            "subject_combination": "物理+化学",
            "primary_direction_id": secondary_direction_payload["id"],
            "secondary_direction_id": alternative_direction_payload["id"],
            "alternative_direction_id": None,
            "priority_focuses_json": ["interest"],
            "preferred_industries_json": ["互联网服务"],
            "preferred_job_types_json": ["产品运营"],
            "target_employment_cities_json": ["深圳"],
            "accepts_postgraduate": False,
            "accepts_public_service": True,
            "accepts_certificate": False,
            "accepts_long_training": True,
            "student_rank_override": 30500,
            "note": "重新排序后保存",
            "selected_rule": preview_payload["applicable_rules"][0],
            "items": [
                {
                    "order": 1,
                    "plan_id": preview_payload["candidates"][1]["plan_id"],
                    "candidate": preview_payload["candidates"][1],
                }
            ],
        },
    )
    assert update_draft_response.status_code == 200
    update_draft_payload = update_draft_response.json()
    assert update_draft_payload["name"] == "张三-本科批二次调整"
    assert len(update_draft_payload["items"]) == 1
    assert update_draft_payload["score_input_mode"] == "score_range"
    assert update_draft_payload["score_range_min"] == 570
    assert update_draft_payload["risk_preference"] == "conservative"
    assert update_draft_payload["items"][0]["candidate"]["plan_id"] == preview_payload["candidates"][1]["plan_id"]
    assert update_draft_payload["primary_direction_id"] == secondary_direction_payload["id"]
    assert update_draft_payload["accepts_public_service"] is True

    export_draft_response = client.post(
        "/api/reports/export",
        json={
            "report_type": "volunteer_draft_summary",
            "draft_id": draft_payload["id"],
        },
    )
    assert export_draft_response.status_code == 200
    export_draft_payload = export_draft_response.json()
    assert export_draft_payload["report_type"] == "volunteer_draft_summary"
    assert export_draft_payload["download_url"].startswith("/api/reports/exports/")

    download_draft_response = client.get(export_draft_payload["download_url"])
    assert download_draft_response.status_code == 200
    assert download_draft_response.content
    export_workbook = load_workbook(BytesIO(download_draft_response.content))
    export_summary_sheet = export_workbook["志愿草稿概况"]
    export_boundary_sheet = export_workbook["边界概览"]
    export_sheet = export_workbook["志愿草稿"]
    assert export_summary_sheet.cell(row=5, column=1).value == "目标年份"
    assert export_summary_sheet.cell(row=5, column=2).value == 2026
    assert export_summary_sheet.cell(row=8, column=1).value == "分数模式"
    assert export_summary_sheet.cell(row=8, column=2).value == "分数区间"
    assert export_summary_sheet.cell(row=9, column=1).value == "模拟说明"
    assert "2026届二模" in str(export_summary_sheet.cell(row=9, column=2).value)
    assert "历史映射" in str(export_summary_sheet.cell(row=9, column=2).value)
    assert export_boundary_sheet.cell(row=1, column=1).value == "类型"
    export_boundary_rows = [
        (
            export_boundary_sheet.cell(row=index, column=1).value,
            export_boundary_sheet.cell(row=index, column=2).value,
            export_boundary_sheet.cell(row=index, column=3).value,
        )
        for index in range(2, export_boundary_sheet.max_row + 1)
    ]
    assert (
        "类别专用规则口径",
        "1 条已选志愿当前按普通类专用规则解释",
        "当前命中的省份规则已细分到普通类；同省同年其他类别可能适用不同的志愿上限、单位结构和选科口径。",
    ) in export_boundary_rows
    assert export_sheet.cell(row=1, column=4).value == "院校代码"
    assert export_sheet.cell(row=1, column=6).value == "专业代码"
    assert export_sheet.cell(row=1, column=15).value == "命中规则"
    assert export_sheet.cell(row=1, column=16).value == "录取口径"
    assert export_sheet.cell(row=1, column=17).value == "边界说明"
    assert export_sheet.cell(row=1, column=24).value == "章程链接"
    assert export_sheet.cell(row=1, column=25).value == "章程状态"
    assert export_sheet.cell(row=1, column=26).value == "校区备注"
    assert export_sheet.cell(row=1, column=27).value == "章程备注"
    assert "命中规则：" in str(export_sheet.cell(row=2, column=15).value)
    assert "参考" in str(export_sheet.cell(row=2, column=16).value)
    assert export_sheet.cell(row=2, column=24).value == "https://chapter.example/lingnan"
    assert export_sheet.cell(row=2, column=25).value == "confirmed_manual_review"
    assert export_sheet.cell(row=2, column=26).value == "广州校区"
    assert export_sheet.cell(row=2, column=27).value == "部分专业需关注校区安排。"
    exported_college_codes = {
        export_sheet.cell(row=row_index, column=4).value for row_index in range(2, export_sheet.max_row + 1)
    }
    exported_major_codes = {
        export_sheet.cell(row=row_index, column=6).value for row_index in range(2, export_sheet.max_row + 1)
    }
    assert exported_college_codes == {"10561"}
    assert exported_major_codes == {update_draft_payload["items"][0]["candidate"]["major_code_snapshot"]}

    delete_draft_response = client.delete(f"/api/recommendations/volunteer-drafts/{draft_payload['id']}")
    assert delete_draft_response.status_code == 200
    assert delete_draft_response.json()["message"] == "志愿草稿已删除"

    list_after_delete_response = client.get(f"/api/recommendations/volunteer-drafts?student_id=1&exam_id={exam_id}")
    assert list_after_delete_response.status_code == 200
    assert list_after_delete_response.json() == []

    student_detail_response = client.get("/api/students/1")
    assert student_detail_response.status_code == 200
    student_payload = student_detail_response.json()
    update_student_response = client.put(
        "/api/students/1",
        json={
            **student_payload,
            "origin_province": "广东",
        },
    )
    assert update_student_response.status_code == 200
    assert update_student_response.json()["origin_province"] == "广东"

    general_response = client.post(
        "/api/recommendations/generate",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "score_input_mode": "estimated_score_and_rank",
            "student_rank_override": 31000,
            "comprehensive_score": 582,
            "reference_exam_name": "2026届一模",
            "use_historical_mapping": True,
            "risk_preference": "balanced",
        },
    )
    assert general_response.status_code == 200
    general_payload = general_response.json()
    assert general_payload["result_count"] == 3
    assert len(general_payload["challenge"]) == 1
    assert len(general_payload["steady"]) == 1
    assert len(general_payload["safe"]) == 1
    assert general_payload["challenge"][0]["college_name"] == "岭南科技大学"
    assert general_payload["steady"][0]["college_name"] == "湾区信息大学"
    assert general_payload["safe"][0]["college_name"] == "南方应用大学"
    assert general_payload["challenge"][0]["career_match_strength"] == "high"
    assert general_payload["challenge"][0]["matched_direction_names_json"] == ["软件平台工程"]
    assert "资格证" in (general_payload["challenge"][0]["career_match_summary"] or "")
    assert "simulation_mode" in (general_payload["challenge"][0]["risk_flags_json"] or [])
    assert general_payload["challenge"][0]["snapshot_json"]["score_input_label"] == "预估分数 + 预估位次"
    assert general_payload["challenge"][0]["snapshot_json"]["score_confidence"] == "estimated"
    assert general_payload["challenge"][0]["snapshot_json"]["chapter_url"] == "https://chapter.example/lingnan"
    assert general_payload["challenge"][0]["snapshot_json"]["chapter_campus_note"] == "广州校区"
    assert general_payload["challenge"][0]["snapshot_json"]["chapter_review_status"] == "confirmed_manual_review"
    assert general_payload["challenge"][0]["snapshot_json"]["chapter_language_requirement"] == "英语语种考生优先。"
    assert general_payload["challenge"][0]["snapshot_json"]["chapter_single_subject_requirement"] == "数学单科不低于 100 分。"
    assert "chapter_special_requirement" in (general_payload["challenge"][0]["risk_flags_json"] or [])
    assert "校区备注：广州校区。" in general_payload["challenge"][0]["reason_text"]
    assert "章程备注：部分专业需关注校区安排。" in general_payload["challenge"][0]["reason_text"]
    assert "语言要求：英语语种考生优先。" in general_payload["challenge"][0]["reason_text"]
    assert "单科要求：数学单科不低于 100 分。" in general_payload["challenge"][0]["reason_text"]
    assert "参考考试：2026届一模。" in general_payload["challenge"][0]["reason_text"]

    export_recommendation_response = client.post(
        "/api/reports/export",
        json={
            "report_type": "recommendation_summary",
            "scheme_id": general_payload["scheme_id"],
            "student_id": 1,
        },
    )
    assert export_recommendation_response.status_code == 200
    export_recommendation_payload = export_recommendation_response.json()
    assert export_recommendation_payload["download_url"].startswith("/api/reports/exports/")

    download_recommendation_response = client.get(export_recommendation_payload["download_url"])
    assert download_recommendation_response.status_code == 200
    recommendation_workbook = load_workbook(BytesIO(download_recommendation_response.content))
    recommendation_summary_sheet = recommendation_workbook["推荐概况"]
    recommendation_risk_sheet = recommendation_workbook["风险概览"]
    recommendation_detail_sheet = recommendation_workbook["推荐结果"]
    assert recommendation_summary_sheet.cell(row=5, column=1).value == "目标年份"
    assert recommendation_summary_sheet.cell(row=5, column=2).value == 2026
    assert recommendation_summary_sheet.cell(row=6, column=1).value == "分数模式"
    assert recommendation_summary_sheet.cell(row=6, column=2).value == "预估分数 + 预估位次"
    assert recommendation_summary_sheet.cell(row=7, column=1).value == "模拟说明"
    assert "2026届一模" in str(recommendation_summary_sheet.cell(row=7, column=2).value)
    assert recommendation_summary_sheet.cell(row=8, column=1).value == "目标方向"
    assert "软件平台工程" in str(recommendation_summary_sheet.cell(row=8, column=2).value)
    assert recommendation_summary_sheet.cell(row=9, column=1).value == "可接受路径"
    assert "考公" in str(recommendation_summary_sheet.cell(row=9, column=2).value)
    assert recommendation_risk_sheet.cell(row=1, column=1).value == "分组"
    assert recommendation_risk_sheet.cell(row=1, column=2).value == "类型"
    recommendation_risk_rows = [
        (
            recommendation_risk_sheet.cell(row=index, column=1).value,
            recommendation_risk_sheet.cell(row=index, column=2).value,
            recommendation_risk_sheet.cell(row=index, column=3).value,
            recommendation_risk_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, recommendation_risk_sheet.max_row + 1)
    ]
    assert any(row[1] == "统一风险口径" and "2026 正式招生计划" in str(row[3]) for row in recommendation_risk_rows)
    assert any(row[1] == "模拟结果" and "预估分数 + 预估位次" in str(row[2]) and "2026届一模" in str(row[3]) for row in recommendation_risk_rows)
    assert recommendation_detail_sheet.cell(row=1, column=8).value == "录取结果年份"
    assert recommendation_detail_sheet.cell(row=1, column=9).value == "计划年份"
    assert recommendation_detail_sheet.cell(row=1, column=10).value == "批次"
    assert recommendation_detail_sheet.cell(row=1, column=11).value == "选科要求"
    assert recommendation_detail_sheet.cell(row=1, column=12).value == "职业匹配"
    assert recommendation_detail_sheet.cell(row=1, column=18).value == "章程链接"
    assert recommendation_detail_sheet.cell(row=1, column=19).value == "章程状态"
    assert recommendation_detail_sheet.cell(row=1, column=20).value == "章程限制状态"
    assert recommendation_detail_sheet.cell(row=1, column=21).value == "校区备注"
    assert recommendation_detail_sheet.cell(row=1, column=22).value == "章程备注"
    assert recommendation_detail_sheet.cell(row=2, column=13).value == "软件平台工程"
    assert "资格证" in str(recommendation_detail_sheet.cell(row=2, column=14).value)
    assert recommendation_detail_sheet.cell(row=2, column=18).value == "https://chapter.example/lingnan"
    assert recommendation_detail_sheet.cell(row=2, column=19).value == "confirmed_manual_review"
    assert "语种" in str(recommendation_detail_sheet.cell(row=2, column=20).value)
    assert recommendation_detail_sheet.cell(row=2, column=21).value == "广州校区"
    assert recommendation_detail_sheet.cell(row=2, column=22).value == "部分专业需关注校区安排。"

    initial_history_response = client.get("/api/recommendations/history?student_id=1")
    assert initial_history_response.status_code == 200
    assert len(initial_history_response.json()) == 1

    failed_response = client.post(
        "/api/recommendations/generate",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "student_rank_override": 31000,
            "major_keyword": "不存在的专业方向",
        },
    )
    assert failed_response.status_code == 404
    assert failed_response.json()["detail"] == "当前条件下暂无可推荐的录取数据"

    history_after_failure = client.get("/api/recommendations/history?student_id=1")
    assert history_after_failure.status_code == 200
    assert len(history_after_failure.json()) == 1

    settings_response = client.put(
        "/api/recommendations/settings",
        json={
            "safe_ratio_max": 0.85,
            "steady_ratio_max": 1.0,
            "rush_ratio_max": 1.15,
            "whitelist_college_ids": [],
            "blacklist_college_ids": [general_payload["challenge"][0]["college_id"]],
        },
    )
    assert settings_response.status_code == 200
    settings_payload = settings_response.json()
    assert settings_payload["blacklist_college_ids"] == [general_payload["challenge"][0]["college_id"]]
    assert settings_payload["strategy_presets"] == []

    preset_response = client.post(
        "/api/recommendations/strategy-presets",
        json={
            "name": "保守模板",
            "note": "优先稳保",
            "safe_ratio_max": 0.8,
            "steady_ratio_max": 0.95,
            "rush_ratio_max": 1.05,
            "whitelist_college_ids": [general_payload["safe"][0]["college_id"]],
            "blacklist_college_ids": [general_payload["challenge"][0]["college_id"]],
        },
    )
    assert preset_response.status_code == 200
    preset_payload = preset_response.json()
    assert len(preset_payload["strategy_presets"]) == 1
    assert preset_payload["strategy_presets"][0]["name"] == "保守模板"

    delete_preset_response = client.delete(
        f"/api/recommendations/strategy-presets/{preset_payload['strategy_presets'][0]['id']}"
    )
    assert delete_preset_response.status_code == 200
    assert delete_preset_response.json()["strategy_presets"] == []

    filtered_response = client.post(
        "/api/recommendations/generate",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "student_rank_override": 31000,
            "name": "黑名单过滤后",
        },
    )
    assert filtered_response.status_code == 200
    filtered_payload = filtered_response.json()
    filtered_colleges = {
        item["college_name"] for key in ("challenge", "steady", "safe") for item in filtered_payload[key]
    }
    assert "岭南科技大学" not in filtered_colleges

    art_response = client.post(
        "/api/recommendations/generate",
        json={
            "student_id": 3,
            "exam_id": exam_id,
            "province": "广东",
            "score_input_mode": "estimated_score",
            "culture_score": 360,
            "professional_score": 270,
        },
    )
    assert art_response.status_code == 200
    art_payload = art_response.json()
    assert art_payload["result_count"] == 1
    assert art_payload["steady"][0]["college_name"] == "岭南艺术学院"
    assert "art_recommendation" in (art_payload["steady"][0]["risk_flags_json"] or [])

    history_response = client.get("/api/recommendations/history?student_id=1")
    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert len(history_payload) == 2
    assert history_payload[0]["scheme_name"] == "黑名单过滤后"
    assert history_payload[0]["result_count"] == 2
    assert history_payload[1]["result_count"] == 3

    export_recommendation_compare_response = client.post(
        "/api/reports/export",
        json={
            "report_type": "recommendation_summary",
            "scheme_id": general_payload["scheme_id"],
            "student_id": 1,
        },
    )
    assert export_recommendation_compare_response.status_code == 200
    export_recommendation_compare_payload = export_recommendation_compare_response.json()
    download_recommendation_compare_response = client.get(export_recommendation_compare_payload["download_url"])
    assert download_recommendation_compare_response.status_code == 200
    recommendation_compare_workbook = load_workbook(BytesIO(download_recommendation_compare_response.content))
    recommendation_compare_risk_sheet = recommendation_compare_workbook["风险概览"]
    recommendation_compare_rows = [
        (
            recommendation_compare_risk_sheet.cell(row=index, column=1).value,
            recommendation_compare_risk_sheet.cell(row=index, column=2).value,
            recommendation_compare_risk_sheet.cell(row=index, column=3).value,
            recommendation_compare_risk_sheet.cell(row=index, column=4).value,
        )
        for index in range(2, recommendation_compare_risk_sheet.max_row + 1)
    ]
    assert any(row[0] == "历史对照摘要" and row[1] == "历史方案差异" for row in recommendation_compare_rows)
    assert any("黑名单过滤后" in str(row[2]) or "黑名单过滤后" in str(row[3]) for row in recommendation_compare_rows)

    result_list_response = client.get(
        f"/api/recommendations/history/{general_payload['scheme_id']}/results?student_id=1"
    )
    assert result_list_response.status_code == 200
    assert len(result_list_response.json()) == 3


def test_batch_recommendation_uses_each_student_origin_province(client) -> None:
    exam_id = create_exam_with_scores(client)

    import_response = client.post(
        "/api/admissions/import",
        files={
            "file": (
                "cross-province-admissions.xlsx",
                build_cross_province_admission_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["success_rows"] == 2
    assert import_payload["created_college_count"] == 1
    assert import_payload["created_major_count"] == 1

    student_1_payload = client.get("/api/students/1")
    assert student_1_payload.status_code == 200
    update_student_1_response = client.put(
        "/api/students/1",
        json={
            **student_1_payload.json(),
            "origin_province": "山东",
        },
    )
    assert update_student_1_response.status_code == 200
    assert update_student_1_response.json()["origin_province"] == "山东"

    student_2_payload = client.get("/api/students/2")
    assert student_2_payload.status_code == 200
    update_student_2_response = client.put(
        "/api/students/2",
        json={
            **student_2_payload.json(),
            "origin_province": "河北",
        },
    )
    assert update_student_2_response.status_code == 200
    assert update_student_2_response.json()["origin_province"] == "河北"

    batch_response = client.post(
        "/api/recommendations/batch-generate",
        json={
            "student_ids": [1, 2],
            "exam_id": exam_id,
            "name": "跨省批量推荐",
            "target_year": 2026,
        },
    )
    assert batch_response.status_code == 200
    batch_payload = batch_response.json()
    assert batch_payload["message"] == "批量推荐完成，已按 2 个生源地分别生成"
    assert len(batch_payload["scheme_ids"]) == 2
    assert batch_payload["result_count"] == 2

    items_by_student = {item["student_id"]: item for item in batch_payload["items"]}
    assert set(items_by_student) == {1, 2}
    assert items_by_student[1]["student_name"] == "张三"
    assert items_by_student[1]["province"] == "山东"
    assert items_by_student[1]["result_count"] == 1
    assert items_by_student[2]["student_name"] == "李四"
    assert items_by_student[2]["province"] == "河北"
    assert items_by_student[2]["result_count"] == 1

    student_1_history_response = client.get("/api/recommendations/history?student_id=1")
    assert student_1_history_response.status_code == 200
    student_1_history_payload = student_1_history_response.json()
    assert len(student_1_history_payload) == 1
    assert student_1_history_payload[0]["scheme_id"] == items_by_student[1]["scheme_id"]
    assert student_1_history_payload[0]["province"] == "山东"

    student_2_history_response = client.get("/api/recommendations/history?student_id=2")
    assert student_2_history_response.status_code == 200
    student_2_history_payload = student_2_history_response.json()
    assert len(student_2_history_payload) == 1
    assert student_2_history_payload[0]["scheme_id"] == items_by_student[2]["scheme_id"]
    assert student_2_history_payload[0]["province"] == "河北"

    student_1_results_response = client.get(
        f"/api/recommendations/history/{items_by_student[1]['scheme_id']}/results?student_id=1"
    )
    assert student_1_results_response.status_code == 200
    student_1_results = student_1_results_response.json()
    assert len(student_1_results) == 1
    assert student_1_results[0]["college_name"] == "华北信息大学"
    assert student_1_results[0]["major_name"] == "计算机科学与技术"
    assert student_1_results[0]["reference_rank"] == 12000

    student_2_results_response = client.get(
        f"/api/recommendations/history/{items_by_student[2]['scheme_id']}/results?student_id=2"
    )
    assert student_2_results_response.status_code == 200
    student_2_results = student_2_results_response.json()
    assert len(student_2_results) == 1
    assert student_2_results[0]["college_name"] == "华北信息大学"
    assert student_2_results[0]["major_name"] == "计算机科学与技术"
    assert student_2_results[0]["reference_rank"] == 22000


def test_volunteer_workbench_preview_supports_exam_mode_compatibility(client) -> None:
    exam_id = create_exam_with_scores(client)

    admission_import_response = client.post(
        "/api/admissions/import",
        files={
            "file": (
                "admissions.xlsx",
                build_admission_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert admission_import_response.status_code == 200

    plan_import_response = client.post(
        "/api/enrollment-plans/import",
        files={
            "file": (
                "enrollment-plans.xlsx",
                build_enrollment_plan_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert plan_import_response.status_code == 200

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "3+1+2",
            "candidate_type": "general",
            "score_input_mode": "actual_rank",
            "student_rank_override": 31000,
            "subject_combination": "物理+化学",
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["applicable_rule_count"] == 2
    assert preview_payload["candidate_count"] == 2
    assert {item["exam_mode"] for item in preview_payload["applicable_rules"]} == {"历史类", "物理类"}
    assert {item["exam_mode"] for item in preview_payload["candidates"]} == {"物理类"}
    assert any(item["code"] == "compatible_exam_mode_fallback" for item in preview_payload["rule_alerts"])
    assert all(item["matched_rule_exam_mode"] == "物理类" for item in preview_payload["candidates"])
    assert all(item["matched_rule_batch"] == "本科批" for item in preview_payload["candidates"])
    assert all(item["matched_rule_candidate_type"] == "" for item in preview_payload["candidates"])
    assert all(item["matched_rule_is_baseline"] is True for item in preview_payload["candidates"])
    assert all(item["reference_record_count"] == 1 for item in preview_payload["candidates"])
    assert all(item["reference_years_json"] == [2025] for item in preview_payload["candidates"])
    assert all(item["reference_source_notes_json"] == ["近年数据"] for item in preview_payload["candidates"])
    assert any(item["reference_scope"] == "major" for item in preview_payload["candidates"])
    assert any(item["reference_scope"] == "college" for item in preview_payload["candidates"])
    assert all("兼容模式命中" in item["match_tags_json"] for item in preview_payload["candidates"])
    assert any("专业线参考" in item["match_tags_json"] for item in preview_payload["candidates"])
    assert any("院校线参考" in item["match_tags_json"] for item in preview_payload["candidates"])
    assert all("基线规则命中" in item["match_tags_json"] for item in preview_payload["candidates"])
    assert all("已带选科条件" in item["match_tags_json"] for item in preview_payload["candidates"])
    assert all("按位次分层" in item["match_tags_json"] for item in preview_payload["candidates"])
    assert any("当前按 广东 2026 物理类 本科批 规则解释。" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert any("当前请求模式为“3+1+2”" in note for note in preview_payload["candidates"][0]["match_notes_json"])
    assert any(
        "当前未配置“3+1+2”精确规则，先按兼容模式" in item and "物理类" in item
        for item in preview_payload["input_notes"]
    )


def test_volunteer_workbench_preview_truncates_large_candidate_pool(client) -> None:
    exam_id = create_exam_with_scores(client)

    plan_import_response = client.post(
        "/api/enrollment-plans/import",
        files={
            "file": (
                "large-enrollment-plans.xlsx",
                build_large_enrollment_plan_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert plan_import_response.status_code == 200
    assert plan_import_response.json()["success_rows"] == 305

    admission_import_response = client.post(
        "/api/admissions/import",
        files={
            "file": (
                "large-admissions.xlsx",
                build_large_admission_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert admission_import_response.status_code == 200
    assert admission_import_response.json()["success_rows"] == 305

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "actual_rank",
            "student_rank_override": 31000,
            "subject_combination": "物理+化学",
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["candidate_count"] == 305
    assert preview_payload["returned_candidate_count"] == 300
    assert preview_payload["is_candidate_truncated"] is True
    assert len(preview_payload["candidates"]) == 300
    assert any(item["code"] == "candidate_result_truncated" for item in preview_payload["rule_alerts"])


def test_volunteer_workbench_preview_explains_missing_rule_year(client) -> None:
    exam_id = create_exam_with_scores(client)

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2025,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "actual_rank",
            "student_rank_override": 31000,
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["applicable_rule_count"] == 0
    assert preview_payload["candidate_count"] == 0
    assert any(item["code"] == "missing_rule_year" for item in preview_payload["rule_alerts"])
    assert any(
        "当前未找到 广东 2025 年省份规则；该省现有 2026 年规则" in item
        for item in preview_payload["input_notes"]
    )


def test_volunteer_draft_detail_reuses_live_rule_alerts(client) -> None:
    exam_id = create_exam_with_scores(client)

    admission_import_response = client.post(
        "/api/admissions/import",
        files={
            "file": (
                "admissions.xlsx",
                build_admission_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert admission_import_response.status_code == 200

    plan_import_response = client.post(
        "/api/enrollment-plans/import",
        files={
            "file": (
                "enrollment-plans.xlsx",
                build_enrollment_plan_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert plan_import_response.status_code == 200

    preview_response = client.post(
        "/api/recommendations/volunteer-workbench/preview",
        json={
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2026,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "actual_rank",
            "student_rank_override": 31000,
            "subject_combination": "物理+化学",
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["candidate_count"] >= 1

    create_draft_response = client.post(
        "/api/recommendations/volunteer-drafts",
        json={
            "name": "张三-规则告警复用",
            "student_id": 1,
            "exam_id": exam_id,
            "province": "广东",
            "target_year": 2025,
            "batch": "本科批",
            "exam_mode": "物理类",
            "candidate_type": "general",
            "score_input_mode": "actual_rank",
            "use_historical_mapping": False,
            "risk_preference": "balanced",
            "target_regions_json": [],
            "school_level_tags_json": [],
            "major_keyword": "",
            "subject_combination": "物理+化学",
            "priority_focuses_json": [],
            "preferred_industries_json": [],
            "preferred_job_types_json": [],
            "target_employment_cities_json": [],
            "accepts_postgraduate": False,
            "accepts_public_service": False,
            "accepts_certificate": False,
            "accepts_long_training": False,
            "selected_rule": None,
            "items": [
                {
                    "order": 1,
                    "plan_id": preview_payload["candidates"][0]["plan_id"],
                    "candidate": preview_payload["candidates"][0],
                }
            ],
        },
    )
    assert create_draft_response.status_code == 200
    draft_payload = create_draft_response.json()
    assert any(item["code"] == "missing_rule_year" for item in draft_payload["rule_alerts"])
    assert draft_payload["applicable_rules"] == []

    detail_response = client.get(f"/api/recommendations/volunteer-drafts/{draft_payload['id']}")
    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert any(item["code"] == "missing_rule_year" for item in detail_payload["rule_alerts"])
    assert detail_payload["applicable_rules"] == []
