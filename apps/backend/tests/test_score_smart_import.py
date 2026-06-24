from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook


def _create_exam(client, name: str = "2026届高一智能导入测试") -> int:
    exam_response = client.post(
        "/api/exams",
        json={
            "name": name,
            "exam_type": "月考",
            "exam_date": "2026-05-01",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]
    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            },
            {
                "subject_id": 2,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 2,
                "is_active": True,
            },
        ],
    )
    assert subject_response.status_code == 200
    return exam_id


def _wide_platform_workbook(rows: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["七天网络成绩单"])
    sheet.append(["导出时间", "2026-05-01"])
    sheet.append(["班级", "考号", "姓名", "语文成绩", "数学成绩", "总分", "班级名次", "年级名次"])
    for row in rows or [
        ["1班", "2026001", "张三", 118, 125, 243, 1, 1],
        ["1班", "2026002", "李四", 110, 120, 230, 2, 2],
    ]:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wide_conversion_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "赋分报告"
    sheet.append(["平台赋分报告"])
    sheet.append(["班级", "考号", "姓名", "语文成绩", "数学原始分", "数学赋分", "总分", "班级名次", "年级名次"])
    sheet.append(["1班", "2026001", "张三", 118, 72, 86, 204, 1, 1])
    sheet.append(["1班", "2026002", "李四", 110, 75, 82, 192, 2, 2])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _missing_student_no_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["姓名", "班级", "语文成绩"])
    sheet.append(["张三", "1班", 118])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_wide_score_sheet_preview_import_and_profile_reuse(client) -> None:
    exam_id = _create_exam(client)
    content = _wide_platform_workbook()

    preview_response = client.post(
        f"/api/exams/{exam_id}/scores/import/preview",
        files={"file": ("qitian.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert preview["layout_type"] == "wide"
    assert preview["header_row"] == 3
    assert preview["import_ready"] is True
    assert preview["detected_record_count"] == 4
    assert preview["mapping"]["field_mapping"]["student_no"] == "考号"
    assert preview["mapping"]["subject_mapping"]["语文成绩"] == "语文"
    assert preview["mapping"]["metadata_mapping"]["source_total_score"] == "总分"
    assert preview["mapping"]["metadata_mapping"]["source_class_rank"] == "班级名次"
    assert preview["mapping"]["metadata_mapping"]["source_school_rank"] == "年级名次"

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={
            "strategy": "overwrite",
            "rebuild": "true",
            "mapping_json": json.dumps(preview["mapping"]),
            "save_profile_name": "七天网络宽表",
        },
        files={"file": ("qitian.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_response.status_code == 200
    payload = import_response.json()
    assert payload["success_rows"] == 4
    assert payload["failed_rows"] == 0
    assert payload["profile_id"] is not None
    assert payload["detection_summary"]["layout_type"] == "wide"

    analytics_response = client.get(f"/api/analytics/students/1?exam_id={exam_id}")
    assert analytics_response.status_code == 200
    assert analytics_response.json()["total_score"] == 243

    audit_response = client.get(f"/api/exams/{exam_id}/score-rank-audit")
    assert audit_response.status_code == 200
    audit_payload = audit_response.json()
    assert audit_payload["rank_scope_label"] == "校内名次（本次有效导入样本）"
    assert audit_payload["mapping_rate"] == 1
    assert audit_payload["source_rank_count"] == 2

    profiles_response = client.get("/api/exams/score-import-profiles")
    assert profiles_response.status_code == 200
    profiles = profiles_response.json()
    saved_profile = next(item for item in profiles if item["name"] == "七天网络宽表")

    next_exam_id = _create_exam(client, name="2026届高一智能导入复用测试")
    reuse_response = client.post(
        f"/api/exams/{next_exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true", "profile_id": str(saved_profile["id"])},
        files={
            "file": (
                "qitian-next.xlsx",
                _wide_platform_workbook(rows=[["1班", "2026001", "张三", 120, 128, 248, 1, 1]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert reuse_response.status_code == 200
    assert reuse_response.json()["success_rows"] == 2


def test_wide_score_sheet_groups_original_and_converted_scores(client) -> None:
    exam_id = _create_exam(client, name="2026届高一赋分导入测试")
    content = _wide_conversion_workbook()

    preview_response = client.post(
        f"/api/exams/{exam_id}/scores/import/preview",
        files={"file": ("conversion.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert preview["layout_type"] == "wide"
    assert preview["detected_record_count"] == 4
    assert preview["mapping"]["subject_mapping"]["数学原始分"] == "数学"
    assert preview["mapping"]["subject_mapping"]["数学赋分"] == "数学"
    assert preview["mapping"]["subject_score_types"]["数学原始分"] == "original"
    assert preview["mapping"]["subject_score_types"]["数学赋分"] == "converted"
    math_preview = next(item for item in preview["normalized_preview"] if item["姓名"] == "张三" and item["科目"] == "数学")
    assert math_preview["分数"] == 86
    assert math_preview["原始分"] == 72
    assert math_preview["赋分"] == 86
    assert math_preview["成绩口径"] == "converted"

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true", "mapping_json": json.dumps(preview["mapping"])},
        files={"file": ("conversion.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_response.status_code == 200
    assert import_response.json()["success_rows"] == 4

    analytics_response = client.get(f"/api/analytics/students/1?exam_id={exam_id}")
    assert analytics_response.status_code == 200
    analytics = analytics_response.json()
    assert analytics["total_score"] == 204
    assert analytics["score_value_type"] == "converted"
    math_subject = next(item for item in analytics["subjects"] if item["subject_name"] == "数学")
    assert math_subject["score"] == 86
    assert math_subject["original_score"] == 72
    assert math_subject["converted_score"] == 86
    assert math_subject["score_value_label"] == "赋分"

    profile_response = client.get("/api/students/1/profile")
    assert profile_response.status_code == 200
    trend = next(item for item in profile_response.json()["exam_trends"] if item["exam_id"] == exam_id)
    assert trend["score_value_label"] == "赋分"
    assert next(item for item in trend["subjects"] if item["subject_name"] == "数学")["score"] == 86


def test_score_rank_uses_exam_source_class_after_current_class_changes(client) -> None:
    exam_id = _create_exam(client, name="2026届高一源班级排名测试")
    content = _wide_platform_workbook(
        rows=[
            ["2班", "2026001", "张三", 118, 125, 243, 1, 1],
            ["2班", "2026002", "李四", 110, 120, 230, 2, 2],
        ]
    )

    preview_response = client.post(
        f"/api/exams/{exam_id}/scores/import/preview",
        files={"file": ("source-class.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={
            "strategy": "overwrite",
            "rebuild": "true",
            "mapping_json": json.dumps(preview["mapping"]),
        },
        files={"file": ("source-class.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_response.status_code == 200

    class_two_response = client.get(f"/api/analytics/classes/2?exam_id={exam_id}")
    assert class_two_response.status_code == 200
    assert class_two_response.json()["student_count"] == 2

    class_one_response = client.get(f"/api/analytics/classes/1?exam_id={exam_id}")
    assert class_one_response.status_code == 404

    student_response = client.get(f"/api/analytics/students/1?exam_id={exam_id}")
    assert student_response.status_code == 200
    student_payload = student_response.json()
    assert student_payload["class_rank"] == 1
    assert student_payload["grade_rank"] == 1

    audit_response = client.get(f"/api/exams/{exam_id}/score-rank-audit")
    assert audit_response.status_code == 200
    audit_payload = audit_response.json()
    assert audit_payload["mapping_rate"] == 1
    assert audit_payload["class_mappings"][0]["source_class_name"] == "2班"
    assert audit_payload["class_mappings"][0]["mapped_class_name"] == "2班"


def test_score_target_lines_feed_grade_decision_analytics(client) -> None:
    exam_id = _create_exam(client, name="2026届高一目标线分析测试")
    content = _wide_platform_workbook()
    preview_response = client.post(
        f"/api/exams/{exam_id}/scores/import/preview",
        files={"file": ("target-lines.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={
            "strategy": "overwrite",
            "rebuild": "true",
            "mapping_json": json.dumps(preview["mapping"]),
        },
        files={"file": ("target-lines.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_response.status_code == 200

    target_response = client.put(
        f"/api/exams/{exam_id}/score-target-lines",
        json=[
            {
                "name": "本科线",
                "line_type": "score",
                "score_value": 240,
                "rank_value": None,
                "near_margin_score": 15,
                "near_margin_rank": None,
                "sort_order": 1,
                "note": "",
                "is_active": True,
            },
            {
                "name": "前1名线",
                "line_type": "rank",
                "score_value": None,
                "rank_value": 1,
                "near_margin_score": None,
                "near_margin_rank": 1,
                "sort_order": 2,
                "note": "",
                "is_active": True,
            },
        ],
    )
    assert target_response.status_code == 200

    grade_response = client.get(f"/api/analytics/grades/1?exam_id={exam_id}")
    assert grade_response.status_code == 200
    grade_payload = grade_response.json()
    summaries = {item["line_name"]: item for item in grade_payload["target_line_summaries"]}
    assert summaries["本科线"]["reached_count"] == 1
    assert summaries["本科线"]["near_below_count"] == 1
    assert summaries["前1名线"]["reached_count"] == 1
    assert grade_payload["critical_students"]
    assert grade_payload["class_breakdown"][0]["target_line_counts"]["本科线"] == 1
    assert grade_payload["class_contributions"][0]["target_line_counts"]["本科线"] == 1


def test_long_csv_score_sheet_imports_after_preview(client) -> None:
    exam_id = _create_exam(client, name="2026届高一CSV智能导入测试")
    content = "\ufeff平台成绩明细\n学号,姓名,班级,科目,成绩\n2026001,张三,1班,语文,119\n2026001,张三,1班,数学,126\n".encode()

    preview_response = client.post(
        f"/api/exams/{exam_id}/scores/import/preview",
        files={"file": ("scores.csv", content, "text/csv")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert preview["layout_type"] == "long"
    assert preview["header_row"] == 2
    assert preview["mapping"]["field_mapping"]["score"] == "成绩"

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true", "mapping_json": json.dumps(preview["mapping"])},
        files={"file": ("scores.csv", content, "text/csv")},
    )
    assert import_response.status_code == 200
    assert import_response.json()["success_rows"] == 2


def test_smart_score_import_blocks_files_without_unique_student_number(client) -> None:
    exam_id = _create_exam(client, name="2026届高一缺学号智能导入测试")
    content = _missing_student_no_workbook()

    preview_response = client.post(
        f"/api/exams/{exam_id}/scores/import/preview",
        files={"file": ("missing-no.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert preview["import_ready"] is False
    assert any("唯一编号" in message for message in preview["messages"])

    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true", "mapping_json": json.dumps(preview["mapping"])},
        files={"file": ("missing-no.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_response.status_code == 400
    assert "唯一编号" in import_response.json()["detail"]
