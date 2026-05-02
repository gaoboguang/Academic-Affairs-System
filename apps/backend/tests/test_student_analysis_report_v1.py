from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import Exam, ScoreSubjectSnapshot, ScoreTotalSnapshot, Semester, Student


def _build_score_workbook(exam_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    for row in rows:
        sheet.append([exam_name, *row])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_question_workbook(exam_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "科目", "题号", "题目满分", "学生得分", "知识点", "题型", "能力层级", "错因标签", "错因备注", "备注"])
    for row in rows:
        normalized_row = list(row)
        if len(normalized_row) == 10:
            normalized_row = [*normalized_row[:9], "", "", normalized_row[9]]
        sheet.append([exam_name, *normalized_row])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _create_student(client, index: int) -> int:
    response = client.post(
        "/api/students",
        json={
            "student_no": f"20261{index:03d}",
            "name": f"测试生{index}",
            "gender": "男",
            "admission_year": 2024,
            "current_grade_id": 1,
            "current_class_id": 1 if index <= 6 else 2,
            "status": "active",
            "student_type": "general",
            "is_active": True,
        },
    )
    assert response.status_code == 200
    return response.json()["id"]


def _create_exam(client, name: str, exam_date: str, rows: list[list[object]]) -> int:
    exam_response = client.post(
        "/api/exams",
        json={
            "name": name,
            "exam_type": "阶段测试",
            "exam_date": exam_date,
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert exam_response.status_code == 200
    exam_id = exam_response.json()["id"]
    subject_response = client.post(
        f"/api/exams/{exam_id}/subjects",
        json=[
            {
                "subject_id": 1,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 1,
                "is_active": True,
            },
            {
                "subject_id": 2,
                "full_score": 150,
                "is_in_total": True,
                "excellent_line": 110,
                "pass_line": 90,
                "sort_order": 2,
                "is_active": True,
            },
        ],
    )
    assert subject_response.status_code == 200
    import_response = client.post(
        f"/api/exams/{exam_id}/scores/import",
        data={"strategy": "overwrite", "rebuild": "true"},
        files={
            "file": (
                f"{name}.xlsx",
                _build_score_workbook(name, rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    return exam_id


def _rows_for_exam(math_scores: dict[int, int], chinese_scores: dict[int, int]) -> list[list[object]]:
    rows: list[list[object]] = []
    for index in sorted(math_scores):
        student_no = f"20261{index:03d}"
        student_name = f"测试生{index}"
        class_name = "1班" if index <= 6 else "2班"
        rows.append([student_no, student_name, class_name, "语文", chinese_scores[index], "", ""])
        rows.append([student_no, student_name, class_name, "数学", math_scores[index], "", ""])
    return rows


def _seed_exam_total_snapshots(app, count: int) -> int:
    with app.state.db.session_scope() as session:
        semester = session.scalar(select(Semester).where(Semester.is_current.is_(True)))
        assert semester is not None
        exam = Exam(
            name="大样本学生名单测试",
            exam_type="阶段测试",
            exam_date=date(2026, 5, 2),
            semester_id=semester.id,
            grade_scope_json=[1],
            status="published",
        )
        session.add(exam)
        session.flush()

        students = [
            Student(
                student_no=f"ANALYZABLE{i:04d}",
                name=f"可分析生{i:03d}",
                current_grade_id=1,
                current_class_id=1,
                status="active",
            )
            for i in range(1, count + 1)
        ]
        session.add_all(students)
        session.flush()
        session.add_all(
            [
                ScoreTotalSnapshot(
                    exam_id=exam.id,
                    student_id=student.id,
                    total_score=float(700 - index),
                    class_rank=index,
                    grade_rank=index,
                    grade_percentile=round((count - index) / count, 4),
                )
                for index, student in enumerate(students, start=1)
            ]
        )
        return exam.id


def test_exam_analyzable_students_returns_all_total_snapshot_students(client, app) -> None:
    exam_id = _seed_exam_total_snapshots(app, 205)

    response = client.get(f"/api/analytics/exams/{exam_id}/students")
    assert response.status_code == 200
    payload = response.json()

    assert payload["exam_id"] == exam_id
    assert payload["total"] == 205
    assert len(payload["items"]) == 205
    assert payload["items"][0]["name"] == "可分析生001"
    assert payload["items"][0]["grade_rank"] == 1
    assert payload["items"][-1]["name"] == "可分析生205"


def test_exam_analyzable_students_falls_back_to_subject_snapshots(client, app) -> None:
    with app.state.db.session_scope() as session:
        semester = session.scalar(select(Semester).where(Semester.is_current.is_(True)))
        assert semester is not None
        exam = Exam(
            name="仅分科快照测试",
            exam_type="阶段测试",
            exam_date=date(2026, 5, 3),
            semester_id=semester.id,
            grade_scope_json=[1],
            status="published",
        )
        student = Student(
            student_no="SUBJECTONLY001",
            name="仅分科学生",
            current_grade_id=1,
            current_class_id=1,
            status="active",
        )
        session.add_all([exam, student])
        session.flush()
        session.add(
            ScoreSubjectSnapshot(
                exam_id=exam.id,
                student_id=student.id,
                subject_id=1,
                score=88,
                grade_rank=7,
            )
        )
        exam_id = exam.id

    response = client.get(f"/api/analytics/exams/{exam_id}/students")
    assert response.status_code == 200
    payload = response.json()

    assert payload["total"] == 1
    assert payload["items"][0]["name"] == "仅分科学生"
    assert payload["items"][0]["total_score"] is None


def test_exam_analyzable_students_returns_empty_for_exam_without_scores(client) -> None:
    response = client.post(
        "/api/exams",
        json={
            "name": "无成绩考试",
            "exam_type": "阶段测试",
            "exam_date": "2026-05-04",
            "semester_id": 2,
            "grade_scope_json": [1],
            "is_trend_enabled": True,
            "status": "published",
            "note": "",
            "is_active": True,
        },
    )
    assert response.status_code == 200
    exam_id = response.json()["id"]

    students_response = client.get(f"/api/analytics/exams/{exam_id}/students")
    assert students_response.status_code == 200
    assert students_response.json() == {"exam_id": exam_id, "total": 0, "items": []}


def test_student_knowledge_import_builds_learning_plan(client) -> None:
    student_ids = [_create_student(client, index) for index in range(1, 4)]
    exam_name = "知识点诊断测试"
    exam_id = _create_exam(
        client,
        exam_name,
        "2026-05-05",
        _rows_for_exam(
            {1: 60, 2: 120, 3: 125},
            {1: 120, 2: 100, 3: 95},
        ),
    )
    question_rows = [
        ["20261001", "测试生1", "数学", "12", 10, 2, "函数单调性", "选择题", "理解", ""],
        ["20261001", "测试生1", "数学", "18", 10, 4, "函数单调性、导数应用", "解答题", "综合应用", ""],
        ["20261001", "测试生1", "数学", "4", 4, 1, "集合", "选择题", "识记", ""],
        ["20261002", "测试生2", "数学", "12", 10, 8, "函数单调性", "选择题", "理解", ""],
        ["20261002", "测试生2", "数学", "18", 10, 9, "函数单调性、导数应用", "解答题", "综合应用", ""],
        ["20261003", "测试生3", "数学", "12", 10, 9, "函数单调性", "选择题", "理解", ""],
        ["20261003", "测试生3", "数学", "18", 10, 10, "函数单调性、导数应用", "解答题", "综合应用", ""],
    ]

    import_response = client.post(
        f"/api/exams/{exam_id}/score-questions/import",
        data={"strategy": "overwrite"},
        files={
            "file": (
                "questions.xlsx",
                _build_question_workbook(exam_name, question_rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["success_rows"] == 7
    assert import_payload["failed_rows"] == 0
    assert import_payload["snapshot_count"] >= 5

    response = client.get(f"/api/analytics/students/{student_ids[0]}?exam_id={exam_id}")
    assert response.status_code == 200
    payload = response.json()
    knowledge_points = payload["knowledge_points"]

    assert knowledge_points[0]["knowledge_point_name"] == "函数单调性"
    assert knowledge_points[0]["subject_name"] == "数学"
    assert knowledge_points[0]["score_rate"] < knowledge_points[0]["grade_average_rate"]
    assert knowledge_points[0]["lost_score"] > 0
    assert knowledge_points[0]["priority_score"] > 0
    assert knowledge_points[0]["diagnosis_label"] == "优先补弱"
    assert set(knowledge_points[0]["question_numbers"]) == {"12", "18"}
    assert any(item["title"] == "知识点清单" for item in payload["action_suggestions"])

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "student_knowledge_plan", "exam_id": exam_id, "student_id": student_ids[0]},
    )
    assert export_response.status_code == 200
    assert export_response.json()["report_name"] == "学生知识点学习清单"
    download_response = client.get(export_response.json()["download_url"])
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    assert {"学习清单", "知识点诊断", "下一步建议"} <= set(workbook.sheetnames)
    assert workbook["知识点诊断"].cell(row=2, column=2).value == "函数单调性"


def test_student_knowledge_trends_rank_stable_weakness_and_exports(client) -> None:
    student_ids = [_create_student(client, index) for index in range(1, 4)]
    exam_specs = [
        ("知识点趋势一", "2026-03-01", 2, 8, 9),
        ("知识点趋势二", "2026-04-01", 3, 8, 9),
        ("知识点趋势三", "2026-05-01", 4, 9, 10),
    ]
    current_exam_id = 0
    for exam_name, exam_date, target_score, second_score, third_score in exam_specs:
        current_exam_id = _create_exam(
            client,
            exam_name,
            exam_date,
            _rows_for_exam(
                {1: 70 + target_score, 2: 120, 3: 130},
                {1: 120, 2: 100, 3: 95},
            ),
        )
        rows = [
            ["20261001", "测试生1", "数学", "12", 10, target_score, "函数单调性", "选择题", "理解", ""],
            ["20261002", "测试生2", "数学", "12", 10, second_score, "函数单调性", "选择题", "理解", ""],
            ["20261003", "测试生3", "数学", "12", 10, third_score, "函数单调性", "选择题", "理解", ""],
        ]
        import_response = client.post(
            f"/api/exams/{current_exam_id}/score-questions/import",
            data={"strategy": "overwrite"},
            files={
                "file": (
                    f"{exam_name}.xlsx",
                    _build_question_workbook(exam_name, rows),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert import_response.status_code == 200

    response = client.get(f"/api/analytics/students/{student_ids[0]}?exam_id={current_exam_id}")
    assert response.status_code == 200
    payload = response.json()
    trends = payload["knowledge_trends"]

    assert trends[0]["knowledge_point_name"] == "函数单调性"
    assert trends[0]["trend_exam_count"] == 3
    assert trends[0]["weak_exam_count"] == 3
    assert trends[0]["trend_label"] == "持续薄弱"
    assert trends[0]["priority_score"] > 0
    assert [point["score_rate"] for point in trends[0]["points"]] == [0.2, 0.3, 0.4]
    assert any(item["category"] == "knowledge_trend_focus" for item in payload["action_suggestions"])

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "student_knowledge_plan", "exam_id": current_exam_id, "student_id": student_ids[0]},
    )
    assert export_response.status_code == 200
    download_response = client.get(export_response.json()["download_url"])
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    assert "连续趋势" in workbook.sheetnames
    trend_sheet = workbook["连续趋势"]
    trend_headers = {cell.value: index for index, cell in enumerate(trend_sheet[1], start=1)}
    assert trend_sheet.cell(row=2, column=trend_headers["知识点"]).value == "函数单调性"
    assert trend_sheet.cell(row=2, column=trend_headers["趋势标签"]).value == "持续薄弱"


def test_student_knowledge_trends_detect_improvement_and_fluctuation(client) -> None:
    student_ids = [_create_student(client, index) for index in range(1, 3)]
    exam_specs = [
        ("知识点趋势波动一", "2026-03-02", 2, 2),
        ("知识点趋势波动二", "2026-04-02", 5, 9),
        ("知识点趋势波动三", "2026-05-02", 8, 3),
    ]
    current_exam_id = 0
    for exam_name, exam_date, improving_score, fluctuating_score in exam_specs:
        current_exam_id = _create_exam(
            client,
            exam_name,
            exam_date,
            _rows_for_exam(
                {1: 90 + improving_score, 2: 90 + fluctuating_score},
                {1: 120, 2: 110},
            ),
        )
        rows = [
            ["20261001", "测试生1", "数学", "12", 10, improving_score, "导数应用", "解答题", "综合应用", ""],
            ["20261002", "测试生2", "数学", "12", 10, fluctuating_score, "导数应用", "解答题", "综合应用", ""],
        ]
        import_response = client.post(
            f"/api/exams/{current_exam_id}/score-questions/import",
            data={"strategy": "overwrite"},
            files={
                "file": (
                    f"{exam_name}.xlsx",
                    _build_question_workbook(exam_name, rows),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert import_response.status_code == 200

    improving_response = client.get(f"/api/analytics/students/{student_ids[0]}?exam_id={current_exam_id}")
    assert improving_response.status_code == 200
    improving_trend = improving_response.json()["knowledge_trends"][0]
    assert improving_trend["knowledge_point_name"] == "导数应用"
    assert improving_trend["trend_label"] == "正在改善"
    assert improving_trend["trend_delta"] == 0.6

    fluctuation_response = client.get(f"/api/analytics/students/{student_ids[1]}?exam_id={current_exam_id}")
    assert fluctuation_response.status_code == 200
    fluctuation_trend = fluctuation_response.json()["knowledge_trends"][0]
    assert fluctuation_trend["knowledge_point_name"] == "导数应用"
    assert fluctuation_trend["trend_label"] == "波动反复"


def test_student_knowledge_trends_require_multiple_question_detail_exams(client) -> None:
    student_id = _create_student(client, 1)
    exam_id = _create_exam(
        client,
        "单次题分趋势测试",
        "2026-05-08",
        _rows_for_exam({1: 90}, {1: 100}),
    )
    import_response = client.post(
        f"/api/exams/{exam_id}/score-questions/import",
        data={"strategy": "overwrite"},
        files={
            "file": (
                "single-question.xlsx",
                _build_question_workbook(
                    "单次题分趋势测试",
                    [["20261001", "测试生1", "数学", "12", 10, 4, "函数单调性", "选择题", "理解", ""]],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200

    response = client.get(f"/api/analytics/students/{student_id}?exam_id={exam_id}")
    assert response.status_code == 200
    assert response.json()["knowledge_trends"] == []


def test_student_knowledge_trends_window_uses_student_snapshot_exams(client) -> None:
    student_id = _create_student(client, 1)
    current_exam_id = 0
    snapshot_specs = {
        1: ("趋势窗口一", 3),
        6: ("趋势窗口六", 4),
    }
    for index in range(1, 7):
        exam_name, score = snapshot_specs.get(index, (f"趋势窗口{index}", None))
        current_exam_id = _create_exam(
            client,
            exam_name,
            f"2026-0{index}-01",
            _rows_for_exam({1: 80 + index}, {1: 100}),
        )
        if score is None:
            continue
        import_response = client.post(
            f"/api/exams/{current_exam_id}/score-questions/import",
            data={"strategy": "overwrite"},
            files={
                "file": (
                    f"{exam_name}.xlsx",
                    _build_question_workbook(
                        exam_name,
                        [["20261001", "测试生1", "数学", "12", 10, score, "函数单调性", "选择题", "理解", ""]],
                    ),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert import_response.status_code == 200

    response = client.get(f"/api/analytics/students/{student_id}?exam_id={current_exam_id}")
    assert response.status_code == 200
    trends = response.json()["knowledge_trends"]
    assert trends[0]["trend_label"] == "持续薄弱"
    assert trends[0]["trend_exam_count"] == 2
    assert [point["exam_name"] for point in trends[0]["points"]] == ["趋势窗口一", "趋势窗口六"]


def test_question_score_import_reports_duplicates(client) -> None:
    _create_student(client, 1)
    exam_name = "题分重复测试"
    exam_id = _create_exam(
        client,
        exam_name,
        "2026-05-06",
        _rows_for_exam({1: 90}, {1: 100}),
    )

    response = client.post(
        f"/api/exams/{exam_id}/score-questions/import",
        data={"strategy": "overwrite"},
        files={
            "file": (
                "duplicate-questions.xlsx",
                _build_question_workbook(
                    exam_name,
                    [
                        ["20261001", "测试生1", "数学", "12", 10, 6, "函数单调性", "选择题", "理解", ""],
                        ["20261001", "测试生1", "数学", "12", 10, 8, "函数单调性", "选择题", "理解", ""],
                    ],
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["success_rows"] == 1
    assert payload["failed_rows"] == 1
    assert "重复" in payload["error_preview"][0]


def test_student_knowledge_plan_requires_question_details(client) -> None:
    student_id = _create_student(client, 1)
    exam_id = _create_exam(
        client,
        "无题分知识点测试",
        "2026-05-07",
        _rows_for_exam({1: 90}, {1: 100}),
    )

    response = client.get(f"/api/analytics/students/{student_id}?exam_id={exam_id}")
    assert response.status_code == 200
    assert response.json()["knowledge_points"] == []

    export_response = client.post(
        "/api/reports/export",
        json={"report_type": "student_knowledge_plan", "exam_id": exam_id, "student_id": student_id},
    )
    assert export_response.status_code == 404
    assert export_response.json()["detail"] == "该学生暂无本次考试知识点题分明细"


def test_student_analysis_report_v1_metrics(client) -> None:
    student_ids = [_create_student(client, index) for index in range(1, 13)]
    target_student_id = student_ids[5]

    _create_exam(
        client,
        "高三诊断一",
        "2026-02-01",
        _rows_for_exam(
            {index: value for index, value in enumerate([105, 104, 103, 102, 101, 95, 96, 94, 93, 92, 91, 90], start=1)},
            {index: value for index, value in enumerate([90, 92, 94, 96, 98, 150, 88, 86, 84, 82, 80, 78], start=1)},
        ),
    )
    _create_exam(
        client,
        "高三诊断二",
        "2026-03-01",
        _rows_for_exam(
            {index: value for index, value in enumerate([106, 105, 104, 103, 102, 70, 97, 96, 95, 94, 93, 92], start=1)},
            {index: value for index, value in enumerate([91, 93, 95, 97, 99, 100, 89, 87, 85, 83, 81, 79], start=1)},
        ),
    )
    exam_id = _create_exam(
        client,
        "高三诊断三",
        "2026-04-01",
        _rows_for_exam(
            {index: value for index, value in enumerate([145, 144, 143, 142, 101, 80, 100, 99, 98, 97, 96, 95], start=1)},
            {index: value for index, value in enumerate([90, 92, 94, 96, 98, 150, 88, 86, 84, 82, 80, 78], start=1)},
        ),
    )

    line_response = client.put(
        f"/api/exams/{exam_id}/score-target-lines",
        json=[
            {
                "name": "本科线",
                "line_type": "score",
                "score_value": 230,
                "near_margin_score": 20,
                "sort_order": 1,
                "is_active": True,
            }
        ],
    )
    assert line_response.status_code == 200

    response = client.get(f"/api/analytics/students/{target_student_id}?exam_id={exam_id}")
    assert response.status_code == 200
    payload = response.json()

    assert payload["total_score"] == 230
    assert payload["grade_percentile"] > 0.6
    assert payload["grade_rank_delta"] > 0
    assert payload["overview_sentence"]
    assert payload["target_line_gaps"][0]["line_name"] == "本科线"
    assert payload["target_line_gaps"][0]["gap_score"] == 0
    assert len(payload["trend_points"]) == 3

    math_subject = next(item for item in payload["subjects"] if item["subject_name"] == "数学")
    chinese_subject = next(item for item in payload["subjects"] if item["subject_name"] == "语文")

    assert math_subject["rank_deviation"] < 0
    assert math_subject["diagnosis"] in {"重点补弱", "严重拖后腿"}
    assert math_subject["z_score"] is not None
    assert math_subject["t_score"] is not None
    assert math_subject["peer_sample_count"] == 12
    assert math_subject["peer_average_delta"] < 0
    assert math_subject["trend_exam_count"] == 3
    assert math_subject["trend_rank_stddev"] is not None
    assert math_subject["primary_effective_line_name"] == "本科线"
    assert math_subject["primary_effective_score_gap"] < 0

    assert chinese_subject["rank_deviation"] > 0
    assert chinese_subject["diagnosis"] == "优势学科"
    assert chinese_subject["primary_effective_score_gap"] > 0

    suggestion_titles = {item["title"] for item in payload["action_suggestions"]}
    assert {"保优", "补弱", "临界预警"} <= suggestion_titles
