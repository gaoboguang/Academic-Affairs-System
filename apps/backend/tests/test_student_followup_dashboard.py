from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from app.models import StudentGrowthRecord, StudentPlanningGoal, StudentPlanningTask
from test_exam_workflow import create_published_exam_with_scores


def _seed_growth_and_planning(app, *, due_date: date | None = None) -> None:
    with app.state.db.session_scope() as session:
        session.add(
            StudentGrowthRecord(
                student_id=1,
                occurred_on=date.today(),
                record_type="stage_review",
                title="阶段学习复盘",
                content="围绕近期成绩波动形成跟进记录。",
                owner_name="李语文",
            )
        )
        session.add(
            StudentPlanningGoal(
                student_id=1,
                target_year=2026,
                pathway_code="gaokao_regular",
                pathway_name="普通类高考",
                status="in_progress",
                priority="high",
            )
        )
        session.add(
            StudentPlanningTask(
                student_id=1,
                source_type="manual",
                task_type="score_review",
                title="复核阶段成绩波动",
                status="in_progress",
                priority="high",
                due_date=due_date or (date.today() - timedelta(days=1)),
                related_route="/analytics",
            )
        )


def test_student_followup_dashboard_and_reports(app, client) -> None:
    exam_id, _payload = create_published_exam_with_scores(client)
    _seed_growth_and_planning(app)

    dashboard_response = client.get(
        f"/api/analytics/adviser-dashboard?class_id=1&exam_id={exam_id}"
    )
    assert dashboard_response.status_code == 200
    dashboard = dashboard_response.json()
    assert dashboard["overview"]["student_count"] == 2
    assert dashboard["overview"]["score_sample_count"] == 2
    assert dashboard["overview"]["growth_record_count"] == 1
    assert dashboard["overview"]["open_task_count"] == 1
    assert dashboard["overview"]["overdue_task_count"] == 1
    assert "attendance_summary" not in dashboard
    assert "behavior_summary" not in dashboard

    risk_by_name = {item["student_name"]: item for item in dashboard["risk_students"]}
    assert risk_by_name["张三"]["risk_level"] == "urgent"
    assert "升学规划逾期任务" in " ".join(risk_by_name["张三"]["reasons"])

    risk_response = client.get(f"/api/analytics/student-risk/1?exam_id={exam_id}")
    assert risk_response.status_code == 200
    risk_payload = risk_response.json()
    assert risk_payload["planning_summary"]["overdue_task_count"] == 1
    assert risk_payload["growth_summary"]["record_count"] == 1
    assert "attendance_summary" not in risk_payload
    assert "behavior_summary" not in risk_payload

    weekly_report_response = client.post(
        "/api/reports/export",
        json={
            "report_type": "adviser_weekly_summary",
            "class_id": 1,
            "exam_id": exam_id,
        },
    )
    assert weekly_report_response.status_code == 200
    weekly_download = client.get(weekly_report_response.json()["download_url"])
    assert weekly_download.status_code == 200
    weekly_book = load_workbook(BytesIO(weekly_download.content))
    weekly_summary = weekly_book["班主任周报"]
    labels = [weekly_summary.cell(row=index, column=1).value for index in range(1, weekly_summary.max_row + 1)]
    assert "成长档案" in labels
    assert "规划任务" in labels
    assert "考勤概况" not in labels

    followup_report_response = client.post(
        "/api/reports/export",
        json={
            "report_type": "student_followup_package",
            "student_id": 1,
            "exam_id": exam_id,
        },
    )
    assert followup_report_response.status_code == 200
    followup_download = client.get(followup_report_response.json()["download_url"])
    assert followup_download.status_code == 200
    followup_book = load_workbook(BytesIO(followup_download.content))
    followup_summary = followup_book["学生跟进包"]
    followup_labels = [followup_summary.cell(row=index, column=1).value for index in range(1, followup_summary.max_row + 1)]
    assert "规划任务" in followup_labels
    assert "成长档案" in followup_labels
    assert "考勤摘要" not in followup_labels


def test_attendance_and_behavior_routes_are_removed(client) -> None:
    assert client.get("/api/attendance/records").status_code == 404
    assert client.get("/api/attendance/template").status_code == 404
    assert client.get("/api/behavior/records").status_code == 404
    assert client.get("/api/behavior/template").status_code == 404


def test_removed_import_templates_are_not_generated(test_settings) -> None:
    from app.exporters.templates import generate_import_templates

    generated = {path.name for path in generate_import_templates(test_settings)}
    assert "attendance_import_template.xlsx" not in generated
    assert "behavior_import_template.xlsx" not in generated


def test_class_profile_followup_summary_uses_score_growth_and_planning(app, client) -> None:
    exam_id, _payload = create_published_exam_with_scores(client)
    _seed_growth_and_planning(app)

    response = client.get(f"/api/classes/overview?grade_id=1&semester_id=2&exam_id={exam_id}")
    assert response.status_code == 200
    grade_group = next(item for item in response.json()["grades"] if item["grade_id"] == 1)
    class_one = next(item for item in grade_group["classes"] if item["class_id"] == 1)
    assert class_one["risk_summary"]["planning_risk_count"] == 1
    assert class_one["risk_summary"]["growth_record_count"] == 1
    assert "attendance_risk_count" not in class_one["risk_summary"]
    assert "behavior_risk_count" not in class_one["risk_summary"]
