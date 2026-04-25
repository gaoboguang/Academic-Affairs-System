from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.importers.students import StudentImporter
from app.models import Grade, SchoolClass, Student


def build_student_workbook(
    *,
    headers: list[str] | None = None,
    rows: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        headers
        or ["学号", "姓名", "性别", "出生日期", "入学年份", "年级", "班级", "学生状态", "学生类别", "艺体方向", "生源地省份", "联系电话", "家庭住址", "备注"]
    )
    for row in (
        rows
        or [
            ["2026101", "导入学生", "男", "2009-04-01", "2024", "高一", "10班", "在读", "普通生", "", "广东", "13800001111", "测试地址", ""],
            ["", "缺学号学生", "女", "", "2024", "高一", "10班", "在读", "普通生", "", "湖南", "", "", ""],
        ]
    ):
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_student_importer_creates_error_report(app, test_settings) -> None:
    with app.state.db.session_scope() as session:
        session.add(
            SchoolClass(
                grade_id=1,
                name="10班",
                class_type="normal",
                student_count=0,
            )
        )
        session.flush()

        importer = StudentImporter(session, test_settings)
        result = importer.execute(
            filename="students.xlsx",
            content=build_student_workbook(),
            strategy="skip_existing",
        )
        imported_student = session.scalar(select(Student).where(Student.student_no == "2026101"))
        imported_origin_province = imported_student.origin_province if imported_student else None

    assert imported_student is not None

    assert result.total_rows == 2
    assert result.status == "partially_failed"
    assert result.success_rows == 1
    assert result.failed_rows == 1
    assert result.created_rows == 1
    assert result.error_report_path is not None
    assert result.error_preview == ["第 3 行：学号不能为空"]
    assert imported_origin_province == "广东"
    error_workbook = load_workbook(test_settings.project_root / result.error_report_path)
    error_sheet = error_workbook["错误报告"]
    assert [error_sheet.cell(row=1, column=index).value for index in range(1, 7)] == [
        "行号",
        "列名",
        "字段名",
        "原始值",
        "错误原因",
        "建议修复",
    ]
    assert error_sheet.cell(row=2, column=2).value == "学号"
    assert error_sheet.cell(row=2, column=5).value == "学号不能为空"
    assert "补齐" in str(error_sheet.cell(row=2, column=6).value)


def test_student_importer_accepts_common_header_aliases(app, test_settings) -> None:
    with app.state.db.session_scope() as session:
        session.add(
            SchoolClass(
                grade_id=1,
                name="10班",
                class_type="normal",
                student_count=0,
            )
        )
        session.flush()

        importer = StudentImporter(session, test_settings)
        result = importer.execute(
            filename="students.xlsx",
            content=build_student_workbook(
                headers=[
                    "省学籍辅号",
                    "姓名",
                    "性别",
                    "出生日期",
                    "入学年份",
                    "年级",
                    "班级",
                    "学生状态",
                    "学生类别",
                    "艺体方向",
                    "生源地省份",
                    "联系电话",
                    "现住地址",
                    "备注",
                ],
                rows=[
                    ["2026201", "别名表头学生", "女", "2009-05-01", "2024", "高一", "10班", "在读", "普通生", "", "山东", "13800002222", "别名地址", ""],
                ],
            ),
            strategy="skip_existing",
        )
        imported_student = session.scalar(select(Student).where(Student.student_no == "2026201"))

    assert result.success_rows == 1
    assert result.failed_rows == 0
    assert result.status == "success"
    assert result.created_rows == 1
    assert result.error_preview == []
    assert result.notice_preview == []
    assert imported_student is not None
    assert imported_student.address == "别名地址"


def test_students_import_api_returns_template_detail(client) -> None:
    response = client.post(
        "/api/students/import",
        files={
            "file": (
                "students.xlsx",
                build_student_workbook(headers=["错误学号", "姓名", "性别", "出生日期", "入学年份", "年级", "班级", "学生状态", "学生类别", "艺体方向", "生源地省份", "联系电话", "家庭住址", "备注"]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"strategy": "skip_existing"},
    )

    assert response.status_code == 400
    assert "学生导入模板表头不匹配" in response.json()["detail"]
    assert "第 1 列应为“学号”，实际为“错误学号”" in response.json()["detail"]


def test_student_importer_auto_creates_missing_classes(app, test_settings) -> None:
    with app.state.db.session_scope() as session:
        importer = StudentImporter(session, test_settings)
        result = importer.execute(
            filename="students.xlsx",
            content=build_student_workbook(
                rows=[
                    ["2026301", "自动建班学生A", "男", "", "2024", "高三", "202301", "在读", "普通生", "", "山东", "", "地址A", ""],
                    ["2026302", "自动建班学生B", "女", "", "2024", "高三", "202301", "在读", "普通生", "", "山东", "", "地址B", ""],
                ],
            ),
            strategy="skip_existing",
        )
        created_class = session.scalar(
            select(SchoolClass).join(SchoolClass.grade).where(Grade.name == "高三", SchoolClass.name == "202301")
        )
        created_count = session.scalar(
            select(func.count()).select_from(SchoolClass).join(SchoolClass.grade).where(Grade.name == "高三", SchoolClass.name == "202301")
        )
        imported_students = session.scalars(
            select(Student).where(Student.student_no.in_(["2026301", "2026302"])).order_by(Student.student_no)
        ).all()

    assert result.success_rows == 2
    assert result.failed_rows == 0
    assert "已自动创建班级 1 个" in result.message
    assert result.notice_preview == ["已自动创建班级：高三 / 202301。"]
    assert created_class is not None
    assert created_class.class_type == "normal"
    assert created_count == 1
    assert [item.current_class_id for item in imported_students] == [created_class.id, created_class.id]


def test_student_importer_builds_notice_preview_for_common_import_issues(app, test_settings) -> None:
    with app.state.db.session_scope() as session:
        session.add(
            SchoolClass(
                grade_id=1,
                name="10班",
                class_type="normal",
                student_count=0,
            )
        )
        session.flush()

        importer = StudentImporter(session, test_settings)
        result = importer.execute(
            filename="students.xlsx",
            content=build_student_workbook(
                rows=[
                    ["2026401", "年级错误学生", "男", "", "2024", "高四", "", "在读", "普通生", "", "山东", "", "", ""],
                    ["2026402", "状态错误学生", "女", "", "2024", "高一", "10班", "借读中", "普通生", "", "山东", "", "", ""],
                    ["2026403", "类别错误学生", "女", "", "2024", "高一", "10班", "在读", "国际生", "", "山东", "", "", ""],
                ],
            ),
            strategy="skip_existing",
        )

    assert result.success_rows == 0
    assert result.failed_rows == 3
    assert result.notice_preview == [
        "未匹配年级：高四（1 行）。",
        "学生状态待核对：借读中（1 行）。",
        "学生类别待核对：国际生（1 行）。",
    ]
