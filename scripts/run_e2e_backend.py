from __future__ import annotations

import os
from pathlib import Path
import shutil
import sys
import tempfile
from io import BytesIO
from datetime import date

import uvicorn
from openpyxl import Workbook
from sqlalchemy import select

ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "apps" / "backend"

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.bootstrap import ensure_runtime_directories
from app.core.config import Settings
from app.core.security import hash_password
from app.db.session import DatabaseManager
from app.exporters.templates import generate_import_templates
from app.main import create_app
from app.models import AcademicYear, AppUser, Base, Exam, Semester, TeachingAssignment
from app.services.bootstrap import seed_demo_data, seed_reference_data
from app.services import exams as exams_service
from app.services import evaluation as evaluation_service
from app.services import workload as workload_service
from app.schemas.exam import ExamPayload, ExamSubjectPayload
from app.schemas.evaluation import AdviserQuantRecordPayload
from app.schemas.workload import WorkloadCalculatePayload


def build_settings() -> Settings:
    data_dir_env = os.environ.get("LOCAL_EDU_E2E_DATA_DIR")
    data_dir = Path(data_dir_env).expanduser() if data_dir_env else Path(tempfile.gettempdir()) / "local-edu-tool-e2e"
    port = int(os.environ.get("LOCAL_EDU_E2E_BACKEND_PORT", "8001"))
    return Settings(
        data_dir=data_dir,
        db_path=data_dir / "app.db",
        allowed_origins=[f"http://127.0.0.1:{port}", "http://127.0.0.1:4173", "http://127.0.0.1:5173"],
        debug=False,
        host="127.0.0.1",
        port=port,
    )


def build_timetable_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["学期", "星期", "节次", "教师", "班级", "学科", "课程类型", "周次规则", "备注"])
    sheet.append(["2025-2026 下学期", 1, 1, "李语文", "1班", "语文", "正课", "全周", ""])
    sheet.append(["2025-2026 下学期", 2, 2, "王数学", "1班", "数学", "正课", "全周", ""])
    sheet.append(["2025-2026 下学期", 3, 1, "李语文", "1班", "语文", "早读", "单周", ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_evaluation_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["模板名称", "教师", "班级", "题目", "分值", "评价对象类型"])
    rows = [
        ["通用课堂评教模板", "李语文", "1班", "教学目标清晰，重难点明确", 5, "学生"],
        ["通用课堂评教模板", "李语文", "1班", "课堂节奏合理，互动有效", 4, "学生"],
        ["通用课堂评教模板", "李语文", "1班", "作业讲评及时，反馈有针对性", 5, "学生"],
        ["通用课堂评教模板", "王数学", "2班", "教学目标清晰，重难点明确", 4, "学生"],
        ["通用课堂评教模板", "王数学", "2班", "课堂节奏合理，互动有效", 4, "学生"],
        ["通用课堂评教模板", "王数学", "2班", "作业讲评及时，反馈有针对性", 3, "学生"],
    ]
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_score_workbook(exam_name: str, chinese: int, math: int, second_chinese: int, second_math: int, third_chinese: int, third_math: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["考试名称", "学号", "姓名", "班级", "科目", "分数", "缺考标记", "备注"])
    sheet.append([exam_name, "2026001", "张三", "1班", "语文", str(chinese), "", ""])
    sheet.append([exam_name, "2026001", "张三", "1班", "数学", str(math), "", ""])
    sheet.append([exam_name, "2026002", "李四", "1班", "语文", str(second_chinese), "", ""])
    sheet.append([exam_name, "2026002", "李四", "1班", "数学", str(second_math), "", ""])
    sheet.append([exam_name, "2026003", "王五", "2班", "语文", str(third_chinese), "", ""])
    sheet.append([exam_name, "2026003", "王五", "2班", "数学", str(third_math), "", ""])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def seed_panorama_exams(session, settings: Settings) -> None:
    previous_year = session.scalar(
        select(AcademicYear).where(AcademicYear.name == "2024-2025")
    )
    if previous_year is None:
        previous_year = AcademicYear(
            name="2024-2025",
            start_date=date(2024, 9, 1),
            end_date=date(2025, 8, 31),
            is_current=False,
        )
        session.add(previous_year)
        session.flush()

    previous_semester = session.scalar(
        select(Semester).where(
            Semester.academic_year_id == previous_year.id,
            Semester.name == "下学期",
        )
    )
    if previous_semester is None:
        previous_semester = Semester(
            academic_year_id=previous_year.id,
            name="下学期",
            start_date=date(2025, 2, 15),
            end_date=date(2025, 7, 15),
            week_count=20,
            is_current=False,
        )
        session.add(previous_semester)
        session.flush()

    previous_assignment = session.scalar(
        select(TeachingAssignment).where(
            TeachingAssignment.teacher_id == 1,
            TeachingAssignment.semester_id == previous_semester.id,
            TeachingAssignment.class_id == 1,
            TeachingAssignment.subject_id == 1,
            TeachingAssignment.course_type == "regular",
        )
    )
    if previous_assignment is None:
        session.add(
            TeachingAssignment(
                teacher_id=1,
                semester_id=previous_semester.id,
                grade_id=1,
                class_id=1,
                subject_id=1,
                course_type="regular",
                weekly_periods_manual=5,
            )
        )

    previous_math_assignment = session.scalar(
        select(TeachingAssignment).where(
            TeachingAssignment.teacher_id == 2,
            TeachingAssignment.semester_id == previous_semester.id,
            TeachingAssignment.class_id == 1,
            TeachingAssignment.subject_id == 2,
            TeachingAssignment.course_type == "regular",
        )
    )
    if previous_math_assignment is None:
        session.add(
            TeachingAssignment(
                teacher_id=2,
                semester_id=previous_semester.id,
                grade_id=1,
                class_id=1,
                subject_id=2,
                course_type="regular",
                weekly_periods_manual=6,
            )
        )

    for exam_name, semester_id, exam_date, scores in [
        ("2025届高一4月月考", previous_semester.id, date(2025, 4, 10), (108, 115, 102, 110, 88, 92)),
        ("2026届高一4月月考", 2, date(2026, 4, 10), (118, 125, 110, 120, 98, 96)),
    ]:
        exam = session.scalar(select(Exam).where(Exam.name == exam_name))
        if exam is None:
            created = exams_service.create_exam(
                session,
                ExamPayload(
                    name=exam_name,
                    exam_type="月考",
                    exam_date=exam_date,
                    semester_id=semester_id,
                    grade_scope_json=[1],
                    is_trend_enabled=True,
                    status="published",
                    note=None,
                    is_active=True,
                ),
            )
            exam_id = created.id
            exams_service.replace_exam_subjects(
                session,
                exam_id,
                [
                    ExamSubjectPayload(
                        subject_id=1,
                        full_score=150,
                        is_in_total=True,
                        excellent_line=110,
                        pass_line=90,
                        sort_order=1,
                        is_active=True,
                    ),
                    ExamSubjectPayload(
                        subject_id=2,
                        full_score=150,
                        is_in_total=True,
                        excellent_line=110,
                        pass_line=90,
                        sort_order=2,
                        is_active=True,
                    ),
                ],
            )
            exams_service.import_scores(
                session,
                settings,
                exam_id=exam_id,
                filename=f"{exam_name}.xlsx",
                content=build_score_workbook(exam_name, *scores),
                strategy="overwrite",
                rebuild=True,
            )


def seed_extended_demo_data(settings: Settings, db_manager: DatabaseManager) -> None:
    with db_manager.session_scope() as session:
        seed_panorama_exams(session, settings)
        batch = workload_service.import_timetable(
            session,
            settings,
            semester_id=2,
            filename="e2e_timetable.xlsx",
            content=build_timetable_workbook(),
            remark="E2E workload seed",
        )
        rule_version = next(item for item in workload_service.list_rule_versions(session) if item.is_default)
        workload_service.calculate_workload(
            session,
            WorkloadCalculatePayload(
                semester_id=2,
                rule_version_id=rule_version.id,
                batch_id=batch.batch_id,
            ),
        )

        template = evaluation_service.list_evaluation_templates(session)[0]
        evaluation_service.import_evaluation_batch(
            session,
            settings,
            template_id=template.id,
            semester_id=2,
            filename="e2e_evaluation.xlsx",
            content=build_evaluation_workbook(),
        )

        adviser_rule_version = evaluation_service.list_adviser_rule_versions(session)[0]
        adviser_rule_item = next(
            item
            for item in evaluation_service.list_adviser_rule_items(session, adviser_rule_version.id)
            if not item.requires_attachment
        )
        evaluation_service.create_adviser_record(
            session,
            AdviserQuantRecordPayload(
                teacher_id=1,
                class_id=1,
                semester_id=2,
                rule_item_id=adviser_rule_item.id,
                record_month="2026-04",
                score=4.0,
                description="E2E 班主任量化样例",
                attachment_file_ids=[],
                is_active=True,
            ),
        )


def prepare_demo_data(settings: Settings) -> None:
    if settings.data_dir.exists():
        shutil.rmtree(settings.data_dir)

    ensure_runtime_directories(settings)
    generate_import_templates(settings)

    db_manager = DatabaseManager(settings.database_url, settings.debug)
    db_manager.initialize()
    Base.metadata.create_all(db_manager.engine)
    with db_manager.session_scope() as session:
        seed_reference_data(session)
        seed_demo_data(session)
        session.add(
            AppUser(
                username="admin",
                display_name="E2E 管理员",
                role="admin",
                password_hash=hash_password("AdminPass123!"),
                must_change_password=False,
            )
        )
    seed_extended_demo_data(settings, db_manager)
    db_manager.dispose()


def main() -> None:
    settings = build_settings()
    prepare_demo_data(settings)
    app = create_app(settings)
    uvicorn.run(app, host=settings.host, port=settings.port, log_level="warning")


if __name__ == "__main__":
    main()
